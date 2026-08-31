# -*- coding: utf-8 -*-
"""
================================================================================
 온라인팀 미니 ERP & 매출 분석 대시보드
================================================================================
회사 ERP '매출 로우데이터'(엑셀/CSV)를 DB에 누적 적재하고, 팀원이 브라우저에서
연차·아이템·시즌·브랜드별 전년비교(대표님 보고 프레임)를 조회하는 팀 전용 미니 ERP.

데이터 저장소는 두 가지를 자동 지원한다.
  · Streamlit secrets에 [postgres] 가 있으면  → Supabase(Postgres)  (배포용, 영속)
  · 없으면                                     → SQLite 파일          (로컬 개발용)

  [A] ETL      : 로우데이터 정제 + STCO 품번코드 해독
  [B] DATABASE : SQLAlchemy 누적적재 + 스키마 자동확장 + 중복방지 (Postgres/SQLite 공용)
  [C] ANALYSIS : 종합 대시보드 + 플래그십(연차·아이템별 전년비교)

공통룰(2026-07-31 추가 · 룰11): 모든 조회 표에는 엑셀 다운로드 버튼을 기본 제공한다.
  · 예외: 주간회의 보고자료 '메인 표'는 팀 주간보고 양식(weekly_template.xlsx)
    특별 템플릿 다운로드를 그대로 유지(변경 금지).

실행:  streamlit run app.py
================================================================================
"""

import io
import os
import gc
import hmac
import math
import hashlib
from datetime import datetime, timezone, timedelta
from urllib.parse import quote_plus

# 260811: 배포 서버(Streamlit Cloud)가 UTC로 도는 경우가 있어 datetime.now()를 화면 표시용으로 쓰면
# 한국 시간보다 9시간 늦게 나오는 문제가 있었음(예: 실제 14:xx인데 05:xx로 표시). 서버 로컬 설정과
# 무관하게 항상 정확한 한국 시간(KST, UTC+9 고정 — 서머타임 없음)을 돌려주는 헬퍼.
KST = timezone(timedelta(hours=9))


def now_kst():
    """화면에 표시하거나 파일명·날짜 기본값으로 쓸 '지금'은 이 함수로 구한다(서버 시간대 무관하게 KST)."""
    return datetime.now(timezone.utc).astimezone(KST)

import numpy as np
import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
import plotly.express as px
import plotly.graph_objects as go
from sqlalchemy import create_engine, text

DB_PATH = "sales_data.db"
TABLE = "sales"
ROW_KEY = "_row_key"

# ==============================================================================
# SECTION A. ETL  ─ 로우데이터 정제 + STCO 품번코드 해독
# ==============================================================================
BRAND_MAP = {
    "0": "ZERO LOUNGE", "9": "QUIKSILVER", "A": "CODI GALLERY", "C": "이월",
    "D": "DIEMS", "J": "GENTLEMENS PHILOSOPHY", "L": "GENDERLESS", "M": "맞춤",
    "N": "NORATED", "P": "PAUL&LOUIS", "R": "ROOM,ET", "S": "STCO",
    "T": "MARKET_TEST", "U": "UMEORA",
}
BRAND_CODE_MAP = dict(BRAND_MAP); BRAND_CODE_MAP["C"] = "이월"

ITEM_MAP = {
    "PA": ("바지-바지", "팬츠", "이너웨어"), "DM": ("바지-청바지", "팬츠", "이너웨어"),
    "GP": ("골프-바지", "팬츠", "이너웨어"), "WP": ("RENEW WORK 팬츠", "팬츠", "이너웨어"),
    "HP": ("바지-하프팬츠", "팬츠", "이너웨어"),
    "DS": ("셔츠-드레스셔츠", "셔츠류", "이너웨어"), "WD": ("RENEW WORK 셔츠", "셔츠류", "이너웨어"),
    "KT": ("스웨터-니트", "니트/티셔츠류", "이너웨어"), "TS": ("티셔츠-티셔츠", "니트/티셔츠류", "이너웨어"),
    "KG": ("스웨터-가디건", "니트/티셔츠류", "이너웨어"), "KV": ("스웨터-조끼", "니트/티셔츠류", "이너웨어"),
    "IT": ("이너웨어 티셔츠", "니트/티셔츠류", "이너웨어"), "GT": ("골프-티셔츠", "니트/티셔츠류", "이너웨어"),
    "GK": ("골프-스웨터니트", "니트/티셔츠류", "이너웨어"), "WI": ("RENEW WORK 스웨터", "니트/티셔츠류", "이너웨어"),
    "WS": ("RENEW WORK 티셔츠", "니트/티셔츠류", "이너웨어"),
    "CT": ("코트-코트", "아우터", "아우터"), "JP": ("점퍼-점퍼", "아우터", "아우터"),
    "JA": ("자켓-자켓", "아우터", "아우터"), "DJ": ("데님 점퍼", "아우터", "아우터"),
    "WO": ("RENEW WORK 코트", "아우터", "아우터"), "PV": ("베스트-패딩 베스트", "아우터", "아우터"),
    "WJ": ("RENEW WORK 점퍼", "아우터", "아우터"), "WK": ("RENEW WORK 자켓", "아우터", "아우터"),
    "GE": ("골프-패딩베스트", "아우터", "아우터"), "GJ": ("점퍼-기타", "아우터", "아우터"),
    "SJ": ("정장-수트상의", "수트류", "수트류"), "SL": ("정장-수트하의", "수트류", "수트류"),
    "SP": ("정장-단품정장", "수트류", "수트류"),
    "EJ": ("세트-셋업 자켓", "수트류", "수트류"), "EP": ("세트-셋업 팬츠", "수트류", "수트류"),
    "JV": ("베스트-우븐조끼", "수트류", "수트류"),
    "LJ": ("점퍼-가죽점퍼", "아우터", "아우터"), "TR": ("팬티", "이너웨어", "이너웨어"),
    "WA": ("지갑", "ACC", "액세서리"), "HA": ("모자", "ACC", "액세서리"),
    "FW": ("신발-신발", "신발", "슈즈"),
    "NT": ("넥타이", "ACC", "액세서리"), "BE": ("벨트-벨트", "ACC", "액세서리"),
    "BA": ("가방-가방", "ACC", "액세서리"), "MF": ("머플러", "ACC", "액세서리"),
    "SC": ("양말", "ACC", "액세서리"), "GL": ("장갑", "ACC", "액세서리"),
    "MU": ("머플러", "ACC", "액세서리"),
}
YEAR_MAP = {"O": 2017, "P": 2018, "Q": 2019, "R": 2020, "S": 2021, "T": 2022,
            "U": 2023, "V": 2024, "W": 2025, "X": 2026, "Y": 2027, "Z": 2028}
SEASON_MAP = {"A": "봄", "B": "여름", "C": "가을", "D": "겨울", "E": "RUNNING", "Z": "공통"}
SEASON_GROUP = {"봄": "S/S", "여름": "S/S", "가을": "F/W", "겨울": "F/W",
                "공통": "상시/ACC", "RUNNING": "상시/ACC"}

# 아이템 → 아이템그룹 (구분자 기준 + 팀 요청: ACC에서 신발·넥타이·벨트·양말 분리)
# 260803부터: 진짜 소스는 DB 아이템 마스터(item_master, "AI 마스터파일" 최종본 업로드분) — get_itemgroup_map()이
# item_master를 우선 사용하고, 마스터에 없는 코드만 아래 _ITEMGROUP_RAW(구 하드코딩)로 폴백한다.
# (item_master가 아직 비어있으면 전체가 이 폴백으로 동작 — 기존과 동일)
_ITEMGROUP_RAW = {
    "수트류": ["SJ", "SL", "EJ", "EP", "JV", "SP"],
    "셔츠류": ["DS", "WD"],
    "팬츠류": ["PA", "HP", "DM", "GP", "WP"],
    "아우터": ["DJ", "JA", "JP", "CT", "WJ", "GJ", "WO", "PV", "GE", "LJ"],
    "니트류": ["KT", "GK", "KG", "KV", "WK", "WI"],
    "티셔츠": ["TS", "GT", "WS", "IT"],
    "신발":   ["FW"], "넥타이": ["NT"], "벨트": ["BE"], "양말": ["SC"],
    "ACC":    ["BA", "WA", "HA", "MF", "GL", "MU", "TR"],
}
_ITEMGROUP_MAP_FALLBACK = {c: g for g, codes in _ITEMGROUP_RAW.items() for c in codes}
# 260803: 아이템 마스터 기준 니트류·티셔츠류가 "니트/티셔츠류" 한 중카테고리로 합쳐짐(모집단 통일 결정).
# 260804 수정(반품률 분석 탭 필터에 '셔츠류'가 안 보이는 버그 리포트로 발견): 아이템 마스터의 실제
# 중카테고리 값은 "셔츠류"·"팬츠류"(둘 다 '류'가 붙음, DS/WD·PA/DM/GP/WP/HP 코드) — 이 리스트가 예전엔
# '류' 없는 "셔츠"·"팬츠"로 적혀 있어서, 마스터 로딩 후 실제 데이터 값("셔츠류"/"팬츠류")과 안 맞아
# ITEMGROUP_ORDER 교집합 필터에서 통째로 빠지는 사고가 있었음 — 마스터 표기와 동일하게 통일.
ITEMGROUP_ORDER = ["수트류", "아우터", "셔츠류", "팬츠류", "니트/티셔츠류",
                   "신발", "넥타이", "벨트", "양말", "ACC", "기타"]

NUMERIC_COLS = ["매장수수료율", "할인율", "최초가", "현판가", "판매수량",
                "최초판매금액", "현판매금액", "실판매금액", "공급금액", "판가율",
                "원가(VAT+)", "판매원가(실판가)", "배수(실판가)", "계",
                "SKT", "상품권", "사용포인트", "마일리지상품권", "임의할인"]
REVENUE_CANDIDATES = ["실판매금액", "현판매금액", "최초판매금액"]
QTY_COL = "판매수량"


def year_age_label(sale_year, product_year):
    """연차: (기준=판매연도) − 상품년도. -1↓=내년신상, 0=신상, 1↑=N년차."""
    try:
        n = int(sale_year) - int(product_year)
    except (TypeError, ValueError):
        return None
    if n <= -1:
        return "내년신상"
    if n == 0:
        return "신상"
    return f"{n}년차"


def year_age_series(sale_year, product_year):
    """year_age_label의 벡터화 버전. 대용량에서 행별 호출 대신 사용(결과 동일).

    비교는 numpy float(NaN→False)로 수행하고, 유효하지 않은 행은 마지막에 None 처리.
    """
    sy = pd.to_numeric(sale_year, errors="coerce")
    py = pd.to_numeric(product_year, errors="coerce")
    n = sy - py
    nyoncha = (n.astype("Int64").astype(str) + "년차").to_numpy()   # "1년차" … (결측은 뒤에서 마스킹)
    nf = n.to_numpy(dtype="float64")
    lab = np.where(nf <= -1, "내년신상",
                   np.where(nf == 0, "신상", nyoncha))
    out = pd.Series(lab, index=sy.index, dtype="object")
    return out.where(sy.notna() & py.notna(), None)


def _make_columns_unique(cols):
    seen, out = {}, []
    for c in cols:
        c = str(c).strip()
        if c in seen:
            seen[c] += 1
            out.append(f"{c}_{seen[c]}")
        else:
            seen[c] = 1
            out.append(c)
    return out


def _find_header_row(raw):
    for i in range(min(10, len(raw))):
        vals = [str(v).strip() for v in raw.iloc[i].tolist()]
        if "판매일자" in vals and "품번" in vals:
            return i
    return 0


# ── 보안: 고객 개인정보 컬럼 차단 ─────────────────────────────────────
# 고객코드 등 '고객' 관련 컬럼은 개인정보라 DB 적재·외부반출 금지.
# 로우데이터를 읽는 즉시 제거하여, 실수로 포함돼 올라와도 저장소·화면에 절대 남지 않게 한다.
PII_KEYWORDS = ("고객",)


def _drop_pii_cols(df):
    """컬럼명에 PII 키워드('고객' 등)가 들어간 컬럼을 모두 제거."""
    pii = [c for c in df.columns if any(k in str(c) for k in PII_KEYWORDS)]
    if pii:
        df = df.drop(columns=pii)
    return df


def read_raw_file(uploaded_file):
    name = uploaded_file.name.lower()
    if name.endswith(".csv"):
        raw = pd.read_csv(uploaded_file, header=None, dtype=str, keep_default_na=False)
    else:
        raw = pd.read_excel(uploaded_file, header=None, dtype=str)
    hrow = _find_header_row(raw)
    header = _make_columns_unique(raw.iloc[hrow].tolist())
    df = raw.iloc[hrow + 1:].copy()
    df.columns = header
    df = df.reset_index(drop=True).dropna(how="all")
    df = df[~df.apply(lambda r: all((str(v).strip() == "" or str(v) == "nan") for v in r), axis=1)]
    # ERP 다운로드 맨 아래 '전체 합계' 행 제거 — 날짜·품번 없이 숫자만 있는 행은 실제 거래가 아님
    if "판매일자" in df.columns:
        dd = df["판매일자"].astype(str).str.strip().str.lower()
        df = df[dd.ne("") & ~dd.isin(["nan", "none", "nat"])]
    df = _drop_pii_cols(df)   # 보안: 고객코드 등 개인정보 컬럼은 읽는 즉시 제거
    return df


def _brand_name(col_val, code):
    if col_val and str(col_val).strip() in BRAND_MAP:
        return BRAND_MAP[str(col_val).strip()]
    if code:
        return BRAND_CODE_MAP.get(str(code)[0].upper())
    return None


def decode_stco(code, cols=None):
    code = str(code).strip().upper()
    cols = cols or {}
    res = {"브랜드명": None, "아이템명": None, "중카테고리": None, "대카테고리": None,
           "연도": None, "시즌명": None, "시즌그룹": None, "순번": None}
    res["브랜드명"] = _brand_name(cols.get("브랜드"), code)
    item_code = str(cols.get("아이템") or (code[1:3] if len(code) >= 3 else "")).strip().upper()
    item = ITEM_MAP.get(item_code)
    if item:
        res["아이템명"], res["중카테고리"], res["대카테고리"] = item
    year_code = str(cols.get("년도") or (code[3] if len(code) >= 4 else "")).strip().upper()
    res["연도"] = YEAR_MAP.get(year_code)
    season_code = str(cols.get("시즌") or (code[4] if len(code) >= 5 else "")).strip().upper()
    season = SEASON_MAP.get(season_code)
    res["시즌명"] = season
    res["시즌그룹"] = SEASON_GROUP.get(season)
    res["순번"] = str(cols.get("순번") or (code[5:7] if len(code) >= 7 else "")).strip()
    return res


def _col_or_code(df, colname, code, idx):
    """해당 컬럼값이 있으면 그것을, 비었거나 컬럼이 없으면 품번코드의 자리(idx)를 사용(벡터화)."""
    code_part = code.str[idx]
    if colname in df.columns:
        s = df[colname].astype(str).str.strip().str.upper()
        return s.where(s.ne("") & s.ne("NAN") & s.ne("NONE"), code_part)
    return code_part


def enrich(df):
    """로우데이터 정제 + STCO 품번코드 해독 (벡터화 · 대용량 메모리 최적화)."""
    for col in NUMERIC_COLS:
        if col in df.columns:
            s = df[col].astype(str).str.replace(",", "", regex=False).str.strip()
            df[col] = pd.to_numeric(s, errors="coerce")

    if "품번" in df.columns:
        code = df["품번"].astype(str).str.strip().str.upper()
        if "브랜드" in df.columns:
            bname = df["브랜드"].astype(str).str.strip().map(BRAND_MAP)
        else:
            bname = pd.Series(np.nan, index=df.index, dtype="object")
        df["브랜드명"] = bname.fillna(code.str[0].map(BRAND_CODE_MAP))
        ic = _col_or_code(df, "아이템", code, slice(1, 3))
        df["아이템명"] = ic.map({k: v[0] for k, v in ITEM_MAP.items()})
        df["중카테고리"] = ic.map({k: v[1] for k, v in ITEM_MAP.items()})
        df["대카테고리"] = ic.map({k: v[2] for k, v in ITEM_MAP.items()})
        df["연도"] = _col_or_code(df, "년도", code, 3).map(YEAR_MAP)
        season = _col_or_code(df, "시즌", code, 4).map(SEASON_MAP)
        df["시즌명"] = season
        df["시즌그룹"] = season.map(SEASON_GROUP)
        df["순번"] = _col_or_code(df, "순번", code, slice(5, 7))
        df["_아이템코드"] = ic

    if "판매일자" in df.columns:
        dt = pd.to_datetime(df["판매일자"], errors="coerce")
        df["_판매일"] = dt
        df["판매연도"] = dt.dt.year
        df["년월"] = dt.dt.strftime("%Y-%m")
        df["주차"] = dt.dt.strftime("%G-W%V")

    if "_아이템코드" in df.columns:
        item_code = df["_아이템코드"]
    elif "아이템" in df.columns:
        item_code = df["아이템"].astype(str).str.strip().str.upper()
    elif "품번" in df.columns:
        item_code = df["품번"].astype(str).str.strip().str.upper().str[1:3]
    else:
        item_code = None
    if item_code is not None:
        df["아이템그룹"] = item_code.map(get_itemgroup_map()).fillna("기타")

    if "판매연도" in df.columns and "연도" in df.columns:
        df["연차"] = year_age_series(df["판매연도"], df["연도"])

    rev = next((c for c in REVENUE_CANDIDATES if c in df.columns), None)
    df["_매출액"] = df[rev] if rev else 0
    df["_최초가매출"] = df["최초판매금액"] if "최초판매금액" in df.columns else 0
    df["_수량"] = pd.to_numeric(df["판매수량"], errors="coerce") if "판매수량" in df.columns else 0
    df["_채널"] = df["매장명"] if "매장명" in df.columns else df.get("매장코드", "기타")
    if "_아이템코드" in df.columns:
        df.drop(columns=["_아이템코드"], inplace=True)
    return df


def add_row_key(df):
    """중복 방지용 행 키(md5) 생성 — 벡터화로 문자열 결합 후 해시.

    결측은 "nan"으로 통일(기존 apply 방식의 str(nan)과 동일)하여, 이미 적재된
    DB의 키와 완전히 같은 값을 생성한다(누적/중복방지 호환).
    """
    key_cols = [c for c in ["판매일자", "매장코드", "판매번호", "판매연번", "품번"] if c in df.columns]
    if not key_cols:
        df[ROW_KEY] = [hashlib.md5(str(i).encode()).hexdigest() for i in range(len(df))]
        return df

    def _col_str(c):
        return df[c].astype("string").fillna("nan").astype(str)

    base = _col_str(key_cols[0])
    for c in key_cols[1:]:
        base = base.str.cat(_col_str(c), sep="|")
    df[ROW_KEY] = [hashlib.md5(s.encode("utf-8")).hexdigest() for s in base]
    return df


# ==============================================================================
# SECTION B. DATABASE  ─ SQLAlchemy (Postgres/SQLite 공용)
# ==============================================================================
@st.cache_resource
def get_engine():
    """secrets에 [postgres] 있으면 Supabase, 없으면 SQLite."""
    try:
        pg = st.secrets.get("postgres", None)
    except Exception:
        pg = None
    if pg:
        url = (f"postgresql+psycopg2://{pg['user']}:{quote_plus(str(pg['password']))}"
               f"@{pg['host']}:{pg.get('port',5432)}/{pg.get('dbname','postgres')}?sslmode=require")
        return create_engine(url, pool_pre_ping=True, pool_recycle=300)
    return create_engine(f"sqlite:///{DB_PATH}")


def backend_name():
    try:
        return "Supabase(Postgres)" if st.secrets.get("postgres", None) else "SQLite(로컬)"
    except Exception:
        return "SQLite(로컬)"


def _table_columns(conn):
    insp = conn.exec_driver_sql(
        "SELECT column_name FROM information_schema.columns WHERE table_name=%s"
        if conn.engine.dialect.name == "postgresql" else
        f'PRAGMA table_info("{TABLE}")',
        (TABLE,) if conn.engine.dialect.name == "postgresql" else ()
    )
    rows = insp.fetchall()
    if conn.engine.dialect.name == "postgresql":
        return [r[0] for r in rows]
    return [r[1] for r in rows]


def ensure_table(conn, df):
    cols = _table_columns(conn)
    q = '"'
    if not cols:
        defs = ", ".join([f'{q}{c}{q} TEXT' for c in df.columns if c != ROW_KEY])
        conn.exec_driver_sql(f'CREATE TABLE {q}{TABLE}{q} ({q}{ROW_KEY}{q} TEXT PRIMARY KEY, {defs})')
        return
    for c in df.columns:
        if c not in cols:
            conn.exec_driver_sql(f'ALTER TABLE {q}{TABLE}{q} ADD COLUMN {q}{c}{q} TEXT')


def append_to_db(df):
    """정제·키 생성된 df를 누적 적재. 중복(ROW_KEY)은 건너뜀. 파생(_) 컬럼 제외.

    메모리 최적화: 저장 대상 컬럼만 추린 뒤 파일 내 중복부터 제거하고,
    실제 신규 행에 대해서만 문자열화/적재를 수행한다.
    """
    save = [c for c in df.columns
            if (not c.startswith("_") or c == ROW_KEY)
            and not any(k in str(c) for k in PII_KEYWORDS)]   # 보안: 고객 개인정보 컬럼 적재 제외
    out = df[save].drop_duplicates(subset=[ROW_KEY])
    eng = get_engine()
    with eng.begin() as conn:
        ensure_table(conn, out)
        before = conn.exec_driver_sql(f'SELECT COUNT(*) FROM "{TABLE}"').scalar()
        existing = set(r[0] for r in conn.exec_driver_sql(f'SELECT "{ROW_KEY}" FROM "{TABLE}"').fetchall())
        new = out[~out[ROW_KEY].isin(existing)]
        n_new = len(new)
        if n_new:
            new = new.astype(object).where(new.notna(), None)  # 결측→None (신규 행에만)
            # DB별 파라미터 한도(Postgres 65535 / SQLite 32766) 안전하게: chunk×cols < 30000
            chunk = max(1, 30000 // max(1, len(new.columns)))
            new.to_sql(TABLE, conn, if_exists="append", index=False, method="multi", chunksize=chunk)
        after = before + n_new
    return {"inserted": n_new, "skipped": len(out) - n_new, "total_after": after}


def delete_dates(date_isos):
    """덮어쓰기 모드용: 지정한 날짜(['YYYY-MM-DD', ...])의 기존 행을 sales에서 삭제하고 삭제 건수 반환.

    저장된 '판매일자' 텍스트의 구분자(-, /, .)를 제거해 'YYYYMMDD'로 정규화한 뒤 매칭하므로,
    2026-07-27 / 2026/07/27 / 20260727 / 뒤에 시간이 붙은 형태까지 모두 같은 날로 잡아 삭제한다.
    """
    if not date_isos:
        return 0
    eng = get_engine()
    total = 0
    with eng.begin() as conn:
        exists = conn.exec_driver_sql(
            "SELECT 1 FROM information_schema.tables WHERE table_name=%s"
            if eng.dialect.name == "postgresql" else
            "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
            (TABLE,)).fetchone()
        if not exists:
            return 0
        for d in date_isos:
            compact = str(d).replace("-", "") + "%"   # 'YYYYMMDD%'
            r = conn.execute(
                text(f'DELETE FROM "{TABLE}" '
                     f'WHERE replace(replace(replace("판매일자", \'-\', \'\'), \'/\', \'\'), \'.\', \'\') LIKE :c'),
                {"c": compact})
            total += (r.rowcount or 0)
    return total


# 분석 화면이 실제로 쓰는 컬럼만 로드 (49만 행 × 60여 컬럼 전체 로드 시 메모리 초과 → OOM)
LOAD_COLS = ["판매일자", "브랜드명", "시즌명", "시즌그룹", "아이템", "아이템명", "품명",
             "연도", "판매연도", "년월", "최초판매금액", "실판매금액", "현판매금액",
             "판매수량", "매장명", "매장코드", "품번",
             "순번", "주문번호", "판매번호", "판매연번",
             "반품판매일자", "반품판매번호", "반품판매연번",
             "원가(VAT+)"]
# 원가(VAT+)(2026-08-07 재작업): 품번별 상세 드릴다운용. 지난번엔 이 컬럼과 "최초가"를 동시에
# 추가했다가 배포 직후 크래시가 나서 전량 롤백했음(claude/배포대기_현황.md 8-2 참고, 원인 미확정
# — OOM 추정이나 재현 확인은 못함). 이번엔 신규 컬럼을 원가(VAT+) 1개로 줄임 — "최초가"는 별도
# 컬럼 로드 없이 기존에 이미 로드 중인 _최초가매출÷_수량(가중평균)으로 대체 계산해서 메모리
# 증가폭을 최소화함(_pn_detail 참고). 원가(VAT+)는 실판매금액처럼 거래줄 단위 금액이라 기간
# 합산(SUM)이 맞음 — append_to_db가 원본 파일 컬럼을 그대로 저장해왔으므로(스키마 자동확장)
# 과거 업로드분도 DB에 이미 값이 있을 가능성이 높아 재업로드(백필) 없이 우선 확인해볼 것.
# 품명(2026-08-04 추가): 매출 로우데이터에 품번 바로 옆에 있는 실제 상품명 컬럼(중태님 확인) —
# DB(sales 테이블)에 이 컬럼이 없는 경우(과거 업로드분 등)에도 load_db()가 존재하는 컬럼만
# 골라 쓰므로(위 use = [c for c in LOAD_COLS if c in have]) 에러 없이 안전하게 빠짐.
# 순번·주문번호·반품판매일자/번호/연번(2026-08-04 추가): "SET/단품 판매 분석" 탭 전용 — 세트
# 그룹핑(주문번호+디자인키)과 반품↔원판매 매칭에 필요. 과거 업로드분에 이 컬럼이 없어도(구버전
# 원본 파일) load_db()가 존재하는 컬럼만 골라 쓰므로 에러 없이 안전하게 빠짐(그 경우 이 탭만 데이터 부족).
LOAD_NUM = ["최초판매금액", "실판매금액", "현판매금액", "판매수량", "판매연도", "연도",
            "원가(VAT+)"]
LOAD_CAT = ["브랜드명", "시즌명", "시즌그룹", "아이템", "아이템명", "년월", "매장명", "매장코드"]


def _existing_columns(conn, eng):
    if eng.dialect.name == "postgresql":
        rows = conn.exec_driver_sql(
            "SELECT column_name FROM information_schema.columns WHERE table_name=%s", (TABLE,)).fetchall()
    else:
        rows = conn.exec_driver_sql(f'PRAGMA table_info("{TABLE}")').fetchall()
        return [r[1] for r in rows]
    return [r[0] for r in rows]


@st.cache_data(ttl=21600)
def load_db():
    """필요한 컬럼만 청크 단위로 읽어 category/downcast로 적재 (대용량 메모리 최적화)."""
    eng = get_engine()
    try:
        with eng.connect() as conn:
            exists = conn.exec_driver_sql(
                "SELECT 1 FROM information_schema.tables WHERE table_name=%s"
                if eng.dialect.name == "postgresql" else
                "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
                (TABLE,)).fetchone()
            if not exists:
                return pd.DataFrame()
            have = _existing_columns(conn, eng)
            use = [c for c in LOAD_COLS if c in have]
            if not use:
                use = have
            q = "SELECT " + ", ".join(f'"{c}"' for c in use) + f' FROM "{TABLE}"'
            parts = []
            for ch in pd.read_sql(q, conn, chunksize=50000):
                for c in LOAD_NUM:
                    if c in ch.columns:
                        ch[c] = pd.to_numeric(
                            ch[c].astype(str).str.replace(",", "", regex=False),
                            errors="coerce", downcast="float")
                for c in LOAD_CAT:
                    if c in ch.columns:
                        ch[c] = ch[c].astype("category")
                parts.append(ch)
            if not parts:
                return pd.DataFrame()
            df = pd.concat(parts, ignore_index=True)
            del parts
            gc.collect()
    except Exception:
        return pd.DataFrame()

    # concat 후 category 재정리(청크별 카테고리 합집합)
    for c in LOAD_CAT:
        if c in df.columns and str(df[c].dtype) != "category":
            df[c] = df[c].astype("category")

    rev = next((c for c in REVENUE_CANDIDATES if c in df.columns), None)
    df["_매출액"] = df[rev] if rev else 0.0
    df["_최초가매출"] = df["최초판매금액"] if "최초판매금액" in df.columns else 0.0
    df["_수량"] = df[QTY_COL] if QTY_COL in df.columns else 0.0
    df["_채널"] = df["매장명"] if "매장명" in df.columns else df.get("매장코드", "기타")
    if "판매일자" in df.columns:
        df["_판매일"] = pd.to_datetime(df["판매일자"], errors="coerce")
        df = df[df["_판매일"].notna()].copy()   # 합계행 등 날짜 없는 행 제외 (대시보드 총액 정합성)
    # 비즈니스 규칙(아이템그룹·연차)은 저장값 대신 항상 최신 기준으로 재계산
    #  → 그룹 정의를 바꿔도 재적재 없이 즉시 반영됨
    if "아이템" in df.columns:
        df["아이템그룹"] = df["아이템"].astype(str).str.strip().str.upper().map(get_itemgroup_map()).fillna("기타")
    if "판매연도" in df.columns and "연도" in df.columns:
        df["연차"] = year_age_series(df["판매연도"], df["연도"])
    return df


def db_row_count():
    try:
        with get_engine().connect() as conn:
            return conn.exec_driver_sql(f'SELECT COUNT(*) FROM "{TABLE}"').scalar()
    except Exception:
        return 0


# ==============================================================================
# SECTION C. ANALYSIS UI
# ==============================================================================
def _won(n):
    try:
        return f"{int(round(float(n))):,} 원"
    except Exception:
        return "-"


def _mm(v):
    try:
        return float(v) / 1e6
    except Exception:
        return 0.0


AGE_RANK = {"내년신상": -1, "신상": 0}


def _age_sort_key(a):
    if a in AGE_RANK:
        return AGE_RANK[a]
    try:
        return int(str(a).replace("년차", ""))
    except Exception:
        return 99


# ---- 전년비교 성과표 (연차 / 아이템그룹 공용) ----
GROUPS = [("실판매금액(백만)", "실판매"), ("판가율", "판가율"), ("비중", "비중"), ("평균단가(원)", "평균단가")]

# 시즌 7행 (2026-07-31 중태님 지시 · 연차별 성과표 전용): G.TOTAL 바로 아래.
#  1~2행 = 시즌그룹 합계(S/S=공통+봄+여름 · F/W=가을+겨울), 3~7행 = 개별 시즌 Z→A→B→C→D.
#  E(RUNNING)는 별도 행 없이 G.TOTAL에만 포함. 비중 분모는 전체(G.TOTAL).
SEASON_ROW_DEFS = [
    ("S/S TOTAL", ["공통", "봄", "여름"]),
    ("F/W TOTAL", ["가을", "겨울"]),
    ("Z (공통)", ["공통"]),
    ("A (봄)", ["봄"]),
    ("B (여름)", ["여름"]),
    ("C (가을)", ["가을"]),
    ("D (겨울)", ["겨울"]),
]


def yoy_frame(cur, prev, dim, order_list=None, season_rows=False, cy=None, extra_rows=None):
    """올해(cur)·전년(prev)을 dim으로 묶어 전년비교 numeric DataFrame(멀티헤더) 반환. G.TOTAL 상단.

    cy=기준연도(예: 2025)를 넘기면 컬럼 라벨이 그 연도 기준으로 동적 표기된다
    (예: cy=2025 → "24년"/"25년"). 안 넘기면 과거 하드코딩 기본값("25년"/"26년") 유지.
    (2026-08-07 버그수정: 예전엔 실제 선택연도와 무관하게 "25년"/"26년"이 고정으로
    찍혀서, 기준연도를 2025로 조회해도 표 헤더는 항상 26년으로 보이는 문제가 있었음.)

    extra_rows=[(라벨, mask_fn), ...] (2026-08-07 추가)이면 **G.TOTAL 바로 아래**에 임의 그룹 소계
    행을 끼워 넣는다. mask_fn(f)는 cur/prev 각각에 적용되는 불리언 마스크 —
    예: 유통채널별 표의 '담당자별 TOTAL' 행, 시즌별/연차별 표의 '브랜드별 TOTAL' 행.

    ⚠️ 260818 순서 변경: 예전엔 extra_rows가 **시즌 7행 다음**이었는데, 시즌별/연차별 한눈에 보기에
    브랜드별 TOTAL을 넣으면서 중태님 지시대로 `G.TOTAL → 브랜드 5행 → 시즌 7행 → 연차행` 순이
    되도록 **extra_rows를 시즌 7행보다 앞으로** 옮겼다. 기존 사용처(유통채널별 표)는 season_rows를
    쓰지 않아서 이 순서 변경의 영향을 전혀 받지 않는다(행 구성 동일).
    """
    cur_lbl = f"{cy % 100:02d}년" if cy is not None else "26년"
    prev_lbl = f"{(cy - 1) % 100:02d}년" if cy is not None else "25년"

    def agg(f):
        if f is None or f.empty:
            return pd.DataFrame(columns=[dim, "rev", "orig", "qty"]).set_index(dim)
        return f.groupby(dim).agg(rev=("_매출액", "sum"), orig=("_최초가매출", "sum"),
                                  qty=("_수량", "sum"))
    c, p = agg(cur), agg(prev)
    keys = list(dict.fromkeys(list(c.index) + list(p.index)))
    if order_list:
        keys = [k for k in order_list if k in keys] + [k for k in keys if k not in order_list]
    else:
        keys = sorted(keys, key=lambda k: -float(c["rev"].get(k, 0)))
    tot_c, tot_p = float(c["rev"].sum()), float(p["rev"].sum())

    def metrics(r26, r25, o26, o25, q26, q25, share_den_c, share_den_p):
        return {
            ("실판매금액(백만)", prev_lbl): r25 / 1e6, ("실판매금액(백만)", cur_lbl): r26 / 1e6,
            ("실판매금액(백만)", "증감율"): ((r26 - r25) / r25) if r25 else None,
            ("판가율", prev_lbl): (r25 / o25) if o25 else 0, ("판가율", cur_lbl): (r26 / o26) if o26 else 0,
            ("판가율", "증감"): ((r26 / o26 if o26 else 0) - (r25 / o25 if o25 else 0)),
            ("비중", prev_lbl): (r25 / share_den_p) if share_den_p else 0,
            ("비중", cur_lbl): (r26 / share_den_c) if share_den_c else 0,
            ("비중", "증감"): ((r26 / share_den_c if share_den_c else 0) - (r25 / share_den_p if share_den_p else 0)),
            ("평균단가(원)", prev_lbl): (r25 / q25) if q25 else 0, ("평균단가(원)", cur_lbl): (r26 / q26) if q26 else 0,
            ("평균단가(원)", "증감"): ((r26 / q26 if q26 else 0) - (r25 / q25 if q25 else 0)),
        }

    rows, index = [], []
    # G.TOTAL 먼저
    rows.append(metrics(tot_c, tot_p, float(c["orig"].sum()), float(p["orig"].sum()),
                        float(c["qty"].sum()), float(p["qty"].sum()), tot_c, tot_p))
    index.append("G.TOTAL")
    # 임의 그룹 소계 (담당자별·브랜드별 TOTAL 등) — G.TOTAL 바로 아래(시즌 7행보다 위, 260818 순서 변경)
    if extra_rows:
        def _esum(f, mask_fn):
            if f is None or f.empty:
                return 0.0, 0.0, 0.0
            sub = f[mask_fn(f)]
            if sub.empty:
                return 0.0, 0.0, 0.0
            return (float(sub["_매출액"].sum()), float(sub["_최초가매출"].sum()),
                    float(sub["_수량"].sum()))
        for lbl, mask_fn in extra_rows:
            r26, o26, q26 = _esum(cur, mask_fn)
            r25, o25, q25 = _esum(prev, mask_fn)
            rows.append(metrics(r26, r25, o26, o25, q26, q25, tot_c, tot_p))
            index.append(lbl)
    # 시즌 7행 (시즌별/연차별 한눈에 보기 전용) — 그 다음, 연차 행들 앞
    if season_rows:
        def _ssum(f, sns):
            if f is None or f.empty or "시즌명" not in f.columns:
                return 0.0, 0.0, 0.0
            sub = f[f["시즌명"].astype(str).isin(sns)]
            return (float(sub["_매출액"].sum()), float(sub["_최초가매출"].sum()),
                    float(sub["_수량"].sum()))
        for lbl, sns in SEASON_ROW_DEFS:
            r26, o26, q26 = _ssum(cur, sns)
            r25, o25, q25 = _ssum(prev, sns)
            rows.append(metrics(r26, r25, o26, o25, q26, q25, tot_c, tot_p))
            index.append(lbl)
    for k in keys:
        rows.append(metrics(float(c["rev"].get(k, 0)), float(p["rev"].get(k, 0)),
                            float(c["orig"].get(k, 0)), float(p["orig"].get(k, 0)),
                            float(c["qty"].get(k, 0)), float(p["qty"].get(k, 0)), tot_c, tot_p))
        index.append(k)
    D = pd.DataFrame(rows, index=index)
    D.columns = pd.MultiIndex.from_tuples(D.columns)
    D.index.name = dim
    return D


def yoy_frame2(cur_m, prev_m, cur_y, prev_y, dim, order_list=None, season_rows=False,
               blk_labels=("당월누계", "연간누계"), cy=None, extra_rows=None):
    """플래그십 2블록 프레임 (2026-07-31 목업 v2 컨펌): 당월누계 + 연간누계.

    현재 헤더 12개 컬럼(실판매금액·판가율·비중·평균단가 × 전년/올해/증감)을 기간별로 복제해
    최상단에 기간 블록(당월누계·연간누계)을 얹는다. 비중은 각 블록 안에서 행÷전체.
    행 순서는 연간누계 기준(당월에만 있는 행은 뒤에 추가, 없는 칸은 '–').
    cy=기준연도를 넘기면 하위 표의 연도 컬럼 라벨("25년"/"26년" 등)이 그 연도 기준으로
    동적 계산된다(2026-08-07 버그수정).
    extra_rows는 두 블록에 동일하게 적용된다(2026-08-07 추가 — 담당자별 TOTAL 등).
    """
    Dm = yoy_frame(cur_m, prev_m, dim, order_list, season_rows=season_rows, cy=cy, extra_rows=extra_rows)
    Dy = yoy_frame(cur_y, prev_y, dim, order_list, season_rows=season_rows, cy=cy, extra_rows=extra_rows)
    idx = list(Dy.index) + [k for k in Dm.index if k not in Dy.index]
    Dm = Dm.reindex(idx)
    Dy = Dy.reindex(idx)
    D = pd.concat([Dm, Dy], axis=1, keys=list(blk_labels))
    D.index.name = dim
    return D


def _fmt_cell(col, v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return "–"
    top, sub = col[-2], col[-1]   # 2단(그룹·항목)/3단(기간·그룹·항목) 컬럼 모두 지원
    if top == "실판매금액(백만)":
        return f"{v:,.1f}" if sub != "증감율" else f"{v*100:+.0f}%"
    if top == "판가율":
        return f"{v*100:.0f}%" if sub != "증감" else f"{v*100:+.1f}%p"
    if top == "비중":
        return f"{v*100:.1f}%" if sub != "증감" else f"{v*100:+.1f}%p"
    if top == "평균단가(원)":
        return f"{v:,.0f}" if sub != "증감" else f"{v:+,.0f}"
    return v


def style_yoy(D):
    disp = D.copy()
    for col in disp.columns:
        disp[col] = [_fmt_cell(col, v) for v in disp[col]]
    delta_cols = [c for c in D.columns if c[-1] in ("증감율", "증감")]

    def color(col):
        vals = D[col]
        return ["color:#c62828;font-weight:600" if (pd.notnull(v) and v < 0)
                else ("color:#1f8a4c;font-weight:600" if pd.notnull(v) and v > 0 else "")
                for v in vals]
    sty = disp.style
    for col in delta_cols:
        sty = sty.apply(lambda s, c=col: color(c), subset=pd.IndexSlice[:, [col]])
    sty = sty.set_properties(**{"text-align": "right"})
    return sty


# ── 룰13 (2026-07-31): 엑셀 다운로드 = 화면에 보이는 컬러·셀서식 그대로 ──────
_XL_SEASON_BOLD = {"S/S TOTAL", "F/W TOTAL"}
_XL_SEASON_SUB = {"Z (공통)", "A (봄)", "B (여름)", "C (가을)", "D (겨울)"}
_XL_DELTA_SUBS = ("증감율", "증감", "편차", "전체 기울기", "최근 기울기")   # 뒤 2개: 채널별 추세 분석(260831)


def styled_excel_bytes(disp, sheet="표", first_block_cols=None, extra_row_labels=None,
                       extra_row_fill="D6F0FA", first_row_total=True):
    """표시용(포맷 문자열) DataFrame을 화면 서식 그대로 엑셀로 변환 (룰13).

    화면과 동일: 헤더 회색+볼드, 구분(인덱스) 연회색, 첫 행 노란 강조(G.TOTAL),
    시즌 TOTAL 블루그레이 / 개별 시즌 연블루, 증감·편차 +초록/-빨강, 숫자 우측정렬,
    전셀 얇은 테두리. first_block_cols=첫 기간블록 컬럼 수 → 경계 두꺼운 세로선(룰12).
    extra_row_labels(2026-08-07 추가)=[라벨, ...]이면 그 행들을 하늘색으로 채움
    (예: 유통채널별 표의 '담당자별 TOTAL' 행 — 화면(perf_table)과 동일 색).
    first_row_total(260831 추가)=False면 첫 행 노란 강조(룰6)를 끈다 — 첫 행이 G.TOTAL/합계가
    아닌 표(예: 채널별 추세 분석 — 첫 행이 그냥 1위 매장)용. 기존 호출부는 전부 기본값 유지.
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        disp.to_excel(w, sheet_name=_safe_name(sheet)[:28] or "표")
        ws = w.book.worksheets[0]
        n_idx = disp.index.nlevels
        n_rows = len(disp)
        data_start = ws.max_row - n_rows + 1          # 헤더/빈줄 이후 첫 데이터 행
        thin = Side(style="thin", color="D9D9D9")
        thick = Side(style="medium", color="555555")
        head_fill = PatternFill("solid", fgColor="F4F4F6")
        idx_fill = PatternFill("solid", fgColor="FAFAFA")
        gt_fill = PatternFill("solid", fgColor="FFF2B8")
        sg_fill = PatternFill("solid", fgColor="E3ECF7")
        ss_fill = PatternFill("solid", fgColor="F4F8FC")
        # extra_rows 강조색 — 담당자별 TOTAL은 하늘색, 브랜드별 TOTAL은 분홍(260818, 화면과 동일)
        xr_fill = PatternFill("solid", fgColor=extra_row_fill)
        _extra_set = set(extra_row_labels or [])
        bcol = (n_idx + first_block_cols + 1) if first_block_cols else None
        subs = [c[-1] if isinstance(c, tuple) else str(c) for c in disp.columns]

        # 1) 헤더 영역 (병합 셀 포함 전체)
        for r in range(1, data_start):
            for k in range(1, ws.max_column + 1):
                cell = ws.cell(r, k)
                cell.fill = head_fill
                cell.font = Font(bold=True, color="111111")
                cell.alignment = Alignment(horizontal="center", vertical="center")
        # 2) 데이터 영역 — 행 성격(첫행 노랑·시즌 블루·extra 하늘색)과 증감 색
        for ri in range(n_rows):
            r = data_start + ri
            ilab = disp.index[ri]
            labs = [str(x) for x in (ilab if isinstance(ilab, tuple) else (ilab,))]
            if ri == 0 and first_row_total:            # 룰6: 첫 행(G.TOTAL/합계) 노란 강조
                fill, bold = gt_fill, True
            elif any(x in _XL_SEASON_BOLD for x in labs):
                fill, bold = sg_fill, True
            elif any(x in _XL_SEASON_SUB for x in labs):
                fill, bold = ss_fill, False
            elif any(x in _extra_set for x in labs):
                fill, bold = xr_fill, False
            else:
                fill, bold = None, False
            for k in range(1, n_idx + 1):              # 구분(인덱스) 셀
                c = ws.cell(r, k)
                c.fill = fill or idx_fill
                c.font = Font(bold=True, color="111111")
                c.alignment = Alignment(horizontal="left", vertical="center")
            for cj in range(len(disp.columns)):        # 데이터 셀
                c = ws.cell(r, n_idx + 1 + cj)
                c.alignment = Alignment(horizontal="right", vertical="center")
                if fill:
                    c.fill = fill
                v = disp.iat[ri, cj]
                if subs[cj] in _XL_DELTA_SUBS and isinstance(v, str) and v[:1] in "+-":
                    c.font = Font(bold=True, color="C62828" if v.startswith("-") else "1F8A4C")
                elif bold:
                    c.font = Font(bold=True, color="111111")
        # 3) 테두리 (경계 컬럼은 왼쪽 두꺼운 선 — 룰12) + 컬럼 폭
        for r in range(1, ws.max_row + 1):
            for k in range(1, ws.max_column + 1):
                ws.cell(r, k).border = Border(
                    left=(thick if (bcol and k == bcol) else thin),
                    right=thin, top=thin, bottom=thin)
        for k in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(k)].width = 14 if k <= n_idx else 10.5
    return buf.getvalue()


def yoy_excel_bytes(D, sheet="분석", first_block_cols=None, extra_row_labels=None,
                    extra_row_fill="D6F0FA"):
    disp = D.copy()
    for col in disp.columns:
        disp[col] = [_fmt_cell(col, v) for v in disp[col]]
    return styled_excel_bytes(disp, sheet, first_block_cols, extra_row_labels,
                              extra_row_fill=extra_row_fill)   # 룰13: 화면 서식 그대로


# ── 공통(룰11 · 2026-07-31): 모든 조회 표 엑셀 다운로드 기본 제공 ──────────────
# 주간회의 보고자료 '메인 표'만 예외(팀 주간보고 양식 템플릿 다운로드 유지 — weekly_excel_bytes).
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _safe_name(s):
    """엑셀 시트명·파일명에 못 쓰는 문자를 '-'로 치환."""
    s = str(s)
    for ch in '\\/:*?"<>|[]':
        s = s.replace(ch, "-")
    return s.strip()


def table_excel_bytes(disp, sheet="표", first_block_cols=None):
    """화면 표시용(포맷된) DataFrame을 엑셀로 변환 — 일반 표 공용 다운로드(룰11).
    룰13: 화면 서식(색·볼드·정렬·테두리·블록 경계선) 그대로 반영."""
    return styled_excel_bytes(disp, sheet, first_block_cols)


def _money_note():
    """룰1: 표 오른쪽 상단 [금액: 백만원 / VAT+] 표기. (레거시 — 현재는 _NOTE_FLOAT를 제목 줄에 붙임)"""
    st.markdown(
        "<div style='text-align:right;color:#888;font-size:0.78rem;margin:8px 0 3px 0;'>"
        "[금액: 백만원 / VAT+]</div>", unsafe_allow_html=True)


# 룰1 표기를 '표 제목 줄 오른쪽'에 붙여, 제목과 표 사이에 별도 줄이 안 생기게 함(여백 축소).
_NOTE_FLOAT = ("<span style='float:right;color:#888;font-weight:400;font-size:0.78rem;"
               "white-space:nowrap;'>[금액: 백만원 / VAT+]</span>")


# 공통 표 CSS: 옵션A 여백(3px 9px) + 헤더·구분 검정 + G.TOTAL(첫 행) 노란 강조 + 증감 색 유지
_TBL_CSS = """
<style>
.erp-wrap{overflow-x:auto;margin:0 0 10px;background:#fff;
    border:1px solid #e8e8ed;border-radius:12px;}
table.erp-tbl{border-collapse:collapse;font-size:0.82rem;
    font-variant-numeric:tabular-nums;font-feature-settings:"tnum";}
table.erp-tbl th, table.erp-tbl td{padding:4px 9px;border:1px solid #f0f0f3;white-space:nowrap;}
table.erp-tbl thead th{color:#1d1d1f;font-weight:600;background:#f5f5f7;text-align:center;
    border-bottom:1px solid #e8e8ed;}
table.erp-tbl tbody th{color:#1d1d1f;font-weight:600;text-align:left;background:#fbfbfd;}
table.erp-tbl td{color:#1d1d1f;text-align:right;}
table.erp-tbl tbody tr:first-child th, table.erp-tbl tbody tr:first-child td{
    background:#fff4cc !important;font-weight:700;}
</style>
"""


# 시즌 7행 강조 CSS (연차별 성과표 전용 — erp-season 클래스가 붙은 표에만 적용)
#  2~3행(S/S·F/W TOTAL)=진한 블루그레이+볼드 · 4~8행(개별 시즌)=연한 톤+들여쓰기
def _season_css_class(key):
    """표마다 고유한 CSS 클래스명 — 같은 화면에 시즌표가 2개 이상 있을 때 서로 간섭하지 않게."""
    safe = "".join(ch if (ch.isalnum() or ch in "-_") else "-" for ch in str(key))
    return f"erp-season-{safe}"


def _extra_row_css(n, cls, color):
    """extra_rows(담당자별·브랜드별 TOTAL) n행을 색칠하는 CSS — **맨 왼쪽 라벨 칸(th)까지** 칠한다.

    extra_rows는 항상 G.TOTAL 바로 아래에 붙으므로 위치는 2 ~ (1+n)으로 고정이다.
    Styler.set_properties로는 값 칸(td)만 칠해지고 인덱스 th가 빠지던 문제(260818 실측)와,
    시즌행 CSS의 !important에 밀려 색이 안 보이던 문제를 한 번에 해결한다.
    """
    if not n:
        return ""
    return f"""
<style>
table.{cls} tbody tr:nth-child(n+2):nth-child(-n+{1 + n}) th,
table.{cls} tbody tr:nth-child(n+2):nth-child(-n+{1 + n}) td{{
    background:{color} !important;}}
</style>
"""


def _season_row_css(offset=0, cls="erp-season"):
    """시즌 7행 강조 CSS. offset = 시즌행 **앞에 끼워 넣은 추가 행 수**(예: 브랜드별 TOTAL 5행).

    260818: 시즌별/연차별 한눈에 보기에 브랜드별 TOTAL 5행이 G.TOTAL 바로 아래로 들어가면서
    시즌 7행의 위치(nth-child)가 그만큼 밀렸다. 위치를 고정값으로 두면 엉뚱한 행에 색이
    칠해지므로 offset을 받아 계산한다(offset=0이면 기존과 완전히 동일).

    ⚠️ cls(표별 고유 클래스)를 반드시 넘길 것 — CSS는 클래스 단위라, 예전처럼 공통 `erp-season`에
    걸면 **같은 화면의 다른 시즌표 CSS가 서로를 덮어써서** 엉뚱한 행에 색이 칠해진다(260818 실측:
    offset이 다른 표 2개를 나란히 두니 뒤 표의 offset=0 규칙이 앞 표까지 덮어써 브랜드행이
    시즌색으로 물들었음). 또 이 규칙은 `!important`라 Styler의 인라인 배경색(브랜드행 분홍)보다
    우선하므로, 범위가 어긋나면 분홍색이 아예 안 보인다.
    """
    a, b = 2 + offset, 3 + offset            # S/S TOTAL · F/W TOTAL
    c, d = 4 + offset, 8 + offset            # 개별 시즌 5행 (Z·A·B·C·D)
    return f"""
<style>
table.{cls} tbody tr:nth-child({a}) th, table.{cls} tbody tr:nth-child({a}) td,
table.{cls} tbody tr:nth-child({b}) th, table.{cls} tbody tr:nth-child({b}) td{{
    background:#e5eefb !important;font-weight:700;}}
table.{cls} tbody tr:nth-child(n+{c}):nth-child(-n+{d}) th,
table.{cls} tbody tr:nth-child(n+{c}):nth-child(-n+{d}) td{{background:#f4f8fd !important;}}
table.{cls} tbody tr:nth-child(n+{c}):nth-child(-n+{d}) th{{padding-left:18px;font-weight:500;}}
</style>
"""


_SEASON_ROW_CSS = _season_row_css()   # 하위 호환용(추가 행이 없을 때의 기본값)


# ─────────────────────────────────────────────────────────────────────────────
# 전역 테마 (2026-08-05) — Apple 스토어 톤: 옅은 회색 배경 · SF Pro 계열 타이포 ·
#   라운드 카드/알약 버튼 · 애플 블루(#0071e3) 포인트 · 사이드바 라디오를 메뉴처럼.
#   ※ 폰트는 맥=SF Pro, 윈도우=Pretendard→맑은 고딕 순으로 자연 대체된다.
# ─────────────────────────────────────────────────────────────────────────────
_APPLE_CSS = """
<style>
/* 타이포 · 배경 --------------------------------------------------------- */
html, body, .stApp, [data-testid="stAppViewContainer"], [data-testid="stSidebar"]{
  font-family:-apple-system,BlinkMacSystemFont,"SF Pro Display","SF Pro Text",
    "Apple SD Gothic Neo","Pretendard","Noto Sans KR","Malgun Gothic",sans-serif;
  letter-spacing:-0.019em;-webkit-font-smoothing:antialiased;}
.stApp{background:#f5f5f7;}
[data-testid="stHeader"]{background:transparent;}
[data-testid="stMarkdownContainer"] h1, [data-testid="stMarkdownContainer"] h2,
[data-testid="stMarkdownContainer"] h3, [data-testid="stMarkdownContainer"] h4,
[data-testid="stMarkdownContainer"] h5{color:#1d1d1f;letter-spacing:-0.03em;}
h1, [data-testid="stMarkdownContainer"] h1{font-weight:700;letter-spacing:-0.035em;}
/* 여백 축소(기존 규칙 유지) */
[data-testid="stVerticalBlock"]{gap:0.4rem;}
[data-testid="stMarkdownContainer"] h3,
[data-testid="stMarkdownContainer"] h5{margin-bottom:0.35rem;padding-bottom:0;}

/* 버튼 : 알약형 --------------------------------------------------------- */
.stButton>button, [data-testid="stFormSubmitButton"]>button,
[data-testid="stDownloadButton"]>button{
  border-radius:980px;border:1px solid #d2d2d7;background:#fff;color:#1d1d1f;
  font-weight:600;letter-spacing:-0.01em;transition:.15s;}
.stButton>button:hover, [data-testid="stFormSubmitButton"]>button:hover,
[data-testid="stDownloadButton"]>button:hover{
  border-color:#0071e3;color:#0071e3;background:#fff;}
.stButton>button[kind="primary"], [data-testid="stFormSubmitButton"]>button[kind="primary"]{
  background:#0071e3;border-color:#0071e3;color:#fff;}

/* 사이드바 -------------------------------------------------------------- */
[data-testid="stSidebar"]{background:#fbfbfd;border-right:1px solid #e8e8ed;}
[data-testid="stSidebar"] [data-testid="stMetric"]{
  background:#fff;border:1px solid #e8e8ed;border-radius:14px;padding:10px 14px;}
[data-testid="stMetricValue"]{font-weight:600;letter-spacing:-0.03em;}

/* 사이드바 조회 메뉴(라디오) → 메뉴 리스트처럼 ------------------------- */
[data-testid="stSidebar"] [role="radiogroup"]{gap:2px;}
[data-testid="stSidebar"] [role="radiogroup"] label{
  width:100%;padding:7px 12px;border-radius:11px;margin:0;transition:.12s;}
[data-testid="stSidebar"] [role="radiogroup"] label:hover{background:#eceef1;}
[data-testid="stSidebar"] [role="radiogroup"] label p{
  font-size:0.88rem;font-weight:500;letter-spacing:-0.02em;}
/* 라디오 동그라미 숨김 (구조: label > span(input) > div > div > div:first-child = 동그라미) */
[data-testid="stSidebar"] [role="radiogroup"] label>div>div>div:first-child{display:none;}
/* 선택 항목 = 애플 블루 알약 (data-selected 우선, 미지원 브라우저는 :has 폴백) */
[data-testid="stSidebar"] [role="radiogroup"] label[data-selected="true"]{background:#0071e3;}
[data-testid="stSidebar"] [role="radiogroup"] label[data-selected="true"] p{
  color:#fff;font-weight:600;}
@supports selector(:has(*)){
  [data-testid="stSidebar"] [role="radiogroup"] label:has(input:checked){background:#0071e3;}
  [data-testid="stSidebar"] [role="radiogroup"] label:has(input:checked) p{
    color:#fff;font-weight:600;}
}

/* 입력·컨테이너 --------------------------------------------------------- */
[data-testid="stFileUploader"] section{
  border-radius:12px;border:1px dashed #d2d2d7;background:#fff;}
[data-testid="stExpander"]{
  border-radius:14px;border:1px solid #e8e8ed;background:#fff;}
[data-baseweb="select"]>div, .stTextInput input, .stNumberInput input, .stDateInput input{
  border-radius:10px;border-color:#d2d2d7;}
[data-testid="stAlert"]{border-radius:14px;}
a{color:#0071e3;text-decoration:none;}

/* 표 제목 줄(perf_table) 전용: 제목+엑셀버튼 컬럼 행의 높이가 버튼 높이(약 40px) 기준으로
   맞춰지면서 제목 글자 아래로 빈 공간이 생겨 "표와 멀어 보이는" 문제(2026-08-06 중태님 지적)
   → 이 행만 아래쪽 정렬로 바꿔서 제목 글자가 버튼과 같은 라인의 '아래쪽'에 붙게 함
   (.perf-title 마커가 있는 행에만 적용 — 다른 필터/컬럼 레이아웃에는 영향 없음). */
div[data-testid="stHorizontalBlock"]:has(.perf-title){align-items:flex-end;}
</style>
"""


# 사이드바 조회 메뉴 (2026-08-05: 탭 → 사이드바 전환. 순서·명칭 = 사용자 확정본)
#   ※ 탭 방식은 안 보는 탭까지 매 실행마다 전부 계산돼서 느렸다.
#      사이드바 메뉴는 '선택된 1개'만 실행되므로 메뉴가 늘어도 속도가 유지된다.
MENU_DASH = "📊 종합 대시보드"
MENU_WEEK = "📋 주간현황 분석"   # 260820 명칭 변경 (구 '주간회의 보고자료')
MENU_FLAG = "📅 연차·아이템 세부분석"
MENU_CHAN = "📈 유통별 세부 분석"
MENU_CATMIX = "🧵 복종별 판매비중 분석"
MENU_INV  = "🏷️ 재고 가공"
MENU_TRND = "📉 추세분석"
MENU_RTN  = "🔄 반품률 분석"
MENU_SET  = "🧩 SET/단품 판매 분석"
MENU_PRICE = "💰 최저가 관리"   # 260820 신설 — 외부몰 최저가 행사 원장·캘린더·최저가 체크·네이버 체크
MENUS = [MENU_DASH, MENU_WEEK, MENU_FLAG, MENU_CHAN, MENU_CATMIX,
         MENU_INV, MENU_TRND, MENU_RTN, MENU_SET, MENU_PRICE]

# 사이드바 메뉴 3개 카테고리 구분 (2026-08-20, 중태님 확정 이미지 기준)
#   대시보드는 최상단 단독 → 📁 Analysis / 📁 노가다 금지 / 📁 궁금한 것
#   그룹별 radio 1개씩 두고, 선택은 세션 "nav_menu" 한 곳에만 보관(디스패치 로직은 기존 그대로).
MENU_GROUPS = [
    (None,        [MENU_DASH]),
    ("Analysis",  [MENU_WEEK, MENU_FLAG, MENU_CHAN, MENU_CATMIX]),
    ("노가다 금지", [MENU_INV, MENU_PRICE]),
    ("궁금한 것",  [MENU_TRND, MENU_RTN, MENU_SET]),
]


def _nav_pick(gi):
    """그룹 gi의 radio가 바뀌면 전역 선택을 갱신하고 다른 그룹 radio는 선택 해제."""
    val = st.session_state.get(f"nav_g{gi}")
    if val is None:
        return
    st.session_state["nav_menu"] = val
    for j in range(len(MENU_GROUPS)):
        if j != gi:
            st.session_state[f"nav_g{j}"] = None


def render_nav_menu():
    """사이드바 조회 메뉴(3개 카테고리) 렌더 → 현재 선택된 메뉴명 반환."""
    if st.session_state.get("nav_menu") not in MENUS:
        st.session_state["nav_menu"] = MENU_DASH
    cur = st.session_state["nav_menu"]
    for gi, (title, items) in enumerate(MENU_GROUPS):
        key = f"nav_g{gi}"
        if key not in st.session_state:          # 첫 렌더: index 인자 대신 세션값으로 초기화(경고 방지)
            st.session_state[key] = cur if cur in items else None
        if title:
            st.markdown(
                f"<div style='font-weight:700;font-size:1.05rem;color:#1d1d1f;"
                f"margin:10px 0 2px;'>📁 {title}</div>",
                unsafe_allow_html=True)
        st.radio(title or "대시보드", items, key=key,
                 label_visibility="collapsed", on_change=_nav_pick, args=(gi,))
    return st.session_state["nav_menu"]


def block_border(sty, n):
    """룰12 (2026-07-31): 당월/연간 기간블록 경계에 두꺼운 세로선 — n=첫 블록 컬럼 수.

    pandas Styler가 셀에 붙이는 col{n} 클래스(헤더 th·데이터 td 공통)를 이용해
    해당 표에만 스코프된 CSS로 경계선을 그린다(다른 표 영향 없음).
    """
    return sty.set_table_styles(
        [{"selector": f"th.col{n}", "props": [("border-left", "3px solid #555555")]},
         {"selector": f"td.col{n}", "props": [("border-left", "3px solid #555555")]}],
        overwrite=False)


def render_styled_table(sty, extra_class="", extra_css=""):
    """Styler를 HTML 표로 렌더(가로여백 축소·헤더검정·G.TOTAL 노란강조). 증감 빨강/초록은 Styler가 유지."""
    cls = ("erp-tbl " + extra_class).strip()
    sty = sty.set_table_attributes(f'class="{cls}"')
    st.markdown(_TBL_CSS + extra_css + f'<div class="erp-wrap">{sty.to_html()}</div>',
                unsafe_allow_html=True)


_XR_FILL_SKY = "#d6f0fa"     # 담당자별 TOTAL 등 (2026-08-07)
_XR_FILL_PINK = "#fbe4e8"    # 브랜드별 TOTAL (260818, 중태님 목업의 분홍)


def _fs_brand_rows():
    """시즌별/연차별 한눈에 보기의 '브랜드별 TOTAL' 5행 (260818 신설, 중태님 목업 그대로).

    라벨·구분 기준은 주간보고 브랜드행·드릴다운3 브랜드 필터와 **같은 소스(_BRAND_MASKS)**를 쓴다
    — 두 화면의 브랜드 숫자가 어긋나지 않게 하려는 것. 특히 **S/D/L = 브랜드 코드 S·D·L 3개
    (STCO·DIEMS·GENDERLESS) 합산**(SDL_BRANDS)이다.

    함수로 감싼 이유: _BRAND_MASKS가 이 파일 아래쪽(주간보고 섹션)에서 정의돼 있어서, 모듈 상수로
    만들면 import 시점에 아직 없는 이름을 참조해 NameError가 난다. 호출 시점에 읽으면 안전하다.
    """
    return [
        ("S/D/L",     _BRAND_MASKS["S/D/L"]),
        ("CODI",      _BRAND_MASKS["CODI GALLERY"]),
        ("ZERO",      _BRAND_MASKS["ZERO LOUNGE"]),
        ("Gentlemen", _BRAND_MASKS["GENTLEMENS"]),
        ("NORATED",   _BRAND_MASKS["NORATED"]),
    ]


def _sort_perf_rows(D, sort_spec, extra_rows, two_blk, cy=None):
    """260824(중태님 요청): 표 행 재정렬 — A. 유통채널별 표 정렬 필터용.

    sort_spec=(블록라벨, 지표그룹, 하위항목, 오름차순여부).
    하위항목 "CUR"는 기준연도 라벨("26년" 등)로 자동 치환 — cy에 따라 연도 라벨이 동적이라
    호출부에서 미리 못 박을 수 없어서 여기서 치환한다(yoy_frame의 cur_lbl 계산과 동일 규칙).

    행 재정렬 규칙:
    - G.TOTAL은 항상 맨 위 고정.
    - extra_rows(담당자별 TOTAL 등) 블록은 G.TOTAL 바로 아래 '블록 위치'를 유지한 채,
      블록 안에서만 같은 기준으로 재정렬 — perf_table의 extra_rows 강조 CSS가
      "항상 2 ~ (1+개수) 행" 위치를 전제하므로(nth-child) 블록 위치를 흩으면 안 된다.
    - 나머지 개별 행(매장 등)도 같은 기준으로 재정렬.
    - 값이 없는 칸("–", 예: 전년 매출 0이라 증감율 없음)은 오름/내림 무관하게 항상 맨 아래.
    """
    blk, top, sub, ascending = sort_spec
    if sub == "CUR":
        sub = f"{cy % 100:02d}년" if cy is not None else "26년"
    col = (blk, top, sub) if two_blk else (top, sub)
    if col not in D.columns:
        return D                      # 방어: 라벨 불일치 시 기존 순서 그대로(크래시 금지)
    extra_set = {lbl for lbl, _ in extra_rows} if extra_rows else set()
    mids = [k for k in D.index if k in extra_set and k != "G.TOTAL"]
    rest = [k for k in D.index if k != "G.TOTAL" and k not in extra_set]

    def _key(k):
        v = D.at[k, col]
        try:
            v = float(v)
        except (TypeError, ValueError):
            v = None
        if v is None or pd.isna(v):
            return (1, 0.0)
        return (0, v if ascending else -v)
    head = ["G.TOTAL"] if "G.TOTAL" in D.index else []
    return D.reindex(head + sorted(mids, key=_key) + sorted(rest, key=_key))


def _min_amount_perf_rows(D, blk, thr_mm, extra_rows, two_blk, cy=None):
    """260831(중태님 요청): 금액 필터 — 선택한 기간 블록의 실판매금액(기준연도)이 문턱(백만 단위)
    미만인 '개별 매장 행'만 표에서 숨긴다. A. 유통채널별 표 전용.

    - G.TOTAL·extra_rows(담당자별 TOTAL 등 집계 행)는 항상 유지 — 행을 '숨길' 뿐 집계에서
      빼는 게 아니므로 합계 숫자는 필터 전과 완전히 동일하다.
    - 값이 없는 칸("–", 전년만 매출 있는 매장 등)은 0으로 간주 → 문턱>0이면 숨겨진다.
    - 문턱이 0 이하면 아무것도 안 함(필터 해제). 컬럼 라벨 불일치 시 기존 그대로 반환(방어).
    """
    if not thr_mm or thr_mm <= 0:
        return D
    sub = f"{cy % 100:02d}년" if cy is not None else "26년"
    col = (blk, "실판매금액(백만)", sub) if two_blk else ("실판매금액(백만)", sub)
    if col not in D.columns:
        return D                      # 방어: 라벨 불일치 시 필터 없이 그대로(크래시 금지)
    extra_set = {lbl for lbl, _ in extra_rows} if extra_rows else set()

    def _val(k):
        try:
            v = float(D.at[k, col])
        except (TypeError, ValueError):
            return 0.0
        return 0.0 if pd.isna(v) else v
    keep = [k for k in D.index
            if k == "G.TOTAL" or k in extra_set or _val(k) >= thr_mm]
    return D.reindex(keep)


def perf_table(cur, prev, dim, order_list, title, key, extra=None, season_rows=False,
               month=None, blk_labels=("당월누계", "연간누계"), preview=False, big_title=False,
               cy=None, extra_rows=None, extra_row_color=None, sort_spec=None, min_spec=None):
    """제목 + 우측 엑셀버튼 + 전년비교 표 렌더.

    extra_row_color(260818 추가): extra_rows 강조 색. 안 주면 기존 하늘색(_XR_FILL_SKY).
    시즌별/연차별 한눈에 보기의 '브랜드별 TOTAL'은 분홍(_XR_FILL_PINK)을 쓴다.

    extra=(컬럼명, {행라벨: 값})이면 표 맨 앞(행이름 바로 옆)에 텍스트 컬럼을 삽입
    — 예: 유통채널별 표의 '담당자'. 엑셀 다운로드에도 그대로 포함된다.
    (2026-08-07: month와의 병용 제한 해제 — 컬럼 레벨 수에 맞춰 자동으로 삽입 키를 맞춘다.)
    season_rows=True면 G.TOTAL 아래 시즌 7행(S/S·F/W TOTAL + Z·A·B·C·D) 삽입.
    month=(cur_m, prev_m)이면 당월누계+연간누계 2블록 표(플래그십 탭 전용, 2026-07-31).
    extra_rows=[(라벨, mask_fn), ...] (2026-08-07 추가)면 G.TOTAL 바로 아래에 임의 그룹 소계
    행을 끼워 넣는다 — 예: 유통채널별 표의 '담당자별 TOTAL'. month와 함께 쓰면 두 블록에
    동일하게 적용된다(yoy_frame2 참고).
    preview=True(2026-08-06 추가)면 '🔍 조회 누르기 전' 안내용 — cur/prev가 빈 DataFrame이라
    yoy_frame이 전부 0/"–"로 채운 스켈레톤을 반환하는 걸 이용해, 실제 계산 없이 이 화면에서
    나올 표의 헤더·행 구조만 미리 보여준다(엑셀 다운로드 버튼은 숨김 — 아직 실데이터가 아니므로).
    big_title=True(2026-08-06 추가)면 제목 글자를 1.5배로 키운다 — 바로 위에 별도 큰 섹션
    제목(### ...)이 이미 있는 표(그룹의 첫 표)는 False로 두고, 그런 제목 없이 표만 연달아
    나오는 나머지 표들(신상+내년신상·1년차·2년차·3년차 등)에 True를 준다(중태님 컨펌 완료,
    2026-08-06). 제목-자기표 간격은 좁히고 이전표-제목 간격은 넓혀서 "이 제목이 어느 표
    것인지" 헷갈리지 않게 한다(가운데 컨펌 캡처 기준).
    cy=기준연도를 넘기면 표 안의 연도 컬럼 라벨("25년"/"26년" 등)이 그 연도 기준으로
    동적 계산된다 — 안 넘기면 과거 하드코딩 기본값 유지(2026-08-07 버그수정).
    sort_spec=(블록라벨, 지표그룹, 하위항목, 오름차순여부) (260824 추가, 중태님 요청)이면
    기본 정렬(연간누계 실판매금액 내림차순) 대신 그 기준으로 행을 재정렬한다 —
    _sort_perf_rows 참고. 현재 사용처: 유통별 세부 분석 A. 유통채널별 표.
    min_spec=(블록라벨, 문턱_백만) (260831 추가, 중태님 요청)이면 그 블록의 실판매금액이
    문턱 미만인 개별 매장 행을 숨긴다(집계 행·합계 숫자는 불변) — _min_amount_perf_rows 참고.
    """
    if month is not None:
        cur_m, prev_m = month
        D = yoy_frame2(cur_m, prev_m, cur, prev, dim, order_list,
                       season_rows=season_rows, blk_labels=blk_labels, cy=cy, extra_rows=extra_rows)
    else:
        D = yoy_frame(cur, prev, dim, order_list, season_rows=season_rows, cy=cy, extra_rows=extra_rows)
    # 260824(중태님 요청): 정렬 필터 — 선택한 (기간블록×지표×방향)으로 행 재정렬.
    # 미리보기(preview)는 스켈레톤이라 정렬 의미가 없어 건너뜀(전부 0이라 순서 변화도 없음).
    if sort_spec and not preview:
        D = _sort_perf_rows(D, sort_spec, extra_rows, month is not None, cy=cy)
    # 260831(중태님 요청): 금액 필터 — min_spec=(블록라벨, 문턱_백만)이면 그 블록 실판매금액이
    # 문턱 미만인 개별 매장 행을 숨긴다(G.TOTAL·extra_rows 집계 행은 유지, 합계 숫자 불변).
    # preview(스켈레톤)는 전부 0이라 필터하면 표가 통째로 비므로 건너뜀.
    if min_spec and not preview:
        D = _min_amount_perf_rows(D, min_spec[0], min_spec[1], extra_rows,
                                  month is not None, cy=cy)
    if extra:
        _name, _map = extra
        # 2026-08-07: D.columns가 2단(단일블록)/3단(month 2블록) 어느 쪽이든 삽입 키의
        # 레벨 수를 자동으로 맞춰준다 — 예전엔 2단 튜플로 고정돼 있어 month와 병용 시 에러났음.
        _extra_key = (_name,) + ("",) * (D.columns.nlevels - 1)
        D.insert(0, _extra_key,
                 ["" if k == "G.TOTAL" else str(_map.get(str(k), "") or "") for k in D.index])
    nblk = sum(1 for c in D.columns if c[0] == blk_labels[0]) if month is not None else None
    if nblk and extra:
        # 2026-08-07: extra 컬럼('담당자' 등)이 맨 앞에 삽입돼 있으면 실제 경계선 위치(컬럼
        # 순번)가 1 밀린다 — 안 더해주면 block_border가 블록1 마지막 칸에 선을 그어버린다.
        nblk += 1
    # 2026-08-07 추가: extra_rows(담당자별 TOTAL 등)로 끼워 넣은 행은 화면·엑셀 모두 하늘색으로
    # 구분 표시 — 바로 아래 개별 매장행과 헷갈리지 않게. 실제로 D에 남아있는 라벨만 사용.
    _extra_lbls = [lbl for lbl, _ in extra_rows if lbl in D.index] if extra_rows else []
    _row_color = extra_row_color or _XR_FILL_SKY
    _row_fill_xl = _row_color.lstrip("#").upper()   # 엑셀용(룰13: 화면 서식 그대로)
    h1, h2 = st.columns([4, 1])
    # 2026-08-06 (중태님 컨펌): 제목↔자기표 간격은 좁히고, 이전표↔제목 간격은 넓혀서
    # "이 제목이 바로 아래 표 것"임을 분명하게 함. big_title=True면 글자도 1.5배.
    # class="perf-title"가 있으면 _APPLE_CSS의 규칙이 이 행을 아래쪽 정렬로 바꿔
    # (엑셀버튼 높이 때문에 생기던) 제목 아래 빈 공간을 없앤다 — 실측 기반 확정(2026-08-06).
    _tsz = "1.5rem" if big_title else "1rem"
    _tmt = "0px" if preview else "22px"   # 이전 표와의 간격 (미리보기는 화면 첫 요소라 0)
    _tstyle = (f"margin:{_tmt} 0 12px;font-weight:700;font-size:{_tsz};"
               "letter-spacing:-0.01em;color:#1d1d1f;line-height:1.3;")
    if preview:
        h1.markdown(f"<div class='perf-title' style='{_tstyle}'>{title}"
                     "<span style='color:#888;font-weight:400;font-size:0.78rem;'>"
                     " — 미리보기(조회 전) · 실제 숫자는 🔍 조회 후 표시돼요</span></div>",
                     unsafe_allow_html=True)
    else:
        h1.markdown(f"<div class='perf-title' style='{_tstyle}'>{title}{_NOTE_FLOAT}</div>",
                   unsafe_allow_html=True)
        h2.download_button("⬇ 엑셀", yoy_excel_bytes(D, title[:28], first_block_cols=nblk,
                                                      extra_row_labels=_extra_lbls,
                                                      extra_row_fill=_row_fill_xl),
                           file_name=f"{_safe_name(title)[:24]}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           key=f"dl_{key}", use_container_width=True)
    sty = style_yoy(D)
    if nblk:
        sty = block_border(sty, nblk)   # 룰12: 당월/연간 경계 두꺼운 선
    # 이 표에만 걸리는 고유 클래스 — 같은 화면의 다른 표와 CSS가 서로 간섭하지 않게(260818)
    _tcls = _season_css_class(key)
    _css = ""
    if _extra_lbls:
        # extra_rows 행 강조 — 담당자별 TOTAL은 하늘색(2026-08-07), 브랜드별 TOTAL은 분홍(260818).
        # ⚠️ Styler.set_properties는 값 칸(td)만 칠하고 **맨 왼쪽 라벨 칸(인덱스 th)은 못 칠한다**
        # (260818 실측). 게다가 시즌행 CSS가 !important라 인라인 스타일보다 세다 — 그래서
        # extra_rows 행도 같은 방식(고유 클래스 + nth-child + !important)으로 칠한다.
        # extra_rows는 항상 G.TOTAL 바로 다음이라 위치는 2 ~ (1+개수)로 고정.
        _css += _extra_row_css(len(_extra_lbls), _tcls, _row_color)
    if season_rows:
        # 시즌 7행 색·들여쓰기 위치는 앞에 끼워 넣은 extra_rows 수만큼 밀어준다(260818)
        _css += _season_row_css(len(_extra_lbls), _tcls)
        render_styled_table(sty, extra_class=f"erp-season {_tcls}", extra_css=_css)
    else:
        render_styled_table(sty, extra_class=_tcls, extra_css=_css)   # 룰3·4 + G.TOTAL 노란강조


def _need_search(flag_key, submitted):
    """조회 버튼 게이트 (2026-08-06 메모리 개선).

    필터를 st.form으로 묶으면 폼 안 위젯은 아무리 바꿔도 화면이 다시 계산되지 않고,
    '🔍 조회'(form_submit_button)를 눌러야 그때 1번만 계산된다 — 회사 ERP의
    [기간 선택 → 조회 버튼] 방식과 동일. 이 함수는 '화면 첫 진입 때 아직 조회를
    한 번도 안 눌렀으면 계산을 건너뛰는' 공용 게이트다. 한 번 조회한 뒤에는
    마지막으로 조회했던 조건으로 결과가 계속 표시된다.
    """
    if submitted:
        st.session_state[flag_key] = True
    if not st.session_state.get(flag_key):
        st.info("👆 기간과 조건을 고른 뒤 **🔍 조회** 버튼을 눌러 주세요 — 조회 전에는 계산하지 않아요.")
        return True
    return False


# ── 품번별 드릴다운 (2026-08-07 재작업) ───────────────────────────────────────
# 표의 합계 숫자(예: 44.1백만원) 아래에 [아이템그룹/연차 선택 + 기간 + 연도 + 🔍 상세보기]
# 컨트롤을 두고, 누르면 그 조건의 품번별 상세(품번/품명/기간판매수량/기간총실판가/
# 원가(VAT+)/최초가/평균판매가/판가율)를 팝업(st.dialog)으로 보여준다.
# 1차 배포 때 이 기능 추가 직후 앱이 통째로 크래시 나서(트레이스백 없음, OOM 추정) 전량
# 롤백했었음(claude/배포대기_현황.md 8-2). 이번 재작업은 신규 로드 컬럼을 "원가(VAT+)" 1개로
# 줄이고(LOAD_COLS/LOAD_NUM 참고), "최초가"는 별도 컬럼을 새로 로드하지 않고 기존에 이미
# 메모리에 있는 _최초가매출÷_수량 가중평균으로 계산해서 메모리 증가폭을 최소화했다.
# st.dialog에 width= 같은 부가 kwarg는 일부러 안 씀 — 지난번 크래시 원인 후보였다가 아니었던
# 것으로 확인된 지점이라, 표면적을 최대한 줄이는 쪽으로 설계.

def _dialog_or_expander(title, on_dismiss=None):
    """st.dialog가 없는 구버전 streamlit 대비 안전판 — 있으면 진짜 팝업, 없으면 expander.

    width="large"(2026-08-07 추가): 컬럼 8개가 스크롤 없이 한 화면에 보이도록 팝업을 넓게.
    지난 크래시 때 width kwarg를 의심해 방어코드를 넣었다가 뺀 적이 있는데, 그때도 크래시가
    안 풀렸던 걸로 봐서 width 자체는 원인이 아니었던 것으로 확인됨 — try/except로만 안전하게 사용.

    on_dismiss(2026-08-07 2단계 드릴다운 추가): X로 팝업을 닫을 때 세션 상태(어느 단계
    팝업을 보여줄지)를 정리하는 콜백. 안 넘기면(None) st.dialog 기본 동작(on_dismiss="ignore")
    그대로 — 상태 정리가 필요 없는 단순 팝업(예: 빈 expander 폴백)에서는 안 써도 무방.
    """
    if hasattr(st, "dialog"):
        kwargs = {"width": "large"}
        if on_dismiss is not None:
            kwargs["on_dismiss"] = on_dismiss
        try:
            return st.dialog(title, **kwargs)
        except TypeError:
            return st.dialog(title)

    def _fallback(fn):
        def _inner(*a, **kw):
            with st.expander(title, expanded=True):
                return fn(*a, **kw)
        return _inner
    return _fallback


def _agg_detail(sub, group_col, label_col="품명"):
    """조건에 맞는 원본 거래행(sub)을 group_col(품번 또는 매장코드) 기준으로 묶어 상세 DataFrame을 만든다.

    2026-08-07 2단계 드릴다운 추가: 원래 품번 전용이던 _pn_detail을 group_col로 일반화해서
    같은 로직을 매장코드 단위 집계에도 재사용(품번별 상세 팝업에서 행을 클릭하면 그 품번의
    매장별 상세를 같은 컬럼 구성으로 다시 보여주는 용도).

    label_col(2026-08-07 추가): group_col 옆에 보여줄 설명 컬럼 — 품번별 표는 "품명"(상품명),
    매장코드별 표는 "매장명"(중태님 요청: 상품명이 다 똑같이 반복되는 대신 매장 이름이 보여야 함).
    """
    if sub is None or sub.empty or group_col not in sub.columns:
        return pd.DataFrame()
    g = sub.groupby(group_col, observed=True)
    qty = g["_수량"].sum() if "_수량" in sub.columns else pd.Series(dtype="float64")
    if qty.empty:
        return pd.DataFrame()
    rev = g["_매출액"].sum() if "_매출액" in sub.columns else pd.Series(0.0, index=qty.index)
    orig = g["_최초가매출"].sum() if "_최초가매출" in sub.columns else pd.Series(0.0, index=qty.index)
    label = g[label_col].first() if label_col in sub.columns else pd.Series("", index=qty.index)
    cost_amt = (g["원가(VAT+)"].sum() if "원가(VAT+)" in sub.columns
                else pd.Series(np.nan, index=qty.index))

    out = pd.DataFrame({"기간판매수량": qty, "기간총실판가": rev, "_orig": orig,
                         label_col: label, "_cost_amt": cost_amt})
    out.index.name = group_col
    out = out.reset_index()
    out["기간판매수량"] = pd.to_numeric(out["기간판매수량"], errors="coerce").astype("float64")
    out["기간총실판가"] = pd.to_numeric(out["기간총실판가"], errors="coerce").astype("float64")
    out["_orig"] = pd.to_numeric(out["_orig"], errors="coerce").astype("float64")
    q = out["기간판매수량"].replace(0, np.nan)
    orig_safe = out["_orig"].replace(0, np.nan)
    out["최초가"] = out["_orig"] / q
    out["평균판매가"] = out["기간총실판가"] / q
    if out["_cost_amt"].notna().any():
        out["원가(VAT+)"] = pd.to_numeric(out["_cost_amt"], errors="coerce") / q
    else:
        out["원가(VAT+)"] = np.nan
    out["판가율"] = out["기간총실판가"] / orig_safe
    out = out.sort_values("기간총실판가", ascending=False)
    return out[[group_col, label_col, "기간판매수량", "기간총실판가", "원가(VAT+)", "최초가", "평균판매가", "판가율"]]


def _pn_detail(sub):
    """조건에 맞는 원본 거래행(sub)을 받아 품번별 상세 DataFrame을 만든다."""
    return _agg_detail(sub, "품번", label_col="품명")


def _show_pn_dialog(title, sub_title, detail, group_col="품번", key_prefix="pn",
                     on_row_click=None, on_dismiss=None, on_back=None):
    """품번별(또는 매장코드별) 상세 DataFrame을 팝업(또는 expander)으로 렌더링.

    on_row_click(선택값): 지정하면 표 행 선택을 켜서(단일행), 행을 클릭했을 때 그 행의
    group_col 값으로 콜백을 호출한다 — 품번별 표에서 행 클릭 → 매장별 2차 팝업 연결용.
    on_dismiss: 팝업을 X로 닫을 때 세션 상태를 정리하는 콜백(안 넘기면 st.dialog 기본 동작).
    on_back(2026-08-10 추가): 지정하면 팝업 맨 위에 "← 뒤로가기" 버튼을 보여주고, 누르면
    콜백을 호출한다 — 매장별(3단계) 팝업에서 품번별(2단계) 팝업으로 돌아가는 용도.
    """
    @_dialog_or_expander(title, on_dismiss=on_dismiss)
    def _popup():
        if on_back is not None:
            if st.button("← 뒤로가기", key=f"{key_prefix}_back"):
                on_back()
        st.caption(sub_title)
        if detail.empty:
            st.info("해당 조건에 판매 데이터가 없어요.")
            return
        disp = detail.copy()
        disp["기간총실판가"] = disp["기간총실판가"].apply(lambda v: f"{_mm(v):,.1f}백만")
        for c in ("원가(VAT+)", "최초가", "평균판매가"):
            disp[c] = disp[c].apply(lambda v: f"{v:,.0f}" if pd.notna(v) else "–")
        disp["판가율"] = disp["판가율"].apply(lambda v: f"{v*100:.0f}%" if pd.notna(v) else "–")
        if on_row_click is not None:
            st.caption("💡 행을 클릭하면 매장별 상세를 볼 수 있어요.")
            ev = st.dataframe(disp, hide_index=True, use_container_width=True,
                               on_select="rerun", selection_mode="single-row",
                               key=f"{key_prefix}_df")
            rows = list(ev.selection.rows) if ev is not None and getattr(ev, "selection", None) else []
            if rows:
                on_row_click(detail.iloc[rows[0]][group_col])
        else:
            st.dataframe(disp, hide_index=True, use_container_width=True, key=f"{key_prefix}_df")
        if detail["원가(VAT+)"].isna().all():
            st.caption("⚠️ 원가(VAT+)가 과거 업로드분에는 없어 전부 '–'로 나올 수 있어요 — "
                       "이후 로우데이터부터 채워집니다.")
    _popup()


def pn_drilldown(cur, prev, cur_m, prev_m, dim, dim_values, title_prefix, key_prefix, cy):
    """표 아래 [🔍 상세보기 + 아이템그룹/연차 선택 + 기간 + 연도] 컨트롤 + 2단계 팝업 연결.

    dim: "아이템그룹" 또는 "연차" — 이 표가 어떤 기준으로 나뉘는지.
    dim_values: 셀렉트박스 옵션(표에 실제로 나오는 값들).
    cur/prev: 연간누계 기준으로 이미 필터된 데이터, cur_m/prev_m: 당월누계 기준.

    1단계(품번별 상세) 팝업에서 행을 클릭하면 세션 상태에 선택 품번을 저장하고 전체 재실행
    (st.rerun)해서, 다음 실행에서는 같은 조건으로 2단계(매장코드별 상세) 팝업을 대신 띄운다.
    st.dialog는 중첩 호출이 안 되므로(Streamlit 제약) 두 팝업을 동시에 열지 않고 이렇게
    "단계 전환" 방식으로 이어붙임 — 사용자 입장에선 품번 클릭 → 매장별 팝업으로 바로 이어지는
    것처럼 보인다.
    """
    if not dim_values:
        return
    # 2026-08-07: 버튼을 맨 왼쪽으로 이동(중태님 요청) — 나머지 셀렉트는 뒤로.
    c1, c2, c3, c4 = st.columns([1, 1.4, 1.3, 0.9])
    c1.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
    go = c1.button("🔍 상세보기", key=f"{key_prefix}_go", use_container_width=True)
    sel_v = c2.selectbox(dim, dim_values, key=f"{key_prefix}_dv")
    period = c3.selectbox("기간", ["당월누계", "연간누계"], key=f"{key_prefix}_pd")
    yr = c4.selectbox("연도", [cy, cy - 1], key=f"{key_prefix}_yr")

    stage_key, pn_key, pn_df_key = f"{key_prefix}_stage", f"{key_prefix}_pnsel", f"{key_prefix}_df"
    if go:
        st.session_state[stage_key] = "pn"
        st.session_state.pop(pn_key, None)
        st.session_state.pop(pn_df_key, None)  # 이전 행 선택 상태 초기화(안 지우면 재조회 즉시 3단계로 다시 넘어감)
    stage = st.session_state.get(stage_key)
    if not stage:
        return

    base_cur, base_prev = (cur_m, prev_m) if period == "당월누계" else (cur, prev)
    src = base_cur if yr == cy else base_prev
    if src is not None and not src.empty and dim in src.columns:
        sub = src[src[dim].astype(str) == str(sel_v)]
    else:
        sub = pd.DataFrame()

    def _close_all():
        st.session_state.pop(stage_key, None)
        st.session_state.pop(pn_key, None)

    def _back_to_pn():
        # 2026-08-10 추가(중태님 요청): 매장별(3단계) 팝업 → 품번별(2단계) 팝업으로 복귀.
        # pn_key(선택했던 품번)는 남겨두지 않음 — 다시 목록에서 고르도록.
        # pn_df_key(품번별 표의 행 선택 상태)도 같이 지워야 함 — 안 지우면 이전에 클릭했던 행이
        # 여전히 "선택됨" 상태로 남아 있어서 되돌아가자마자 곧바로 다시 3단계로 넘어가버림.
        st.session_state[stage_key] = "pn"
        st.session_state.pop(pn_key, None)
        st.session_state.pop(pn_df_key, None)
        st.rerun()

    if stage == "store":
        pn_sel = st.session_state.get(pn_key)
        if pn_sel is not None and "품번" in sub.columns:
            store_sub = sub[sub["품번"].astype(str) == str(pn_sel)]
        else:
            store_sub = pd.DataFrame()
        # 매장별 표는 두번째 컬럼에 품명(어차피 다 같은 값) 대신 매장명을 보여줌(중태님 요청).
        store_detail = _agg_detail(store_sub, "매장코드", label_col="매장명")
        total_rev = store_detail["기간총실판가"].sum() if not store_detail.empty else 0
        pn_name = (store_sub["품명"].iloc[0]
                   if (not store_sub.empty and "품명" in store_sub.columns) else "")
        pn_label = f"{pn_sel}({pn_name})" if pn_name else str(pn_sel)
        # 타이틀 순서(중태님 확정): 연도 · 매장별상세 · 기간 · 아이템그룹/연차 · 품번(상품명)
        _show_pn_dialog(f"{yr}년 매장별상세 · {period} · {sel_v} · {pn_label}",
                         f"실판매금액 큰 순 정렬 · 합계 {_mm(total_rev):,.1f}백만원",
                         store_detail, group_col="매장코드", key_prefix=f"{key_prefix}_st",
                         on_dismiss=_close_all, on_back=_back_to_pn)
        return

    detail = _pn_detail(sub)
    total_rev = detail["기간총실판가"].sum() if not detail.empty else 0

    def _drill_to_store(pn_val):
        st.session_state[pn_key] = pn_val
        st.session_state[stage_key] = "store"
        st.rerun()

    _show_pn_dialog(f"{title_prefix} · {sel_v} · {period} · {yr}년 품번별 상세",
                     f"실판매금액 큰 순 정렬 · 합계 {_mm(total_rev):,.1f}백만원",
                     detail, group_col="품번", key_prefix=key_prefix,
                     on_row_click=_drill_to_store, on_dismiss=_close_all)


def render_flagship(df):
    st.subheader("📅 연차 · 아이템별 전년 대비 분석")
    if df.empty or "_판매일" not in df.columns or df["_판매일"].notna().sum() == 0:
        st.info("데이터를 먼저 적재하세요.")
        return
    d = df[df["_판매일"].notna()].copy()
    years = sorted(d["_판매일"].dt.year.dropna().astype(int).unique(), reverse=True)

    st.caption("올해 vs 전년 '동기간'(같은 날짜범위) 비교 · 금액 단위 백만원 · 판가율=실판가÷최초가(가중)")
    # ── 조건 폼 (2026-08-06): 안의 위젯은 아무리 바꿔도 계산이 안 돌고, 🔍 조회를 눌러야 1번 계산 ──
    with st.form("fs_form"):
        f1, f2 = st.columns([1, 2.4])
        with f1:
            cy = st.selectbox("기준연도", years, index=0, key="fs_y")
        cur_all = d[d["_판매일"].dt.year == cy]
        dmin, dmax = cur_all["_판매일"].min().date(), cur_all["_판매일"].max().date()
        with f2:
            rng = st.date_input(f"기준기간 (전년 {cy-1} 동기간 자동)", value=(dmin, dmax),
                                min_value=d["_판매일"].min().date(), max_value=d["_판매일"].max().date(),
                                key="fs_rng")
        # 공통 필터 (주간보고 방식) — 브랜드별 → 연차별 → 시즌별 · 빈칸=전체
        fb1, fb2, fb3 = st.columns(3)
        brands = sorted(d["브랜드명"].dropna().unique()) if "브랜드명" in d.columns else []
        ages = sorted(d["연차"].dropna().unique(), key=_age_sort_key) if "연차" in d.columns else []
        seasons = sorted(d["시즌명"].dropna().unique()) if "시즌명" in d.columns else []
        selb = fb1.multiselect("브랜드별", brands, default=[], placeholder="전체", key="fs_b")
        sela = fb2.multiselect("연차별", ages, default=[], placeholder="전체", key="fs_a")
        sels = fb3.multiselect("시즌별", seasons, default=[], placeholder="전체", key="fs_s")
        chans = sorted(d["_채널"].dropna().unique()) if "_채널" in d.columns else []
        selc = st.multiselect("매장/채널", chans, default=[], placeholder="전체", key="fs_c")
        st.caption("※ 기준연도를 바꿨다면 기준기간 날짜도 그 연도로 맞춘 뒤 🔍 조회를 눌러 주세요. "
                   "(실제 집계는 기준기간 날짜를 따라요)")
        run = st.form_submit_button("🔍 조회", type="primary")
    if _need_search("fs_go", run):
        # 2026-08-06: 조회 전에도 "이 화면이 뭘 보여주는 표인지" 헤더만 미리 보여줌 —
        # 빈 DataFrame으로 perf_table을 부르면 yoy_frame이 G.TOTAL+시즌7행을 전부 "–"로
        # 채운 스켈레톤을 만들어서, 실제 51만 건 계산 없이 표 구조만 공짜로 보인다.
        _empty = pd.DataFrame()
        perf_table(_empty, _empty, "연차", None, "시즌별/연차별 한눈에 보기", "fs_preview",
                   season_rows=True, month=(_empty, _empty), preview=True, cy=cy,
                   extra_rows=_fs_brand_rows(), extra_row_color=_XR_FILL_PINK)
        return

    if not (isinstance(rng, (list, tuple)) and len(rng) == 2):
        st.info("기간(시작~끝)을 선택한 뒤 🔍 조회를 눌러 주세요.")
        return
    s, e = pd.to_datetime(rng[0]), pd.to_datetime(rng[1])
    base = d.copy()
    if selb and "브랜드명" in base:
        base = base[base["브랜드명"].isin(selb)]
    if sela and "연차" in base:
        base = base[base["연차"].isin(sela)]
    if sels and "시즌명" in base:
        base = base[base["시즌명"].isin(sels)]
    if selc and "_채널" in base:
        base = base[base["_채널"].isin(selc)]
    cur = base[(base["_판매일"] >= s) & (base["_판매일"] <= e)]
    prev = base[(base["_판매일"] >= s - pd.DateOffset(years=1)) & (base["_판매일"] <= e - pd.DateOffset(years=1))]
    # 당월누계 (2026-07-31 목업 v2 컨펌): 기준기간 끝날짜가 속한 달의 1일 → 끝날짜 · 전년 동범위 비교
    ms = e.replace(day=1)
    cur_m = base[(base["_판매일"] >= ms) & (base["_판매일"] <= e)]
    prev_m = base[(base["_판매일"] >= ms - pd.DateOffset(years=1)) & (base["_판매일"] <= e - pd.DateOffset(years=1))]
    blk = (f"당월누계 ({ms.month:02d}/{ms.day:02d}→{e.month:02d}/{e.day:02d})",
           f"연간누계 ({s.month:02d}/{s.day:02d}→{e.month:02d}/{e.day:02d})")

    tot_c = cur["_매출액"].sum()
    tot_p = prev["_매출액"].sum()
    orig_c = cur["_최초가매출"].sum() if "_최초가매출" in cur.columns else 0
    orig_p = prev["_최초가매출"].sum() if "_최초가매출" in prev.columns else 0
    pg_c = (tot_c / orig_c) if orig_c else None
    pg_p = (tot_p / orig_p) if orig_p else None
    k1, k2, k3, k4, k5, k6 = st.columns(6, gap="small")
    k1.metric(f"{cy} 매출(백만)", f"{_mm(tot_c):,.0f}")
    k2.metric(f"{cy-1} 매출(백만)", f"{_mm(tot_p):,.0f}")
    g = ((tot_c - tot_p) / tot_p) if tot_p else None
    k3.metric("전년비 성장률", "신규/–" if g is None else f"{g*100:+.1f}%")
    k4.metric(f"{cy} 판가율", f"{pg_c*100:.1f}%" if pg_c is not None else "–")
    k5.metric(f"{cy-1} 판가율", f"{pg_p*100:.1f}%" if pg_p is not None else "–")
    pg_diff = (pg_c - pg_p) if (pg_c is not None and pg_p is not None) else None
    k6.metric("판가율 증감", "–" if pg_diff is None else f"{pg_diff*100:+.1f}%p")
    if not tot_p:
        st.warning(f"전년({cy-1}) 동기간 데이터가 없어요. {cy-1}년 로우데이터를 적재하면 채워집니다.")

    # 연차 순서
    age_order = sorted([a for a in base["연차"].dropna().unique()], key=_age_sort_key)
    # 표 이름 변경 (2026-07-31 컨펌): "연차별 성과표" → "시즌별/연차별 한눈에 보기"
    # 시즌 7행 포함 (목업 v3 컨펌): S/S TOTAL · F/W TOTAL + Z·A·B·C·D
    # 260818(중태님 목업): G.TOTAL 바로 아래에 브랜드별 TOTAL 5행 추가(분홍) — 시즌 7행보다 위
    st.markdown("### 시즌별/연차별 한눈에 보기")
    perf_table(cur, prev, "연차", age_order, "시즌별/연차별 한눈에 보기", "age",
               season_rows=True, month=(cur_m, prev_m), blk_labels=blk, cy=cy,
               extra_rows=_fs_brand_rows(), extra_row_color=_XR_FILL_PINK)
    st.caption("※ 분홍색 행 = 브랜드별 TOTAL(비중은 G.TOTAL 대비). **S/D/L은 브랜드 코드 S·D·L "
               "3개(STCO·DIEMS·GENDERLESS) 합산**이에요. 위 '브랜드' 필터를 걸면 그 조건 안에서만 집계돼요.")
    pn_drilldown(cur, prev, cur_m, prev_m, "연차", age_order,
                 "시즌별/연차별 한눈에 보기", "pn_age", cy)

    st.markdown("### 아이템그룹별 성과표 (전연차 토탈 + 연차별)")
    grp_present = [g for g in ITEMGROUP_ORDER if g in cur["아이템그룹"].unique()] if "아이템그룹" in cur.columns else ITEMGROUP_ORDER
    perf_table(cur, prev, "아이템그룹", ITEMGROUP_ORDER, "아이템그룹별 성과표 (전연차)", "grp_all",
               month=(cur_m, prev_m), blk_labels=blk, cy=cy)
    pn_drilldown(cur, prev, cur_m, prev_m, "아이템그룹", grp_present,
                 "아이템그룹별 성과표 (전연차)", "pn_grp_all", cy)
    # 연차별 버킷
    buckets = []
    sinsang = [a for a in ["신상", "내년신상"] if a in age_order]
    if sinsang:
        buckets.append(("신상+내년신상", sinsang))
    for a in age_order:
        if a.endswith("년차"):
            buckets.append((a, [a]))
    for name, ages in buckets:
        curb = cur[cur["연차"].isin(ages)]
        prevb = prev[prev["연차"].isin(ages)]
        curb_m = cur_m[cur_m["연차"].isin(ages)]
        prevb_m = prev_m[prev_m["연차"].isin(ages)]
        perf_table(curb, prevb, "아이템그룹", ITEMGROUP_ORDER,
                   f"아이템그룹별 성과표 ({name})", f"grp_{name}",
                   month=(curb_m, prevb_m), blk_labels=blk, big_title=True, cy=cy)
        grp_present_b = ([g for g in ITEMGROUP_ORDER if g in curb["아이템그룹"].unique()]
                          if "아이템그룹" in curb.columns else ITEMGROUP_ORDER)
        pn_drilldown(curb, prevb, curb_m, prevb_m, "아이템그룹", grp_present_b,
                     f"아이템그룹별 성과표 ({name})", f"pn_grp_{name}", cy)


# ── 대시보드 채널 통합 (2026-07-31): 수수료 조건 때문에 2개로 나눠 등록한 매장을 실제 채널로 합산 ──
CH_MERGE = {"SD185": "쿠팡토탈", "SD184": "쿠팡토탈",
            "SD165": "네이버토탈", "SD174": "네이버토탈"}

# 채널 리그 정의 (매장 마스터 '리그구분' 값과 동일해야 함)
LEAGUES = [("1부리그", "🏆", "#2f4d7d", "메이저 채널"),
           ("2부리그", "🚀", "#4a7ab5", "도전 채널"),
           ("꿈나무리그", "🌱", "#3f9464", "루키 채널")]

# 🆕신상판매 1위 배지 산출 제외 (2026-08-01 중태님 지시): 본사 직영 온라인 통합몰(SD065)은
# 신상 매출이 구조적으로 항상 1등이라 재미가 없음 → 이 배지 후보에서만 제외.
# ※ 랭킹 순위·다른 배지(성장/판가율/평균단가)에는 그대로 참여. 매장코드 기준이라 매장명이 바뀌어도 유지.
SINSANG_BADGE_EXCL = {"SD065"}

# 260803 추가 (중태님 지시 · 재확인: 품번이 아니라 '채널 랭킹'과 동일하게 매장 TOP10):
#  채널 랭킹 아래에, 카테고리별로 실판매금액이 많은 "매장(채널)" TOP10 보드 3종을 같은 컨셉으로 노출.
#  · 슈트류=아이템그룹(중카테고리) '수트류' · 온라인셔츠=브랜드 J(GENTLEMENS PHILOSOPHY) 전체 상품 합
#  · FW+NT+BE=아이템그룹 '신발'+'넥타이'+'벨트' 3종 합산 · 순위 기준=실판매금액 · 기간=채널랭킹 토글 연동
#  · 매장 통합(쿠팡토탈·네이버토탈)·'26년 미운영' 제외 등은 위 채널 랭킹과 동일 기준(CH_MERGE·excl_codes)
STORE_RANK_BOARDS = [
    ("슈트류 판매 우수 매장 TOP10", "👔", "#5c4a8a", "중카테고리 기준 · 수트류",
     lambda x: x["아이템그룹"].astype(str) == "수트류"),
    ("온라인셔츠 판매 매장 TOP10", "👕", "#2f6bb0", "브랜드 J (GENTLEMENS PHILOSOPHY) 전체",
     lambda x: x["브랜드명"].astype(str) == "GENTLEMENS PHILOSOPHY"),
    ("FW+NT+BE 판매 매장 TOP10", "👞", "#3f9464", "신발+넥타이+벨트 합산",
     lambda x: x["아이템그룹"].astype(str).isin(["신발", "넥타이", "벨트"])),
]


def _build_store_rank_entries(fc, fp, mask, name_of, excl_codes):
    """카테고리 랭킹 보드용: 조건(mask)에 맞는 매출만 걸러 매장(채널)별 매출 합계·전년비교 entries.

    채널 랭킹(_build_chan)과 동일 기준으로 쿠팡토탈·네이버토탈 합산, '26년 미운영' 매장 제외.
    배지(성장1위·판가율1위·평균단가1위) 계산용으로 최초가매출(o)·판매수량(q)도 함께 집계한다.
    [(채널명, {c, p, o, q, codes}), ...] 를 올해(c) 매출 내림차순으로 반환(TOP10만).
    """
    def _prep(f):
        if f is None or f.empty or "매장코드" not in f.columns:
            return None
        sub = f[mask(f)]
        return sub if not sub.empty else None

    c_df, p_df = _prep(fc), _prep(fp)

    def _grp(f, col):
        if f is None or col not in f.columns:
            return pd.Series(dtype="float64")
        return f.groupby(f["매장코드"].astype(str).str.strip(), observed=True)[col].sum()

    cs, ps = _grp(c_df, "_매출액"), _grp(p_df, "_매출액")
    os_ = _grp(c_df, "_최초가매출")   # 판가율1위 배지용(올해 기준)
    qs = _grp(c_df, "_수량")          # 평균단가1위 배지용(올해 기준)
    out = {}
    for c in set(cs.index) | set(ps.index):
        if c in excl_codes:
            continue
        key = CH_MERGE.get(c, name_of.get(c, c))
        e = out.setdefault(key, {"c": 0.0, "p": 0.0, "o": 0.0, "q": 0.0, "codes": []})
        e["c"] += float(cs.get(c, 0.0))
        e["p"] += float(ps.get(c, 0.0))
        e["o"] += float(os_.get(c, 0.0))
        e["q"] += float(qs.get(c, 0.0))
        e["codes"].append(c)
    entries = [(k, e) for k, e in out.items() if e["c"] > 0 or e["p"] > 0]
    entries.sort(key=lambda t: -t[1]["c"])
    return entries[:10]


def _store_rank_board_html(title, icon, hcolor, subtitle, entries, show_n=10):
    """카테고리별 매장 랭킹 보드 HTML — 채널 리그 보드(_league_board_html)와 동일 컨셉·스타일.

    entries=[(채널명, {c,p,o,q,codes}), ...] 매출(올해=c) 내림차순, 최대 show_n개.
    매장명 아래 ⚡성장1위·💎판가율1위·💰평균단가1위 배지 표시(2026-08-03 추가, 채널 랭킹과 동일 3종).
    매장코드(합산 매장 목록)는 표시 안 함.
    """
    head = (f"<div style='background:{hcolor};color:#fff;padding:9px 12px;font-weight:800;"
            f"font-size:0.92rem;display:flex;justify-content:space-between;align-items:center;'>"
            f"<span>{icon} {title}</span>"
            f"<span style='font-weight:400;font-size:0.7rem;opacity:0.85;'>{subtitle}</span></div>")
    if not entries:
        return ("<div style='border:1px solid #e3e6ea;border-radius:10px;overflow:hidden;'>" + head +
                "<div style='padding:14px;color:#999;font-size:0.8rem;'>매출 있는 매장이 없어요.</div></div>")
    emap = dict(entries)
    prev_rank = {k: i + 1 for i, k in enumerate(
        sorted([k for k, e in entries if e["p"] > 0], key=lambda k: -emap[k]["p"]))}
    growths = {k: (e["c"] - e["p"]) / e["p"] for k, e in entries if e["p"] > 0}
    pgr = {k: e["c"] / e["o"] for k, e in entries if e.get("o", 0) > 0}
    unit = {k: e["c"] / e["q"] for k, e in entries if e.get("q", 0) > 0}
    _winners = [
        (max(growths, key=growths.get) if growths else None, ("⚡성장1위", "#fff3cd", "#ffe08a", "#8a6d00")),
        (max(pgr, key=pgr.get) if pgr else None, ("💎판가율1위", "#f3ecff", "#d3bdf5", "#5b3d99")),
        (max(unit, key=unit.get) if unit else None, ("💰평균단가1위", "#e6f6f1", "#a8dfcd", "#0f6b4f")),
    ]
    star_badges = {}
    for k, (txt, bg, bd, fg) in _winners:
        if k is None:
            continue
        star_badges.setdefault(k, []).append(
            f"<span style='background:{bg};border:1px solid {bd};color:{fg};border-radius:8px;"
            f"font-size:0.6rem;font-weight:800;padding:0 4px;margin-left:3px;white-space:nowrap;'>{txt}</span>")
    max_c = entries[0][1]["c"] or 1.0
    medals = {1: "🥇", 2: "🥈", 3: "🥉"}

    def row_html(rk, k, e):
        g = ((e["c"] - e["p"]) / e["p"]) if e["p"] > 0 else None
        gtxt = (f"<span style='color:{'#1f8a4c' if g >= 0 else '#c62828'};'>{g*100:+.1f}%</span>"
                if g is not None else "<span style='color:#1f8a4c;'>신규</span>")
        if e["p"] > 0:
            pr = prev_rank.get(k)
            dv = (pr - rk) if pr else 0
            if dv > 0:
                mv = f"<span style='color:#1f8a4c;font-weight:800;'>▲{dv}</span>"
            elif dv < 0:
                mv = f"<span style='color:#c62828;font-weight:800;'>▼{-dv}</span>"
            else:
                mv = "<span style='color:#999;'>—</span>"
        else:
            mv = "<span style='color:#1f8a4c;font-weight:800;'>NEW</span>"
        # 매장코드(쿠팡토탈=SD184+SD185 등 합산 목록)는 표시하지 않음 — 배지만 이름 아래 표시(2026-08-03)
        badge = "".join(star_badges.get(k, []))
        return (
            "<div style='display:flex;align-items:center;gap:7px;padding:6px 10px;"
            "border-top:1px solid #f0f2f4;font-size:0.78rem;'>"
            f"<span style='width:26px;text-align:center;font-weight:800;flex:none;'>{medals.get(rk, rk)}</span>"
            f"<span style='width:96px;flex:none;font-weight:600;line-height:1.25;'>{k}{badge}</span>"
            f"<span style='flex:1;'><span style='display:block;height:12px;border-radius:2px;"
            f"background:{hcolor};width:{max(e['c']/max_c*100, 1):.0f}%;'></span></span>"
            f"<span style='flex:none;width:92px;text-align:right;font-size:0.74rem;font-weight:700;"
            f"line-height:1.3;'>{e['c']/1e6:,.1f}<br>{gtxt}</span>"
            f"<span style='flex:none;width:34px;text-align:center;font-size:0.7rem;'>{mv}</span></div>")

    rows_main = "".join(row_html(i + 1, k, e) for i, (k, e) in enumerate(entries[:show_n]))
    tot_c = sum(e["c"] for _, e in entries[:show_n])
    tot_p = sum(e["p"] for _, e in entries[:show_n])
    yoyt = f" · 전년비 {(tot_c - tot_p)/tot_p*100:+.1f}%" if tot_p else ""
    foot = (f"<div style='padding:6px 12px;background:#fafafc;border-top:1px solid #eee;"
            f"font-size:0.7rem;color:#888;'>TOP{show_n} 합계 {tot_c/1e6:,.1f}{yoyt}</div>")
    return ("<div style='border:1px solid #e3e6ea;border-radius:10px;overflow:hidden;'>"
            + head + rows_main + foot + "</div>")


def _league_board_html(lg_name, icon, hcolor, subtitle, entries, show_n=11):
    """채널 리그 보드 HTML (목업 v3 컨펌). entries=[(채널명, {c,p,codes}) 누계매출 내림차순].

    · 매출 있는 채널 11위까지 기본 노출(경쟁심!), 12위 이하는 <details> 펼쳐보기
    · 막대=리그 내 상대 크기 · 🥇🥈🥉 · 전년비 색 · 순위변동(전년 동기간 리그 내 순위) · ⚡성장1위 배지
    """
    head = (f"<div style='background:{hcolor};color:#fff;padding:9px 12px;font-weight:800;"
            f"font-size:0.92rem;display:flex;justify-content:space-between;align-items:center;'>"
            f"<span>{icon} {lg_name}</span>"
            f"<span style='font-weight:400;font-size:0.7rem;opacity:0.85;'>{subtitle} · 매출 {len(entries)}개</span></div>")
    if not entries:
        return ("<div style='border:1px solid #e3e6ea;border-radius:10px;overflow:hidden;'>" + head +
                "<div style='padding:14px;color:#999;font-size:0.8rem;'>이 리그에 매출 있는 매장이 없어요.</div></div>")
    emap = dict(entries)
    prev_rank = {k: i + 1 for i, k in enumerate(
        sorted([k for k, e in entries if e["p"] > 0], key=lambda k: -emap[k]["p"]))}
    growths = {k: (e["c"] - e["p"]) / e["p"] for k, e in entries if e["p"] > 0}
    # 리그별 1위 배지 4종 (2026-07-31 추가): 매출 1등이 아니어도 주인공이 나오게
    _pgr = {k: e["c"] / e["o"] for k, e in entries if e.get("o", 0) > 0}
    _unit = {k: e["c"] / e["q"] for k, e in entries if e.get("q", 0) > 0}
    _sins = {k: e.get("s", 0.0) for k, e in entries
             if e.get("s", 0.0) > 0
             and not (set(e.get("codes", [])) & SINSANG_BADGE_EXCL)}   # 통합몰(SD065) 제외
    _winners = [
        (max(growths, key=growths.get) if growths else None, ("⚡성장1위", "#fff3cd", "#ffe08a", "#8a6d00")),
        (max(_pgr, key=_pgr.get) if _pgr else None, ("💎판가율1위", "#f3ecff", "#d3bdf5", "#5b3d99")),
        (max(_unit, key=_unit.get) if _unit else None, ("💰평균단가1위", "#e6f6f1", "#a8dfcd", "#0f6b4f")),
        (max(_sins, key=_sins.get) if _sins else None, ("🆕신상판매1위", "#e5f1fd", "#a9cdf2", "#1a5da8")),
    ]
    star_badges = {}   # 같은 채널이 여러 배지를 휩쓸 수도 있음 → 리스트로 누적
    for k, (txt, bg, bd, fg) in _winners:
        if k is None:
            continue
        star_badges.setdefault(k, []).append(
            f"<span style='background:{bg};border:1px solid {bd};color:{fg};border-radius:8px;"
            f"font-size:0.6rem;font-weight:800;padding:0 4px;margin-left:3px;white-space:nowrap;'>{txt}</span>")
    max_c = entries[0][1]["c"] or 1.0
    medals = {1: "🥇", 2: "🥈", 3: "🥉"}

    def row_html(rk, k, e):
        g = growths.get(k)
        gtxt = (f"<span style='color:{'#1f8a4c' if g >= 0 else '#c62828'};'>{g*100:+.1f}%</span>"
                if g is not None else "<span style='color:#1f8a4c;'>신규</span>")
        if e["p"] > 0:
            pr = prev_rank.get(k)
            dv = (pr - rk) if pr else 0
            if dv > 0:
                mv = f"<span style='color:#1f8a4c;font-weight:800;'>▲{dv}</span>"
            elif dv < 0:
                mv = f"<span style='color:#c62828;font-weight:800;'>▼{-dv}</span>"
            else:
                mv = "<span style='color:#999;'>—</span>"
        else:
            mv = "<span style='color:#1f8a4c;font-weight:800;'>NEW</span>"
        badge = "".join(star_badges.get(k, []))
        sub = ""
        if k in set(CH_MERGE.values()):
            sub = ("<small style='display:block;color:#999;font-weight:400;font-size:0.66rem;'>"
                   + "+".join(sorted(set(e["codes"]))) + " 합산</small>")
        return (
            "<div style='display:flex;align-items:center;gap:7px;padding:6px 10px;"
            "border-top:1px solid #f0f2f4;font-size:0.78rem;'>"
            f"<span style='width:26px;text-align:center;font-weight:800;flex:none;'>{medals.get(rk, rk)}</span>"
            f"<span style='width:96px;flex:none;font-weight:600;line-height:1.25;'>{k}{badge}{sub}</span>"
            f"<span style='flex:1;'><span style='display:block;height:12px;border-radius:2px;"
            f"background:#2f6bb0;width:{max(e['c']/max_c*100, 1):.0f}%;'></span></span>"
            f"<span style='flex:none;width:92px;text-align:right;font-size:0.74rem;font-weight:700;"
            f"line-height:1.3;'>{e['c']/1e6:,.1f}<br>{gtxt}</span>"
            f"<span style='flex:none;width:34px;text-align:center;font-size:0.7rem;'>{mv}</span></div>")

    rows_main = "".join(row_html(i + 1, k, e) for i, (k, e) in enumerate(entries[:show_n]))
    rows_rest = ""
    if len(entries) > show_n:
        inner = "".join(row_html(i + 1, k, e) for i, (k, e) in enumerate(entries) if i >= show_n)
        rows_rest = (f"<details><summary style='padding:6px 12px;font-size:0.72rem;color:#666;"
                     f"cursor:pointer;'>{show_n + 1}위 이하 {len(entries) - show_n}개 보기</summary>{inner}</details>")
    tot_c = sum(e["c"] for _, e in entries)
    tot_p = sum(e["p"] for _, e in entries)
    yoyt = f" · 전년비 {(tot_c - tot_p)/tot_p*100:+.1f}%" if tot_p else ""
    foot = (f"<div style='padding:6px 12px;background:#fafafc;border-top:1px solid #eee;"
            f"font-size:0.7rem;color:#888;'>리그 합계 {tot_c/1e6:,.1f}{yoyt}</div>")
    return ("<div style='border:1px solid #e3e6ea;border-radius:10px;overflow:hidden;'>"
            + head + rows_main + rows_rest + foot + "</div>")


def render_priority_banner():
    """종합 대시보드 — 온라인팀 우선순위(당월·금주) 표시 (260830 신규 · 월별 매출 표 바로 위).

    사이드바에서 '온라인팀 우선순위 업로드 FORM'을 올려두면 당월/금주 두 표를 나란히 보여준다.
    업로드 전(데이터 없음)에는 아무것도 그리지 않는다(대시보드 깔끔 유지).
    """
    try:
        pri = load_priority()
    except Exception:
        return
    if pri is None or pri.empty:
        return

    def _sec_html(title, sub):
        cats = list(dict.fromkeys(sub["category"]))
        items = {c: sub[sub["category"] == c].sort_values("no") for c in cats}
        nrow = max((len(v) for v in items.values()), default=0)
        # 260830 3차: 번호 컬럼을 <colgroup>으로 딱 숫자 폭(34px)만큼 고정 — table-layout:fixed에서
        # 첫 행이 colspan 헤더라 td width가 안 먹고 4컬럼 균등분배되던 문제 해결. 넓어진 내용 칸에
        # 맞춰 글자도 0.82rem → 0.95rem으로 키움(중태님 지시: 번호 칸 축소 + 글씨 확대).
        th = ("padding:6px 10px;background:#f5f5f7;border:1px solid #e3e6ea;"
              "font-weight:700;text-align:center;font-size:0.9rem;color:#1d1d1f;")
        td = ("padding:6px 12px;border:1px solid #e3e6ea;font-size:0.95rem;"
              "text-align:left;color:#1d1d1f;line-height:1.5;")
        tdn = td + "text-align:center;color:#888;padding:6px 4px;"
        h = (f"<div style='font-weight:700;font-size:0.95rem;margin:8px 0 6px;'>{title}</div>"
             "<table style='border-collapse:collapse;width:100%;table-layout:fixed;'>")
        h += "<colgroup>" + "<col style='width:34px'><col>" * len(cats) + "</colgroup>"
        h += "<tr>" + "".join(f"<th style='{th}' colspan='2'>{c}</th>" for c in cats) + "</tr>"
        for i in range(nrow):
            h += "<tr>"
            for c in cats:
                rows_c = items[c]
                if i < len(rows_c):
                    r = rows_c.iloc[i]
                    h += f"<td style='{tdn}'>{r['no']}</td><td style='{td}'>{r['content']}</td>"
                else:
                    h += f"<td style='{tdn}'></td><td style='{td}'></td>"
            h += "</tr>"
        return h + "</table>"

    st.markdown("### 온라인팀 우선순위")
    # 260830 2차: 좌우 2컬럼 → 상하 배치(당월 위·금주 아래)로 변경 — 칸이 가로로 넓어져
    # 내용이 한 줄에 들어가므로 결과적으로 세로 길이도 크게 안 늘어남(중태님 지시).
    for sec, icon in (("당월", "🗓️"), ("금주", "📌")):
        sub = pri[pri["section"] == sec]
        if not sub.empty:
            st.markdown(_sec_html(f"{icon} {sec} 우선순위", sub), unsafe_allow_html=True)
    st.caption("※ 사이드바 **온라인팀 우선순위 업로드**로 교체할 수 있어요 — "
               "주간현황 분석 '⬇ 엑셀'의 당월·금주 우선순위 칸에도 같은 내용이 채워져요.")
    st.divider()


def render_dashboard(df):
    """종합 대시보드 (2026-07-31 전면 개편 · 목업 컨펌) — 전체 그림 + 자동 인사이트.

    ① KPI 6장: 누계 매출 전년비 · 연간 진도율(vs 시간진도) · 당월 매출 전년비 · 당월 진도율 · 누계 판가율 · S/D/L 신상매출(항목22)
    ② 월별 매출(올해 vs 전년 vs 사업계획 점선) ③ 자동 인사이트(성장/부진 채널·아이템·신상 비중·판가율)
    ④ 아이템그룹 증감액 + 연차 구성 변화 ⑤ 채널 TOP10(색=전년비).
    필터 없음 = 회사 전체 기준(사업계획 진도율 정합성). 금액 백만원(룰1)·연도 2자리(룰2).
    """
    if df is None or df.empty or "_판매일" not in df.columns or df["_판매일"].notna().sum() == 0:
        st.info("데이터를 먼저 적재하세요.")
        return
    d = df[df["_판매일"].notna()]
    asof = d["_판매일"].max()
    cy, py = int(asof.year), int(asof.year) - 1
    sy, sc = str(py)[-2:], str(cy)[-2:]
    y_start = asof.replace(month=1, day=1)
    m_start = asof.replace(day=1)
    cur_y = d[(d["_판매일"] >= y_start) & (d["_판매일"] <= asof)]
    prev_y = d[(d["_판매일"] >= y_start - pd.DateOffset(years=1)) & (d["_판매일"] <= asof - pd.DateOffset(years=1))]
    cur_m = d[(d["_판매일"] >= m_start) & (d["_판매일"] <= asof)]
    prev_m = d[(d["_판매일"] >= m_start - pd.DateOffset(years=1)) & (d["_판매일"] <= asof - pd.DateOffset(years=1))]

    st.caption(f"기준일 **{asof.date()}** (데이터 마지막 날) · 올해({sc}) vs 전년({sy}) 동기간 · "
               "필터 없음 = 회사 전체 그림 · [금액: 백만원 / VAT+]")

    # ── 사업계획 (매장 시트의 G.TOTAL 월별 목표) ──
    plan = load_plan()
    pmon = {}
    if plan is not None and not plan.empty:
        pg = plan[(plan["dim"] == "store") & (plan["code"] == "G.TOTAL")]
        pmon = {int(k): float(v) for k, v in pg.groupby("month")["amount"].sum().to_dict().items()}
    plan_annual = sum(pmon.values()) if pmon else None
    plan_month = pmon.get(int(asof.month)) if pmon else None

    ry, rpy = float(cur_y["_매출액"].sum()), float(prev_y["_매출액"].sum())
    rm, rpm = float(cur_m["_매출액"].sum()), float(prev_m["_매출액"].sum())
    oy, opy = float(cur_y["_최초가매출"].sum()), float(prev_y["_최초가매출"].sum())
    pg26 = (ry / oy) if oy else None
    pg25 = (rpy / opy) if opy else None

    # ── ① KPI 6장 (항목22 · 260821: S/D/L 신상매출 카드 추가) ──
    _sdl_new = lambda x: x[x["브랜드명"].isin(SDL_BRANDS) & x["연차"].isin(["신상", "내년신상"])]
    sdl_y = float(_sdl_new(cur_y)["_매출액"].sum())
    sdl_py = float(_sdl_new(prev_y)["_매출액"].sum())
    k1, k2, k3, k4, k5, k6 = st.columns(6)
    k1.metric(f"{sc}년 누계 매출(백만)", f"{_mm(ry):,.0f}",
              f"{(ry-rpy)/rpy*100:+.1f}% vs 전년 {_mm(rpy):,.0f}" if rpy else None)
    tprog_y = float(asof.dayofyear) / (366.0 if asof.is_leap_year else 365.0)
    if plan_annual:
        prog_y = ry / plan_annual
        k2.metric("연간 사업계획 진도율", f"{prog_y*100:.0f}%",
                  f"{(prog_y-tprog_y)*100:+.1f}%p vs 시간진도 {tprog_y*100:.0f}%")
        k2.progress(min(prog_y, 1.0))
        k2.caption(f"계획 {plan_annual/1e6:,.0f} · 잔여 {max(plan_annual-ry, 0)/1e6:,.0f}")
    else:
        k2.metric("연간 사업계획 진도율", "–")
        k2.caption("사업계획 업로드 시 표시돼요")
    k3.metric(f"당월({asof.month}월) 매출(백만)", f"{_mm(rm):,.1f}",
              f"{(rm-rpm)/rpm*100:+.1f}% vs 전년 {_mm(rpm):,.1f}" if rpm else None)
    tprog_m = float(asof.day) / float(asof.days_in_month)
    if plan_month:
        prog_m = rm / plan_month
        k4.metric("당월 계획 진도율", f"{prog_m*100:.0f}%",
                  f"{(prog_m-tprog_m)*100:+.1f}%p vs 시간진도 {tprog_m*100:.0f}%")
        k4.progress(min(prog_m, 1.0))
        k4.caption(f"{asof.month}월 계획 {plan_month/1e6:,.1f}")
    else:
        k4.metric("당월 계획 진도율", "–")
        k4.caption("사업계획 업로드 시 표시돼요")
    k5.metric("누계 판가율", f"{pg26*100:.1f}%" if pg26 is not None else "–",
              f"{(pg26-pg25)*100:+.1f}%p vs 전년 {pg25*100:.1f}%"
              if (pg26 is not None and pg25 is not None) else None)
    k6.metric("S/D/L 신상매출(백만)", f"{_mm(sdl_y):,.0f}",
              f"{(sdl_y-sdl_py)/sdl_py*100:+.1f}% vs 전년 {_mm(sdl_py):,.0f}" if sdl_py else None)
    k6.caption(f"{sc}년 누계 · STCO/DIEMS/GENDERLESS · 연차 신상+내년신상")

    # ── 온라인팀 우선순위 (260830 신규 · 업로드돼 있으면 월별 매출 위에 표시) ──
    render_priority_banner()

    # ── ② 월별 매출: 올해 vs 전년 vs 사업계획 ──
    st.markdown(f"### 월별 매출 — {sc}년 vs {sy}년 vs 사업계획")
    dm = d[d["_판매일"].dt.year.isin([cy, py])]
    gsum = dm.groupby([dm["_판매일"].dt.year, dm["_판매일"].dt.month])["_매출액"].sum().to_dict()
    months = list(range(1, 13))
    xlab = [f"{m}월" for m in months]
    v25 = [gsum.get((py, m), 0.0) / 1e6 for m in months]
    v26 = [(gsum.get((cy, m), 0.0) / 1e6) if m <= asof.month else None for m in months]
    fig = go.Figure()
    fig.add_bar(x=xlab, y=v25, name=f"{sy}년 실적", marker_color="#b8bec7")
    fig.add_bar(x=xlab, y=v26, name=f"{sc}년 실적", marker_color="#2f6bb0")
    if pmon:
        fig.add_scatter(x=xlab, y=[pmon.get(m, 0.0) / 1e6 for m in months],
                        name=f"{sc} 사업계획", mode="lines",
                        line=dict(color="#444444", width=2.5, dash="dash"))
    fig.update_layout(barmode="group", height=330, margin=dict(t=10, b=0, l=0, r=0),
                      legend=dict(orientation="h", y=1.1), yaxis_title="백만원")
    st.plotly_chart(fig, use_container_width=True)
    st.caption("※ 남은 달은 계획 점선만 보여요 — 앞으로 채울 목표. 막대에 마우스를 올리면 값이 떠요.")

    # ── ④ 자동 인사이트 (항목23 · 260821: 카테고리별 매장 랭킹 위로 이동 · 부진채널=1·2부리그 한정) ──
    def _render_insights():
        """자동 인사이트 렌더 — 아이템그룹 증감표(ig)를 반환(아래 ④ 차트에서 재사용)."""
        def _topdiff(cur, prev, col):
            a = cur.groupby(col, observed=True)["_매출액"].sum()
            b = prev.groupby(col, observed=True)["_매출액"].sum()
            m = pd.concat([a, b], axis=1, keys=["c", "p"]).fillna(0.0)
            m = m[(m["c"] != 0) | (m["p"] != 0)]
            m["d"] = m["c"] - m["p"]
            return m.sort_values("d", ascending=False)

        def _ent(name, r):
            pct = f" ({r['d']/r['p']*100:+.1f}%)" if r["p"] else " (신규)"
            color = "#1f8a4c" if r["d"] >= 0 else "#c62828"
            return f"<b>{name}</b> <span style='color:{color};font-weight:700'>{r['d']/1e6:+,.1f}</span>{pct}"

        lines = []
        if chan and not prev_y.empty:
            # 채널 인사이트도 쿠팡토탈·네이버토탈 통합 기준으로 계산
            chdf = pd.DataFrame([(k, e["c"], e["p"], e.get("lg")) for k, e in chan.items()],
                                columns=["k", "c", "p", "lg"]).set_index("k")
            chdf = chdf[(chdf["c"] != 0) | (chdf["p"] != 0)]
            chdf["d"] = chdf["c"] - chdf["p"]
            ch = chdf.sort_values("d", ascending=False)
            ups = ch[ch["d"] > 0].head(2)
            # 부진 채널(항목23 · 260821): 1부리그·2부리그 매장만 대상 — 꿈나무리그·리그 미지정 제외
            _has_lg = ch["lg"].notna().any()
            ch_dn = ch[ch["lg"].isin(["1부리그", "2부리그"])] if _has_lg else ch
            dns = ch_dn[ch_dn["d"] < 0].tail(2).iloc[::-1]
            if not ups.empty:
                lines.append("<b>성장 견인 채널</b>: " + " · ".join(_ent(str(i), r) for i, r in ups.iterrows()))
            if not dns.empty:
                lines.append("<b>부진 채널</b>" + ("(1·2부리그)" if _has_lg else "") + ": "
                             + " · ".join(_ent(str(i), r) for i, r in dns.iterrows()))
        ig = _topdiff(cur_y, prev_y, "아이템그룹") if (not prev_y.empty and "아이템그룹" in d.columns) else pd.DataFrame()
        if not ig.empty:
            top, bot = ig.iloc[0], ig.iloc[-1]
            seg = "<b>아이템</b>: " + _ent(str(ig.index[0]), top) + " 성장 1위"
            if bot["d"] < 0:
                seg += " / " + _ent(str(ig.index[-1]), bot) + " 부진 1위"
            lines.append(seg)
        if "연차" in d.columns and ry and rpy:
            s26 = float(cur_y[cur_y["연차"].isin(["신상", "내년신상"])]["_매출액"].sum()) / ry
            s25 = float(prev_y[prev_y["연차"].isin(["신상", "내년신상"])]["_매출액"].sum()) / rpy
            lines.append(f"<b>연차 구성</b>: 신상 비중 <b>{s26*100:.1f}%</b> "
                         f"(전년 {s25*100:.1f}%, {(s26-s25)*100:+.1f}%p)")
        if pg26 is not None and pg25 is not None:
            dv = (pg26 - pg25) * 100
            if dv <= -1:
                lines.append(f"<b>주의</b>: 누계 판가율 <span style='color:#c62828;font-weight:700'>{dv:+.1f}%p</span>"
                             " — 매출 대비 할인 폭이 커지는 추세")
            else:
                lines.append(f"<b>판가율</b>: 전년비 {dv:+.1f}%p")
        if lines:
            st.markdown("### 📌 자동 인사이트")
            body = "<br>".join(f"{i+1}. {t}" for i, t in enumerate(lines))
            st.markdown("<div style='background:#f7f9fc;border:1px solid #dde6f0;border-radius:8px;"
                        f"padding:12px 16px;font-size:0.9rem;line-height:1.9;'>{body}</div>",
                        unsafe_allow_html=True)

        return ig

    _ins_done = False

    # ── ③ 채널 랭킹 (2026-07-31 목업 v3 컨펌 + 수정: 연간/월간 토글 · '26년 미운영' 제외) ──
    #    쿠팡토탈(SD185+SD184)·네이버토탈(SD165+SD174) 합산 = 대시보드 채널 공통 기준(인사이트 포함)
    #    '26년 미운영' 매장(마스터 담당자 속성)은 운영 종료 → 랭킹·채널 인사이트 모두 제외
    chan = {}
    if "매장코드" in d.columns:
        if "매장명" in d.columns:
            _nm = d[["매장코드", "매장명"]].astype(str)
            name_of = dict(zip(_nm["매장코드"].str.strip(), _nm["매장명"]))
        else:
            name_of = {}
        mast = load_master()
        has_league = (mast is not None) and (not mast.empty) and ("리그구분" in mast.columns)
        lg_of_code, excl_codes = {}, set()
        if mast is not None and not mast.empty:
            for _, mr in mast.iterrows():
                _code = str(mr.get("매장코드", "")).strip()
                if has_league:
                    _lg = str(mr.get("리그구분", "")).strip()
                    if _lg and _lg.lower() not in ("nan", "none"):
                        lg_of_code[_code] = _lg
                if str(mr.get("담당자", "")).strip() == "26년 미운영":
                    excl_codes.add(_code)   # 운영 종료 매장 제외

        def _build_chan(fc, fp):
            """기간(fc=올해, fp=전년 동기간)별 채널 집계 — 통합·리그·배지 지표 포함."""
            c6 = fc["매장코드"].astype(str).str.strip()
            c5 = fp["매장코드"].astype(str).str.strip()
            _cs = fc.groupby(c6, observed=True)["_매출액"].sum()
            _ps = fp.groupby(c5, observed=True)["_매출액"].sum()
            _os = fc.groupby(c6, observed=True)["_최초가매출"].sum()   # 판가율 1위용
            _qs = fc.groupby(c6, observed=True)["_수량"].sum()         # 평균단가 1위용
            if "연차" in fc.columns:                                     # 신상판매 1위용
                _si = fc[fc["연차"].isin(["신상", "내년신상"])]
                _ss = _si.groupby(_si["매장코드"].astype(str).str.strip(), observed=True)["_매출액"].sum()
            else:
                _ss = pd.Series(dtype="float64")
            out = {}
            for c in set(_cs.index) | set(_ps.index):
                if c in excl_codes:
                    continue
                key = CH_MERGE.get(c, name_of.get(c, c))
                e = out.setdefault(key, {"c": 0.0, "p": 0.0, "o": 0.0, "q": 0.0, "s": 0.0, "codes": []})
                e["c"] += float(_cs.get(c, 0.0))
                e["p"] += float(_ps.get(c, 0.0))
                e["o"] += float(_os.get(c, 0.0))
                e["q"] += float(_qs.get(c, 0.0))
                e["s"] += float(_ss.get(c, 0.0))
                e["codes"].append(c)
            for key, e in out.items():
                members = sorted(e["codes"], key=lambda x: -float(_cs.get(x, 0.0)))
                e["lg"] = next((lg_of_code[x] for x in members if x in lg_of_code), None)
            return out

        t1, t2 = st.columns([3, 2])
        t1.markdown("### 🏟️ 채널 랭킹")
        mode = t2.radio("랭킹 기준", ["연간랭킹", "월간랭킹"], index=1, horizontal=True,
                        label_visibility="collapsed", key="dash_rank_mode")   # 기본=월간랭킹(항목23 · 260821)
        fcur, fprev = (cur_y, prev_y) if mode == "연간랭킹" else (cur_m, prev_m)
        rng_txt = (f"연간누계 {y_start.date()} → {asof.date()}" if mode == "연간랭킹"
                   else f"당월 {m_start.date()} → {asof.date()}")
        chan_rank = _build_chan(fcur, fprev)
        # 자동 인사이트의 채널 계산은 토글과 무관하게 항상 '연간누계' 기준 유지
        chan = chan_rank if mode == "연간랭킹" else _build_chan(cur_y, prev_y)

        if not has_league:
            st.info("매장 마스터에 **'리그구분'** 컬럼(값: 1부리그/2부리그/꿈나무리그)을 추가해 "
                    "업로드하면 리그 보드가 채워져요. 사이드바 → 매장 기준정보 업로드.")
        else:
            cols3 = st.columns(3)
            for (lg_name, icon, hcolor, subtitle), colx in zip(LEAGUES, cols3):
                entries = [(k, e) for k, e in chan_rank.items() if e.get("lg") == lg_name and e["c"] > 0]
                entries.sort(key=lambda t: -t[1]["c"])
                with colx:
                    st.markdown(_league_board_html(lg_name, icon, hcolor, subtitle, entries),
                                unsafe_allow_html=True)
            n_un = sum(1 for e in chan_rank.values() if e["c"] > 0 and not e.get("lg"))
            note = (f"**{mode}** ({rng_txt} · 전년 동기간 비교) · 리그당 매출 있는 채널 11위까지 기본 노출"
                    "(12위 이하는 펼쳐보기) · 막대=리그 내 상대 크기 · 순위변동=전년 동기간 리그 내 순위 대비 · "
                    "쿠팡토탈=SD185+SD184 · 네이버토탈=SD165+SD174 합산 · '26년 미운영' 매장 제외 · "
                    "리그별 1위 배지 4종: ⚡성장(전년비)·💎판가율·💰평균단가·🆕신상판매(신상+내년신상)")
            if n_un:
                note += f" · 리그 미지정 {n_un}개 채널은 랭킹 미포함(마스터에 리그구분 입력 시 반영)"
            st.caption(note)

        ig = _render_insights(); _ins_done = True   # 자동 인사이트 → 카테고리별 매장 랭킹 위쪽에 표시

        # ── ③-2 카테고리별 매장 랭킹 (2026-08-03 추가): 채널 랭킹과 동일한 토글(연간/월간)·기간에 연동 ──
        #    슈트류(중카테고리) · 온라인셔츠(브랜드 J 전체) · FW+NT+BE(신발+넥타이+벨트) — 각각 매출 많은 매장 TOP10.
        #    순위 기준=실판매금액. 채널 랭킹과 동일하게 쿠팡토탈·네이버토탈 합산, '26년 미운영' 매장 제외.
        if "아이템그룹" in d.columns and "브랜드명" in d.columns:
            st.markdown("### 🏆 카테고리별 매장 랭킹")
            scols = st.columns(3)
            for (stitle, sicon, shcolor, ssubtitle, smask), scol in zip(STORE_RANK_BOARDS, scols):
                sentries = _build_store_rank_entries(fcur, fprev, smask, name_of, excl_codes)
                with scol:
                    st.markdown(_store_rank_board_html(stitle, sicon, shcolor, ssubtitle, sentries),
                                unsafe_allow_html=True)
            st.caption(f"**{mode}** ({rng_txt} · 전년 동기간 비교) · 순위 기준=실판매금액 · "
                       "매출 상위 10개 매장만 노출 · 막대=보드 내 상대 크기 · "
                       "쿠팡토탈=SD185+SD184 · 네이버토탈=SD165+SD174 합산 · '26년 미운영' 매장 제외.")

    if not _ins_done:
        ig = _render_insights()

    # ── ④ 아이템그룹 증감액 + 연차 구성 변화 ──
    c1, c2 = st.columns(2)
    with c1:
        st.markdown(f"**아이템그룹별 전년비 증감액 (누계 · {sc} vs {sy})**")
        if not ig.empty:
            names = [str(i) for i in ig.index]
            vals = [float(v) / 1e6 for v in ig["d"]]
            txt = [f"{v:+,.1f}" + (f" ({r['d']/r['p']*100:+.1f}%)" if r["p"] else "")
                   for v, (_, r) in zip(vals, ig.iterrows())]
            figb = go.Figure(go.Bar(
                x=vals[::-1], y=names[::-1], orientation="h",
                marker_color=["#1f8a4c" if v >= 0 else "#c62828" for v in vals[::-1]],
                text=txt[::-1], textposition="outside"))
            figb.update_layout(height=340, margin=dict(t=10, b=0, l=0, r=10), xaxis_title="증감액(백만)")
            st.plotly_chart(figb, use_container_width=True)
        else:
            st.info("전년 데이터가 있어야 증감 분석이 가능해요.")
    with c2:
        st.markdown("**연차 구성 변화 (매출 비중)**")

        def _bucket(a):
            if a is None or (isinstance(a, float) and pd.isna(a)):
                return None
            if a in ("신상", "내년신상"):
                return "신상"
            if a in ("1년차", "2년차"):
                return a
            return "3년차↑"

        rows = []
        for lbl, f, tot in ((f"{sy}년", prev_y, rpy), (f"{sc}년", cur_y, ry)):
            if tot and "연차" in f.columns:
                sser = f.assign(_b=f["연차"].map(_bucket)).groupby("_b")["_매출액"].sum() / tot
                rows.append((lbl, sser))
        if rows:
            order = ["신상", "1년차", "2년차", "3년차↑"]
            colors = {"신상": "#2f6bb0", "1년차": "#5a8ec7", "2년차": "#8fb3d9", "3년차↑": "#c3d4e6"}
            figc = go.Figure()
            for bkt in order:
                figc.add_bar(y=[r[0] for r in rows],
                             x=[float(r[1].get(bkt, 0)) * 100 for r in rows],
                             name=bkt, orientation="h", marker_color=colors[bkt],
                             text=[f"{float(r[1].get(bkt, 0))*100:.1f}%" for r in rows],
                             textposition="inside")
            figc.update_layout(barmode="stack", height=340, margin=dict(t=10, b=0, l=0, r=0),
                               legend=dict(orientation="h", y=1.15), xaxis_title="비중(%)")
            st.plotly_chart(figc, use_container_width=True)
        else:
            st.info("전년 데이터가 있어야 구성 비교가 가능해요.")

    # (기존 '채널 TOP10' 차트는 채널 리그 랭킹 보드가 대체 — 2026-07-31)


# ── B. 채널별 추세 분석 (260831, 중태님 요청) ─────────────────────────────────
# A. 유통채널별 표 아래 신설(기존 브랜드별 표는 C로 재번호). 최근 X구간(주간/월간 토글)의
# 흐름으로 "추세가 좋은 매장 / 나쁜 매장"을 두 가지 잣대로 나눠 보여준다:
#   (1) 매출 순위 변동 — 구간마다 대상 매장끼리 실판매금액 순위(1위=최대)를 매기고,
#       전반부 평균순위 − 후반부 평균순위(+면 순위가 올라가는 중 = 개선).
#   (2) 매출 추세 기울기 — 매출을 2구간 이동평균으로 스무딩한 곡선의 회귀 기울기를
#       매장 평균 매출로 나눈 %/구간 기준(+면 상승 추세). 260831 후속3에서 전년동기
#       신장율 방식(후속2)을 이 방식으로 교체 — 신규 매장 제외·% 노이즈 문제, 함수 안
#       주석 참고. 전년 데이터 불필요 → 신규 매장 포함. TOP 표 아래 이동평균 라인차트.
# 판정 방식(AskUserQuestion 확정, 260831): 전반부 vs 후반부 평균 비교.
#   후반부 = 뒤 k개 구간(k = "후반부 구간 수" 위젯, 260831 후속 — 디폴트 X//2, 1까지 좁혀
#   "가장 최근 흐름"만 볼 수 있음) / 전반부 = 후반부를 뺀 앞쪽 전체(X−k개).
# 출력(확정): 분석별 '개선 TOP N'·'악화 TOP N' 두 표(변화 0인 매장은 어느 쪽에도 안 나옴).
# 엑셀(룰11·13)은 TOP 잘라내기 전 '전체 대상 매장' 추세표(개선폭 내림차순)를 내려준다.
# 위젯은 조회 폼 밖 → 바꾸는 즉시 반영(A표 정렬·금액 필터와 동일 패턴). 기준 종료일 =
# 조회기간 종료일(e). 브랜드/연차/시즌/담당 필터는 A표와 동일하게 적용된 상태(base)를 쓴다.


def _trend_buckets(e, unit, n):
    """추세 분석용 최근 n개 구간 [(라벨, 시작, 끝)] — 과거→최근 순.

    주간: 조회 종료일 e로 끝나는 7일 단위 n개(라벨 = 그 주 종료일 "~MM/DD").
    월간: e가 속한 달을 마지막 구간(1일~e, 진행분)으로 하는 달력월 n개
    (라벨 = "M월", 연도가 다르면 "YY.M월").
    """
    e = pd.to_datetime(e).normalize()
    out = []
    if unit == "주간":
        for i in range(n):                       # i=0 이 가장 과거
            end = e - pd.Timedelta(days=7 * (n - 1 - i))
            out.append((f"~{end:%m/%d}", end - pd.Timedelta(days=6), end))
    else:
        cur_m0 = e.replace(day=1)
        for i in range(n):
            m0 = cur_m0 - pd.DateOffset(months=n - 1 - i)
            m_end = e if m0 == cur_m0 else (m0 + pd.DateOffset(months=1)
                                            - pd.Timedelta(days=1))
            lbl = f"{m0.month}월" if m0.year == e.year else f"{str(m0.year)[-2:]}.{m0.month}월"
            out.append((lbl, m0, m_end))
    return out


def _trend_matrices(base, buckets, with_prev=True):
    """구간별 매장 매출 매트릭스 (cur, prev) — index=매장(_채널), columns=구간 라벨.

    prev는 각 구간을 정확히 1년 전으로 시프트한 같은 날짜범위(앱 전역의 '전년 동기간'
    관행 — render_channel_brand 등과 동일). 올해 매출이 전혀 없는 매장(전년만 있는
    매장)은 대상에서 제외되도록 prev를 cur.index로 reindex한다.
    with_prev=False(260831 후속3)면 전년 계산을 건너뛰고 prev로 빈 프레임을 반환 —
    현재 화면(순위 변동·이동평균 기울기)은 전년 매출이 필요 없어서 51만 행 마스킹·집계
    X회를 아낀다(신장율류 분석을 되살릴 때만 True로 호출할 것).
    """
    cur_cols, prev_cols = {}, {}
    dt = base["_판매일"]
    for lbl, s, t in buckets:
        cur_cols[lbl] = base.loc[(dt >= s) & (dt <= t)].groupby("_채널")["_매출액"].sum()
        if with_prev:
            s1, t1 = s - pd.DateOffset(years=1), t - pd.DateOffset(years=1)
            prev_cols[lbl] = base.loc[(dt >= s1) & (dt <= t1)].groupby("_채널")["_매출액"].sum()
    cur = pd.DataFrame(cur_cols).fillna(0.0)
    if not with_prev:
        return cur, pd.DataFrame()
    prev = pd.DataFrame(prev_cols).reindex(cur.index).fillna(0.0) if not cur.empty \
        else pd.DataFrame(prev_cols)
    return cur, prev


def _trend_split(labels, back_n=None):
    """전반부/후반부 라벨 분할.

    back_n(후반부 구간 수, 260831 후속 — 중태님 요청) 지정 시: 뒤 back_n개가 후반부,
    나머지 앞 전체가 전반부 — "최근 1~2구간으로 좁혀서" 최신 흐름만 따로 보는 용도.
    (범위 밖 값은 1 ~ len-1로 클램프.)
    미지정(None) 시: 반반(h=len//2) — 앞 h개 vs 뒤 h개(홀수면 가운데 1개 제외).
    """
    if back_n:
        k = max(1, min(int(back_n), len(labels) - 1))
        return list(labels[:len(labels) - k]), list(labels[len(labels) - k:])
    h = max(1, len(labels) // 2)
    return list(labels[:h]), list(labels[-h:])


def _trend_render_table(disp, key, empty_msg):
    """추세 분석 소표 렌더 — perf_table과 같은 erp-tbl 룩. 단 첫 행은 G.TOTAL이 아니라
    그냥 1위 매장이므로, 표 고유 클래스로 _TBL_CSS의 첫 행 노란 강조(룰6)를 되돌린다."""
    if disp.empty:
        st.caption(empty_msg)
        return
    cls = _season_css_class(key)

    def _delta_color(col):
        return ["color:#c62828;font-weight:600" if str(v).startswith("-")
                else ("color:#1f8a4c;font-weight:600" if str(v).startswith("+") else "")
                for v in disp[col]]
    sty = disp.style.set_properties(**{"text-align": "right"})
    if "담당자" in disp.columns:
        sty = sty.set_properties(subset=pd.IndexSlice[:, ["담당자"]],
                                 **{"text-align": "left"})
    for _dc in ("증감", "전체 기울기", "최근 기울기"):
        if _dc in disp.columns:
            # 260831 버그수정: subset에 람다 안에서만 유효한 c를 쓰다 NameError(라이브 크래시)
            # — 함수 스코프 변수 _dc로 통일.
            sty = sty.apply(lambda s, c=_dc: _delta_color(c),
                            subset=pd.IndexSlice[:, [_dc]])
    _css = f"""
<style>
table.{cls} tbody tr:first-child th{{background:#fbfbfd !important;font-weight:600;}}
table.{cls} tbody tr:first-child td{{background:#fff !important;font-weight:400;}}
</style>
"""
    render_styled_table(sty, extra_class=cls, extra_css=_css)


def _render_channel_trend(base, e, chan_mgr):
    """B. 채널별 추세 분석 본체 — 위 블록 주석 참고. base=필터 적용된 데이터, e=조회 종료일."""
    st.markdown("### B. 채널별 추세 분석")
    tc1, tc2, tc3, tc4, tc5 = st.columns([1.1, 1.1, 1.1, 1.1, 1.1])
    _unit = tc1.selectbox("기간 단위", ("주간", "월간"), key="cb_tr_unit",
                          help="주간=조회 종료일로 끝나는 7일 단위, 월간=달력월"
                               "(마지막 달은 조회 종료일까지 진행분).")
    # 구간 수 위젯 key를 단위별로 분리(금액 필터의 기간별 key 분리와 같은 패턴) —
    # 단위를 전환하면 그 단위의 디폴트(주간 8·월간 3)로 시작하되, 각 단위에서 직접
    # 고친 값은 세션 안에서 따로 기억된다.
    if _unit == "주간":
        _n = tc2.number_input("최근 구간 수(주)", min_value=2, max_value=26, value=8,
                              step=1, key="cb_tr_n_wk")
    else:
        _n = tc2.number_input("최근 구간 수(개월)", min_value=2, max_value=12, value=3,
                              step=1, key="cb_tr_n_mo")
    _n = int(_n)
    # 260831 후속(중태님 요청): 후반부 구간 수 필터 — 후반부를 항상 반(X//2)으로 두지 않고,
    # 최근 1~2구간 등 원하는 만큼 좁혀서 "가장 최신 흐름"만 따로 비교할 수 있게.
    # 후반부 = 뒤 k개 구간, 전반부 = 나머지 앞 전체(X−k개). 디폴트 k=X//2(기존과 동일 감각).
    # 구간 수(X)를 줄여 기존에 기억된 k가 범위(1~X−1)를 벗어나면 위젯 생성 전에 클램프
    # (안 하면 Streamlit이 value out of range 에러를 낸다).
    _bk_key = "cb_tr_back_n"
    if _bk_key in st.session_state:
        try:
            st.session_state[_bk_key] = max(1, min(int(st.session_state[_bk_key]), _n - 1))
        except Exception:
            del st.session_state[_bk_key]
    _back_n = tc3.number_input("후반부 구간 수", min_value=1, max_value=_n - 1,
                               value=max(1, _n // 2), step=1, key=_bk_key,
                               help="① 순위 변동에서 후반부(최근 쪽)로 묶을 구간 수예요. "
                                    "1이면 가장 최근 1구간만, 2면 최근 2구간 평균을 후반부로 "
                                    "봐요. 전반부는 후반부를 뺀 앞쪽 전체 구간의 평균. "
                                    "(② 매출 추세 기울기에는 적용되지 않아요.)")
    _back_n = int(_back_n)
    _topn = tc4.number_input("TOP 매장 수", min_value=3, max_value=30, value=10,
                             step=1, key="cb_tr_topn")
    _thr = tc5.number_input("대상 최소금액(만원)", min_value=0, step=100, value=100,
                            key="cb_tr_min_amt",
                            help="최근 X구간 실판매금액 합계가 이 금액(만원) 미만인 매장은 "
                                 "추세 분석 대상에서 제외해요(소액 매장의 순위 노이즈 방지). "
                                 "0을 넣으면 매출 있는 전체 매장이 대상.")
    buckets = _trend_buckets(e, _unit, _n)
    labels = [b[0] for b in buckets]
    front, back = _trend_split(labels, _back_n)
    # 260831 후속3: 전년 매출은 더 이상 안 씀(①순위·②기울기 모두 올해 매출만 필요)
    cur, _ = _trend_matrices(base, buckets, with_prev=False)
    if cur.empty:
        st.info("현재 조회조건에서 추세를 계산할 매출 데이터가 없어요.")
        return
    tot = cur.sum(axis=1)
    keep = tot > 0
    if _thr > 0:
        keep &= tot >= _thr * 10000.0          # 만원 → 원
    cur_u = cur.loc[keep]
    if len(cur_u) < 2:
        st.info("추세 분석 대상 매장이 2개 미만이에요 — 최소금액 문턱을 낮추거나 기간을 조정해 보세요.")
        return
    st.caption(f"대상 {len(cur_u)}개 매장(최근 {_n}{'주' if _unit == '주간' else '개월'} 합계 "
               f"{_thr:,}만원 이상{', 0=전체' if not _thr else ''}) · "
               f"① 순위 변동 = 전반부 {len(front)}구간(앞) vs 후반부 {len(back)}구간(최근) 평균 비교 · "
               "② 매출 추세 = 2구간 이동평균 곡선의 기울기 · "
               f"기준 종료일 = 조회기간 종료일({pd.to_datetime(e):%m/%d}) · "
               "단위·구간 수·후반부·TOP·문턱은 바꾸는 즉시 반영 — 🔍 조회를 다시 누를 필요 없어요.")

    def _mgr(ch):
        return chan_mgr.get(str(ch), "")

    # ── (1) 매출 순위 변동 ────────────────────────────────────────────────────
    ranks = cur_u.rank(axis=0, ascending=False, method="min").astype(int)
    rf, rb = ranks[front].mean(axis=1), ranks[back].mean(axis=1)
    rd = rf - rb                                # +면 순위 상승(개선)

    def _fmt_d(v, suf=""):
        return "0.0" + suf if abs(v) < 1e-9 else f"{v:+.1f}{suf}"

    def _rank_disp(idx):
        rows = []
        for ch in idx:
            row = {"담당자": _mgr(ch)}
            for lbl in labels:
                row[lbl] = str(ranks.at[ch, lbl])
            row["전반평균"] = f"{rf[ch]:.1f}"
            row["후반평균"] = f"{rb[ch]:.1f}"
            row["증감"] = _fmt_d(rd[ch])
            rows.append(row)
        return pd.DataFrame(rows, index=pd.Index(idx, name="매장"))

    _imp_idx = sorted(rd[rd > 1e-9].index, key=lambda c: (-rd[c], rb[c]))
    _wor_idx = sorted(rd[rd < -1e-9].index, key=lambda c: (rd[c], rb[c]))
    _all_rank_idx = sorted(rd.index, key=lambda c: (-rd[c], rb[c]))
    rh1, rh2 = st.columns([4, 1])
    rh1.markdown("#### ① 매출 순위 변동 — 구간별 매출 순위(1위=최대)의 흐름")
    rh2.download_button("⬇ 엑셀",
                        styled_excel_bytes(_rank_disp(_all_rank_idx), "매출순위 변동",
                                           first_row_total=False),
                        file_name="채널추세_매출순위변동.xlsx", mime=XLSX_MIME,
                        key="dl_cb_tr_rank", use_container_width=True)
    rc1, rc2 = st.columns(2)
    with rc1:
        st.markdown(f"**🔼 순위 개선 TOP {min(_topn, len(_imp_idx))}** — 전반부보다 후반부 순위가 올라간 매장")
        _trend_render_table(_rank_disp(_imp_idx[:_topn]), "cb_tr_rank_up",
                            "순위가 개선된 매장이 없어요.")
    with rc2:
        st.markdown(f"**🔽 순위 악화 TOP {min(_topn, len(_wor_idx))}** — 전반부보다 후반부 순위가 내려간 매장")
        _trend_render_table(_rank_disp(_wor_idx[:_topn]), "cb_tr_rank_dn",
                            "순위가 악화된 매장이 없어요.")
    st.caption("숫자는 각 구간의 매출 순위(대상 매장끼리, 1위=최대) · 증감 = 전반평균 − 후반평균 "
               "(+면 순위 상승) · 증감 0인 매장은 어느 표에도 안 나와요 · 엑셀엔 전체 대상 매장이 담겨요.")

    # ── (2) 매출 추세 기울기 (2구간 이동평균) ────────────────────────────────
    # 260831 후속3(중태님 제안·확정): 전년동기 신장율 방식(후속2)을 이 분석으로 교체.
    # 교체 배경 — 신장율 방식은 ⓐ 전년 데이터가 없는 신규 매장이 통째로 판정에서 빠지고
    # (실화면 대상 23개 중 11개 제외 — 정작 추세가 제일 궁금한 매장들), ⓑ % 숫자투성이라
    # 팀원이 읽기 어려웠음. 새 방식(중태님 아이디어 + 설계 확정):
    #   ⓐ 매장별 구간 매출을 연속 2구간 이동평균으로 스무딩 → X−1개 점의 곡선
    #      (한 구간 튀는 매출이 절반으로 눌려 곡선이 매끈해짐).
    #   ⓑ 전체 기울기 = 그 곡선 전체의 회귀(추세선) 기울기 ÷ 매장 평균 매출 → %/구간.
    #      매장 크기와 무관하게 '가파름'을 비교하기 위한 정규화 — 안 하면 큰 매장이 항상 이김.
    #   ⓒ 최근 기울기 = (마지막 점 − 직전 점) ÷ 평균 매출 — 최신 모멘텀 참고용 컬럼.
    #   정렬(AskUserQuestion 확정) = 전체 기울기(가장 노이즈에 강한 기준). 상승/하락 TOP
    #   표 아래에 이동평균 곡선 라인차트를 같이 그림(기울기를 눈으로 확인).
    # 전년 데이터 불필요 → 신규 매장 포함 전체가 판정 대상. 평균 매출 ≤ 0(반품 초과 등
    # 희귀 케이스)만 판정 불가 "–". 시즌 효과(가을 진입 등 전 매장 동반 상승)도 '상승'으로
    # 잡히는 관점임을 캡션에 명시 — 전년 대비 관점은 A표 증감율이 담당.
    ma = cur_u.T.rolling(2).mean().T.iloc[:, 1:]     # X−1개 점, 라벨 = 뒤쪽 구간
    ma_lbls = list(ma.columns)
    sh1, sh2 = st.columns([4, 1])
    sh1.markdown("#### ② 매출 추세 기울기 — 2구간 이동평균 곡선이 오르는/내리는 기세")
    if len(ma_lbls) < 2:
        st.info("이동평균 기울기를 계산하려면 최근 구간 수를 3 이상으로 해주세요.")
        return
    _m = len(ma_lbls)
    _tc = np.arange(_m) - (_m - 1) / 2.0             # 시점 중심화 → 회귀 기울기 벡터 계산
    slope_abs = ma.mul(_tc, axis=1).sum(axis=1) / float((_tc ** 2).sum())   # 원/구간
    _lv = cur_u.mean(axis=1)
    _lv = _lv.where(_lv > 0)                         # 평균 매출 ≤ 0 → 판정 불가
    slope = slope_abs / _lv * 100.0                  # 전체 기울기 (%/구간)
    recent = (ma[ma_lbls[-1]] - ma[ma_lbls[-2]]) / _lv * 100.0   # 최근 기울기 (%/구간)
    _s_valid = slope.notna()

    def _fmt_s(v):
        return "–" if pd.isna(v) else _fmt_d(v, "%")

    def _slope_disp(idx):
        rows = []
        for ch in idx:
            row = {"담당자": _mgr(ch)}
            for lbl in ma_lbls:
                row[lbl] = f"{_mm(ma.at[ch, lbl]):,.1f}"
            row["전체 기울기"] = _fmt_s(slope[ch])
            row["최근 기울기"] = _fmt_s(recent[ch])
            rows.append(row)
        return pd.DataFrame(rows, index=pd.Index(idx, name="매장"))

    def _slope_chart(idx, title, key):
        fig = go.Figure()
        for ch in idx:
            fig.add_trace(go.Scatter(x=ma_lbls, y=[_mm(ma.at[ch, l]) for l in ma_lbls],
                                     mode="lines+markers", name=str(ch)))
        fig.update_layout(height=340, margin=dict(l=10, r=10, t=34, b=10),
                          title=dict(text=title, font=dict(size=13)),
                          yaxis_title="이동평균 매출(백만)",
                          legend=dict(orientation="h", yanchor="top", y=-0.15,
                                      font=dict(size=10)))
        st.plotly_chart(fig, use_container_width=True, key=key)

    sv = slope[_s_valid]
    _sup_idx = sorted(sv[sv > 1e-9].index, key=lambda c: (-slope[c], -recent[c]))
    _sdn_idx = sorted(sv[sv < -1e-9].index, key=lambda c: (slope[c], recent[c]))
    _all_s_idx = (sorted(sv.index, key=lambda c: (-slope[c], -recent[c]))
                  + sorted(slope.index[~_s_valid]))
    sh2.download_button("⬇ 엑셀",
                        styled_excel_bytes(_slope_disp(_all_s_idx), "매출 추세 기울기",
                                           first_row_total=False),
                        file_name="채널추세_기울기.xlsx", mime=XLSX_MIME,
                        key="dl_cb_tr_slope", use_container_width=True)
    _s_excl = int((~_s_valid).sum())
    st.caption(f"판정 대상 {int(_s_valid.sum())}개 매장 — 전년 데이터가 필요 없어 "
               "올해 신규 입점 매장도 전부 포함돼요"
               + (f" (판정 불가 {_s_excl}개 — 평균 매출이 0 이하, 엑셀 맨 아래 \"–\")."
                  if _s_excl else "."))
    sc1, sc2 = st.columns(2)
    with sc1:
        st.markdown(f"**🔼 상승 추세 TOP {min(_topn, len(_sup_idx))}** — 이동평균 곡선이 가장 가파르게 오르는 매장")
        _trend_render_table(_slope_disp(_sup_idx[:_topn]), "cb_tr_sl_up",
                            "상승 추세 매장이 없어요.")
        if _sup_idx:
            _slope_chart(_sup_idx[:_topn], "상승 추세 TOP — 이동평균 곡선", "ch_cb_tr_sl_up")
    with sc2:
        st.markdown(f"**🔽 하락 추세 TOP {min(_topn, len(_sdn_idx))}** — 이동평균 곡선이 가장 가파르게 내리는 매장")
        _trend_render_table(_slope_disp(_sdn_idx[:_topn]), "cb_tr_sl_dn",
                            "하락 추세 매장이 없어요.")
        if _sdn_idx:
            _slope_chart(_sdn_idx[:_topn], "하락 추세 TOP — 이동평균 곡선", "ch_cb_tr_sl_dn")
    st.caption("표 숫자 = 연속 2구간 이동평균 매출(백만원, 각 라벨은 뒤쪽 구간 기준) · "
               "전체 기울기 = 이동평균 곡선 전체 추세선의 기울기 ÷ 그 매장 평균 매출(%/구간 — "
               "매장 크기와 무관하게 가파름 비교, 정렬 기준) · 최근 기울기 = 마지막 두 점 차이 "
               "기준 최신 모멘텀(참고용) · 기울기 0인 매장은 어느 표에도 안 나와요 · "
               "※ 시즌 흐름으로 전 매장이 같이 오르내리는 것도 '추세'로 잡혀요 — 전년 대비 "
               "관점은 A표 증감율로 확인하세요.")


def render_channel_brand(df):
    """매주 대표님 보고 B: 유통채널별 · 브랜드별 매출현황 (전년 동기간 비교)."""
    st.subheader("📈 유통별 세부 분석 (전년 동기간 비교)")
    if df.empty or "_판매일" not in df.columns or df["_판매일"].notna().sum() == 0:
        st.info("데이터를 먼저 적재하세요.")
        return
    d = df[df["_판매일"].notna()].copy()
    # 매장 담당자 매핑 (매장 마스터의 담당자 기준) — 담당자 컬럼·담당 필터용
    master = load_master()
    if not master.empty and "담당자" in master.columns:
        _mgr_map = dict(zip(master["매장코드"].astype(str).str.strip(),
                            master["담당자"].astype(str).str.strip()))
        d["_담당자"] = d["매장코드"].astype(str).str.strip().map(_mgr_map)
    else:
        d["_담당자"] = None
    dmin, dmax = d["_판매일"].min().date(), d["_판매일"].max().date()
    default_start = (pd.to_datetime(dmax) - pd.Timedelta(days=6)).date()

    st.caption("올해 vs 전년 '동기간'(같은 날짜범위) 비교 · 금액 백만원 · 판가율=실판가÷최초가(가중) · 기본기간=최근 1주")
    # ── 조건 폼 (2026-08-06): 조건 변경 중엔 계산 안 함, 🔍 조회 때 1번만 ──
    with st.form("cb_form"):
        rng = st.date_input("조회기간 (기본: 최근 1주)", value=(default_start, dmax),
                            min_value=dmin, max_value=dmax, key="cb_rng")
        # 공통 필터 (주간보고 방식) — 브랜드별 → 연차별 → 시즌별 → 매장 담당 · 빈칸=전체
        cb1, cb2, cb3, cb4 = st.columns(4)
        brands = sorted(d["브랜드명"].dropna().unique()) if "브랜드명" in d.columns else []
        ages = sorted(d["연차"].dropna().unique(), key=_age_sort_key) if "연차" in d.columns else []
        seasons = sorted(d["시즌명"].dropna().unique()) if "시즌명" in d.columns else []
        _mans = sorted({str(m).strip() for m in d["_담당자"].dropna().astype(str)
                        if str(m).strip() and str(m).strip().lower() not in ("nan", "none")})
        # 260807 추가: 유통채널별 표 미리보기용 '담당자별 TOTAL' 행 스켈레톤(실계산 없음)
        # ※ '26년 미운영'·'직원구매'는 담당 필터(_mans)엔 남기되, TOTAL 행 스켈레톤에서는 제외
        #   (실제 조회 시의 _ch_mans 로직과 동일하게 맞춤).
        _CH_MGR_TOTAL_EXCL = {"26년 미운영", "직원구매"}
        _preview_mgr_rows = [
            (f"{m} TOTAL", (lambda name: (lambda x: x["_담당자"].astype(str).str.strip() == name))(m))
            for m in _mans if m not in _CH_MGR_TOTAL_EXCL]
        selb = cb1.multiselect("브랜드별", brands, default=[], placeholder="전체", key="cb_brand")
        sela = cb2.multiselect("연차별", ages, default=[], placeholder="전체", key="cb_age")
        sels = cb3.multiselect("시즌별", seasons, default=[], placeholder="전체", key="cb_season")
        selm = cb4.multiselect("매장 담당", _mans, default=[], placeholder="전체", key="cb_mgr")
        if not _mans:
            st.caption("※ 매장 기준정보(담당자)가 없어 담당자 컬럼·필터가 비어 있어요 — 사이드바에서 매장 기준정보를 업로드하면 채워져요.")
        run = st.form_submit_button("🔍 조회", type="primary")
    if _need_search("cb_go", run):
        # 2026-08-06: 조회 전에도 표 헤더(구조)만 미리 보여줌 — 실제 계산 없음(perf_table 참고).
        _empty = pd.DataFrame()
        perf_table(_empty, _empty, "_채널", None, "유통채널별 매출현황", "cb_ch_preview",
                   extra=("담당자", {}), month=(_empty, _empty),
                   blk_labels=("조회기간", "연간누계"), extra_rows=_preview_mgr_rows, preview=True)
        # 260831: B. 채널별 추세 분석은 실계산이 필요해 미리보기 스켈레톤이 없음 — 안내만.
        st.markdown("### B. 채널별 추세 분석")
        st.caption("최근 X주/X개월의 매출 순위 변동과 이동평균 매출 추세 기울기로 추세가 좋은 "
                   "매장과 나쁜 매장을 나눠 보여드려요 — 🔍 조회 후 표시돼요.")
        perf_table(_empty, _empty, "브랜드명", None, "브랜드별 매출현황", "cb_br_preview",
                   preview=True)
        return

    if not (isinstance(rng, (list, tuple)) and len(rng) == 2):
        st.info("기간(시작~끝)을 선택하세요.")
        return
    s, e = pd.to_datetime(rng[0]), pd.to_datetime(rng[1])
    cy_cb = int(e.year)   # 2026-08-07 버그수정: 표 연도 라벨을 조회기간 종료일 기준으로 동적 계산
    base = d
    if selb and "브랜드명" in base.columns:
        base = base[base["브랜드명"].isin(selb)]
    if sela and "연차" in base.columns:
        base = base[base["연차"].isin(sela)]
    if sels and "시즌명" in base.columns:
        base = base[base["시즌명"].isin(sels)]
    if selm and "_담당자" in base.columns:
        base = base[base["_담당자"].astype(str).str.strip().isin(selm)]
    cur = base[(base["_판매일"] >= s) & (base["_판매일"] <= e)]
    prev = base[(base["_판매일"] >= s - pd.DateOffset(years=1)) & (base["_판매일"] <= e - pd.DateOffset(years=1))]
    # 260807 추가(중태님 지시): 유통채널별 표에 '연간누계' 블록을 조회기간 오른쪽에 병기
    y_start = e.replace(month=1, day=1)
    cur_y = base[(base["_판매일"] >= y_start) & (base["_판매일"] <= e)]
    prev_y = base[(base["_판매일"] >= y_start - pd.DateOffset(years=1)) & (base["_판매일"] <= e - pd.DateOffset(years=1))]

    tot_c, tot_p = cur["_매출액"].sum(), prev["_매출액"].sum()
    k1, k2, k3 = st.columns(3)
    k1.metric("기간 매출(백만)", f"{_mm(tot_c):,.0f}")
    k2.metric("전년 동기(백만)", f"{_mm(tot_p):,.0f}")
    g = ((tot_c - tot_p) / tot_p) if tot_p else None
    k3.metric("전년비 신장률", "신규/–" if g is None else f"{g*100:+.1f}%")
    if not tot_p:
        st.warning("전년 동기간 데이터가 없어요. 기간을 조정하거나 전년 로우데이터를 적재하세요.")

    # ── 260824(중태님 요청): A표 정렬 필터 — 기간(연간누계/조회기간) × 지표(실판매금액/
    #    매출 증감율/판가율) × 방향(내림/오름). 조회 폼 밖·표 바로 위에 두어 🔍 조회를 다시
    #    누르지 않아도 바꾸는 즉시 재정렬된다(행 순서만 바뀌고 숫자·집계는 그대로).
    #    기본값(연간누계·실판매금액·내림차순)은 기존 표와 100% 동일한 순서.
    _CB_SORT_METRICS = {"실판매금액": ("실판매금액(백만)", "CUR"),
                        "매출 증감율": ("실판매금액(백만)", "증감율"),
                        "판가율": ("판가율", "CUR")}
    sc1, sc_amt, sc2, sc3, _sc_sp = st.columns([1.1, 1.1, 1.1, 1.1, 1.1])
    _s_blk = sc1.selectbox("정렬 기준 기간", ("연간누계", "조회기간"), key="cb_sort_blk")
    # 260831(중태님 요청): 금액 필터 — 정렬 기준 기간의 실판매금액이 이 값(만원) 미만인 개별
    # 매장 행은 숨김(G.TOTAL·담당자별 TOTAL 등 집계 행과 합계 숫자는 그대로).
    # 디폴트: 연간누계 1,000만원 / 조회기간 100만원. 위젯 key를 기간별로 분리해서, 기간을
    # 전환하면 그 기간의 디폴트로 시작하되 각 기간에서 직접 고친 값은 세션 안에서 따로
    # 기억된다(전환한다고 수정값이 날아가지 않음). 0을 넣으면 필터 해제(전체 표시).
    _s_amt = sc_amt.number_input(
        "금액 필터(만원 이상만 보기)", min_value=0, step=100,
        value=1000 if _s_blk == "연간누계" else 100,
        key=f"cb_min_amt_{'yr' if _s_blk == '연간누계' else 'per'}",
        help="정렬 기준 기간의 실판매금액이 이 금액(만원) 이상인 매장만 표시해요. "
             "0을 넣으면 전체 표시. G.TOTAL·담당자별 TOTAL과 합계 숫자는 그대로예요.")
    _s_met = sc2.selectbox("정렬 지표", list(_CB_SORT_METRICS), key="cb_sort_met")
    _s_dir = sc3.selectbox("정렬 방향", ("내림차순 ↓", "오름차순 ↑"), key="cb_sort_dir")
    _s_asc = _s_dir.startswith("오름")
    _cb_sort_spec = (_s_blk, *_CB_SORT_METRICS[_s_met], _s_asc)
    _cb_min_spec = (_s_blk, _s_amt / 100.0)   # 만원 → 백만 (표 내부 단위)
    _amt_tag = f" · {_s_blk} {_s_amt:,}만원 이상" if _s_amt else ""
    st.markdown(f"### A. 유통채널별 ({_s_blk} {_s_met} "
                f"{'오름차순 ↑' if _s_asc else '내림차순 ↓'}{_amt_tag})")
    # 매장명(행) → 담당자 매핑: 표 맨 앞 '담당자' 컬럼으로 표시
    _cm = d[["_채널", "_담당자"]].astype(str).drop_duplicates(subset=["_채널"])
    # 260805: pandas 3.x에서는 astype(str) 후에도 결측이 float(nan)으로 남아 .strip()이 터진다.
    #         매장 기준정보 미업로드 시 담당자가 전부 결측이므로 str()로 한 번 더 감싼다.
    chan_mgr = {c: ("" if str(m).strip().lower() in ("nan", "none", "") else str(m).strip())
                for c, m in zip(_cm["_채널"], _cm["_담당자"])}
    # 260807 추가(중태님 지시): G.TOTAL 아래 '담당자별 TOTAL' 행 삽입 — 현재 조회조건(base)에
    # 매출이 있는 매장의 담당자 전원, 연간누계(cur_y) 매출 큰 순으로 정렬(표 전체 정렬 기준과 통일).
    # 260807 추가 수정: '26년 미운영'·'직원구매'는 실제 담당자가 아니라 TOTAL 행 자체를 안 보여줌
    # (해당 매장은 여전히 G.TOTAL·개별 매장행엔 그대로 포함, 담당자별 TOTAL만 제외).
    _CH_MGR_TOTAL_EXCL = {"26년 미운영", "직원구매"}
    _ch_mans = sorted({str(m).strip() for m in base["_담당자"].dropna().astype(str)
                       if str(m).strip() and str(m).strip().lower() not in ("nan", "none")
                       and str(m).strip() not in _CH_MGR_TOTAL_EXCL})
    if "_담당자" in cur_y.columns and not cur_y.empty:
        _mgr_rev = cur_y.groupby(cur_y["_담당자"].astype(str).str.strip())["_매출액"].sum()
    else:
        _mgr_rev = pd.Series(dtype="float64")
    _ch_mans.sort(key=lambda m: -float(_mgr_rev.get(m, 0.0)))
    mgr_extra_rows = [
        (f"{m} TOTAL", (lambda name: (lambda x: x["_담당자"].astype(str).str.strip() == name))(m))
        for m in _ch_mans]
    perf_table(cur_y, prev_y, "_채널", None, "유통채널별 매출현황", "cb_ch",
               extra=("담당자", chan_mgr), month=(cur, prev), blk_labels=("조회기간", "연간누계"),
               extra_rows=mgr_extra_rows, cy=cy_cb, sort_spec=_cb_sort_spec,
               min_spec=_cb_min_spec)
    st.caption("※ 채널을 자사몰/외부몰 등 그룹으로 묶으려면 '채널 기준정보(매핑)'가 필요해요 — 준비되면 그룹 집계도 추가해드릴게요. "
               "G.TOTAL 아래 담당자별 TOTAL → 개별 매장 순으로, 두 구간 모두 위에서 고른 정렬 기준을 따라요 "
               "(값이 없는 행(–)은 항상 맨 아래). 정렬·금액 필터는 바꾸는 즉시 반영 — 🔍 조회를 다시 누를 필요 없어요. "
               "금액 필터에 걸러진 매장도 G.TOTAL·담당자별 TOTAL 합계엔 그대로 포함돼요(표시만 숨김). "
               "담당자 미지정 매장은 담당자별 TOTAL 어디에도 안 잡히지만 G.TOTAL엔 포함돼요.")

    # ── 260831(중태님 요청): B. 채널별 추세 분석 — A표 아래 신설, 기존 브랜드별은 C로 재번호 ──
    _render_channel_trend(base, e, chan_mgr)

    st.markdown("### C. 브랜드별")
    perf_table(cur, prev, "브랜드명", None, "브랜드별 매출현황", "cb_br", cy=cy_cb)


# ==============================================================================
# 복종별 판매비중 분석 (2026-08-08 신규, 같은 날 수정1~3) — 매장별 아이템(복종) 판매 구성비 현황표.
#   과거 "매장별 아이템 비중 분석"(반자동 Python 스크립트 + 엑셀 3시트 산출물) 프로젝트의 시트1
#   (원본+집계 데이터 표)만 이식한 것 — 시트2(베스트/워스트 정리)·시트3(매장별 제안 문구)은 이번
#   범위 밖(중태님 확정, 2026-08-08 "일단 첫번째 시트의 기본 복종별 판매 비중 현황표만").
#   카테고리 매핑은 그 스크립트의 자체 9분류(CATEGORY_MAP) 대신, 이 ERP의 상품마스터 기반
#   중카테고리/소카테고리(item_master, get_itemgroup_map 계열)를 그대로 재사용해서 재고 가공·
#   판매분석 화면과 기준을 통일했다(claude/쇼핑몰재고모니터링_전달사항_260803 원칙과 동일).
#   색상 규칙(중태님 확정, 2026-08-08):
#     ① SD065(온라인통합몰) 행 = 주황색 — 다른 매장이 참조하는 비교 기준선이라 항상 눈에 띄게.
#     ② 복종별 '금액' 열에서 매출 top5 매장(SD065·매장담당별평균 제외) = 녹색.
#     ③ 복종별 '%' 열에서 상위5(SD065·매장담당별평균 제외) = 빨간색 / 하위5(동일 제외) = 파란색.
#        (SD065·매장담당별평균은 '비교 기준선'이지 순위 경쟁 대상이 아니므로 ②·③ 모두에서 제외.)
#     ④ 기간 합계매출 100만원 미만 매장은 G.TOTAL 합계엔 포함하되 목록에는 표시하지 않음.
#   같은 날 후속 수정 3건(중태님):
#     [수정1] 복종 '금액' 열 헤더를 "슈트류 금액(백만)" → "슈트류"로 간결화(단위는 표 제목 옆
#             [금액: 백만원 / VAT+] 표기로 이미 안내됨).
#     [수정2] G.TOTAL과 개별 매장행 사이에 '{담당자} 평균' 행을 담당자마다 하나씩 삽입 — 그
#             담당자가 맡은 매장(표시 대상 중) 평균값. 하늘색(#d6f0fa, 유통별 세부분석의
#             담당자별 TOTAL 행과 동일 색)으로 구분, 순위 경쟁(①·②·③)에서는 제외.
#     [수정3] 복종(금액·% 2열 묶음) 사이 경계에 진한 회색 세로선을 넣어 복종별 구분을 명확히
#             (기존 block_border 재사용 — 룰12와 동일 메커니즘, 복종 수만큼 반복 호출).
#     [수정4] 위 ②(아이템별 매출 top5) 강조색을 분홍 → 녹색으로 변경(2026-08-09).
#   2026-08-09 추가 수정 2건:
#     [수정5] '%' 열 헤더를 "슈트류 %"처럼 복종명을 붙이지 않고 그냥 "%"로 간결화 — 표1·표2(아래
#             신규 YoY 표) 둘 다 적용. 내부 컬럼명(랭킹·구분선 계산에 쓰는 실제 키)은 그대로 두고
#             Styler.format_index()/엑셀 헤더 셀 값만 표시용으로 바꿔치기(로직에 영향 없음).
#     [수정6] "담당별 전년대비 복종 비중 변화" 표 신규 — 표1과 동일한 행 구성(G.TOTAL·매장담당별
#             평균·SD065·개별 매장)을 "{전년} 라벨"/"{올해} 라벨"/"ㄴ증감" 3행씩 쌓아 전년 동기간
#             대비 변화를 보여줌. 전년 동기간=조회기간을 그대로 1년 시프트(이 앱의 기존 YoY 관행과
#             동일). ㄴ증감 행: 합계·복종 금액 열=증감액(백만) (증감율%) 둘 다 표기, %열=%p 차이,
#             양수 초록/음수 빨강(이 앱의 기존 증감 색 관행과 동일). 순위는 연도별로 그 해 자체
#             매출 100만원 이상 매장 모집단 안에서 따로 계산(전년에 데이터 없으면 "–").
#     [수정7] "담당별" 필터 추가(브랜드/연차/시즌과 동일한 형태, 총 4개) — 상단 필터에서 담당자를
#             고르면 표1·표2 모두 그 담당자 소관 매장만으로 줄어드는 드릴다운(2026-08-09). 구현:
#             row-level(base/prev_base)에 _담당자 컬럼을 매장코드→담당자 매핑으로 미리 붙여두고
#             selm(멀티셀렉트)을 다른 필터와 동일하게 .isin()으로 적용 — 이후 G.TOTAL/shown/piv/
#             매장담당별 평균이 전부 자동으로 좁혀지므로 표1·표2 자체 로직은 수정 불필요.
# ==============================================================================
_CATMIX_SD065 = "SD065"
_CATMIX_FLOOR = 1_000_000    # 원 단위 — 기간 합계매출 이 미만인 매장은 목록에서 제외(총계엔 포함)
_CATMIX_GREEN = "background-color:#c8e6c9;font-weight:600"   # ② 아이템별 매출 top5
_CATMIX_RED = "background-color:#ffcdd2;font-weight:600"     # ③ 복종 비중 상위5
_CATMIX_BLUE = "background-color:#bbdefb;font-weight:600"    # ③ 복종 비중 하위5
_CATMIX_MGR_BG = "#d6f0fa"    # 수정2: 매장담당별 평균 행 — 유통별 세부분석의 담당자별 TOTAL과 동일 색


def _catmix_style(disp, num, sd_label, amt_cols, pct_cols, mgr_labels=None, n_meta=3):
    """복종별 판매비중 표 전용 Styler — 아이템별 top5(녹색)·복종비중 상하위5(빨강/파랑)·SD065(주황)·
    매장담당별 평균(하늘색, 수정2) · 복종 사이 진한 회색 구분선(수정3).

    disp=화면 표시용(포맷 문자열) DataFrame, num=랭킹 계산용 원본 숫자 DataFrame(동일 index·columns).
    G.TOTAL·SD065·매장담당별 평균 행은 순위 경쟁(top5/bottom5/녹색)에서 제외한다.
    n_meta=매장코드·합계·순위 등 복종 앞에 오는 메타 컬럼 수(복종 구분선 위치 계산용).
    """
    mgr_labels = mgr_labels or []
    excl = [lbl for lbl in (["G.TOTAL", sd_label] + list(mgr_labels)) if lbl and lbl in num.index]

    def _green(col):
        pool = num[col].drop(index=excl, errors="ignore")
        top = set(pool[pool > 0].nlargest(5).index)
        return [_CATMIX_GREEN if idx in top else "" for idx in num.index]

    def _redblue(col):
        pool = num[col].drop(index=excl, errors="ignore")
        top = set(pool.nlargest(5).index)
        bot = set(pool.nsmallest(5).index)
        out = []
        for idx in num.index:
            if idx in top:
                out.append(_CATMIX_RED)
            elif idx in bot:
                out.append(_CATMIX_BLUE)
            else:
                out.append("")
        return out

    sty = disp.style
    for col in amt_cols:
        sty = sty.apply(lambda s, c=col: _green(c), subset=pd.IndexSlice[:, [col]])
    for col in pct_cols:
        sty = sty.apply(lambda s, c=col: _redblue(c), subset=pd.IndexSlice[:, [col]])
    _mgr_in = [lbl for lbl in mgr_labels if lbl in disp.index]
    if _mgr_in:
        # 수정2: 매장담당별 평균 행 — 유통별 세부분석의 담당자별 TOTAL과 동일하게 하늘색만(볼드 없음)
        sty = sty.set_properties(subset=pd.IndexSlice[_mgr_in, :], **{"background-color": _CATMIX_MGR_BG})
    if sd_label and sd_label in disp.index:
        sty = sty.set_properties(subset=pd.IndexSlice[[sd_label], :],
                                  **{"background-color": "#ffe0b2", "font-weight": "700"})
    sty = sty.set_properties(**{"text-align": "right"})
    # 수정3: 복종(금액·% 2열 묶음) 사이 경계에 진한 회색 세로선 — 메타 컬럼 뒤부터 복종마다 반복
    for i in range(len(amt_cols)):
        sty = block_border(sty, n_meta + 2 * i)
    # 수정5: '%' 열 헤더를 "슈트류 %" → "%"로 간결화(내부 컬럼명은 그대로 유지 — 랭킹·구분선 계산에 영향 없음)
    _pct_set = set(pct_cols)
    sty = sty.format_index(lambda c: "%" if c in _pct_set else c, axis=1)
    return sty


def _catmix_excel_bytes(disp, sd_label, amt_cols, pct_cols, num, sheet="복종별판매비중",
                         mgr_labels=None, n_meta=3):
    """룰13: 화면 서식(G.TOTAL 노랑·SD065 주황·담당별평균 하늘색·아이템top5 녹색·복종비중 상하위5
    빨강파랑·복종 구분 진한 회색 세로선) 그대로 엑셀 반영."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    mgr_labels = mgr_labels or []
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        disp.to_excel(w, sheet_name=_safe_name(sheet)[:28] or "표")
        ws = w.book.worksheets[0]
        n_idx = disp.index.nlevels
        n_rows = len(disp)
        data_start = ws.max_row - n_rows + 1
        thin = Side(style="thin", color="D9D9D9")
        thick = Side(style="medium", color="555555")   # 수정3: 복종 구분 진한 회색 세로선
        head_fill = PatternFill("solid", fgColor="F4F4F6")
        idx_fill = PatternFill("solid", fgColor="FAFAFA")
        gt_fill = PatternFill("solid", fgColor="FFF2B8")     # G.TOTAL 노랑
        sd_fill = PatternFill("solid", fgColor="FFE0B2")     # ① SD065 주황
        mgr_fill = PatternFill("solid", fgColor="D6F0FA")    # 수정2: 매장담당별 평균 하늘색
        green_fill = PatternFill("solid", fgColor="C8E6C9")  # ② 아이템별 매출 top5
        red_fill = PatternFill("solid", fgColor="FFCDD2")    # ③ 복종비중 상위5
        blue_fill = PatternFill("solid", fgColor="BBDEFB")   # ③ 복종비중 하위5

        _mgr_set = set(mgr_labels)
        excl = [lbl for lbl in (["G.TOTAL", sd_label] + list(mgr_labels)) if lbl and lbl in num.index]
        top5_amt = {c: set(num[c].drop(index=excl, errors="ignore").pipe(lambda s: s[s > 0]).nlargest(5).index)
                    for c in amt_cols}
        top5_pct = {c: set(num[c].drop(index=excl, errors="ignore").nlargest(5).index) for c in pct_cols}
        bot5_pct = {c: set(num[c].drop(index=excl, errors="ignore").nsmallest(5).index) for c in pct_cols}
        bcols = {n_idx + (n_meta + 2 * i) + 1 for i in range(len(amt_cols))}   # 수정3: 진한 세로선 절대열번호

        for r in range(1, data_start):
            for k in range(1, ws.max_column + 1):
                cell = ws.cell(r, k)
                cell.fill = head_fill
                cell.font = Font(bold=True, color="111111")
                cell.alignment = Alignment(horizontal="center", vertical="center")
        # 수정5: '%' 열 헤더 셀 텍스트를 "슈트류 %" → "%"로 간결화(화면과 동일)
        _pct_set = set(pct_cols)
        for cj, col in enumerate(disp.columns):
            if col in _pct_set:
                ws.cell(data_start - 1, n_idx + 1 + cj).value = "%"

        for ri in range(n_rows):
            r = data_start + ri
            lbl = disp.index[ri]
            row_fill, row_bold = None, False
            if ri == 0:
                row_fill, row_bold = gt_fill, True
            elif lbl in _mgr_set:
                row_fill, row_bold = mgr_fill, False
            elif lbl == sd_label:
                row_fill, row_bold = sd_fill, True
            for k in range(1, n_idx + 1):
                c = ws.cell(r, k)
                c.fill = row_fill or idx_fill
                c.font = Font(bold=True, color="111111")
                c.alignment = Alignment(horizontal="left", vertical="center")
            for cj, col in enumerate(disp.columns):
                c = ws.cell(r, n_idx + 1 + cj)
                c.alignment = Alignment(horizontal="right", vertical="center")
                cell_fill = row_fill
                if col in amt_cols and lbl in top5_amt.get(col, ()):
                    cell_fill = green_fill
                elif col in pct_cols and lbl in top5_pct.get(col, ()):
                    cell_fill = red_fill
                elif col in pct_cols and lbl in bot5_pct.get(col, ()):
                    cell_fill = blue_fill
                if cell_fill:
                    c.fill = cell_fill
                if row_bold:
                    c.font = Font(bold=True, color="111111")
        for r in range(1, ws.max_row + 1):
            for k in range(1, ws.max_column + 1):
                ws.cell(r, k).border = Border(left=(thick if k in bcols else thin),
                                               right=thin, top=thin, bottom=thin)
        for k in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(k)].width = 14 if k <= n_idx else 11
    return buf.getvalue()


# ── 담당별 전년대비 복종 비중 변화 (2026-08-09 신규, [수정5][수정6]) ──────────────
# 표1(복종별 판매비중)과 같은 행 구성(G.TOTAL·매장담당별 평균·SD065·개별 매장)이되, 각 행을
# "{전년} {라벨}" / "{올해} {라벨}" / "ㄴ증감" 3개 행으로 쌓아 전년 동기간 대비 변화를 보여준다.
# 전년 동기간 = 기존 앱 전역 관행과 동일(예: render_channel_brand)하게 s−1년~e−1년으로 계산.
# ⚠ pandas Styler.apply/.map은 "인덱스(행 라벨)가 유일하지 않으면" 에러를 낸다(KeyError) — 그런데
#   'ㄴ증감' 라벨은 매장/담당자 수만큼 반복돼 그 자체로는 유일하지 않다. 그래서 내부적으로는
#   f"{_CATMIX_DIFF_PREFIX}{매장명}"처럼 매장명을 붙여 유일한 키로 쓰고, 화면 표시만
#   format_index()로 "ㄴ증감"으로 통일해 보여준다(엑셀도 동일 방식으로 헤더 텍스트만 치환).
_CATMIX_DIFF_LABEL = "ㄴ증감"
_CATMIX_DIFF_PREFIX = "ㄴ증감__"


def _catmix_yoy_style(disp2, sign2, gt_labels, sd_labels, mgr_labels, amt_cols, pct_cols, n_meta=3):
    """담당별 전년대비 복종 비중 변화 표 전용 Styler.

    disp2=표시용(포맷 문자열), sign2=증감 부호 판정용 원본 숫자(ㄴ증감 행만 값 있음, 나머진 NaN).
    disp2.index의 ㄴ증감 행은 내부적으로 f"ㄴ증감__{매장명}"(유일 키) — 화면엔 format_index로
    "ㄴ증감"만 보임. G.TOTAL/SD065/매장담당별 평균 행은 연도 쌍(전년·올해) 모두 표1과 동일 배경색,
    ㄴ증감 행은 양수=초록/음수=빨강(이 앱의 기존 증감 색 관행과 동일) + 행 라벨 자체는 옅은 빨강
    이탤릭으로 구분.
    """
    sty = disp2.style

    def _delta_color(col):
        vals = sign2[col]
        out = []
        for v in vals:
            if pd.isna(v) or v == 0:
                out.append("")
            elif v > 0:
                out.append("color:#1f8a4c;font-weight:700")
            else:
                out.append("color:#c62828;font-weight:700")
        return out

    for col in ["합계(백만)"] + amt_cols + pct_cols:
        sty = sty.apply(lambda s, c=col: _delta_color(c), subset=pd.IndexSlice[:, [col]])
    if gt_labels:
        sty = sty.set_properties(subset=pd.IndexSlice[gt_labels, :],
                                  **{"background-color": "#fff2b8", "font-weight": "700"})
    if sd_labels:
        sty = sty.set_properties(subset=pd.IndexSlice[sd_labels, :],
                                  **{"background-color": "#ffe0b2", "font-weight": "700"})
    if mgr_labels:
        sty = sty.set_properties(subset=pd.IndexSlice[mgr_labels, :],
                                  **{"background-color": _CATMIX_MGR_BG})
    sty = sty.set_properties(**{"text-align": "right"})
    for i in range(len(amt_cols)):
        sty = block_border(sty, n_meta + 2 * i)
    _pct_set = set(pct_cols)
    sty = sty.format_index(lambda c: "%" if c in _pct_set else c, axis=1)      # 수정5
    # 행 라벨(인덱스) 셀만 스타일/치환 — map_index·format_index는 데이터 셀(td)엔 영향 없어
    # 위 증감 색과 안 섞인다. 유일 키(ㄴ증감__매장명)를 화면엔 "ㄴ증감"으로만 보이게 함.
    sty = sty.map_index(lambda v: ("color:#c0392b;font-style:italic;font-weight:600"
                                    if str(v).startswith(_CATMIX_DIFF_PREFIX) else ""), axis=0)
    sty = sty.format_index(lambda v: (_CATMIX_DIFF_LABEL if str(v).startswith(_CATMIX_DIFF_PREFIX) else v),
                            axis=0)
    return sty


def _catmix_yoy_excel_bytes(disp2, sign2, gt_labels, sd_labels, mgr_labels, amt_cols, pct_cols,
                             sheet="복종비중YoY", n_meta=3):
    """룰13: 화면 서식(G.TOTAL·SD065·담당별평균 배경 + ㄴ증감 행 +초록/-빨강 + 복종 구분 세로선 +
    % 헤더 간결화) 그대로 엑셀 반영."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        disp2.to_excel(w, sheet_name=_safe_name(sheet)[:28] or "표")
        ws = w.book.worksheets[0]
        n_idx = disp2.index.nlevels
        n_rows = len(disp2)
        data_start = ws.max_row - n_rows + 1
        thin = Side(style="thin", color="D9D9D9")
        thick = Side(style="medium", color="555555")
        head_fill = PatternFill("solid", fgColor="F4F4F6")
        idx_fill = PatternFill("solid", fgColor="FAFAFA")
        gt_fill = PatternFill("solid", fgColor="FFF2B8")
        sd_fill = PatternFill("solid", fgColor="FFE0B2")
        mgr_fill = PatternFill("solid", fgColor="D6F0FA")
        gt_set, sd_set, mgr_set = set(gt_labels or []), set(sd_labels or []), set(mgr_labels or [])
        delta_cols = set(["합계(백만)"] + amt_cols + pct_cols)
        pct_set = set(pct_cols)
        bcols = {n_idx + (n_meta + 2 * i) + 1 for i in range(len(amt_cols))}

        hdr_r = data_start - 1
        for k in range(1, ws.max_column + 1):
            cell = ws.cell(hdr_r, k)
            cell.fill = head_fill
            cell.font = Font(bold=True, color="111111")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for cj, col in enumerate(disp2.columns):     # 수정5: % 헤더 간결화
            if col in pct_set:
                ws.cell(hdr_r, n_idx + 1 + cj).value = "%"

        for ri in range(n_rows):
            r = data_start + ri
            lbl = disp2.index[ri]
            is_diff = str(lbl).startswith(_CATMIX_DIFF_PREFIX)
            row_fill, row_bold = None, False
            if lbl in gt_set:
                row_fill, row_bold = gt_fill, True
            elif lbl in sd_set:
                row_fill, row_bold = sd_fill, True
            elif lbl in mgr_set:
                row_fill, row_bold = mgr_fill, False
            for k in range(1, n_idx + 1):
                c = ws.cell(r, k)
                c.fill = row_fill or idx_fill
                c.font = Font(bold=True, color=("C0392B" if is_diff else "111111"), italic=is_diff)
                c.alignment = Alignment(horizontal="left", vertical="center")
                if is_diff:      # 유일 키(ㄴ증감__매장명) 대신 화면과 동일하게 "ㄴ증감"만 표시
                    c.value = _CATMIX_DIFF_LABEL
            for cj, col in enumerate(disp2.columns):
                c = ws.cell(r, n_idx + 1 + cj)
                c.alignment = Alignment(horizontal="right", vertical="center")
                if row_fill:
                    c.fill = row_fill
                if is_diff and col in delta_cols:
                    sv = sign2.iloc[ri][col] if col in sign2.columns else None
                    if pd.notna(sv) and sv != 0:
                        c.font = Font(bold=True, color=("1F8A4C" if sv > 0 else "C62828"))
                elif row_bold:
                    c.font = Font(bold=True, color="111111")
        for r in range(1, ws.max_row + 1):
            for k in range(1, ws.max_column + 1):
                ws.cell(r, k).border = Border(left=(thick if k in bcols else thin),
                                               right=thin, top=thin, bottom=thin)
        for k in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(k)].width = 14 if k <= n_idx else 12
    return buf.getvalue()


# ── [수정8] 매장별 가이드 (2026-08-09 신규, 시트3 "매장별 제안" v2 이관) ──────────────
# 설계 근거: claude/시트3_매장별제안_v2_요약.md (2026-08-08 업로드 v2 이관문서 요약).
# 중태님 확정(2026-08-09): (1) 복종 분류는 원본 9분류 대신 이 ERP 자체 중/소카테고리 재사용
# (표1·표2와 동일 기준 — 시트3_요약.md의 "이식 시 유의점 6" 결정) (2) 화면 표(세로 병합 필요)가
# 아니라 표1 옆 "⬇ 매장별 가이드" 엑셀 다운로드로 구현 — 세로 병합은 openpyxl로 간단하지만 이 앱의
# 기존 pandas Styler/HTML 표 방식으로는 신규 패턴이 필요했던 문제를 우회.
# 모집단: 강점/약점 판정·비교매장 2곳 = shown(현재 조회조건의 100만원 이상 매장, 표1과 동일 모집단).
#         전사 평균(G열)만 예외 = store_tot(같은 조회조건, 100만원 미만 포함 전체 매장) — 원본 스펙의
#         "필터 적용 전 전체 매장"을 이 ERP 맥락으로 매핑(원본엔 브랜드/연차/시즌/담당 필터가 없었으므로
#         "필터"=100만원 매출 플로어만을 의미했던 것으로 해석, claude/시트3_매장별제안_v2_요약.md 참고).
# "기타" 복종은 강점/약점 후보에서 제외(전사평균 G열 계산엔 자연히 포함 — 매장 합계에 이미 반영됨).
# ⚠️ 문서에 명시 안 된 세부 결정(구현 시 임의 보완, 추후 확인 필요):
#   - 약점 복종이 3개 이상일 때 "최대 2개"로 자르는 기준 = cat_rank가 가장 나쁜(큰) 순으로 2개 채택.
#   - 카테고리 내 동순위(rank) 타이브레이크 = pandas rank(method="min") 표준 처리.
_CATMIX_GUIDE_WEAK_FILL = "FDF6F4"   # 약점 행 배경(연한 살구색, 원본 스펙과 동일)


def _rankgap_verdict(rank_total, cat_rank):
    """시트3_요약.md '강점/약점 판정 로직(6장)' — rank_total=매장 전체순위, cat_rank=그 복종 절대금액
    순위(모집단 내). 1~3위 매장은 별도 규칙(중태님 2026-07-01 확정, 임의 단순화 금지)."""
    if rank_total == 1:
        return "강점" if cat_rank == 1 else ("중립" if cat_rank == 2 else "약점")
    if rank_total in (2, 3):
        return "강점" if cat_rank <= 3 else "약점"
    gap = rank_total - cat_rank
    return "강점" if gap > 0 else ("약점" if gap < 0 else "중립")


def _catmix_guide_coach(my_pct):
    """시트3_요약.md K열 코칭 멘트 3단계 분기(규칙 기반, LLM 호출 불필요). 어휘 규칙(2026-08-08
    담당자 피드백): 온라인 쇼핑몰 매장이므로 '진열'·'매대' 금지, '노출'·'상품 등록'·'세트 구성'만 사용."""
    if my_pct <= 1.0:
        return ("판매가 사실상 없는 상태 — '안 팔리는 아이템'이 아니라 '거의 팔고 있지 않은 아이템'입니다. "
                "입점 여부와 카테고리·검색 노출 구성부터 점검해주세요.")
    if my_pct <= 3.0:
        return "보조 아이템으로 최소한만 취급되는 상태로 보여요. 상품 등록·메인/카테고리 노출 여부부터 확인해주세요."
    return "매장 규모 대비 노출이 약한 편이에요. 연관 상품 노출과 세트 구성 제안을 점검해보세요."


def _catmix_guide_rows(shown, piv, store_tot, cats):
    """매장별 가이드 원본 행 데이터 조립(엑셀 작성과 분리 — 단위 테스트 용이하게). 반환: list[dict],
    각 dict는 A~K열 값 + 병합용 메타(_block_first/_block_size)를 담음."""
    cats_rank = [c for c in cats if c != "기타" and c in piv.columns]
    total_all = float(store_tot["합계"].sum())
    cat_total_all = {c: float(piv[c].sum()) for c in cats_rank}
    overall_avg_pct = {c: (cat_total_all[c] / total_all * 100.0 if total_all else 0.0) for c in cats_rank}

    codes = shown["매장코드"].astype(str).tolist()
    piv_codes = piv.reindex(codes).fillna(0.0)
    cat_rank_map = {c: piv_codes[c].rank(ascending=False, method="min") for c in cats_rank}
    # rank()의 인덱스는 piv_codes.index(=codes 순서 그대로) — code→rank 조회용 dict로 변환
    cat_rank_map = {c: dict(zip(codes, ranks.tolist())) for c, ranks in cat_rank_map.items()}

    total_map = dict(zip(shown["매장코드"], shown["합계"]))
    name_map = dict(zip(shown["매장코드"], shown["매장명"]))

    rows = []
    for _, srow in shown.sort_values("순위").iterrows():
        code, name = srow["매장코드"], srow["매장명"]
        rank_total = int(srow["순위"]); my_total = float(srow["합계"])
        cat_amt = piv.loc[code] if code in piv.index else pd.Series(0.0, index=cats)

        strengths, weaknesses = [], []
        for c in cats_rank:
            cr = cat_rank_map[c].get(code)
            if cr is None:
                continue
            cr = int(cr)
            my_pct = (float(cat_amt.get(c, 0.0)) / my_total * 100.0) if my_total else 0.0
            verdict = _rankgap_verdict(rank_total, cr)
            if verdict == "강점":
                strengths.append((c, cr, my_pct))
            elif verdict == "약점":
                weaknesses.append((c, cr, my_pct))

        strengths.sort(key=lambda t: t[1])                       # cat_rank 좋은(작은) 순 — 최대 2개
        top_strengths = strengths[:2]
        strength_txt = ("\n".join(f"{c}({cr}위, {pct:.1f}%)" for c, cr, pct in top_strengths)
                         if top_strengths else "뚜렷한 강점 복종 없음")

        weaknesses.sort(key=lambda t: t[1], reverse=True)        # cat_rank 나쁜(큰) 순 — 최대 2개
        top_weak = weaknesses[:2]

        if not top_weak:
            rows.append({
                "매장코드": code, "매장명": name, "합계순위": rank_total, "강점요약": strength_txt,
                "약점카테고리": "뚜렷한 약점 복종 없음(전 복종 고르게 강세)",
                "내비중": None, "전사평균": None, "격차": None, "비교매장": "",
                "기회금액": None, "코칭멘트": "강점 유지·확대에 집중하세요.",
                "_block_first": True, "_block_size": 1,
            })
            continue

        for wi, (c, cr, my_pct) in enumerate(top_weak):
            avg_pct = overall_avg_pct.get(c, 0.0)
            gap = avg_pct - my_pct
            amt_series = piv_codes[c] if c in piv_codes.columns else pd.Series(0.0, index=codes)
            peers = amt_series.drop(index=[code], errors="ignore").sort_values(ascending=False)
            peer_codes = peers.head(2).index.tolist()
            peer_pcts, peer_strs = [], []
            for pc in peer_codes:
                p_total = total_map.get(pc, 0.0)
                p_amt = float(piv.loc[pc, c]) if (pc in piv.index and c in piv.columns) else 0.0
                p_pct = (p_amt / p_total * 100.0) if p_total else 0.0
                peer_pcts.append(p_pct)
                peer_strs.append(f"{name_map.get(pc, pc)} {p_pct:.1f}%")
            peer_txt = "\n".join(peer_strs)

            opp_amt, coach_prefix = None, ""
            if gap > 0:
                opp_amt = (gap / 100.0) * my_total
            else:
                # 예외 케이스(시트3_요약.md): rank-gap상 약점인데 전사평균 이상 비중 →
                # 비교 매장 중 비중이 더 높은 쪽과의 격차로 대체 계산. 그마저 낮으면 J열 문자 그대로 "-".
                alt_gap = (max(peer_pcts) - my_pct) if peer_pcts else -1.0
                opp_amt = (alt_gap / 100.0) * my_total if alt_gap > 0 else "-"
                coach_prefix = ("전사 평균 대비는 낮지 않지만, 매출 규모가 비슷한 매장들과 비교하면 "
                                 "상대적으로 약한 편이에요. ")

            rows.append({
                "매장코드": code, "매장명": name, "합계순위": rank_total, "강점요약": strength_txt,
                "약점카테고리": f"{c} ({cr}위)", "내비중": my_pct, "전사평균": avg_pct, "격차": gap,
                "비교매장": peer_txt, "기회금액": opp_amt, "코칭멘트": coach_prefix + _catmix_guide_coach(my_pct),
                "_block_first": (wi == 0), "_block_size": len(top_weak),
            })
    return rows


def _catmix_guide_excel_bytes(shown, piv, store_tot, cats, sheet="매장별가이드"):
    """시트3(매장별 제안) v2 구조 — 매장별 강점/약점 복종을 rank-gap 판정으로 뽑아 전사 평균·비교
    매장·기회금액·코칭 멘트까지 묶어 보여주는 담당자용 가이드 엑셀(11열 A~K, A~D 세로 병합)."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import openpyxl

    rows = _catmix_guide_rows(shown, piv, store_tot, cats)
    headers = ["매장코드", "매장명", "합계순위", "강점 요약", "약점 카테고리", "내 비중(%)",
               "전사 평균(%)", "격차(%p)", "비교 매장(비중)", "평균 대비 놓친 매출(원)", "담당자 코칭 멘트"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = _safe_name(sheet)[:28] or "매장별가이드"

    head_fill = PatternFill("solid", fgColor="F2F2F2")
    weak_fill = PatternFill("solid", fgColor=_CATMIX_GUIDE_WEAK_FILL)
    thin = Side(style="thin", color="D9D9D9")

    for cj, h in enumerate(headers, start=1):
        c = ws.cell(1, cj, h)
        c.fill = head_fill
        c.font = Font(name="맑은 고딕", size=9, bold=True, color="404040")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    if not rows:
        ws.freeze_panes = "A2"
        buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

    r = 2
    for row in rows:
        is_weak = row["약점카테고리"] != "뚜렷한 약점 복종 없음(전 복종 고르게 강세)"
        vals = [row["매장코드"], row["매장명"], row["합계순위"], row["강점요약"], row["약점카테고리"],
                row["내비중"], row["전사평균"], row["격차"], row["비교매장"], row["기회금액"], row["코칭멘트"]]
        for cj, v in enumerate(vals, start=1):
            c = ws.cell(r, cj, v)
            c.alignment = Alignment(horizontal=("left" if cj in (2, 4, 5, 9, 11) else "center"),
                                     vertical="center", wrap_text=(cj in (4, 5, 9, 11)))
            if cj == 6 and v is not None:      # F 내비중
                c.number_format = '0.0"%"'
            elif cj == 7 and v is not None:    # G 전사평균
                c.number_format = '0.0"%"'
            elif cj == 8 and v is not None:    # H 격차
                c.number_format = '0.0"%p"'
                if v > 0:
                    c.font = Font(bold=True, color="C0392B")
            elif cj == 10 and isinstance(v, (int, float)):   # J 기회금액(양수 값이 있을 때만 강조)
                c.number_format = '#,##0"원"'
                c.font = Font(bold=True, color="8A5300")
            if is_weak and cj >= 5:            # E~K 약점 행 배경(원본 스펙)
                c.fill = weak_fill
        r += 1

    # A~D 세로 병합(매장 블록 단위)
    r = 2
    for row in rows:
        if row["_block_first"]:
            n = row["_block_size"]
            if n > 1:
                for col in (1, 2, 3, 4):
                    ws.merge_cells(start_row=r, start_column=col, end_row=r + n - 1, end_column=col)
                    ws.cell(r, col).alignment = Alignment(horizontal=("left" if col in (2, 4) else "center"),
                                                           vertical="center", wrap_text=(col == 4))
        r += 1

    for rr in range(1, ws.max_row + 1):
        for cc in range(1, len(headers) + 1):
            ws.cell(rr, cc).border = Border(left=thin, right=thin, top=thin, bottom=thin)
    widths = {1: 10, 2: 14, 3: 9, 4: 22, 5: 16, 6: 9, 7: 10, 8: 9, 9: 24, 10: 16, 11: 42}
    for cj, w in widths.items():
        ws.column_dimensions[get_column_letter(cj)].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO(); wb.save(buf)
    return buf.getvalue()


def render_category_mix(df):
    """🧵 복종별 판매비중 분석 — 매장별 복종(중/소카테고리)별 판매 금액·비중 현황표.

    과거 "매장별 아이템 비중 분석" 프로젝트(엑셀 3시트 산출물)의 시트1(원본+집계 데이터)만 이식한
    것 — 시트2(베스트/워스트)·시트3(매장별 제안)은 이번 범위 밖(중태님 확정, 2026-08-08).
    """
    st.subheader("🧵 복종별 판매비중 분석")
    if df is None or df.empty or "_판매일" not in df.columns or df["_판매일"].notna().sum() == 0:
        st.info("데이터를 먼저 적재하세요.")
        return
    need_cols = {"아이템", "매장코드", "매장명", "_매출액", "_판매일"}
    if not need_cols.issubset(df.columns):
        st.info("이 리포트에 필요한 컬럼(아이템·매장코드·매장명 등)이 없어요.")
        return
    d = df[df["_판매일"].notna()].copy()

    # [수정7, 2026-08-09] 담당별 필터 추가 위해 매장 담당자를 행(거래) 단위로 먼저 매핑 — 이 앱의
    # 기존 관행(render_channel_brand 등)과 동일한 패턴. shown["_담당자"] 산출에도 동일 맵 재사용.
    master = load_master()
    if not master.empty and "담당자" in master.columns:
        _mgr_map = dict(zip(master["매장코드"].astype(str).str.strip(),
                            master["담당자"].astype(str).str.strip()))
        d["_담당자"] = d["매장코드"].astype(str).str.strip().map(_mgr_map)
        d["_담당자"] = d["_담당자"].where(
            d["_담당자"].notna() & d["_담당자"].str.strip().ne("")
            & ~d["_담당자"].str.lower().isin(["nan", "none"]), None)
    else:
        _mgr_map = {}
        d["_담당자"] = None
    _mans = sorted({str(m).strip() for m in d["_담당자"].dropna().astype(str)
                    if str(m).strip() and str(m).strip().lower() not in ("nan", "none")})

    st.caption("매장별로 어떤 복종(아이템군)의 판매 비중이 높고 낮은지 보여줘요. "
               "🟧 온라인통합몰(SD065)은 다른 매장이 참조하는 비교 기준선이라 항상 주황색으로 표시돼요. "
               "🟦 G.TOTAL 바로 아래엔 매장담당별 평균(하늘색)이 담당자마다 한 행씩 나와요. "
               "🟢 복종별 '금액' 열의 매출 상위 5개 매장(SD065·담당별평균 제외)은 녹색, "
               "🔴🔵 '%' 열의 상위 5개는 빨간색·하위 5개는 파란색(둘 다 SD065·담당별평균 제외)으로 "
               "표시돼요. 기간 합계매출 100만원 미만 매장은 총계엔 포함되지만 목록엔 표시하지 않아요. "
               "🔽 담당별 필터에서 담당자를 고르면 두 표(복종별 판매비중·전년대비 변화) 모두 그 "
               "담당자의 매장만으로 줄어들어요(드릴다운, 2026-08-09).")

    dmin, dmax = d["_판매일"].min().date(), d["_판매일"].max().date()
    default_start = max(pd.to_datetime(dmax) - pd.Timedelta(days=6), pd.to_datetime(dmin)).date()
    with st.form("cm_form"):
        rng = st.date_input("조회기간 (시작일~종료일 직접 지정)", value=(default_start, dmax),
                            min_value=dmin, max_value=dmax, key="cm_rng")
        cm0, cm1, cm2, cm3, cm4 = st.columns([1.1, 1, 1, 1, 1])
        level = cm0.radio("카테고리 기준", CATMIX_CAT_LEVELS, horizontal=True, key="cm_level")
        # [수정7] 담당별 필터 추가 — 공통룰10(브랜드/연차/시즌) 4번째 필터, 빈칸=전체
        brands = sorted(d["브랜드명"].dropna().unique()) if "브랜드명" in d.columns else []
        ages = sorted(d["연차"].dropna().unique(), key=_age_sort_key) if "연차" in d.columns else []
        seasons = sorted(d["시즌명"].dropna().unique()) if "시즌명" in d.columns else []
        selb = cm1.multiselect("브랜드별", brands, default=[], placeholder="전체", key="cm_brand")
        sela = cm2.multiselect("연차별", ages, default=[], placeholder="전체", key="cm_age")
        sels = cm3.multiselect("시즌별", seasons, default=[], placeholder="전체", key="cm_season")
        selm = cm4.multiselect("담당별", _mans, default=[], placeholder="전체", key="cm_mgr")
        if not _mans:
            st.caption("※ 매장 기준정보(담당자)가 없어 담당별 필터가 비어 있어요 — 사이드바에서 매장 기준정보를 업로드하면 채워져요.")
        run = st.form_submit_button("🔍 조회", type="primary")
    if _need_search("cm_go", run):
        return
    if not (isinstance(rng, (list, tuple)) and len(rng) == 2):
        st.info("기간(시작~끝)을 선택한 뒤 🔍 조회를 눌러 주세요.")
        return
    s, e = pd.to_datetime(rng[0]), pd.to_datetime(rng[1])
    if e < s:
        st.error("종료일이 시작일보다 앞서요. 기간을 다시 선택해 주세요.")
        return

    base = d[(d["_판매일"] >= s) & (d["_판매일"] <= e)]
    if selb and "브랜드명" in base.columns:
        base = base[base["브랜드명"].isin(selb)]
    if sela and "연차" in base.columns:
        base = base[base["연차"].isin(sela)]
    if sels and "시즌명" in base.columns:
        base = base[base["시즌명"].isin(sels)]
    if selm and "_담당자" in base.columns:      # [수정7] 담당별 드릴다운 — 선택한 담당자의 매장만
        base = base[base["_담당자"].astype(str).str.strip().isin(selm)]
    if base.empty:
        st.info("선택한 조건에 매출 데이터가 없어요.")
        return

    cat_map = get_itemgroup_map() if level == "중카테고리" else get_itemgroup_map_small()
    base = base.assign(_복종=base["아이템"].astype(str).str.strip().str.upper().map(cat_map).fillna("기타"))

    store_tot = base.groupby(["매장코드", "매장명"])["_매출액"].sum().rename("합계").reset_index()
    shown = store_tot[store_tot["합계"] >= _CATMIX_FLOOR].copy()
    hidden_n = len(store_tot) - len(shown)
    if shown.empty:
        st.info("기간 합계매출 100만원 이상인 매장이 없어요.")
        return
    shown = shown.sort_values("합계", ascending=False).reset_index(drop=True)
    shown["순위"] = shown.index + 1

    # 수정2: 매장담당별 평균 — 매장 기준정보(담당자)와 매칭, 표시 대상(shown) 매장만 대상으로 그룹평균
    # (담당자 맵은 위 [수정7]에서 이미 계산해둔 _mgr_map 재사용)
    shown["_담당자"] = shown["매장코드"].astype(str).str.strip().map(_mgr_map)
    shown["_담당자"] = shown["_담당자"].where(
        shown["_담당자"].notna() & shown["_담당자"].astype(str).str.strip().ne("")
        & ~shown["_담당자"].astype(str).str.lower().isin(["nan", "none"]), None)

    piv = base.pivot_table(index="매장코드", columns="_복종", values="_매출액", aggfunc="sum", fill_value=0.0)
    cats = piv.sum(axis=0).sort_values(ascending=False).index.tolist()
    piv = piv.reindex(columns=cats, fill_value=0.0)
    amt_cols = list(cats)              # 수정1: "{복종} 금액(백만)" → "{복종}"으로 간결화
    pct_cols = [f"{c} %" for c in cats]
    n_meta = 3    # 매장코드·합계(백만)·순위 — 복종 구분선(수정3) 위치 계산용

    def _row(cat_amt, total, code_disp, rank_val):
        out = {"매장코드": code_disp, "합계(백만)": total / 1e6, "순위": rank_val}
        for c in cats:
            a = float(cat_amt.get(c, 0.0))
            out[c] = a / 1e6
            out[f"{c} %"] = (a / total * 100) if total else 0.0
        return out

    rows, index = [], []
    total_amt_all = float(store_tot["합계"].sum())
    rows.append(_row(piv.sum(axis=0), total_amt_all, "", None))
    index.append("G.TOTAL")

    # 수정2: 매장담당별 평균 행 — G.TOTAL 바로 아래, 개별 매장행 위. 평균 합계 큰 순 정렬.
    # mgr_avg_tot는 아래 "담당별 전년대비 복종 비중 변화" 표(2026-08-09 신규, 수정6)에서도 재사용
    # 하므로 if 안이 아니라 항상(빈 Series라도) 계산해둔다 — 동작은 기존과 동일(빈 Series면 루프 무실행).
    mgr_labels = []
    mgr_avg_tot = (shown.groupby("_담당자")["합계"].mean().sort_values(ascending=False)
                   if shown["_담당자"].notna().any() else pd.Series(dtype="float64"))
    for mgr, avg_tot in mgr_avg_tot.items():
        codes = shown.loc[shown["_담당자"] == mgr, "매장코드"]
        cat_amt_avg = piv.reindex(codes).fillna(0.0)[cats].mean(axis=0)
        lbl = f"{mgr} 평균"
        rows.append(_row(cat_amt_avg, float(avg_tot), "", None))
        index.append(lbl)
        mgr_labels.append(lbl)

    for _, r in shown.iterrows():
        code = r["매장코드"]
        cat_amt = piv.loc[code] if code in piv.index else pd.Series(0.0, index=cats)
        rows.append(_row(cat_amt, float(r["합계"]), code, int(r["순위"])))
        index.append(r["매장명"])
    num = pd.DataFrame(rows, index=index)
    num.index.name = "매장명"

    sd_row = shown[shown["매장코드"].astype(str).str.strip().str.upper() == _CATMIX_SD065]
    sd_label = sd_row.iloc[0]["매장명"] if not sd_row.empty else None

    disp = num.copy()
    disp["합계(백만)"] = num["합계(백만)"].map(lambda v: f"{v:,.1f}")
    disp["순위"] = num["순위"].map(lambda v: "–" if pd.isna(v) else f"{int(v)}")
    for c in amt_cols:
        disp[c] = num[c].map(lambda v: f"{v:,.1f}")
    for c in pct_cols:
        disp[c] = num[c].map(lambda v: f"{v:.1f}%")

    h1, h2, h2g = st.columns([3.4, 1, 1])
    h1.markdown(f"### 매장별 복종({level}) 판매비중{_NOTE_FLOAT}", unsafe_allow_html=True)
    h2.download_button("⬇ 엑셀", _catmix_excel_bytes(disp, sd_label, amt_cols, pct_cols, num,
                                                       sheet=f"복종별판매비중_{level}",
                                                       mgr_labels=mgr_labels, n_meta=n_meta),
                       file_name=f"복종별판매비중_{level}_{s.date()}_{e.date()}.xlsx", mime=XLSX_MIME,
                       key="cm_dl", use_container_width=True)
    # [수정8, 2026-08-09] 매장별 가이드(시트3 v2 이관) — 표1과 같은 조회조건(shown/piv/store_tot)을
    # 그대로 재사용해 강점/약점·전사평균·비교매장·기회금액·코칭멘트를 엑셀로 내려받음(화면 표는 없음).
    h2g.download_button("⬇ 매장별 가이드", _catmix_guide_excel_bytes(shown, piv, store_tot, cats,
                                                                     sheet=f"매장별가이드_{level}"),
                       file_name=f"매장별가이드_{level}_{s.date()}_{e.date()}.xlsx", mime=XLSX_MIME,
                       key="cm_guide_dl", use_container_width=True)
    sty = _catmix_style(disp, num, sd_label, amt_cols, pct_cols, mgr_labels=mgr_labels, n_meta=n_meta)
    render_styled_table(sty)
    _mgr_note = (f" · 매장담당별 평균 {len(mgr_labels)}명(하늘색, 표시된 매장 기준)" if mgr_labels
                 else " · 매장 기준정보(담당자)가 없어 매장담당별 평균은 표시되지 않았어요")
    st.caption(f"※ 매장 {len(shown)}개 표시(기간 합계매출 100만원 미만 {hidden_n}개 매장은 총계엔 "
               "포함되지만 목록에서는 제외) · 순위=표시된 매장 안에서 매출 큰 순 · "
               "복종 열은 총매출 큰 순으로 정렬돼요" + _mgr_note + ".")

    # ── [수정6] 담당별 전년대비 복종 비중 변화 (2026-08-09 신규) ──────────────
    # 표1과 동일한 행 구성(G.TOTAL·매장담당별 평균·SD065·개별 매장)을 "{전년} 라벨"/"{올해} 라벨"/
    # "ㄴ증감" 3행씩으로 쌓아 전년 동기간 대비 변화를 보여줌. 전년 동기간 계산은 이 앱의 기존 관행
    # (render_channel_brand 등)과 동일하게 조회기간을 그대로 1년 전으로 시프트.
    ps, pe = s - pd.DateOffset(years=1), e - pd.DateOffset(years=1)
    prev_base = d[(d["_판매일"] >= ps) & (d["_판매일"] <= pe)]
    if selb and "브랜드명" in prev_base.columns:
        prev_base = prev_base[prev_base["브랜드명"].isin(selb)]
    if sela and "연차" in prev_base.columns:
        prev_base = prev_base[prev_base["연차"].isin(sela)]
    if sels and "시즌명" in prev_base.columns:
        prev_base = prev_base[prev_base["시즌명"].isin(sels)]
    if selm and "_담당자" in prev_base.columns:  # [수정7] 담당별 드릴다운 — base와 동일하게 적용
        prev_base = prev_base[prev_base["_담당자"].astype(str).str.strip().isin(selm)]

    cy = int(e.year)
    py2, cy2 = (cy - 1) % 100, cy % 100

    h3, h4 = st.columns([4, 1])
    h3.markdown(f"### 담당별 전년대비 복종 비중 변화 (전년 동기간 비교){_NOTE_FLOAT}", unsafe_allow_html=True)
    if prev_base.empty:
        st.info(f"전년 동기간({ps.date()}~{pe.date()}) 매출 데이터가 없어서 전년대비 비교표를 만들 수 없어요.")
        return

    prev_base = prev_base.assign(
        _복종=prev_base["아이템"].astype(str).str.strip().str.upper().map(cat_map).fillna("기타"))
    store_tot_prev = prev_base.groupby(["매장코드", "매장명"])["_매출액"].sum().rename("합계").reset_index()
    prev_total_map = dict(zip(store_tot_prev["매장코드"], store_tot_prev["합계"]))
    shown_prev = store_tot_prev[store_tot_prev["합계"] >= _CATMIX_FLOOR].copy()
    shown_prev = shown_prev.sort_values("합계", ascending=False).reset_index(drop=True)
    shown_prev["순위"] = shown_prev.index + 1
    prev_rank_map = dict(zip(shown_prev["매장코드"], shown_prev["순위"]))
    piv_prev = prev_base.pivot_table(index="매장코드", columns="_복종", values="_매출액",
                                      aggfunc="sum", fill_value=0.0).reindex(columns=cats, fill_value=0.0)

    # 전년 데이터도 표1과 동일한 _row() 헬퍼로 조립 — 컬럼 순서·계산 방식(총액 0이면 %도 0) 일치 보장
    prev_rows = [_row(piv_prev.sum(axis=0), float(store_tot_prev["합계"].sum()), "", None)]
    for mgr, avg_tot in mgr_avg_tot.items():
        codes = shown.loc[shown["_담당자"] == mgr, "매장코드"]
        cat_amt_avg_prev = piv_prev.reindex(codes).fillna(0.0)[cats].mean(axis=0)
        avg_tot_prev = float(pd.Series([prev_total_map.get(c, 0.0) for c in codes]).mean()) if len(codes) else 0.0
        prev_rows.append(_row(cat_amt_avg_prev, avg_tot_prev, "", None))
    for _, r in shown.iterrows():
        code = r["매장코드"]
        cat_amt_prev = piv_prev.loc[code] if code in piv_prev.index else pd.Series(0.0, index=cats)
        # 순위(전년)는 전년도 자체 매출 100만원 이상 매장 모집단 안에서 재계산(연도별로 다를 수 있음,
        # 중태님 확인 완료) — 해당 매장이 전년엔 100만원 미만·데이터 없음이면 "–"로 표시.
        prev_rows.append(_row(cat_amt_prev, float(prev_total_map.get(code, 0.0)), code,
                               prev_rank_map.get(code)))
    prev_num = pd.DataFrame(prev_rows, index=num.index)
    prev_num.index.name = "매장명"

    def _fmt_row2(r):
        out = {"매장코드": r["매장코드"], "합계(백만)": f"{r['합계(백만)']:,.1f}",
               "순위": "–" if pd.isna(r["순위"]) else f"{int(r['순위'])}"}
        for c in cats:
            out[c] = f"{r[c]:,.1f}"
            out[f"{c} %"] = f"{r[f'{c} %']:.1f}%"
        return out

    def _yoy_amt(cv, pv):
        # 중태님 확인(2026-08-09): 증감액(절대값, 백만원)과 증감율(%) 둘 다 함께 표기.
        diff = cv - pv
        if pv == 0:
            return (f"{diff:+,.1f} (신규)" if diff else "–"), diff
        return f"{diff:+,.1f} ({diff / pv * 100:+.1f}%)", diff

    def _yoy_pct(cv, pv):
        diff = cv - pv
        return f"{diff:+.1f}%p", diff

    def _diff_row2(cur_r, prev_r):
        out, sig = {"매장코드": "", "합계(백만)": None, "순위": "–"}, {}
        out["합계(백만)"], sig["합계(백만)"] = _yoy_amt(cur_r["합계(백만)"], prev_r["합계(백만)"])
        for c in cats:
            out[c], sig[c] = _yoy_amt(cur_r[c], prev_r[c])
            pc = f"{c} %"
            out[pc], sig[pc] = _yoy_pct(cur_r[pc], prev_r[pc])
        return out, sig

    rows2, index2, sign_rows2 = [], [], []
    gt_labels2, sd_labels2, mgr_labels2 = [], [], []
    for lbl in num.index:
        cur_r, prev_r = num.loc[lbl], prev_num.loc[lbl]
        py_lbl, cy_lbl = f"{py2} {lbl}", f"{cy2} {lbl}"
        rows2.append(_fmt_row2(prev_r)); index2.append(py_lbl); sign_rows2.append({})
        rows2.append(_fmt_row2(cur_r)); index2.append(cy_lbl); sign_rows2.append({})
        dtxt, dsig = _diff_row2(cur_r, prev_r)
        # 유일 키(라벨별 매장명 접미사) — 표시는 format_index()로 전부 "ㄴ증감"으로 통일(위 함수 참고)
        rows2.append(dtxt); index2.append(f"{_CATMIX_DIFF_PREFIX}{lbl}"); sign_rows2.append(dsig)
        if lbl == "G.TOTAL":
            gt_labels2 += [py_lbl, cy_lbl]
        elif sd_label and lbl == sd_label:
            sd_labels2 += [py_lbl, cy_lbl]
        elif lbl in mgr_labels:
            mgr_labels2 += [py_lbl, cy_lbl]

    disp2 = pd.DataFrame(rows2, index=index2)
    disp2.index.name = "매장명"
    sign2 = pd.DataFrame(sign_rows2, index=index2).reindex(columns=disp2.columns)

    h4.download_button(
        "⬇ 엑셀",
        _catmix_yoy_excel_bytes(disp2, sign2, gt_labels2, sd_labels2, mgr_labels2, amt_cols, pct_cols,
                                 sheet=f"복종비중YoY_{level}", n_meta=n_meta),
        file_name=f"담당별_전년대비_복종비중변화_{level}_{s.date()}_{e.date()}.xlsx", mime=XLSX_MIME,
        key="cm_yoy_dl", use_container_width=True)
    sty2 = _catmix_yoy_style(disp2, sign2, gt_labels2, sd_labels2, mgr_labels2, amt_cols, pct_cols,
                              n_meta=n_meta)
    render_styled_table(sty2)
    st.caption(f"※ {py2}년={ps.date()}~{pe.date()}(전년 동기간) · {cy2}년={s.date()}~{e.date()}(조회기간, "
               "위 표와 동일 대상) · ㄴ증감 행 — 합계·복종 금액 열: 증감액(백만) (증감율%), "
               "% 열: %p 차이 · 초록=증가 빨강=감소 · 순위는 각 연도 자체 매출 100만원 이상 매장 "
               "모집단 안에서 따로 계산(연도별로 달라질 수 있어요, 전년에 없던 매장은 \"–\").")


# ==============================================================================
# 매장(채널) 기준정보 마스터  ─ 업로드 시 전체 교체
# ==============================================================================
MASTER_TABLE = "channel_master"
MASTER_COLS = ["매장코드", "매장명", "담당자", "유통성격", "채널소유", "채널스토리", "리그구분"]
# ↑ 리그구분 (2026-07-31 추가): 1부리그/2부리그/꿈나무리그 — 대시보드 채널 리그 랭킹용


def read_master_file(uploaded_file):
    name = uploaded_file.name.lower()
    if name.endswith(".csv"):
        raw = pd.read_csv(uploaded_file, header=None, dtype=str, keep_default_na=False)
    else:
        raw = pd.read_excel(uploaded_file, header=None, dtype=str)
    hrow = 0
    for i in range(min(10, len(raw))):
        vals = [str(v).strip() for v in raw.iloc[i].tolist()]
        if "매장코드" in vals:
            hrow = i
            break
    header = [str(v).strip() for v in raw.iloc[hrow].tolist()]
    m = raw.iloc[hrow + 1:].copy()
    m.columns = header
    m = m.dropna(how="all")
    keep = [c for c in MASTER_COLS if c in m.columns]
    m = m[keep].copy()
    for c in keep:
        m[c] = m[c].astype(str).str.strip()
    m = m[m["매장코드"].ne("") & ~m["매장코드"].str.lower().isin(["nan", "none"])]
    if "유통성격" in m.columns:
        m["유통성격"] = m["유통성격"].replace({"벤더": "밴더"})  # 표기 통일
    return m.reset_index(drop=True)


def replace_master(m):
    eng = get_engine()
    with eng.begin() as conn:
        m.astype(str).to_sql(MASTER_TABLE, conn, if_exists="replace", index=False)
    return len(m)


@st.cache_data(ttl=21600)
def load_master():
    eng = get_engine()
    try:
        with eng.connect() as conn:
            exists = conn.exec_driver_sql(
                "SELECT 1 FROM information_schema.tables WHERE table_name=%s"
                if eng.dialect.name == "postgresql" else
                "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
                (MASTER_TABLE,)).fetchone()
            if not exists:
                return pd.DataFrame()
            m = pd.read_sql(f'SELECT * FROM "{MASTER_TABLE}"', conn)
    except Exception:
        return pd.DataFrame()
    for c in m.columns:
        m[c] = m[c].astype(str).str.strip()
    return m


def master_row_count():
    try:
        with get_engine().connect() as conn:
            return conn.exec_driver_sql(f'SELECT COUNT(*) FROM "{MASTER_TABLE}"').scalar()
    except Exception:
        return 0


# ── 사업계획(월별 목표) 마스터 : 매장별 + 브랜드별(연차 포함) ──
PLAN_TABLE = "plan_master"


def read_plan_file(uploaded_file):
    """사업계획 엑셀(매장별 시트 + 브랜드별 시트) → long DF[dim, code, sub, month, amount]."""
    import re
    def _num(v):
        try:
            f = float(v)
            return 0.0 if pd.isna(f) else f
        except Exception:
            return 0.0
    xls = pd.read_excel(uploaded_file, sheet_name=None, header=None, dtype=object)
    out = []
    for _, raw in xls.items():
        if raw is None or len(raw) == 0:
            continue
        hrow = kind = None
        for i in range(min(10, len(raw))):
            vals = [str(v).strip() for v in raw.iloc[i].tolist()]
            if "매장코드" in vals:
                hrow, kind = i, "store"; break
            if "브랜드" in vals:
                hrow, kind = i, "brand"; break
        if hrow is None:
            continue
        header = [str(v).strip() for v in raw.iloc[hrow].tolist()]
        mcols = {}
        for ci, h in enumerate(header):
            mm = re.match(r"(\d{1,2})\s*월", h)
            if mm:
                mcols[int(mm.group(1))] = ci
        body = raw.iloc[hrow + 1:]
        if kind == "store":
            ci_code = header.index("매장코드")
            for _, r in body.iterrows():
                code = str(r.iloc[ci_code]).strip()
                if code == "" or code.lower() in ("nan", "none"):
                    continue
                for m, ci in mcols.items():
                    out.append(("store", code, "", m, _num(r.iloc[ci])))
        else:
            ci_b = header.index("브랜드")
            ci_a = header.index("연차") if "연차" in header else None
            for _, r in body.iterrows():
                b = str(r.iloc[ci_b]).strip()
                if b == "" or b.lower() in ("nan", "none"):
                    continue
                sub = str(r.iloc[ci_a]).strip() if ci_a is not None else "합계"
                for m, ci in mcols.items():
                    out.append(("brand", b, sub, m, _num(r.iloc[ci])))
    return pd.DataFrame(out, columns=["dim", "code", "sub", "month", "amount"])


def replace_plan(p):
    eng = get_engine()
    with eng.begin() as conn:
        p.to_sql(PLAN_TABLE, conn, if_exists="replace", index=False)
    return len(p)


@st.cache_data(ttl=21600)
def load_plan():
    eng = get_engine()
    try:
        with eng.connect() as conn:
            exists = conn.exec_driver_sql(
                "SELECT 1 FROM information_schema.tables WHERE table_name=%s"
                if eng.dialect.name == "postgresql" else
                "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
                (PLAN_TABLE,)).fetchone()
            if not exists:
                return pd.DataFrame()
            p = pd.read_sql(f'SELECT * FROM "{PLAN_TABLE}"', conn)
    except Exception:
        return pd.DataFrame()
    if not p.empty:
        p["month"] = pd.to_numeric(p["month"], errors="coerce")
        p["amount"] = pd.to_numeric(p["amount"], errors="coerce").fillna(0.0)
        p["code"] = p["code"].astype(str).str.strip()
        p["sub"] = p["sub"].astype(str).str.strip()
    return p


def plan_row_count():
    try:
        with get_engine().connect() as conn:
            return conn.exec_driver_sql(f'SELECT COUNT(*) FROM "{PLAN_TABLE}"').scalar()
    except Exception:
        return 0


# ── 온라인팀 우선순위(당월·금주) 마스터 : 주간 업무 보고 FORM 업로드 (260830 신규) ──
#   사이드바에서 '온라인팀 우선순위 업로드 FORM.xlsx'를 올리면 전체 교체 저장.
#   쓰이는 곳 ① 종합 대시보드 — 월별 매출 표 바로 위에 당월·금주 우선순위 표시
#           ② 주간현황 분석 '⬇ 엑셀' — weekly_template의 당월(5~7행)·금주(35~37행) 빈칸 자동 채움
PRIORITY_TABLE = "weekly_priority"
_PRI_SECTIONS = ("당월", "금주")


def read_priority_file(uploaded_file):
    """온라인팀 우선순위 업로드 FORM(엑셀) → DF[section, category, no, content].

    양식 규칙(업로드 FORM·weekly_template 공통 구조): '당월 우선순위'/'금주 우선순위' 제목 셀
    아래 행이 카테고리 헤더(예: '매출/ 고객 접점 운영' · '내부 기획 및 개발 이슈'), 그 아래부터
    [번호 칸 | 내용 칸(바로 오른쪽)]이 세로로 이어진다. 제목·헤더 위치를 텍스트로 찾으므로
    행이 몇 줄 밀려 있어도, 항목이 3개보다 많거나 적어도 그대로 읽는다.
    """
    from openpyxl import load_workbook
    wb = load_workbook(uploaded_file, data_only=True)
    out = []
    ordmap = {}   # (section, category) → 등장 순서. 카테고리 좌/우 순서를 그대로 보존(정렬용).
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 300)):
            for cell in row:
                v = str(cell.value).strip() if cell.value is not None else ""
                if "우선순위" not in v:
                    continue
                sec = next((s for s in _PRI_SECTIONS if s in v), None)
                if sec is None:
                    continue
                tr = cell.row
                hdr = [(c.column, str(c.value).strip())
                       for c in ws[tr + 1] if c.value not in (None, "")]
                for col, cat in hdr:
                    ordmap.setdefault((sec, cat), len(ordmap))
                    for r in range(tr + 2, min(tr + 12, ws.max_row + 1)):
                        no = ws.cell(r, col).value
                        txt = ws.cell(r, col + 1).value
                        if (no is None or str(no).strip() == "") and \
                           (txt is None or str(txt).strip() == ""):
                            break   # 번호도 내용도 없으면 그 카테고리 끝
                        if txt is None or str(txt).strip() == "":
                            continue   # 번호만 있고 내용이 빈 줄은 건너뜀
                        try:
                            no_i = int(float(str(no).strip()))
                        except (TypeError, ValueError):
                            no_i = r - (tr + 1)
                        out.append((sec, cat, ordmap[(sec, cat)], no_i, str(txt).strip()))
    df = pd.DataFrame(out, columns=["section", "category", "ord", "no", "content"]).drop_duplicates()
    if df.empty:
        raise ValueError("'당월 우선순위'/'금주 우선순위' 제목을 찾지 못했어요 — "
                         "온라인팀 우선순위 업로드 FORM 양식 그대로인지 확인해 주세요.")
    return df.reset_index(drop=True)


def replace_priority(p):
    eng = get_engine()
    with eng.begin() as conn:
        p.astype(str).to_sql(PRIORITY_TABLE, conn, if_exists="replace", index=False)
    return len(p)


@st.cache_data(ttl=21600)
def load_priority():
    eng = get_engine()
    try:
        with eng.connect() as conn:
            exists = conn.exec_driver_sql(
                "SELECT 1 FROM information_schema.tables WHERE table_name=%s"
                if eng.dialect.name == "postgresql" else
                "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
                (PRIORITY_TABLE,)).fetchone()
            if not exists:
                return pd.DataFrame()
            p = pd.read_sql(f'SELECT * FROM "{PRIORITY_TABLE}"', conn)
    except Exception:
        return pd.DataFrame()
    if not p.empty:
        for c in ("section", "category", "content"):
            p[c] = p[c].astype(str).str.strip()
        for c in ("ord", "no"):
            if c in p.columns:
                p[c] = pd.to_numeric(p[c], errors="coerce").fillna(0).astype(int)
        if "ord" not in p.columns:   # 구버전 테이블 호환
            p["ord"] = 0
        p = p.sort_values(["section", "ord", "no"], kind="stable")
    return p


def priority_row_count():
    try:
        with get_engine().connect() as conn:
            return conn.exec_driver_sql(f'SELECT COUNT(*) FROM "{PRIORITY_TABLE}"').scalar()
    except Exception:
        return 0


# 사업계획 → 주간보고 행 매핑
_PLAN_BRAND = {"S/D/L": "S/D/L", "A (CODI GALLERY)": "A", "0 (ZERO LOUNGE)": "0",
               "J (GENTLEMENS)": "J", "N (NORATED)": "N"}
_PLAN_SUB = {"합계": "합계", "신상": "신상+내년신상", "1년차": "1년차", "2년차": "2년차", "3년차": "3년차"}
_PLAN_CH_KW = {"원래직입점": "원래", "웹뜰이관": "웹뜰", "웍스바이이관": "웍스"}


def _plan_cum(plan, master):
    """연간 전체(1~12월) 사업계획. dict{store, brand, cs} 반환. 없으면 None."""
    if plan is None or plan.empty:
        return None
    p = plan
    store = p[p["dim"] == "store"].groupby("code")["amount"].sum().to_dict()
    brand = p[p["dim"] == "brand"].groupby(["code", "sub"])["amount"].sum().to_dict()
    cs = {"원래": [], "웹뜰": [], "웍스": []}
    if master is not None and not master.empty and "채널스토리" in master.columns:
        for _, mr in master.iterrows():
            story = str(mr.get("채널스토리", ""))
            code = str(mr.get("매장코드", "")).strip()
            for kw in cs:
                if kw in story:
                    cs[kw].append(code)
    return {"store": store, "brand": brand, "cs": cs}


def _plan_for(key, cum):
    if cum is None:
        return None
    sec, mid, sub = key
    store, brand, cs = cum["store"], cum["brand"], cum["cs"]
    if mid == "G.TOTAL":
        return store.get("G.TOTAL")
    if sec == "유통별":
        if mid == "통합몰":
            return store.get("SD065")
        if mid == "네이버스토어":
            return (store.get("SD165") or 0) + (store.get("SD174") or 0)
        kw = _PLAN_CH_KW.get(mid)
        if kw:
            codes = cs.get(kw, [])
            return sum(store.get(c, 0) for c in codes) if codes else None
    if sec == "브랜드별":
        b = _PLAN_BRAND.get(mid); s = _PLAN_SUB.get(sub)
        if b and s:
            return brand.get((b, s))
    return None


def _store_annual():
    """매장코드 → 연간 사업계획 dict."""
    plan = load_plan()
    if plan is None or plan.empty:
        return {}
    s = plan[plan["dim"] == "store"]
    return s.groupby("code")["amount"].sum().to_dict()


def inject_plan(by, idx, master):
    """누계 블록 by[key]에 '사업계획'(연간)·'진도율' 주입. 진도율 = 26누계실적 ÷ 연간계획."""
    cum = _plan_cum(load_plan(), master)
    for key in idx:
        p = _plan_for(key, cum)
        r = by.get(key)
        if r is None:
            continue
        r["사업계획"] = p
        act = r.get("cy실판가")
        r["진도율"] = (act / p) if (p and act is not None) else None


def inject_plan_manager(by, idx, master):
    """담당별 표 by[key]에 '사업계획'(담당자 매장 연간계획 합)·'진도율' 주입."""
    store = _store_annual()
    mgr_codes = {}
    if master is not None and not master.empty and "담당자" in master.columns:
        for _, mr in master.iterrows():
            m = str(mr.get("담당자", "")).strip()
            c = str(mr.get("매장코드", "")).strip()
            if m:
                mgr_codes.setdefault(m, []).append(c)
    for key in idx:
        _, mid, _ = key
        if mid == "G.TOTAL":
            p = store.get("G.TOTAL")
        else:
            codes = mgr_codes.get(mid, [])
            p = sum(store.get(c, 0) for c in codes) if (codes and store) else None
        r = by.get(key)
        if r is None:
            continue
        r["사업계획"] = p
        act = r.get("cy실판가")
        r["진도율"] = (act / p) if (p and act is not None) else None


# ==============================================================================
# 주간회의 보고자료  ─ 당월실적 / 연간누계 (전년 동기간 비교)
# ==============================================================================
SDL_BRANDS = ["STCO", "DIEMS", "GENDERLESS"]
WK_MONEY = ["실판가", "사업계획"]

# 유통별 5개 분류 기준 (요약행 + 매장 드릴다운 공용 · 단일 소스)
_CHANNEL_MASKS = {
    "통합몰":      lambda x: x["매장코드"].astype(str).str.strip().isin(["SD065"]),
    "네이버스토어": lambda x: x["매장코드"].astype(str).str.strip().isin(["SD165", "SD174"]),
    "원래직입점":   lambda x: x["_채널스토리"].astype(str).str.contains("원래", na=False),
    "웹뜰이관":     lambda x: x["_채널스토리"].astype(str).str.contains("웹뜰", na=False),
    "웍스바이이관": lambda x: x["_채널스토리"].astype(str).str.contains("웍스", na=False),
}

# 브랜드별 5개 분류 기준 (드릴다운3 "유통/브랜드 선택" 필터 전용 · _wk_rows 브랜드별 행과 동일 기준)
_BRAND_MASKS = {
    "S/D/L":       lambda x: x["브랜드명"].isin(SDL_BRANDS),
    "CODI GALLERY": lambda x: x["브랜드명"] == "CODI GALLERY",
    "ZERO LOUNGE": lambda x: x["브랜드명"] == "ZERO LOUNGE",
    "GENTLEMENS":  lambda x: x["브랜드명"] == "GENTLEMENS PHILOSOPHY",
    "NORATED":     lambda x: x["브랜드명"] == "NORATED",
}

# 드릴다운3 "연차" 필터 버킷 — None은 "위 4개 버킷에 없는 나머지 연차 전부(4년차 이상)"를 뜻하며
# 실행 시점에 실제 데이터의 연차값을 훑어 동적으로 채운다(연차 표기가 "4년차"/"5년차"/"4년차↑" 등
# 데이터마다 다를 수 있어 하드코딩하지 않음).
_AGE_BUCKET_DEFS = [
    ("신상+내년신상", ["신상", "내년신상"]),
    ("1년차", ["1년차"]),
    ("2년차", ["2년차"]),
    ("3년차", ["3년차"]),
    ("4년차↑", None),
]

# 드릴다운3 "아이템 or 매장" 선택지 — '아이템'은 중카테고리(아이템그룹) 기준 breakdown(드릴다운2와
# 동일한 render_weekly_item_drilldown 재사용), '매장별'은 필터에 해당하는 매장 목록(드릴다운1과
# 동일한 render_weekly_drilldown 재사용).
WK_DIM_OPTS = ["아이템", "매장별"]


def _wk_metrics(cur_sub, prev_sub, total_c):
    r26 = float(cur_sub["_매출액"].sum()); r25 = float(prev_sub["_매출액"].sum())
    o26 = float(cur_sub["_최초가매출"].sum()); o25 = float(prev_sub["_최초가매출"].sum())
    pg26 = (r26 / o26) if o26 else 0.0; pg25 = (r25 / o25) if o25 else 0.0
    return {
        "py실판가": r25, "py판가율": pg25,
        "cy실판가": r26, "증감율": ((r26 - r25) / r25) if r25 else None,
        "비중": (r26 / total_c) if total_c else 0.0, "cy판가율": pg26,
        "편차": pg26 - pg25,
    }


def _wk_block(cur, prev, rows):
    total_c = float(cur["_매출액"].sum())
    out = {}
    for key, mask in rows:
        cs = cur[mask(cur)] if len(cur) else cur
        ps = prev[mask(prev)] if len(prev) else prev
        out[key] = _wk_metrics(cs, ps, total_c)
    return out


def _wk_rows():
    def code(x, cs): return x["매장코드"].astype(str).str.strip().isin(cs)
    def story(x, kw): return x["_채널스토리"].astype(str).str.contains(kw, na=False)  # 핵심단어 유연매칭
    def brand(x, ns): return x["브랜드명"].isin(ns)
    def age(x, a): return x["연차"].isin(a)
    return [
        (("전체", "G.TOTAL", "합계"), lambda x: pd.Series(True, index=x.index)),
        (("유통별", "통합몰", "합계"), _CHANNEL_MASKS["통합몰"]),
        (("유통별", "네이버스토어", "합계"), _CHANNEL_MASKS["네이버스토어"]),
        (("유통별", "원래직입점", "합계"), _CHANNEL_MASKS["원래직입점"]),
        (("유통별", "웹뜰이관", "합계"), _CHANNEL_MASKS["웹뜰이관"]),
        (("유통별", "웍스바이이관", "합계"), _CHANNEL_MASKS["웍스바이이관"]),
        (("브랜드별", "S/D/L", "합계"), lambda x: brand(x, SDL_BRANDS)),
        (("브랜드별", "S/D/L", "신상"), lambda x: brand(x, SDL_BRANDS) & age(x, ["신상", "내년신상"])),
        (("브랜드별", "S/D/L", "1년차"), lambda x: brand(x, SDL_BRANDS) & age(x, ["1년차"])),
        (("브랜드별", "S/D/L", "2년차"), lambda x: brand(x, SDL_BRANDS) & age(x, ["2년차"])),
        (("브랜드별", "S/D/L", "3년차"), lambda x: brand(x, SDL_BRANDS) & age(x, ["3년차"])),
        (("브랜드별", "A (CODI GALLERY)", "합계"), lambda x: brand(x, ["CODI GALLERY"])),
        (("브랜드별", "0 (ZERO LOUNGE)", "합계"), lambda x: brand(x, ["ZERO LOUNGE"])),
        (("브랜드별", "J (GENTLEMENS)", "합계"), lambda x: brand(x, ["GENTLEMENS PHILOSOPHY"])),
        (("브랜드별", "N (NORATED)", "합계"), lambda x: brand(x, ["NORATED"])),
    ]


def _wk_fmt(block, sub, v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return "–"
    if sub == "사업계획":
        return f"{v/1e6:,.1f}"   # 백만원 · 소수1자리
    if sub == "진도율":
        return f"{v*100:.0f}%"
    if "실판가" in sub:
        return f"{v/1e6:,.1f}"   # 룰1: 백만원 · 소수1자리
    if "판가율" in sub:
        return f"{v*100:.1f}%"
    if sub == "증감율":
        return f"{v*100:+.1f}%"
    if sub == "비중":
        return f"{v*100:.1f}%"
    if sub == "편차":
        return f"{v*100:+.1f}%p"
    return v


def weekly_excel_bytes(rows, bm, by, asof, cy, py):
    """팀 주간보고 양식(weekly_template.xlsx)을 템플릿으로 열어 매출현황·마감일·특이사항만 채워 반환."""
    import os
    from openpyxl import load_workbook
    tpl = os.path.join(os.path.dirname(os.path.abspath(__file__)), "weekly_template.xlsx")
    wb = load_workbook(tpl)
    ws = wb["주간보고"] if "주간보고" in wb.sheetnames else wb.active

    ws["P1"] = f"마감일: {str(cy)[-2:]}년 {asof.month:02d}월 {asof.day:02d}일"

    row_map = {
        ("전체", "G.TOTAL", "합계"): 12,
        ("유통별", "통합몰", "합계"): 13, ("유통별", "네이버스토어", "합계"): 14,
        ("유통별", "원래직입점", "합계"): 15, ("유통별", "웹뜰이관", "합계"): 16,
        ("유통별", "웍스바이이관", "합계"): 17,
        ("브랜드별", "S/D/L", "합계"): 18, ("브랜드별", "S/D/L", "신상"): 19,
        ("브랜드별", "S/D/L", "1년차"): 20, ("브랜드별", "S/D/L", "2년차"): 21,
        ("브랜드별", "S/D/L", "3년차"): 22,
        ("브랜드별", "A (CODI GALLERY)", "합계"): 23, ("브랜드별", "0 (ZERO LOUNGE)", "합계"): 24,
        ("브랜드별", "J (GENTLEMENS)", "합계"): 25, ("브랜드별", "N (NORATED)", "합계"): 26,
    }

    def setc(r, col, v):
        c = ws.cell(r, col)
        if v is None or (isinstance(v, float) and pd.isna(v)):
            c.value = "–"
        else:
            c.value = float(v)

    for key, r in row_map.items():
        m = bm.get(key, {}); y = by.get(key, {})
        setc(r, 4, m.get("py실판가")); setc(r, 5, m.get("py판가율")); setc(r, 6, m.get("cy실판가"))
        setc(r, 7, m.get("증감율")); setc(r, 8, m.get("비중")); setc(r, 9, m.get("cy판가율"))
        setc(r, 10, m.get("편차"))
        setc(r, 11, y.get("py실판가")); setc(r, 12, y.get("py판가율")); setc(r, 14, y.get("cy실판가"))
        setc(r, 16, y.get("증감율")); setc(r, 17, y.get("비중")); setc(r, 18, y.get("cy판가율"))
        setc(r, 19, y.get("편차"))

    def pc(v):
        return "N/A" if (v is None or (isinstance(v, float) and pd.isna(v))) else f"{v*100:+.0f}%"
    G = bm.get(("전체", "G.TOTAL", "합계"), {})
    TM = bm.get(("유통별", "통합몰", "합계"), {}); NV = bm.get(("유통별", "네이버스토어", "합계"), {})
    j26 = (TM.get("cy실판가") or 0) + (NV.get("cy실판가") or 0)
    j25 = (TM.get("py실판가") or 0) + (NV.get("py실판가") or 0)
    jasa = ((j26 - j25) / j25) if j25 else None
    OW = bm.get(("유통별", "원래직입점", "합계"), {}); WT = bm.get(("유통별", "웹뜰이관", "합계"), {})
    WK = bm.get(("유통별", "웍스바이이관", "합계"), {})
    gt = G.get("증감율")
    trend = "상승" if (gt or 0) >= 0 else "하락"
    ws["A29"] = (f"1. 당월실적 전년대비 {pc(gt)} {trend} 추세\n"
                 f"2. 통합몰은 {pc(TM.get('증감율'))} , 네이버스토어 {pc(NV.get('증감율'))}. "
                 f"자사채널 전체는 {pc(jasa)} 추세\n"
                 f"3. 원래직입점 {pc(OW.get('증감율'))} , 웹뜰이관 {pc(WT.get('증감율'))}, "
                 f"웍스바이이관 {pc(WK.get('증감율'))} 추세")

    # ── 온라인팀 우선순위(당월·금주) 자동 채움 (260830 신규) ─────────────────────
    # 사이드바에 업로드해 둔 우선순위를 템플릿의 '1. 당월 우선순위'·'4. 금주 우선순위' 빈칸에 채운다.
    # 템플릿 쪽도 제목 텍스트로 위치를 찾으므로(고정 행번호 아님) 양식이 몇 줄 밀려도 안전하고,
    # 업로드가 없으면 기존처럼 빈칸 그대로 나간다. 카테고리는 좌/우 등장 순서로 대응시킨다.
    try:
        _pri = load_priority()
    except Exception:
        _pri = None
    if _pri is not None and not _pri.empty:
        for _row in ws.iter_rows(min_col=1, max_col=3):
            _tv = next((str(c.value).strip() for c in _row if c.value not in (None, "")), "")
            if "우선순위" not in _tv:
                continue
            _sec = next((s for s in _PRI_SECTIONS if s in _tv), None)
            if _sec is None:
                continue
            _tr = _row[0].row
            _hdr = [c.column for c in ws[_tr + 1] if c.value not in (None, "")]
            _sub = _pri[_pri["section"] == _sec]
            _cats = list(dict.fromkeys(_sub["category"]))
            for _hi, _col in enumerate(_hdr):
                if _hi >= len(_cats):
                    continue
                _items = _sub[_sub["category"] == _cats[_hi]].sort_values("no")
                # 번호 칸이 미리 채워진 행들(1·2·3…)이 이 카테고리의 슬롯
                _slots = []
                _r = _tr + 2
                while _r <= ws.max_row:
                    _nv = ws.cell(_r, _col).value
                    if _nv is None or str(_nv).strip() == "":
                        break
                    _slots.append(_r)
                    _r += 1
                for _j, (_, _it) in enumerate(_items.iterrows()):
                    if _j >= len(_slots):
                        break   # 템플릿 슬롯(기본 3칸)보다 많은 항목은 엑셀엔 생략
                    ws.cell(_slots[_j], _col).value = int(_it["no"])
                    ws.cell(_slots[_j], _col + 1).value = str(_it["content"])

    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


def _wk_style_table(bm, by, idx, cy, py, click_ns=None, period=None):
    """주간보고 프레임(당월+누계 · 동일 컬럼)으로 (bm,by,idx)를 스타일 표로 변환. 메인표·담당별표 공용.

    반환: (Styler, 표시용 DataFrame) — 표시용 DF는 엑셀 다운로드(룰11)에 사용.

    period(항목26 · 260828 추가): "당월 실적"이면 당월 7컬럼만, "연간누계"면 누계 9컬럼만 렌더.
    None(기본값)이면 기존과 100% 동일(두 블록 모두). 한 블록만 표시 중엔 룰12 경계선을 긋지 않고,
    "연간누계"만 표시 중엔 클릭 대상 셀(당월 올해 실판가)이 없으므로 click_ns 링크 래핑도 생략된다.

    click_ns(2026-08-18 추가): 문자열을 주면 **당월 실적의 '올해 실판가' 컬럼 셀만** 클릭 가능한
    링크(<span class="wk-pnclick">)로 감싼다. 이 표를 그리는 호출부에서 _wk_pn_click_bridge()로
    숨은 버튼과 연결해줘야 실제로 팝업이 뜬다(아이템그룹별 상세표 전용 — 다른 표는 기본값 None
    이라 지금까지와 완전히 동일하게 렌더된다). 엑셀 다운로드에 쓰는 disp는 링크를 씌우기 **전**
    값을 그대로 돌려주므로 다운로드 파일엔 HTML 태그가 섞이지 않는다.
    """
    MON, YTD = "당월 실적", "연간누계"
    sy, sc = str(py)[-2:], str(cy)[-2:]   # 룰2: 연도 2자리
    # 컬럼 순서 (2026-07-31 팀장님 지정): 실판가 25→26 → 증감율 → 비중 → (누계: 사업계획→진도율) → 판가율 25→26 → 편차
    mcols = [(MON, f"{sy}실판가"), (MON, f"{sc}실판가"), (MON, "증감율"), (MON, "비중"),
             (MON, f"{sy}판가율"), (MON, f"{sc}판가율"), (MON, "편차")]
    ycols = [(YTD, f"{sy}실판가"), (YTD, f"{sc}실판가"), (YTD, "증감율"), (YTD, "비중"),
             (YTD, "사업계획"), (YTD, "진도율"),
             (YTD, f"{sy}판가율"), (YTD, f"{sc}판가율"), (YTD, "편차")]
    # 항목26(260828): 표시 기간 필터 — 한 블록만 고르면 그 블록 컬럼만 남긴다
    use_m = period != YTD
    use_y = period != MON
    if not use_m:
        mcols = []
    if not use_y:
        ycols = []

    def cellval(block_res, key, sub):
        r = block_res[key]
        if "실판가" in sub:
            return r["py실판가"] if sub.startswith(sy) else r["cy실판가"]
        if "판가율" in sub:
            return r["py판가율"] if sub.startswith(sy) else r["cy판가율"]
        if sub in ("사업계획", "진도율"):
            return r.get(sub)
        return r[sub]

    data = [[cellval(bm, k, s[1]) for s in mcols] + [cellval(by, k, s[1]) for s in ycols] for k in idx]
    D = pd.DataFrame(data, index=pd.MultiIndex.from_tuples(idx),
                     columns=pd.MultiIndex.from_tuples(mcols + ycols))
    disp = D.copy()
    for col in disp.columns:
        disp[col] = [_wk_fmt(col[0], col[1], v) for v in D[col]]

    def _color(col):
        if col[1] not in ("증감율", "편차"):
            return ["" for _ in D[col]]
        return ["color:#c62828;font-weight:600" if (pd.notnull(v) and v < 0)
                else ("color:#1f8a4c;font-weight:600" if (pd.notnull(v) and v > 0) else "") for v in D[col]]

    # 260818: 클릭 가능 셀(당월 올해 실판가)만 링크로 감싼다 — 엑셀용 disp는 건드리지 않기 위해
    # style 전용 사본(sdisp)을 따로 만든다. Styler는 셀 값을 그대로(escape 없이) HTML에 넣으므로
    # <span>이 그대로 살아난다(pandas 2.x·3.x 양쪽에서 확인).
    sdisp = disp
    if click_ns and use_m:   # 항목26: 당월 블록이 보일 때만 클릭 링크(연간누계만 표시 중엔 대상 셀 없음)
        sdisp = disp.copy()
        _ccol = (MON, f"{sc}실판가")
        sdisp[_ccol] = [
            (v if (v is None or str(v) == "–")
             else f"<span class='wk-pnclick' data-wkpn='{click_ns}#{i}'>{v}</span>")
            for i, v in enumerate(disp[_ccol])
        ]
    sty = sdisp.style
    for col in D.columns:
        if col[1] in ("증감율", "편차"):
            sty = sty.apply(lambda s, c=col: _color(c), subset=pd.IndexSlice[:, [col]])
    sty = sty.set_properties(**{"text-align": "right"})
    if use_m and use_y:   # 룰12: 당월/누계 경계선 — 두 블록이 다 보일 때만 (항목26)
        sty = block_border(sty, len(mcols))
    return sty, disp


# ──────────────────────────────────────────────────────────────────────────────
# 260818 신규 — 아이템그룹별 상세표의 "당월 올해 실판가" 숫자 클릭 → 품번별 상세 팝업
#   (A) 그 조건·그 아이템그룹의 품번별 판매현황
#   (B) 품번별 "판매수량 기준" 상위 3개 매장(매장명·실판가·판가율)
#
# ⚠️ 클릭 방식에 대한 설계 메모 (실측 근거 있음, 바꾸기 전 반드시 읽을 것)
#   메인 표들은 커스텀 HTML(render_styled_table)로 그려져서 Streamlit이 셀 클릭을 직접
#   받지 못한다. 후보가 둘이었는데 로컬 Streamlit + Playwright로 직접 실험해 확인했다.
#     ① <a href="?param=..."> 쿼리파라미터 링크 → 클릭 시 **페이지가 통째로 새로 뜨면서
#        session_state가 전부 초기화됨**(카운터 3 → 0으로 리셋되는 것 확인). 이 앱은
#        로그인·조회게이트(_need_search)·각종 필터가 전부 session_state에 있어서 못 쓴다.
#     ② 숨은 st.button을 만들어 두고, 셀 클릭을 그 버튼 클릭으로 넘기는 브리지(아래 방식)
#        → **리로드 없음·session_state 그대로 유지**(카운터 3 유지, URL 불변, 연속 클릭도 정상).
#   그래서 ②를 채택했다. 브리지는 순수 클라이언트(JS)라 파이썬 쪽 메모리·연산 부담이 0이고,
#   설령 JS가 막히더라도 표는 그대로 보이고 팝업만 안 뜨는 정도로 안전하게 실패한다.
# ──────────────────────────────────────────────────────────────────────────────

# 숨은 버튼 라벨에 붙이는 보이지 않는 표식(U+2063 INVISIBLE SEPARATOR) — JS가 이 표식으로
# 버튼을 찾는다. st.container(key=...)의 st-key- 클래스에 의존하지 않으므로 스트림릿 버전을 타지 않음.
_WK_PN_MARK = "\u2063wkpn\u2063"   # U+2063 INVISIBLE SEPARATOR로 감싼 표식

_WK_PN_CSS = """
<style>
span.wk-pnclick{color:#0071e3;cursor:pointer;border-bottom:1px dashed rgba(0,113,227,.55);}
span.wk-pnclick:hover{background:#eaf2ff;border-radius:3px;}
</style>
"""


# 숨은 버튼을 화면에서만 지우는 CSS(visually-hidden). **display:none을 쓰지 않는다** —
# display:none이면 렌더 트리에서 빠져 innerText가 빈 문자열이 되는 브라우저가 있어, 라벨로 버튼을
#찾던 로직이 조용히 실패한다(260818 "클릭해도 팝업이 안 뜬다" 리포트의 유력 원인 중 하나).
_WK_PN_HIDE_PROPS = ("position:absolute!important;width:1px!important;height:1px!important;"
                     "margin:-1px!important;padding:0!important;border:0!important;"
                     "overflow:hidden!important;clip:rect(0 0 0 0)!important;opacity:0!important;")


def _copy_shortcut_guard():
    """표에서 값을 드래그해 **Ctrl+C(맥은 ⌘+C)로 복사할 때 'Clear caches' 창이 뜨는 것**을 막는다.

    원인: Streamlit 자체 단축키에 `C` = Clear caches, `R` = Rerun 이 있는데, 이 핸들러가
    Ctrl/⌘가 눌린 상태인지 보지 않아서 **복사 단축키까지 자기 단축키로 받아버린다**.
    입력창 안에서는 무시되지만, 우리 표는 커스텀 HTML이라 입력창이 아니어서 그대로 걸린다
    (260818 중태님 리포트: 팝업에서 품번을 복사하려니 Clear caches 확인창이 떴음).

    대응: 부모 창의 keydown을 **캡처 단계에서 가장 먼저** 받아, 아래 둘 중 하나면
    `stopImmediatePropagation()`으로 Streamlit 핸들러에 도달하지 못하게 한다.
      ① Ctrl/⌘ 가 눌린 채로 C·X·V·A·R  → 복사/잘라내기/붙여넣기/전체선택/새로고침
      ② **화면에 드래그로 선택된 텍스트가 있는 상태**에서 C·R
    ②를 넣은 이유: 스트림릿 버전에 따라 Ctrl 감지 자체가 다르게 동작해(로컬 1.61에서는 Ctrl+C가
    안 걸리는데 배포본에서는 걸렸음) ①만으로는 못 막는 경우가 있고, 한글 IME가 켜져 있으면
    `e.key`가 'ㅊ'로 와서 글자 비교도 빗나가기 때문. "값을 선택해 둔 상태 = 복사하려는 중"이라는
    상황 자체로 막는 게 가장 확실하다. 글자 비교도 `e.key`와 **물리 키(`e.code`)를 함께** 본다.

    `preventDefault()`는 **하지 않으므로** 브라우저 기본 동작(복사·붙여넣기·전체선택·새로고침)은
    그대로 동작한다. 아무것도 선택하지 않은 상태에서 그냥 `C`·`R`을 누르는 원래 Streamlit 단축키도
    평소처럼 살아있다.
    """
    components.html(
        """
<script>
(function(){
  var w = window.parent;
  if(!w || w.__wkCopyGuard) return;
  w.__wkCopyGuard = true;
  var KEYS  = ['c', 'x', 'v', 'a', 'r'];
  var CODES = ['KeyC', 'KeyX', 'KeyV', 'KeyA', 'KeyR'];
  function hasSelection(){
    try{
      var s = w.getSelection && w.getSelection();
      return !!(s && String(s).length > 0);
    }catch(err){ return false; }
  }
  function guard(e){
    var k = (e.key || '').toLowerCase();
    var isKey = KEYS.indexOf(k) >= 0 || CODES.indexOf(e.code || '') >= 0;
    if(!isKey) return;
    // ① 복사 계열 단축키  ② 텍스트를 선택해 둔 상태(= 복사하려는 중)
    if((e.ctrlKey || e.metaKey) || hasSelection()){
      e.stopImmediatePropagation();   // Streamlit 단축키만 차단 (기본 동작은 그대로)
    }
  }
  // keydown 하나만 막으면 안 된다 — Streamlit 버전에 따라 keypress/keyup에서 단축키를 처리하기도
  // 해서, 실제로 keydown만 막았을 때 창이 그대로 떴다(260818 로컬 실측). 3가지 모두 캡처한다.
  ['keydown', 'keypress', 'keyup'].forEach(function(t){
    w.addEventListener(t, guard, true);        // ← 캡처 단계: 다른 핸들러보다 먼저 받는다
    if(w.document) w.document.addEventListener(t, guard, true);
  });
})();
</script>
""",
        height=0,
    )


def _wk_pn_hide_css(ns, n):
    """숨은 버튼 n개를 CSS로 즉시 감춘다 — st.button(key=...)이 컨테이너에 붙여주는
    `.st-key-<key>` 클래스를 그대로 쓴다. JS를 기다리지 않아 깜빡임이 없다."""
    sels = ", ".join(f".st-key-wkpnb_{ns}_{i}" for i in range(n))
    return f"<style>{sels} {{{_WK_PN_HIDE_PROPS}}}</style>" if n else ""


def _wk_pn_click_bridge(ns):
    """숫자 셀(span.wk-pnclick[data-wkpn="ns#i"]) 클릭 → 숨은 st.button 클릭으로 연결.

    ⚠️ 260818 2차 수정 — 배포 화면에서 "클릭해도 팝업이 안 뜬다"는 리포트가 있어, 버튼을 찾는
    방법을 통째로 더 튼튼하게 바꿨다. 되돌리기 전에 아래를 꼭 읽을 것.

    버튼 찾기는 **3단계**로 시도한다(앞이 실패하면 뒤로 넘어감):
      1순위 `.st-key-wkpnb_<ns>_<i>` — st.button(key=...)이 컨테이너에 자동으로 붙여주는 클래스.
            보이지 않는 문자·라벨 렌더 방식과 **무관**해서 가장 안전하다(현재 배포 버전에 존재 확인).
      2순위 라벨의 보이지 않는 표식(U+2063) 매칭 — 구버전 Streamlit(st-key 클래스가 없던 시절) 대비.
            읽을 때 innerText가 아니라 **textContent**를 쓴다(렌더 여부와 무관하게 항상 값이 나옴).
      3순위 문서 순서 — 표식이 붙은 버튼들의 순서가 곧 셀 순서(#0, #1, …)라, 라벨을 전혀 못 읽어도
            i번째 버튼을 누르면 된다.

    숨기기는 파이썬이 내보낸 CSS(_wk_pn_hide_css)가 담당하고, 이 스크립트는 st-key 클래스가 없는
    구버전을 위해 한 번 더 인라인 스타일로 감춘다(둘 다 display:none이 아님).

    매 rerun마다 다시 실행되고, setInterval + MutationObserver로 재바인딩해서 Streamlit이 DOM을
    다시 그려도 클릭이 끊기지 않는다. height=0이라 자리도 안 차지한다. 이 스크립트가 통째로
    막히더라도 표는 그대로 보이고, 표 아래 '숫자 클릭이 안 될 때' 폴백 셀렉트로 같은 팝업을 열 수 있다.
    """
    components.html(
        """
<script>
(function(){
  var doc = window.parent && window.parent.document;
  if(!doc) return;
  var MARK = "\\u2063wkpn\\u2063";
  var HIDE = "position:absolute!important;width:1px!important;height:1px!important;" +
             "margin:-1px!important;padding:0!important;border:0!important;" +
             "overflow:hidden!important;clip:rect(0 0 0 0)!important;opacity:0!important;";

  // 표식이 붙은 버튼들을 문서 순서대로 모으고(2·3순위용), 구버전 대비로 한 번 더 감춘다.
  function collect(){
    var list = [];
    var btns = doc.querySelectorAll('button');
    for(var i = 0; i < btns.length; i++){
      var b = btns[i];
      var t = (b.textContent || '').trim();          // innerText 아님(빈값 위험)
      if(t.indexOf(MARK) !== 0) continue;
      list.push({key: t.slice(MARK.length), el: b});
      var box = b.closest('[data-testid="stElementContainer"]') || b.parentElement;
      if(box && box.getAttribute('data-wkpnhidden') !== '1'){
        box.setAttribute('data-wkpnhidden', '1');
        box.setAttribute('style', (box.getAttribute('style') || '') + HIDE);
      }
    }
    return list;
  }

  function fire(want){
    if(!want) return false;
    var at = want.indexOf('#');
    var ns = at < 0 ? want : want.slice(0, at);
    var ix = at < 0 ? -1 : parseInt(want.slice(at + 1), 10);

    // 1순위 — st.button(key=...)이 붙여주는 st-key 클래스로 정확히 지목
    if(ix >= 0){
      var box = doc.querySelector('.st-key-wkpnb_' + ns + '_' + ix);
      var b1 = box && box.querySelector('button');
      if(b1){ b1.click(); return true; }
    }
    var list = collect();
    // 2순위 — 라벨 표식 매칭
    for(var i = 0; i < list.length; i++){
      if(list[i].key === want){ list[i].el.click(); return true; }
    }
    // 3순위 — 문서 순서
    if(ix >= 0 && list[ix]){ list[ix].el.click(); return true; }
    return false;
  }

  function bind(){
    collect();
    var spans = doc.querySelectorAll('span.wk-pnclick');
    for(var i = 0; i < spans.length; i++){
      (function(el){
        if(el.getAttribute('data-wkbound') === '1') return;
        el.setAttribute('data-wkbound', '1');
        el.addEventListener('click', function(){ fire(el.getAttribute('data-wkpn')); });
      })(spans[i]);
    }
  }

  bind();
  if(!window.__wkpnTimer){ window.__wkpnTimer = setInterval(bind, 400); }
  if(!window.__wkpnObs && doc.body){
    window.__wkpnObs = new MutationObserver(function(){ bind(); });
    window.__wkpnObs.observe(doc.body, {childList: true, subtree: true});
  }
})();
</script>
""",
        height=0,
    )


def _wk_pn_top_detail(sub, top_n=3, limit=None):
    """조건에 맞는 원본 거래행(sub)에서 품번별 판매현황 + 품번별 상위 N개 매장을 한 표로 만든다.

    컬럼(중태님 첨부 양식 그대로):
      품번 · 판매수량 · 실판가 · 최초판매가 · 평균실판가 · 판가율
      + 매장명1 · 실판가 · 판가율 / 매장명2 · … / 매장명3 · …

    계산 정의는 앱 전체와 동일하게 맞춤(_agg_detail·_wk_metrics와 같은 원리):
      · 판매수량 = _수량 합 · 실판가 = _매출액 합
      · 최초판매가 = _최초가매출 ÷ _수량 (품번 단위 가중평균 단가)
      · 평균실판가 = _매출액 ÷ _수량      (품번 단위 가중평균 단가)
      · 판가율   = _매출액 ÷ _최초가매출  (가중 판가율, 룰과 동일)
    상위 매장은 **판매수량 기준**(중태님 확정, 2026-08-18)이며, 수량이 같으면 실판가가 큰 쪽이
    앞선다. 정렬은 품번 실판가 큰 순, limit을 주면 화면용으로 상위 N개만 잘라낸다(엑셀은 전체).
    """
    empty = pd.DataFrame()
    if sub is None or sub.empty or "품번" not in sub.columns:
        return empty, 0
    need = {"_수량", "_매출액", "_최초가매출"}
    if not need.issubset(set(sub.columns)):
        return empty, 0

    # 260818 2차 보강: load_db가 메모리 절감을 위해 품번·매장명을 category дtype으로 읽는 경우가 있어,
    # 그대로 groupby/map 하면 카테고리 정렬·매핑이 예상과 다르게 동작할 수 있다. 여기서 한 번
    # 문자열·숫자로 정규화한 작업용 프레임을 만들어 쓴다(원본 sub은 건드리지 않음).
    scol = "매장명" if "매장명" in sub.columns else ("매장코드" if "매장코드" in sub.columns else None)
    work = pd.DataFrame({
        "품번": sub["품번"].astype(str).str.strip(),
        "_수량": pd.to_numeric(sub["_수량"], errors="coerce").fillna(0.0).astype("float64"),
        "_매출액": pd.to_numeric(sub["_매출액"], errors="coerce").fillna(0.0).astype("float64"),
        "_최초가매출": pd.to_numeric(sub["_최초가매출"], errors="coerce").fillna(0.0).astype("float64"),
    })
    if scol is not None:
        work[scol] = sub[scol].astype(str).str.strip()

    g = work.groupby("품번", observed=True)
    base = pd.DataFrame({
        "판매수량": g["_수량"].sum(),
        "실판가": g["_매출액"].sum(),
        "_orig": g["_최초가매출"].sum(),
    })
    base = base[(base["판매수량"] != 0) | (base["실판가"] != 0)]
    if base.empty:
        return empty, 0
    q = pd.to_numeric(base["판매수량"], errors="coerce").replace(0, np.nan)
    o = pd.to_numeric(base["_orig"], errors="coerce").replace(0, np.nan)
    base["최초판매가"] = base["_orig"] / q
    base["평균실판가"] = base["실판가"] / q
    base["판가율"] = base["실판가"] / o
    base = base.sort_values("실판가", ascending=False)
    total_n = len(base)
    if limit is not None and total_n > limit:
        base = base.head(limit)
    base = base.reset_index()

    # ── 품번 × 매장 집계 → 품번별 상위 N개 매장 ────────────────────────────────
    store_cols = {}
    if scol is not None:
        keep = work[work["품번"].isin(set(base["품번"]))]
        sg = keep.groupby(["품번", scol], observed=True).agg(
            q=("_수량", "sum"), r=("_매출액", "sum"), o=("_최초가매출", "sum")).reset_index()
        sg = sg.sort_values(["품번", "q", "r"], ascending=[True, False, False])
        sg["_rank"] = sg.groupby("품번", observed=True).cumcount()
        sg = sg[sg["_rank"] < top_n]
        for i in range(top_n):
            part = sg[sg["_rank"] == i].set_index("품번")
            idx = base["품번"]
            store_cols[f"매장명{i+1}"] = idx.map(part[scol]).values
            store_cols[f"실판가{i+1}"] = idx.map(part["r"]).values
            _o = idx.map(part["o"]).replace(0, np.nan).values
            store_cols[f"판가율{i+1}"] = pd.Series(idx.map(part["r"]).values) / pd.Series(_o)
    else:
        for i in range(top_n):
            store_cols[f"매장명{i+1}"] = np.nan
            store_cols[f"실판가{i+1}"] = np.nan
            store_cols[f"판가율{i+1}"] = np.nan

    out = base[["품번", "판매수량", "실판가", "최초판매가", "평균실판가", "판가율"]].copy()
    for k, v in store_cols.items():
        out[k] = np.asarray(v)
    return out, total_n


def _wk_pn_total_row(sub):
    """팝업 맨 위 '합계' 행 — 화면에 몇 개만 보여주든 항상 **그 조건 전체** 기준으로 계산한다.

    이 합계의 '실판가'는 사용자가 클릭한 표 셀(당월 올해 실판가)의 원래 금액과 같아야 한다 —
    팝업이 그 숫자를 품번으로 쪼갠 것이기 때문. (검증용으로도 쓰라고 일부러 맨 위에 둠)
    """
    if sub is None or sub.empty:
        return None
    qty = float(pd.to_numeric(sub["_수량"], errors="coerce").fillna(0).sum())
    rev = float(pd.to_numeric(sub["_매출액"], errors="coerce").fillna(0).sum())
    org = float(pd.to_numeric(sub["_최초가매출"], errors="coerce").fillna(0).sum())
    return {
        "판매수량": qty, "실판가": rev,
        "최초판매가": (org / qty) if qty else np.nan,
        "평균실판가": (rev / qty) if qty else np.nan,
        "판가율": (rev / org) if org else np.nan,
    }


def _wk_pn_fmt_table(det, top_n=3, total_row=None):
    """_wk_pn_top_detail 결과를 화면·엑셀 공용의 2단 헤더 표시용 DataFrame으로 포맷.

    total_row(dict)를 주면 맨 윗줄에 '합계' 행을 붙인다 — 공통표시룰(룰6)상 표의 첫 행은
    화면·엑셀 모두 노란색으로 강조되므로, 첫 행이 합계여야 다른 표들과 의미가 일관된다.

    ※ 상위 매장 쪽 '실판가'·'판가율' 헤더는 첨부 양식대로 숫자 없이 그대로 보여야 하는데,
      pandas 컬럼은 중복되면 다루기 곤란해서 뒤에 보이지 않는 문자(U+200B)를 1~2개 붙여
      **화면 표기는 동일하게, 내부 키는 유일하게** 만든다.
    """
    G1, G2 = "품번별 판매 현황", f"판매 상위 {top_n}개 매장"
    cols, data = [], {}
    n = len(det)

    def _num(v, pct=False):
        if v is None or pd.isna(v):
            return "–"
        return f"{v*100:.1f}%" if pct else f"{v:,.0f}"

    def put(grp, name, vals, head=None):
        key = (grp, name)
        cols.append(key)
        data[key] = ([head] if total_row is not None else []) + list(vals)

    put(G1, "품번", [str(v) for v in det["품번"]], head="전체")   # 맨 왼쪽 구분칸이 이미 "합계"
    for c in ("판매수량", "실판가", "최초판매가", "평균실판가"):
        put(G1, c, [_num(v) for v in det[c]], head=_num((total_row or {}).get(c)))
    put(G1, "판가율", [_num(v, pct=True) for v in det["판가율"]],
        head=_num((total_row or {}).get("판가율"), pct=True))
    for i in range(1, top_n + 1):
        pad = "\u200b" * (i - 1)      # 화면엔 안 보이는 중복 방지용 꼬리표(ZERO WIDTH SPACE)
        put(G2, f"매장명{i}", ["–" if pd.isna(v) else str(v) for v in det[f"매장명{i}"]], head="–")
        put(G2, f"실판가{pad}", [_num(v) for v in det[f"실판가{i}"]], head="–")
        put(G2, f"판가율{pad}", [_num(v, pct=True) for v in det[f"판가율{i}"]], head="–")
    disp = pd.DataFrame(data, columns=pd.MultiIndex.from_tuples(cols))
    disp.index = (["합계"] if total_row is not None else []) + [str(i) for i in range(1, n + 1)]
    return disp


_WK_PN_SCREEN_LIMIT = 30   # 화면 표시 품번 수(중태님 확정) — 엑셀은 전체


def _wk_pn_popup(sub, title, caption, key_prefix, on_dismiss=None):
    """품번별 판매현황 + 상위 3개 매장 팝업(첨부 양식). 화면=상위 30개 / 엑셀=전체.

    260818 2차: 집계·렌더를 try/except로 감쌌다. 실데이터의 예상 못 한 값(빈 매장명, 이상한
    dtype 등)으로 예외가 나면 **팝업이 조용히 안 뜨는 대신** 팝업 안에 원인을 보여준다 —
    "클릭했는데 아무 일도 안 일어난다"가 가장 진단하기 어려운 상태라 일부러 드러낸다.
    """
    try:
        det_screen, total_n = _wk_pn_top_detail(sub, top_n=3, limit=_WK_PN_SCREEN_LIMIT)
        tot = _wk_pn_total_row(sub)
        err = None
    except Exception as e:                                        # noqa: BLE001
        det_screen, total_n, tot, err = pd.DataFrame(), 0, None, e

    @_dialog_or_expander(title, on_dismiss=on_dismiss)
    def _popup():
        st.caption(caption)
        if err is not None:
            st.error(f"품번별 상세를 만드는 중 오류가 났어요 — 이 문구를 그대로 전달해 주세요.\n\n"
                     f"`{type(err).__name__}: {err}`")
            return
        if det_screen.empty:
            st.info("해당 조건에 판매 데이터가 없어요.")
            return
        try:
            disp = _wk_pn_fmt_table(det_screen, top_n=3, total_row=tot)
            shown = len(det_screen)
            h1, h2 = st.columns([5, 1.3])
            h1.markdown(
                f"<span style='font-size:0.82rem;color:#555;'>품번 {total_n:,}개 중 실판가 큰 순 "
                f"{shown:,}개 표시 · 상위 매장은 <b>판매수량</b> 기준 · 맨 윗줄 합계는 전체 "
                f"{total_n:,}개 기준(클릭한 표 숫자와 같아야 정상)</span>"
                "<span style='float:right;color:#888;font-size:0.78rem;white-space:nowrap;'>"
                "[금액: 원 / VAT+]</span>", unsafe_allow_html=True)
            # 엑셀은 전체 품번 — 화면과 같은 함수·같은 서식(룰13)
            det_all, _ = _wk_pn_top_detail(sub, top_n=3, limit=None)
            xls = _wk_pn_fmt_table(det_all, top_n=3, total_row=tot)
            h2.download_button("⬇ 엑셀(전체)", table_excel_bytes(xls, "품번별 상세", first_block_cols=6),
                               file_name=f"{_safe_name(title)}_품번별상세.xlsx", mime=XLSX_MIME,
                               key=f"{key_prefix}_dl", use_container_width=True)
            sty = block_border(disp.style.set_properties(**{"text-align": "right"}), 6)   # 룰12: 경계선
            render_styled_table(sty)   # 룰6: 첫 행(=합계) 노란 강조 — 다른 표들과 동일한 의미로 맞춤
            if total_n > shown:
                st.caption(f"※ 화면엔 상위 {shown:,}개만 보여요 — 나머지 {total_n - shown:,}개까지 전부 "
                           "보려면 위 '⬇ 엑셀(전체)'를 받아주세요(합계 행은 언제나 전체 기준).")
        except Exception as e:                                    # noqa: BLE001
            st.error(f"표를 그리는 중 오류가 났어요 — 이 문구를 그대로 전달해 주세요.\n\n"
                     f"`{type(e).__name__}: {e}`")
    _popup()


def render_weekly_drilldown(cur_m, prev_m, cur_y, prev_y, label, mask, cy, py, show_plan=True,
                            period=None, big_title=False):
    """선택한 그룹(유통 또는 담당자)의 매장별 상세표 — 주간보고와 동일 형식(당월+누계). 비중=해당 그룹 내.

    period·big_title(항목26 · 260828): 드릴다운1 전용 — period("당월 실적"/"연간누계")면 그 블록만
    표시(경계선·엑셀 경계선도 함께 제거), big_title=True면 결과 타이틀을 ##### 헤딩 크기로 렌더.
    기본값이면 기존 화면(드릴다운 2·3 포함)과 100% 동일."""
    cm, pm = cur_m[mask(cur_m)], prev_m[mask(prev_m)]
    cyd, pyd = cur_y[mask(cur_y)], prev_y[mask(prev_y)]
    if cm.empty and cyd.empty and pm.empty and pyd.empty:
        st.info(f"'{label}'에 해당하는 매장 데이터가 없어요.")
        return

    tot_m = float(cm["_매출액"].sum())    # 당월 유통 total (비중 분모)
    tot_y = float(cyd["_매출액"].sum())   # 누계 유통 total

    codes = pd.Index(pd.concat([cyd["매장코드"], pyd["매장코드"]])
                     .astype(str).str.strip().replace({"nan": None, "none": None}).dropna().unique())
    lbl_src = pd.concat([cur_y, prev_y])
    name_map = {}
    if "매장명" in lbl_src.columns:
        tmp = lbl_src[["매장코드", "매장명"]].astype(str)
        name_map = dict(zip(tmp["매장코드"].str.strip(), tmp["매장명"]))

    def sub(frame, c):
        return frame[frame["매장코드"].astype(str).str.strip() == c]

    store_ann = _store_annual() if show_plan else {}   # 매장코드 → 연간 사업계획 (필터 시 생략)

    def _addplan(y, pl):
        y["사업계획"] = pl
        act = y.get("cy실판가")
        y["진도율"] = (act / pl) if (pl and act is not None) else None
        return y

    store_rows = []
    for c in codes:
        m = _wk_metrics(sub(cm, c), sub(pm, c), tot_m)
        y = _addplan(_wk_metrics(sub(cyd, c), sub(pyd, c), tot_y),
                     store_ann.get(str(c).strip()) if store_ann else None)
        rev_y = float(sub(cyd, c)["_매출액"].sum())
        store_rows.append((name_map.get(c, c) or c, m, y, rev_y))
    store_rows.sort(key=lambda t: -t[3])   # 누계 올해 매출 큰 순

    plan_head = sum(store_ann.get(str(c).strip(), 0) for c in codes) if store_ann else None
    head_y = _addplan(_wk_metrics(cyd, pyd, tot_y), plan_head)
    entries = [(f"{label} (합계)", _wk_metrics(cm, pm, tot_m), head_y)]
    entries += [(r[0], r[1], r[2]) for r in store_rows]

    sy, sc = str(py)[-2:], str(cy)[-2:]
    MON, YTD = "당월 실적", "연간누계"
    # 컬럼 순서 (2026-07-31 팀장님 지정): 실판가 25→26 → 증감율 → 비중 → (누계: 사업계획→진도율) → 판가율 25→26 → 편차
    mcols = [(MON, f"{sy}실판가"), (MON, f"{sc}실판가"), (MON, "증감율"), (MON, "비중"),
             (MON, f"{sy}판가율"), (MON, f"{sc}판가율"), (MON, "편차")]
    ycols = [(YTD, f"{sy}실판가"), (YTD, f"{sc}실판가"), (YTD, "증감율"), (YTD, "비중"),
             (YTD, "사업계획"), (YTD, "진도율"),
             (YTD, f"{sy}판가율"), (YTD, f"{sc}판가율"), (YTD, "편차")]
    # 항목26(260828): 표시 기간 필터 — 한 블록만 고르면 그 블록 컬럼만 남긴다
    use_m = period != YTD
    use_y = period != MON
    if not use_m:
        mcols = []
    if not use_y:
        ycols = []

    def val(metrics, s):
        if "실판가" in s:
            return metrics["py실판가"] if s.startswith(sy) else metrics["cy실판가"]
        if "판가율" in s:
            return metrics["py판가율"] if s.startswith(sy) else metrics["cy판가율"]
        return metrics.get(s)

    idx, data = [], []
    for _row_lbl, m, y in entries:   # ← label(선택 그룹명)을 덮어쓰지 않도록 별도 변수 사용
        idx.append(_row_lbl)
        data.append([val(m, s[1]) for s in mcols] + [val(y, s[1]) for s in ycols])
    D = pd.DataFrame(data, index=idx, columns=pd.MultiIndex.from_tuples(mcols + ycols))

    disp = D.copy()
    for col in disp.columns:
        disp[col] = [_wk_fmt(col[0], col[1], v) for v in D[col]]

    def _color(col):
        if col[1] not in ("증감율", "편차"):
            return ["" for _ in D[col]]
        return ["color:#c62828;font-weight:600" if (pd.notnull(v) and v < 0)
                else ("color:#1f8a4c;font-weight:600" if (pd.notnull(v) and v > 0) else "") for v in D[col]]
    sty = disp.style
    for col in D.columns:
        if col[1] in ("증감율", "편차"):
            sty = sty.apply(lambda s, c=col: _color(c), subset=pd.IndexSlice[:, [col]])
    sty = sty.set_properties(**{"text-align": "right"})
    if use_m and use_y:   # 룰12: 당월/누계 경계선 — 두 블록이 다 보일 때만 (항목26)
        sty = block_border(sty, len(mcols))

    # 룰11: 제목 + 우측 일반 엑셀 다운로드 버튼 (2026-07-31)
    if big_title:   # 항목26: 드릴다운1은 위 필터 박스와 간격을 살짝 띄운다
        st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)
    d1, d2 = st.columns([5, 1])
    _sub = (f"<span style='color:#888;font-size:0.8rem;font-weight:400;'>"
            f"(매장 {len(store_rows)}개 · 비중=해당 그룹 내 · 매출 큰 순)</span>")
    if big_title:   # 항목26: 드릴다운1 결과 타이틀은 다른 드릴다운 타이틀과 동일 크기(##### 헤딩)
        d1.markdown(f"##### 🔍 {label} · 매장별 상세  {_sub}{_NOTE_FLOAT}", unsafe_allow_html=True)
    else:
        d1.markdown(f"**🔍 {label} · 매장별 상세**  {_sub}{_NOTE_FLOAT}", unsafe_allow_html=True)
    d2.download_button("⬇ 엑셀", table_excel_bytes(disp, f"{label} 매장별",
                                                  first_block_cols=len(mcols) if (use_m and use_y) else None),
                       file_name=f"{_safe_name(label)}_매장별상세.xlsx", mime=XLSX_MIME,
                       key=f"wk_dl_drill_{label}", use_container_width=True)
    render_styled_table(sty)


def render_weekly_item_drilldown(cur_m, prev_m, cur_y, prev_y, label, mask, cy, py, click_ns=None,
                                 period=None, big_title=False):
    """선택한 매장(또는 담당자 전체매장)의 아이템그룹별 상세표 — 주간보고 동일 프레임. 비중=해당 그룹 내.

    click_ns(2026-08-18 추가): 표 안의 **당월 올해 실판가** 숫자를 클릭하면 그 행(아이템그룹)의
    품번별 판매현황 + 품번별 상위 3개 매장을 팝업으로 보여준다. 기간은 당월(cm) 기준이고,
    유통/브랜드·연차 등 위쪽 필터는 이미 cm에 적용된 상태라 별도 조건 전달이 필요 없다.
    None(기본값)이면 클릭 기능 없이 지금까지와 완전히 동일하게 렌더된다.
    """
    cm, pm = cur_m[mask(cur_m)], prev_m[mask(prev_m)]
    cyd, pyd = cur_y[mask(cur_y)], prev_y[mask(prev_y)]
    if cm.empty and cyd.empty and pm.empty and pyd.empty:
        st.info(f"'{label}' 데이터가 없어요.")
        return
    present = set(pd.concat([cyd["아이템그룹"], pyd["아이템그룹"], cm["아이템그룹"], pm["아이템그룹"]]).astype(str))
    groups = [g for g in ITEMGROUP_ORDER if g in present] + [g for g in present if g not in ITEMGROUP_ORDER]
    rows = [(("전체", "G.TOTAL", "합계"), lambda x: pd.Series(True, index=x.index))]
    for g in groups:
        rows.append((("아이템", g, "합계"),
                     (lambda gg: (lambda x: x["아이템그룹"].astype(str) == gg))(g)))
    bm = _wk_block(cm, pm, rows)
    by = _wk_block(cyd, pyd, rows)
    sty, disp = _wk_style_table(bm, by, [k for k, _ in rows], cy, py, click_ns=click_ns, period=period)
    if click_ns:
        # 링크 스타일 + 숨은 버튼 감추기 CSS를 표보다 먼저 내보낸다(버튼이 잠깐 보였다 사라지는 것 방지).
        # 항목26: 타이틀 줄보다도 먼저 내보낸다 — 타이틀과 표 사이에 끼면 빈 세로 여백이 생긴다.
        st.markdown(_WK_PN_CSS + _wk_pn_hide_css(click_ns, len(rows)), unsafe_allow_html=True)
    # 룰11: 제목 + 우측 일반 엑셀 다운로드 버튼 (2026-07-31)
    if big_title:   # 항목26: 드릴다운1은 위 필터 박스와 간격을 살짝 띄운다
        st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)
    i1, i2 = st.columns([5, 1])
    _has_m = any(c[0] == "당월 실적" for c in disp.columns)
    _hint = ("· <b>당월 26실판가 숫자를 클릭</b>하면 품번별 상세" if (click_ns and _has_m) else "")
    _sub = (f"<span style='color:#888;font-size:0.8rem;font-weight:400;'>"
            f"(비중=해당 그룹 내 · G.TOTAL=선택 전체 {_hint})</span>")
    if big_title:   # 항목26: 드릴다운1 결과 타이틀은 다른 드릴다운 타이틀과 동일 크기(##### 헤딩)
        i1.markdown(f"##### 🔍 {label} · 아이템그룹별 상세  {_sub}{_NOTE_FLOAT}", unsafe_allow_html=True)
    else:
        i1.markdown(f"**🔍 {label} · 아이템그룹별 상세**  {_sub}{_NOTE_FLOAT}", unsafe_allow_html=True)
    _nm = sum(1 for c in disp.columns if c[0] == "당월 실적")   # 당월 블록 컬럼 수(7)
    _ny = sum(1 for c in disp.columns if c[0] == "연간누계")
    i2.download_button("⬇ 엑셀", table_excel_bytes(disp, f"{label} 아이템",
                                                  first_block_cols=_nm if (_nm and _ny) else None),
                       file_name=f"{_safe_name(label)}_아이템그룹별상세.xlsx", mime=XLSX_MIME,
                       key=f"wk_dl_item_{label}", use_container_width=True)
    render_styled_table(sty)
    if not click_ns:
        return

    # ── 260818: 숫자 클릭 → 품번별 상세 팝업 ──────────────────────────────────
    # 숨은 버튼(JS가 대신 눌러줌) — 라벨의 보이지 않는 표식으로 브리지가 찾아낸다.
    picked = None
    for i, (key, _m) in enumerate(rows):
        if st.button(f"{_WK_PN_MARK}{click_ns}#{i}", key=f"wkpnb_{click_ns}_{i}"):
            picked = i
    _wk_pn_click_bridge(click_ns)

    # 260818 2차: JS 브리지가 어떤 이유로든 안 먹는 환경(사내 보안 확장, iframe 차단 등)을 대비한
    # **항상 동작하는 폴백**. 접힌 상태라 평소엔 눈에 거의 안 띄고, 펴서 아이템을 고르면 숫자 클릭과
    # 똑같은 팝업이 뜬다. (숫자 클릭이 안 뜬다는 리포트가 있어 신설 — 기능이 조용히 실패하지 않도록)
    _row_names = [k[1] for k, _ in rows]
    with st.expander("🔍 숫자 클릭이 안 될 때 — 아이템을 골라서 품번별 상세 보기", expanded=False):
        fb1, fb2 = st.columns([3, 1])
        fb_sel = fb1.selectbox("아이템", _row_names, key=f"wkpn_fb_sel_{click_ns}",
                               label_visibility="collapsed")
        if fb2.button("상세보기", key=f"wkpn_fb_go_{click_ns}", use_container_width=True):
            picked = _row_names.index(fb_sel)

    # 열려 있는 팝업은 "행 인덱스"가 아니라 **아이템그룹 이름**으로 기억한다 — 위쪽 필터(매장·연차 등)를
    # 바꾸면 표의 행 구성이 달라져서, 인덱스로 기억하면 엉뚱한 아이템의 팝업이 뜰 수 있기 때문.
    # 바뀐 조건에 그 아이템그룹이 아예 없어지면 팝업은 조용히 닫힌다.
    open_key = f"wkpn_open_{click_ns}"
    # 조회 조건(label)이 바뀌면 이전에 열어둔 팝업은 닫는다 — 매장·필터를 바꿨는데 지난번 팝업이
    # 혼자 다시 뜨는 걸 막기 위함(팝업 안 숫자는 새 조건 기준이라 더 헷갈림).
    lbl_key = f"wkpn_lbl_{click_ns}"
    if st.session_state.get(lbl_key) != label:
        st.session_state[lbl_key] = label
        if picked is None:
            st.session_state.pop(open_key, None)
    if picked is not None:
        st.session_state[open_key] = rows[picked][0][1]
    opened_name = st.session_state.get(open_key)
    hit = [(k, m) for k, m in rows if k[1] == opened_name]
    if opened_name is None or not hit:
        st.session_state.pop(open_key, None)
        return

    row_key, row_mask = hit[0]
    grp = row_key[1]                       # "G.TOTAL" 또는 아이템그룹명
    sub = cm[row_mask(cm)] if len(cm) else cm
    grp_txt = "전체 아이템" if grp == "G.TOTAL" else grp

    def _close():
        st.session_state.pop(open_key, None)

    _wk_pn_popup(
        sub,
        title=f"{label} · {grp_txt} · 품번별 상세",
        caption=f"{label} · 아이템: {grp_txt} · 기간: 당월({cy}년) — "
                f"품번별 판매현황 + 품번별 판매수량 상위 3개 매장",
        key_prefix=f"wkpn_{click_ns}_{_safe_name(grp)}",
        on_dismiss=_close,
    )


def render_weekly_category_drilldown(cur_m, prev_m, cur_y, prev_y, cy, py):
    """🔍 (드릴다운 1) 유통/브랜드 · 연차 · 아이템/매장별 상세 보기 — 2026-08-10 신규.
    (260828 번호 재정렬: 화면 표기가 드릴다운 1로 바뀜. 코드 내부 주석·키의 "드릴다운3"(click_ns "cat" 등)
     명칭은 세션키 호환을 위해 그대로 두었다 — 내부 드릴다운3 = 화면 드릴다운 1, 내부 1·2 = 화면 2·3.)

    드릴다운1(유통 또는 담당자 '하나')·드릴다운2(매장 또는 담당자 '하나')와 달리, 유통·브랜드
    (다중선택·OR)와 연차(다중선택·OR)로 데이터 자체를 자유롭게 좁힌 뒤, 그 결과를 "아이템"
    (중카테고리 기준) 또는 "매장별"(그 필터에 해당하는 매장 목록)로 펼쳐 본다. 이미 검증된
    render_weekly_item_drilldown/render_weekly_drilldown에 필터링 끝난 데이터 + '항상 참' 마스크를
    그대로 넘겨 재사용 — 헤더(당월실적·연간누계)·엑셀 다운로드·비중 계산 방식이 메인 표·드릴다운1·2와
    완전히 동일하다(중복 로직 없음).
    """
    st.markdown("##### 🔍 (드릴다운 1) 유통/브랜드 · 연차 · 아이템/매장별 상세 보기")   # 260828: 드릴다운 번호 재정렬(화면 순서 기준)
    pool = pd.concat([cur_m, prev_m, cur_y, prev_y])
    if pool.empty:
        st.info("표시할 데이터가 없어요.")
        return

    group_opts = list(_CHANNEL_MASKS.keys()) + list(_BRAND_MASKS.keys())
    age_opts = [lbl for lbl, _ in _AGE_BUCKET_DEFS]
    season_opts = sorted(pool["시즌명"].dropna().astype(str).unique()) if "시즌명" in pool.columns else []
    with st.form("wk_cat_form"):
        # 시즌 필터 추가(항목24 · 260821) · 표시 기간 필터 추가(항목26 · 260828)
        wc0, wc1, wc3, wc4, wc2 = st.columns([1.5, 0.9, 0.9, 1.45, 1.0])
        selg = wc0.multiselect("유통/브랜드 선택", group_opts, default=[], placeholder="전체", key="wk_cat_grp")
        sela = wc1.multiselect("연차", age_opts, default=[], placeholder="전체", key="wk_cat_age")
        sels = wc3.multiselect("시즌", season_opts, default=[], placeholder="전체", key="wk_cat_season")
        selp = wc4.radio("표시 기간", ["전체", "당월 실적", "연간누계"], horizontal=True, key="wk_cat_period")
        dim = wc2.radio("아이템 or 매장", WK_DIM_OPTS, horizontal=True, key="wk_cat_dim")
        run = st.form_submit_button("🔍 상세보기", type="primary")
    if _need_search("wk_cat_go", run):
        return

    def _grp_mask(frame, name):
        fn = _CHANNEL_MASKS.get(name) or _BRAND_MASKS.get(name)
        return fn(frame) if fn is not None else pd.Series(False, index=frame.index)

    def _apply(frame):
        if frame.empty:
            return frame
        out = frame
        if selg:
            comb = _grp_mask(out, selg[0])
            for nm in selg[1:]:
                comb = comb | _grp_mask(out, nm)
            out = out[comb]
        if sela and "연차" in out.columns:
            known_ages = {"신상", "내년신상", "1년차", "2년차", "3년차"}
            all_ages = set(str(a) for a in pool["연차"].dropna().unique()) if "연차" in pool.columns else set()
            rest_ages = all_ages - known_ages
            age_vals = set()
            for lbl in sela:
                bucket = dict(_AGE_BUCKET_DEFS)[lbl]
                age_vals |= set(bucket) if bucket is not None else rest_ages
            out = out[out["연차"].astype(str).isin(age_vals)]
        if sels and "시즌명" in out.columns:
            out = out[out["시즌명"].astype(str).isin(sels)]
        return out

    fcm, fpm, fcy, fpy = _apply(cur_m), _apply(prev_m), _apply(cur_y), _apply(prev_y)
    if fcm.empty and fpm.empty and fcy.empty and fpy.empty:
        st.info("선택한 조건에 해당하는 데이터가 없어요.")
        return

    _selg_txt = "·".join(selg) if selg else "전체"
    _sela_txt = "·".join(sela) if sela else "전체"
    _sels_txt = "·".join(sels) if sels else "전체"
    label = f"유통/브랜드: {_selg_txt} · 연차: {_sela_txt} · 시즌: {_sels_txt}"
    _all_true = lambda x: pd.Series(True, index=x.index)   # 이미 필터링된 데이터를 그대로 통과시킴

    _period = None if selp == "전체" else selp   # 항목26: "당월 실적"/"연간누계"면 그 블록만
    if dim == "아이템":
        # click_ns="cat" → 당월 올해 실판가 숫자 클릭 시 품번별 상세 팝업(260818 신규)
        render_weekly_item_drilldown(fcm, fpm, fcy, fpy, label, _all_true, cy, py, click_ns="cat",
                                     period=_period, big_title=True)
    else:
        render_weekly_drilldown(fcm, fpm, fcy, fpy, label, _all_true, cy, py, show_plan=False,
                                period=_period, big_title=True)
    st.caption("※ 유통/브랜드·연차·시즌은 다중선택(선택한 항목 중 하나라도 해당하면 포함, OR 조건) — "
               "빈칸이면 전체. '아이템'은 중카테고리(아이템그룹) 기준 breakdown, '매장별'은 위 필터에 "
               "해당하는 매장 목록을 보여줘요. 비중=선택 조건 내 비중, 필터가 걸린 상태라 사업계획·진도율은 "
               "'–'로 표시돼요.")


def render_weekly_report(df):
    st.subheader("📋 주간현황 분석 (당월 · 연간누계, 전년 동기간 비교)")
    if df.empty or "_판매일" not in df.columns or df["_판매일"].notna().sum() == 0:
        st.info("데이터를 먼저 적재하세요.")
        return
    d = df[df["_판매일"].notna()].copy()
    master = load_master()
    if not master.empty and "채널스토리" in master.columns:
        cs_map = dict(zip(master["매장코드"].astype(str).str.strip(), master["채널스토리"]))
        d["_채널스토리"] = d["매장코드"].astype(str).str.strip().map(cs_map)
    else:
        d["_채널스토리"] = None
        st.warning("매장 기준정보(채널스토리)가 없어 유통별 3개(원래직입점·웹뜰이관·웍스바이이관)는 0으로 나와요. "
                   "사이드바 **매장 기준정보 업로드**에 마스터 파일을 올리면 채워져요.")
    # 담당자 매핑 (드릴다운 담당별용)
    if not master.empty and "담당자" in master.columns:
        mgr_map = dict(zip(master["매장코드"].astype(str).str.strip(), master["담당자"].astype(str).str.strip()))
        d["_담당자"] = d["매장코드"].astype(str).str.strip().map(mgr_map)
    else:
        d["_담당자"] = None

    dmin, dmax = d["_판매일"].min().date(), d["_판매일"].max().date()
    asof = st.date_input("조회 기준일 (당월·누계의 끝 날짜)", value=dmax, min_value=dmin, max_value=dmax, key="wk_asof")
    asof = pd.to_datetime(asof)
    cy, py = asof.year, asof.year - 1
    st.caption(f"올해({cy}) vs 전년({py}) 동기간 · 실판가=실매출(백만원) · 판가율=실판가÷최초가(가중) · 비중=행÷전체")

    # ── 공통 필터 (이 페이지 모든 표에 적용) — 브랜드별 → 연차별 → 시즌별 ──
    fc1, fc2, fc3 = st.columns(3)
    _brands = sorted(d["브랜드명"].dropna().unique()) if "브랜드명" in d.columns else []
    _ages = sorted(d["연차"].dropna().unique(), key=_age_sort_key) if "연차" in d.columns else []
    _seasons = sorted(d["시즌명"].dropna().unique()) if "시즌명" in d.columns else []
    selb = fc1.multiselect("브랜드별", _brands, default=[], placeholder="전체", key="wk_fb")
    sela = fc2.multiselect("연차별", _ages, default=[], placeholder="전체", key="wk_fa")
    sels = fc3.multiselect("시즌별", _seasons, default=[], placeholder="전체", key="wk_fs")
    fd = d
    if selb:
        fd = fd[fd["브랜드명"].isin(selb)]
    if sela:
        fd = fd[fd["연차"].isin(sela)]
    if sels:
        fd = fd[fd["시즌명"].isin(sels)]
    _filtered = bool(selb or sela or sels)
    if _filtered:
        st.caption("🔎 필터 적용 중 — 실적·판가율·증감·비중은 선택 조건 기준. "
                   "**사업계획·진도율은 시즌/연차 세분화가 없어 필터 시 '–'로 표시**(전체일 때만 계획 표시).")

    m_start = asof.replace(day=1)
    y_start = asof.replace(month=1, day=1)
    cur_m = fd[(fd["_판매일"] >= m_start) & (fd["_판매일"] <= asof)]
    prev_m = fd[(fd["_판매일"] >= m_start - pd.DateOffset(years=1)) & (fd["_판매일"] <= asof - pd.DateOffset(years=1))]
    cur_y = fd[(fd["_판매일"] >= y_start) & (fd["_판매일"] <= asof)]
    prev_y = fd[(fd["_판매일"] >= y_start - pd.DateOffset(years=1)) & (fd["_판매일"] <= asof - pd.DateOffset(years=1))]

    rows = _wk_rows()
    bm = _wk_block(cur_m, prev_m, rows)   # 당월
    by = _wk_block(cur_y, prev_y, rows)   # 누계

    # 표 구성: 행(섹션/구분/세부) × 열(블록×지표) — 공용 프레임 함수
    idx = [k for k, _ in rows]
    if not _filtered:
        inject_plan(by, idx, master)   # 연간 사업계획·진도율 주입 (필터 없을 때만)
    sty, _disp_main = _wk_style_table(bm, by, idx, cy, py)

    h1, h2 = st.columns([5, 1])
    h1.markdown(f"**주간보고 · 기준일 {asof.date()}**  (당월 {m_start.date()} → {asof.date()} · 누계 {y_start.date()} → {asof.date()})"
                f"{_NOTE_FLOAT}", unsafe_allow_html=True)
    # 엑셀 다운로드 — ⚠️ 메인 표 전용 '특별 조건': 팀 주간보고 양식(weekly_template.xlsx) 템플릿에
    # 값을 채워 내려받는 방식. 룰11(일반 엑셀 버튼)의 예외이므로 절대 일반 방식으로 바꾸지 말 것.
    xls_bytes = weekly_excel_bytes(rows, bm, by, asof, cy, py)
    h2.download_button("⬇ 엑셀", xls_bytes, file_name=f"주간보고_{asof.date()}.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                       key="wk_dl", use_container_width=True)
    render_styled_table(sty)   # 룰3·4 + 헤더검정 + G.TOTAL 노란강조
    st.caption("※ 유통별 5개는 주요 채널만 (직영몰·특수채널·K2K이관 등은 G.TOTAL엔 포함, 유통 행엔 미표기). "
               "S/D/L 신상=신상+내년신상, 4년차↑는 합계엔 포함되나 별도 행 없음. 사업계획·진도율은 목표 입력 후 채워짐.")

    st.divider()
    render_weekly_category_drilldown(cur_m, prev_m, cur_y, prev_y, cy, py)   # 2026-08-10, 메인 표 바로 아래로 배치(중태님 지시)

    # ── 매장 담당별 분석 (위 표와 동일 프레임, 행만 담당자) ──
    _MGR_BOTTOM = ["없음", "26년 미운영", "직원구매"]   # 비담당 라벨 → 맨 아래(이 순서)
    managers = []
    if "_담당자" in d.columns:
        _mset = {m for m in d["_담당자"].dropna().astype(str).str.strip()
                 if m and m.lower() not in ("nan", "none")}
        managers = sorted(m for m in _mset if m not in _MGR_BOTTOM) + [m for m in _MGR_BOTTOM if m in _mset]
    if managers:
        st.divider()
        mrows = [(("전체", "G.TOTAL", "합계"), lambda x: pd.Series(True, index=x.index))]
        for nm in managers:
            mrows.append((("담당별", nm, "합계"),
                          (lambda name: (lambda x: x["_담당자"].astype(str).str.strip() == name))(nm)))
        bm2 = _wk_block(cur_m, prev_m, mrows)
        by2 = _wk_block(cur_y, prev_y, mrows)
        if not _filtered:
            inject_plan_manager(by2, [k for k, _ in mrows], master)   # 담당자 매장 연간계획 합
        sty2, disp2 = _wk_style_table(bm2, by2, [k for k, _ in mrows], cy, py)
        # 룰11: 제목 + 우측 일반 엑셀 다운로드 버튼 (2026-07-31)
        g1, g2 = st.columns([5, 1])
        g1.markdown("##### 👤 매장 담당별 분석" + _NOTE_FLOAT, unsafe_allow_html=True)
        _nm2 = sum(1 for c in disp2.columns if c[0] == "당월 실적")   # 당월 블록 컬럼 수(7)
        g2.download_button("⬇ 엑셀", table_excel_bytes(disp2, "매장 담당별 분석", first_block_cols=_nm2),
                           file_name=f"매장담당별분석_{asof.date()}.xlsx", mime=XLSX_MIME,
                           key="wk_dl_mgr", use_container_width=True)
        render_styled_table(sty2)
        st.caption("※ 담당별 = 매장 마스터의 담당자 기준. 담당 미지정 매장은 담당 행엔 미포함(G.TOTAL엔 포함). 비중=행÷전체.")

    st.divider()
    st.markdown("##### 🔍 (드릴다운 2) 유통별/담당별 매장 상세 보기")   # 260828: 드릴다운 번호 재정렬(구 드릴다운1)
    NONE, HEAD_C, HEAD_M = "(선택 안 함)", "─ 유통별 ─", "─ 담당별 ─"
    opts = [NONE, HEAD_C] + list(_CHANNEL_MASKS.keys())
    if managers:
        opts += [HEAD_M] + managers
    sel = st.selectbox("유통 또는 담당자를 선택하면 그 그룹의 매장별 지표가 같은 형식으로 펼쳐져요.",
                       opts, key="wk_drill")
    if sel not in (NONE, HEAD_C, HEAD_M):
        if sel in _CHANNEL_MASKS:
            _mask = _CHANNEL_MASKS[sel]
        else:
            _mask = (lambda nm: (lambda x: x["_담당자"].astype(str).str.strip() == nm))(sel)
        render_weekly_drilldown(cur_m, prev_m, cur_y, prev_y, sel, _mask, cy, py, show_plan=not _filtered)

    st.divider()
    st.markdown("##### 🔍 (드릴다운 3) 매장별/담당별 아이템분석")   # 260828: 드릴다운 번호 재정렬(구 드릴다운2)
    sv = cur_y.assign(_c=cur_y["매장코드"].astype(str).str.strip()).groupby("_c")["_매출액"].sum().sort_values(ascending=False)
    nmap = dict(zip(d["매장코드"].astype(str).str.strip(), d["매장명"].astype(str)))
    labels = [f"{nmap.get(c, c)} ({c})" for c in sv.index]
    code_of = dict(zip(labels, sv.index))
    I_NONE, I_HM, I_HS = "(선택 안 함)", "─ 담당별 ─", "─ 매장별 ─"
    iopts = [I_NONE]
    if managers:
        iopts += [I_HM] + managers
    iopts += [I_HS] + labels
    isel = st.selectbox("담당자(담당 전체매장 기준) 또는 매장을 선택하면 아이템그룹별 지표가 같은 형식으로 펼쳐져요.",
                        iopts, key="wk_item_drill")
    if isel not in (I_NONE, I_HM, I_HS):
        # 260818: 드릴다운2의 아이템그룹별 상세표에도 동일하게 숫자 클릭 → 품번별 상세 팝업 적용
        # (click_ns를 "d2"로 따로 줘서 드릴다운3("cat")과 팝업 상태가 서로 섞이지 않게 함)
        if isel in managers:
            imask = (lambda nm: (lambda x: x["_담당자"].astype(str).str.strip() == nm))(isel)
            render_weekly_item_drilldown(cur_m, prev_m, cur_y, prev_y, f"{isel} (담당 전체매장)", imask,
                                         cy, py, click_ns="d2")
        else:
            code = code_of[isel]
            imask = (lambda c: (lambda x: x["매장코드"].astype(str).str.strip() == c))(code)
            render_weekly_item_drilldown(cur_m, prev_m, cur_y, prev_y, isel, imask, cy, py, click_ns="d2")


# ==============================================================================
# 재고 모니터링 1차 가공  ─ 260731 확정 기준 · v3.5 123열 (2026-08-02 ERP 이식, 260811 사이즈구분 컬럼 + 가격시뮬 + 260826 SET가격 9컬럼 + 260826-2 물량등급 반영)
# ==============================================================================
# 원본: '쇼핑몰재고 모니터링 자료 1차 가공' 프로젝트 process_260731.py (판정 로직 1:1 이식)
#  · 선판정: 수정일='오프라인' → AA·AB·AF '오프라인' / 온라인창고<20 → AA·AF '재고20미만'(AB 미적용)
#    (260811: 판정 소스를 '이관구분' 컬럼에서 '수정일' 컬럼으로 변경 — 로우데이터상 실제 오프라인/
#     온라인/부분이관 값은 '수정일'이라는 이름의 컬럼(raw21)에 들어오고, '이관구분' 컬럼(raw22)은
#     현재 로우데이터에서 공란으로 내려온다. 컬럼명과 실제 내용이 어긋난 회사 ERP 추출 특성.)
#  · AA/AB 5등급: A 상위20% / B 21~50 / C 51~80 / D 81~100 / E 판매0(자동부여)
#  · 등급 모집단 = (중카테고리 × 년도 × 시즌) — 중카테고리는 아이템 마스터(item_master) 기준,
#    260803부터 니트류·티셔츠류가 "니트/티셔츠류" 한 중카테고리로 통일됨(마스터 미등록 시엔 구 분리판 폴백)
#  · AF(AI제안방향) 8룰 매트릭스 (재고 분기 200, 온라인창고 기준)
#  · AC/AD/AE = 사이즈 등급 · SET 상태 구분 · SET 등급 (size-grade-classifier 로직 내장, X·Y 변수)
#  · 260806 세트 판정 기준서(260731 확정) 동기화 4건:
#    ① 매칭표 확장 — 95→74·76 추가, 110→92·94·98 추가(115와 동일 구성)
#    ② 하의 잔여 1.3배 규칙 — 하의총재고÷상의총재고 > 1.3 일 때만 '하의단품'. 1.3 정확히는 미초과.
#       상의 잔여에는 미적용(자켓이 남는 건 곧 상의단품 신호).
#    ③ SET 등급 A → A-1/A-2 분리 — 핵심 2개 상의 '실재고'가 둘 다 Y 이상이면 A-1, 아니면 A-2.
#       (C-2/C-3은 여전히 '세트 가능 수량' 기준 — 두 수량을 혼동하지 말 것)
#    ④ A09↔A09 세트업 지원 + 상/하 판별을 사이즈코드가 아닌 아이템 코드로 변경
#  · 260807 개정(size-grade-classifier 스킬 반영): A09 핵심 사이즈 M·L(2개) → M·L·XL(3개) 확장,
#    빅은 XXL만 남음. 단품 등급(_INV_SYSTEMS["A09"])·SET 등급(_INV_SET_CORE_A09 등) 둘 다 반영.
#    A16(핵심 2개 그대로)은 영향 없음.
#  · ⚠️ 260811 개정 — 사이즈코드 판정 소스 전환('쇼핑몰재고 모니터링 1차' 프로젝트 260811 전달사항 + process_260811.py):
#    로우데이터에 신규 컬럼 '사이즈구분'이 추가됨(사이즈정보 14칸 그룹 바로 앞, raw 79번째 열=0-index 78).
#    이제 이 컬럼값을 사이즈코드(A16/A17/A09/A05/A06/A18)로 "직접" 사용한다 — 더 이상 사이즈 마스터
#    (품번→사이즈코드)를 조회해서 채우지 않는다. 등급 산정 알고리즘(_inv_grade_one/_inv_set_grade 등)
#    자체는 전혀 바뀌지 않았다 — 바뀐 건 "사이즈코드를 어디서 가져오는가" 뿐이다.
#    · 공란(값 없음) → 사이즈 마스터로 되돌아가 보완하지 않고 바로 '해당없음'(미매칭) 처리.
#    · 사이즈 마스터는 이제 필수가 아니라 옵션(참고용) — 있으면 새 컬럼값과의 불일치 건수만 참고 표시.
#    · 로우데이터 열 수 기준 93 → 94열로 상향(신규 컬럼 1개 증가). 94열이 아니면 처리 거부.
#    · 출력도 신규 '사이즈구분' 컬럼 1개가 늘어 106 → 107열(v3.2)로 변경, CN열(92번째)에 위치.
#      구 106열 템플릿을 넣으면 자동으로 컬럼 1개를 삽입해 107열로 보정한다.
#  · 출력: 113열 v3.3 — 서식은 저장소 동봉 템플릿(inventory_template.xlsx = 최신 v3.3 결과물)에서 1:1 복제
#  · 260802 서식 확정: C·K·L·M·AA·AB·AF·AG·AH·AI 노란색 / BA·BG·BH·BI 초록색 / BK~CM 숨김
#    (사이즈구분 신규 컬럼은 로우파일에 이미 있는 값을 그대로 통과시키는 열이라 노란색 대상 아님 — 기본 서식 유지)
#  · ⚠️ 260811 추가 개정 — 가격 시뮬레이션 5컬럼 신설(팀장 지시, 중태님과 채팅으로 계산식 검증 확정):
#    로우데이터의 '몰가격'(raw J열=0-idx9) · '기준판매가'(raw AB열=0-idx27) 두 값으로 결과물의 몰가격
#    컬럼(V열=22번째) 바로 뒤에 신규 가격 5컬럼을 삽입한다 — (네이버)상시가·(쿠폰진행)상시가·
#    (쿠폰진행)행사가·(쿠폰X/무배)상시가·(쿠폰X/무배)행사가. 계산식:
#      · 상시가(네이버) = 몰가격×1.05 / 상시가(쿠폰진행) = 몰가격×1.15 / 행사가(쿠폰진행) = 몰가격×1.1
#      · 상시가(쿠폰X/무배) = 몰가격<30,000이면 (몰가격×1.05)+3,000, 아니면 몰가격×1.05
#      · 행사가(쿠폰X/무배) = 몰가격<30,000이면 몰가격+3,000, 아니면 몰가격
#    공통 규칙: (1) 5개 전부 끝 3자리 기준 3단계 스냅(260813 변경 — 기존 100원단위 반올림 대체):
#    끝 3자리 000이면 그대로 / 001~500이면 500으로 맞춤 / 501~999이면 900으로 맞춤. (2) 기준판매가에
#    값이 있고 계산값이 그 값을 넘으면 → 그 계산 하나만 기준판매가로 대체(중간 단계 없이 1회 비교·
#    캡핑, 기준판매가도 동일한 3단계 스냅 적용 후 비교). (2-1) 260814 추가: 기준판매가가 공란이면 —
#    대신 '최초가'(raw H열=0-idx7, 상품 택 가격표)와 비교해서 계산값이 최초가를 넘으면 그 계산 하나만
#    최초가로 대체(택가격보다 비싸게 팔 수 없다는 안전장치). 기준판매가·최초가 둘 다 공란이면 캡핑
#    없이 계산값 그대로. (3) 몰가격이 공란인 행은 5컬럼 전부 공란.
#    서식: 5컬럼 전부 노란색, 헤더는 그룹행(GROUP_R) 비우고 실제 컬럼명만(사이즈구분과 동일 방식).
#  · ⚠️ 260811 추가 개정(2) — 기준판매가 복제 컬럼 신설(팀장 지시): 몰가격 바로 뒤(가격5컬럼보다도
#    앞)에 기준판매가 값을 그대로 복제한 컬럼 1개를 초록색으로 추가 삽입 — 뒤쪽(패스스루 구간)의
#    원본 기준판매가 컬럼은 그대로 유지, 값만 똑같이 한 번 더 보여주는 것(몰가격 옆에서 바로 비교
#    하기 편하도록). 결과물이 107 → 112 → 113열(v3.3)로 확장된다.
#    구 107열 템플릿을 넣으면 자동으로 6컬럼(가격5+기준판매가복제1)을 삽입해 113열로 보정한다
#    (구 106열 템플릿도 106→107→112→113 3단계 자동 보정).
# ※ 사이즈 마스터(품번→사이즈코드)는 DB(size_master)에 저장 — 260811부터 판정에는 미사용, 참고 보고 전용.
# ※ 아이템 마스터(아이템코드→대/중/소카테고리)는 DB(item_master)에 저장 — 사이드바(관리자)에서 업로드/교체.
#   재고모니터링 중카테고리·판매분석 아이템그룹이 모두 여기서 나온다(단일 소스). 미등록 코드만 구 하드코딩 폴백.
# ※ 재고 로우데이터는 DB에 적재하지 않음(그때그때 가공→엑셀 다운로드만). 가공은 전 팀원 사용 가능.
INV_TEMPLATE_FILE = "inventory_template.xlsx"   # GitHub 저장소에 weekly_template.xlsx처럼 동봉
SIZE_MASTER_TABLE = "size_master"
ITEM_MASTER_TABLE = "item_master"
# 260811: 사이즈구분 컬럼 도입으로 입출력 스펙 변경 — 하드코딩 대신 상수로 관리.
INV_RAW_COLS = 94          # 로우데이터 총 열 수 (구 93 → 94, '사이즈구분' 신규 1열)
INV_RAW_MOLGA_COL = 9      # raw 0-index — 몰가격(J열)
INV_RAW_GIJUN_COL = 27     # raw 0-index — 기준판매가(AB열)
INV_RAW_CHOJOGA_COL = 7    # raw 0-index — 최초가(H열). 260814: 기준판매가가 공란일 때의 대체 캡핑 기준.
INV_RAW_HYUNPAN_COL = 8    # raw 0-index — 현판가(I열). 260826: SET 가격 합산에 사용.
# 260811(가격시뮬): 몰가격(22) 바로 뒤에 ① 기준판매가 복제 1컬럼(녹색) → ② 신규 가격 5컬럼(노란색)
# 순서로 삽입 — 결과물 107 → 112 → 113열(v3.3)로 확장.
INV_GIJUN_COPY_COL = 23    # 기준판매가 복제 컬럼 위치(몰가격 바로 다음, 원본 기준판매가는 뒤쪽 그대로 유지)
INV_PRICE_SIM_COL = 24     # 신규 가격 5컬럼 시작 위치(기준판매가 복제 컬럼 바로 다음)
INV_PRICE_SIM_N = 5
INV_PRICE_SIM_HEADERS = ["(네이버) 상시가", "(쿠폰진행) 상시가", "(쿠폰진행) 행사가",
                         "(쿠폰X/무배) 상시가", "(쿠폰X/무배) 행사가"]
# 260826(SET가격 9컬럼, 중태님 지시 — 수기 샘플 엑셀 "단품가격에 SET 가격 추가.xlsx"와 동일 규격):
# 단품가격 블록(20~28: 최초가·현판가·몰가격·기준판매가·가격시뮬5) 바로 뒤에 "SET 가격" 9컬럼
# (29~37, 같은 구성)을 신설 — 결과물 113 → 122열(v3.4)로 확장. SET품번으로 짝지어진 상의+하의의
# 단품가격 9칸을 각각 합산해(엑셀 수기 수식 =T상의행+T하의행 방식 그대로) 짝 양쪽 행에 동일하게
# 기입한다. 사이즈 매칭 성패(SET구성실패 포함)와 무관하게 상/하 데이터가 다 있으면 계산하고,
# 단품아이템·짝 없는 세트상품 행은 9칸 전부 공란. 구 113열 템플릿은 자동으로 9컬럼을 삽입해
# 122열로 보정한다(구 106·107·108·112 템플릿도 기존 체인을 거쳐 113→122까지 순차 보정).
INV_SET_PRICE_COL = 29     # SET 가격 9컬럼 시작 위치(단품 (쿠폰X/무배) 행사가 바로 다음)
INV_SET_PRICE_N = 9
INV_SET_PRICE_HEADERS = ["최초가", "현판가", "몰가격", "기준판매가"] + INV_PRICE_SIM_HEADERS
# 260826-2(물량등급, 중태님 지시): 최초출고일(40)과 기간판매수량분석 사이(41번째)에 "물량등급"
# 1컬럼 신설 — 결과물 122 → 123열(v3.5). 온라인창고 재고(재고20미만 판정과 같은 소스, raw 40번째
# 열) 숫자를 기준으로 A/B/C/D 4등급: A ≥ vol_a / B ≥ vol_b / C ≥ vol_c / D = vol_c 미만.
# 기준 숫자 3개(vol_a > vol_b > vol_c)는 재고가공 화면에서 변수 X·Y 아랫줄에 직접 입력
# (기본값 500/300/100), D등급 칸은 "C기준-1장 이하"로 자동 표기만 된다. 구 122열 템플릿은
# 자동으로 1컬럼을 삽입해 123열로 보정한다(구 106~113 템플릿도 기존 체인을 거쳐 순차 보정).
INV_VOL_GRADE_COL = 41     # 물량등급 컬럼 위치(최초출고일 바로 다음, 판매진도 그룹의 첫 칸)
INV_TOTAL_COLS = 123       # 결과물 총 열 수 (구 106→107→112→113→122→123)
INV_SIZECODE_COL = 108     # 결과물에서 '사이즈구분' 컬럼 위치 (구 107 + 물량등급 1칸)
_INV_KNOWN_SIZE_CODES = {"A16", "A17", "A09", "A05", "A06", "A18"}

# 260806: 아이템 마스터에도 폴백에도 없는 아이템 코드의 표기값. 예전엔 이런 코드를 만나면 가공을
#         통째로 중단했는데(팀 요청으로 변경), 이제는 해당 상품 행만 이 값으로 표기하고 등급
#         모집단에서 빼서 나머지 상품은 정상 가공한다. 마스터에 코드를 추가한 뒤 다시 돌리면
#         자동으로 정상 등급을 받는다.
_INV_UNMAPPED = "미분류"

# ── 중카테고리 매핑 폴백 (260731 니트류/티셔츠류 분리판 — 아이템 마스터 미등록/미매칭 코드용) ──
# 260803부터 진짜 소스는 DB 아이템 마스터(item_master) — _inv_cat_lookup()이 그걸 우선 쓰고,
# 마스터에 없는 코드만 아래로 폴백한다.
_INV_CAT_FALLBACK = {}
for _cat, _codes in {
    # 260804: "팬츠"→"팬츠류"로 통일(아이템 마스터 실제 표기·ITEMGROUP_ORDER와 동일 기준 맞춤).
    "팬츠류": ["PA", "DM", "GP", "WP", "HP"], "셔츠류": ["DS", "WD"],
    "니트류": ["KT", "KG", "KV", "GK", "WI"], "티셔츠류": ["TS", "IT", "GT", "WS"],
    "아우터": ["CT", "JP", "JA", "DJ", "WO", "PV", "WJ", "WK", "GE", "GJ"],
    "수트류": ["SJ", "SL", "EJ", "EP", "JV"], "신발": ["FW"],
    "ACC": ["NT", "BE", "BA", "MF", "MU", "SC", "GL", "HA", "WA"]}.items():
    for _c in _codes:
        _INV_CAT_FALLBACK[_c] = _cat

# ══════════════════════════════════════════════════════════════════════════
# ▼▼▼ size-grade-classifier 스킬 이식 블록 시작 (SYNC BLOCK) ▼▼▼
# 아래 ~ "▲▲▲ 이식 블록 끝" 까지는 스킬 size-grade-classifier의
# scripts/classify.py + scripts/set_classify.py 로직을 손으로 그대로 옮겨
# 심은 것이다. Streamlit Cloud 배포 환경에는 스킬 파일이 없어 런타임에
# 직접 import하지 못하므로, 스킬이 바뀌어도 이 블록엔 "자동으로" 안 따라온다.
#
# ⚠️ 재고가공 메뉴(사이즈 등급·SET 상태 구분·SET 등급) 로직에 변경이 필요할 때:
#   1) (원본 검증) '쇼핑몰재고 모니터링 1차' 프로젝트에서 실데이터로 규칙을
#      설계·검증하고 size-grade-classifier 스킬 저장까지 마친다.
#   2) (이 프로젝트에서) "스킬 최신본이랑 app.py랑 어긋나는 부분 있는지 확인
#      하고 반영해줘"라고 한 번만 요청한다. 스킬은 계정 전체에 동기화되므로
#      이 세션에서 스킬 폴더를 바로 열어 app.py와 비교할 수 있다 — 별도
#      "전달 자료" 작성은 불필요.
#
# 최근 동기화: 260807 — A09 핵심 사이즈 M·L(2개) → M·L·XL(3개) 확장 반영
#   (_INV_SYSTEMS["A09"] · _INV_SET_CORE_A09 등 상수 + _inv_set_grade() 의
#    nc 기준 일반화. 상수만 바꾸고 함수 로직을 안 고치면 조용히 틀리니 주의.)
# ══════════════════════════════════════════════════════════════════════════
# ── 사이즈 체계 (size-grade-classifier 스킬 정의 내장) ──
_INV_SYSTEMS = {
    "A16": {"core": [5, 7], "small": [3], "big": [9, 10, 11, 13], "all": [3, 5, 7, 9, 10, 11, 13]},
    "A17": {"core": [5, 6, 7], "small": [1, 2, 3], "big": [9, 11, 12, 13],
            "all": [1, 2, 3, 5, 6, 7, 9, 11, 12, 13]},
    # 260807 개정(size-grade-classifier 스킬 반영): 핵심 M·L(2개) → M·L·XL(3개) 확장, 빅은 XXL만 남음.
    "A09": {"core": [4, 5, 6], "small": [2, 3], "big": [7], "all": [2, 3, 4, 5, 6, 7]},
    "A05": {"core": [7, 8, 9, 10], "small": [1, 2, 3, 4, 5, 6], "big": [11, 12],
            "all": list(range(1, 13)), "shoe_rule": True},
}
_INV_TOP = {3: 95, 5: 100, 7: 105, 9: 110, 10: 115, 11: 120, 13: 130}
_INV_BOT = {1: 74, 2: 76, 3: 78, 5: 82, 6: 84, 7: 86, 9: 90, 11: 94, 12: 98, 13: 102}
_INV_TOP_IDX = {v: k for k, v in _INV_TOP.items()}
_INV_BOT_IDX = {v: k for k, v in _INV_BOT.items()}
# 260806 갱신: size-grade-classifier 스킬 최신본 매칭표와 동기화
#   · 95  → 74·76 추가 (기존 78·80·82)
#   · 110 → 92·94·98 추가 (기존 84·86·88·90) → 115와 동일 구성
# 260814 갱신(중태님 지시): 100 → 78 추가 (기존 80·82·84) — 78은 A17 운용 사이즈 목록(_INV_BOT)에
# 이미 있는 값이라 별도 조치 없이 매칭 즉시 반영됨.
_INV_MATCH = {95: [74, 76, 78, 80, 82], 100: [78, 80, 82, 84], 105: [82, 84, 86, 88],
              110: [84, 86, 88, 90, 92, 94, 98],
              115: [84, 86, 88, 90, 92, 94, 98], 120: [86, 88, 90, 92, 94, 98, 102, 106],
              130: [88, 90, 92, 94, 98, 102, 106]}
# 260815 개정(중태님 지시): 구 "하의 잔여 1.3배(그룹 총재고 비율)" 규칙 폐지. 대신 사이즈별 1.5배
# 과다재고 판정으로 대체 — 아래 매칭 루프(for ts in top_ok) 안에서 사용.
#   규칙(1): 특정 상의 사이즈와 SET 매칭이 안 되는 팬츠 사이즈가 X장 이상 남으면 그 자체로 바로
#            '하의단품 별도 판매 필요' (구 비율 게이트 없이 즉시 반영).
#   규칙(2): 상의 사이즈와 매칭되는 후보 팬츠 사이즈 중, 팬츠 재고가 "그 팬츠를 후보로 삼는 모든
#            매칭 상의 사이즈 재고의 합"의 1.3배 이상이면 그 팬츠는 세트 수요 대비 명백한
#            과다재고 → 세트&하의단품.
#            260816 개정(중태님 지시): 팬츠 한 사이즈가 여러 상의 사이즈의 매칭 후보로 동시에
#            걸리는 경우(예: 팬츠 78은 재킷 95·100 둘 다의 후보), 종전에는 그중 하나(=재킷 100,
#            25장)와만 비교해 과다재고로 오판했다 — 팬츠 78 재고 93장이 재킷 100 하나의 1.5배는
#            넘지만, 실제로는 재킷 95(120장)의 수요도 동시에 받는 사이즈라 진짜 과다재고가 아니었음.
#            그래서 "합산기준"으로 전환: 비교 대상을 단일 상의가 아니라 그 팬츠를 후보로 삼는 모든
#            매칭 상의 재고의 합으로 바꾼다. 합산 방식은 분모가 커지므로 배율도 기존 1.5배에서
#            1.3배로 낮춘다 — 1/1.2/1.3/1.5 시뮬레이션 결과, 1.5는 합산기준에서 과소검출(발동
#            사이즈칸 43개, 개별기준 169개 대비 74% 급감)이라 판단, 1.3이 원래 취지(명백한
#            과다재고만 골라내기)와 검출 민감도 균형에 가장 근접.
#   규칙(3): 반대로 상의 재고가, 매칭되는 모든 후보 팬츠 재고 합계의 1.3배 이상이면 상의가 명백한
#            과다재고 → 세트&상의단품.
#            260816 2차 개정(중태님 지시): 규칙(2)를 합산기준·1.3배로 낮춘 뒤, 중태님이 "상의는 원래
#            사이즈 종류가 적으니(보통 4~5개, 하의는 6~7개) 상의 쪽 배율은 하의보다 더 높아야 하지
#            않겠냐"고 문제 제기 — 실측으로 검증함. 실데이터 152개 세트그룹에서 "합산 시 몇 개
#            사이즈를 더하는지"(합산 카운트)를 직접 세어보니 오히려 반대였음: 규칙(2)는 하의 1개당
#            평균 2.05개 상의를 합산(매칭표상 하의 하나가 겹치는 상의 사이즈가 1~2개뿐), 규칙(3)은
#            상의 1개당 평균 2.71개 하의를 합산(재킷 하나가 팬츠 3~8개 사이즈와 넓게 매칭됨 —
#            _INV_MATCH 참고, 110·115는 팬츠 7개와 매칭). 즉 상의 쪽이 "사이즈 종류는 적어도" 매칭
#            폭 자체는 더 넓어서, 규칙(3)의 분모가 이미 규칙(2)보다 구조적으로 더 큼 — 배율까지
#            더 높이면(1.5) 두 보수화 요인이 겹쳐 상의 과다재고를 과소검출하게 됨(1.3 고정 기준 실측:
#            규칙(2) 발동 63칸 vs 규칙(3) 1.5일 때 42칸 — 33% 이상 적게 잡힘, 1.3이면 54칸으로 격차
#            축소). 그래서 규칙(3)도 1.3으로 낮춰 규칙(2)와 통일 — 이미 분모가 큰 규칙(3)에 배율까지
#            높이는 "이중 보수화"를 피한다.
_INV_SET_EXCESS_RATIO = 1.3          # 규칙(3) 전용 — 상의 과다재고 판정 (260816: 1.5→1.3, 규칙2와 통일)
_INV_SET_EXCESS_RATIO_BOT = 1.3      # 260816 신설 — 규칙(2) 전용, 합산기준 하의 과다재고 판정
# 260816 3차 개정(중태님 지시) — 규칙(2)/(3) 합산(분모) 대상의 범위 문제. 팀장님이 실제 사례(재킷
# 100사이즈=27장, 매칭표상 후보는 78·80·82·84인데 78=9장·84=7장은 X문턱(재고 10장) 미만이라 지금까지
# 분모 계산에서 통째로 빠지고 82=17장만 들어갔던 것)를 짚음 — "세트 성립 여부 판정"(X문턱 게이트)과
# "과다재고 분모 계산"을 같은 X문턱으로 묶어 쓰는 게 맞냐는 문제제기. 확인 결과 맞는 지적이라 분리:
#   · "세트 성립"(top_ok/bot_ok/matched/used, tl/bl의 진짜 미매칭 판정)은 지금처럼 X문턱(재고 X장
#     이상)을 그대로 유지 — 이건 안 건드림.
#   · 규칙(2)/(3)의 합산 "분모"만 X문턱과 분리 — 매칭표(MTBL)상의 후보 전체를 대상으로 하되, 처음엔
#     "1장이라도 있으면 포함"을 제안했다가 중태님이 "1장은 너무 적다, 3장 이상이면 포함하자"로 확정.
#     즉 분모용 후보는 "이 사이즈 체계에 실재하는 사이즈이면서 재고 3장 이상"이면 X문턱(보통 10장)
#     미만이라도 합산에 포함한다 — 반품/샘플성 1~2장짜리 재고까지 분모를 부풀리는 건 막으면서도,
#     78(9장)·84(7장)처럼 유의미한 재고는 분모에 정상 반영되게 함.
_INV_SET_EXCESS_DEMAND_MIN = 3       # 260816 신설 — 규칙(2)/(3) 합산 분모에 포함시키는 최소 재고(장)
_INV_SET_CORE = (100, 105)
_INV_SET_SMALL = (95,)
_INV_SET_BIG = (110, 115, 120, 130)

# ── A09(M/L/X) 세트 지원 — 세트판정_기준서_260731 §1·§2·§4 ────────────
_INV_A09 = {2: "XS", 3: "S", 4: "M", 5: "L", 6: "XL", 7: "XXL"}
_INV_A09_IDX = {v: k for k, v in _INV_A09.items()}
_INV_MATCH_A09 = {"XS": ["XS", "S"], "S": ["XS", "S", "M"], "M": ["S", "M", "L"],
                  "L": ["M", "L", "XL"], "XL": ["L", "XL", "XXL"], "XXL": ["XL", "XXL"]}
# 260807 개정: SET 등급용 핵심도 단품과 동일하게 M·L·XL(3개)로 확장, 빅은 XXL만 남음.
_INV_SET_CORE_A09 = ("M", "L", "XL")
_INV_SET_SMALL_A09 = ("XS", "S")
_INV_SET_BIG_A09 = ("XXL",)

# 260815 신규(중태님 지시): A09(상의, 문자 M/L/X) ↔ A17(하의, 숫자 정사이즈) 매칭표 확정 — 그동안
# "문자→숫자 변환표 미정"으로 SET구성실패 처리되던 조합. 재킷이 A09(M/L/XL 등)인데 팬츠가 A17(숫자)인
# 상품(예: EJ+EP 조합)을 세트로 잡을 수 있게 됨. 핵심·스몰·빅 분류는 A09↔A09와 동일하게 상의(A09)
# 쪽 체계를 그대로 쓴다(_INV_SET_CORE_A09 등) — SET 등급은 늘 상의 쪽 핵심 성립 여부로 매기므로.
# XS·S·XXL 행은 중태님이 별도로 지정하지 않아 매칭표에 없음 — 없는 키는 그냥 매칭 후보가 없는 것으로
# 처리되어(빈 리스트) 기존 로직 그대로 미매칭(잔여) 취급된다.
_INV_MATCH_A09_A17 = {"M": [78, 82], "L": [78, 82, 86], "XL": [82, 86, 90]}

# 지원 조합: (상의 사이즈코드, 하의 사이즈코드) → (상의 idx맵, 하의 idx맵, 매칭표, 핵심, 스몰, 빅)
_INV_SET_SYS = {
    ("A16", "A17"): (_INV_TOP, _INV_BOT, _INV_MATCH,
                     _INV_SET_CORE, _INV_SET_SMALL, _INV_SET_BIG),
    ("A09", "A09"): (_INV_A09, _INV_A09, _INV_MATCH_A09,
                     _INV_SET_CORE_A09, _INV_SET_SMALL_A09, _INV_SET_BIG_A09),
    # 260815 신규
    ("A09", "A17"): (_INV_A09, _INV_BOT, _INV_MATCH_A09_A17,
                     _INV_SET_CORE_A09, _INV_SET_SMALL_A09, _INV_SET_BIG_A09),
}

# 상/하 판별은 아이템 코드로 한다 (수트류는 중카테고리가 전부 '수트류'라 카테고리로 못 가름).
# 260814 갱신(중태님 지시): 실제 SET 판매가 가능한 아이템코드는 상의 SJ·EJ, 하의 SL·EP 뿐 —
# 기존에 들어있던 JP·JA·CT·DJ·TS·KT·NT·SH(상의)·PA(하의)는 전부 SET 판매가 애초에 불가능한
# 단품 전용 아이템이라 목록에서 제외한다. 이 코드들은 이제 _inv_set_side()가 top/bot 어느 쪽도
# 아닌 None을 반환해 세트 그룹 안에서 짝을 못 찾고 "단품"으로 빠진다.
_INV_TOP_ITEMS = {"SJ", "EJ"}
_INV_BOT_ITEMS = {"SL", "EP"}


def _inv_set_side(rec):
    """세트 그룹 안에서 이 행이 상의인지 하의인지 — 아이템 코드로만 판별한다.
    260814: 예전엔 아이템코드가 안 걸리면 중카테고리에 '팬츠'·'하의'가 있는지로 보조 판별했는데,
    이러면 SL·EP가 아닌 코드(예: PA)도 카테고리만 맞으면 하의로 잡혀버려 '상의 SJ·EJ, 하의 SL·EP만
    SET 가능'이라는 원칙과 어긋난다 — 보조 판별을 없애고 아이템코드 목록만으로 엄격하게 판별한다.
    """
    it = (rec.get("item") or "").strip().upper()
    if it in _INV_TOP_ITEMS:
        return "top"
    if it in _INV_BOT_ITEMS:
        return "bot"
    return None

# 260802 확정: 컬럼 전체(헤더+데이터) 채우기 색 강제 지정 + 상시 숨김 컬럼
# 260811(가격시뮬 + 기준판매가 복제): 몰가격(22) 뒤에 기준판매가 복제 1컬럼 + 가격5컬럼이 끼어들면서
# 구 27번째 이후 컬럼이 전부 +6 밀림 — 아래 세 상수는 밀린 뒤(113열 기준) 위치. C,K,L,M(그대로) +
# 기준판매가 복제 컬럼(23, 초록색) + 신규 가격5컬럼(24~28, 전부 노란색) +
# AA,AB,AF,AG,AH,AI(구 27,28,32,33,34,35 → +6 = 33,34,38,39,40,41 → 260811 '수정일' 재배치로
# AA~변경후할인율 블록이 한 칸씩 앞당겨져 최종 32,33,37,38,39,40).
# 260826(SET가격 9컬럼): SET 가격 블록(29~37)이 끼어들면서 구 29번째 이후 컬럼이 전부 +9 밀림.
# SET 블록 안의 색은 단품가격 블록과 동일 규칙 — 몰가격·기준판매가(31,32)=초록, 가격시뮬 5칸(33~37)=노랑,
# 최초가·현판가(29,30)=무채색(템플릿 서식 그대로).
# 260826-2(물량등급): 신규 물량등급(41, 노랑 — 판매진도 그룹의 첫 칸)이 끼어들면서 구 41 이후가 +1 밀림.
_INV_YELLOW_COLS = ({3, 11, 12, 13} | set(range(INV_PRICE_SIM_COL, INV_PRICE_SIM_COL + INV_PRICE_SIM_N))
                    | set(range(INV_SET_PRICE_COL + 4, INV_SET_PRICE_COL + INV_SET_PRICE_N))
                    | {INV_VOL_GRADE_COL, 42, 43, 44, 48, 49, 50})
# 260815(헤더 개편): AC·AD·AE(단품 사이즈 컨디션·SET 가능여부·SET 사이즈 컨디션) 3컬럼 = 연분홍,
# 수정일 = 회색. AI제안방향(AF)이 노란색 그룹에 합류. (260826 SET가격 +9 · 260826-2 물량등급 +1 밀림)
_INV_PINK_COLS = {45, 46, 47}
_INV_GRAY_COLS = {51}
_INV_GREEN_COLS = {INV_GIJUN_COPY_COL, INV_SET_PRICE_COL + 2, INV_SET_PRICE_COL + 3,
                   69, 75, 76, 77}        # 기준판매가 복제(23) + SET 몰가격·기준판매가(31,32) + BA,BG,BH,BI(구 68,74,75,76 → +1)
# 260816 개정(중태님 지시, "결과물파일 셀 숨기기 기준자료.xlsx" 첨부) — 구 BK~CM(69~97) 일괄
# 숨김을 폐지하고, 팀장님이 실제 결과 파일에서 직접 지정한 개별 숨김열 18개로 전면 교체한다.
# 첨부 파일을 openpyxl로 열어 column_dimensions.hidden을 직접 읽어 그대로 옮겨온 값 — 헤더 위치
# 기준으로 F(순번)·G(라인)·J(품번)·N(색상)·P(몰상품명2)·U(현판가)·AC(할인율)·AD(최초입고일)·
# AE(최초출고일)·AP(이관구분)·AY(누계입고량)·BA(누계판매량)·BB(판매율)·BH(90퍼센트 창고)·
# BI(온라인반품창고)·BJ(원가금액)·BM(기간판매수량 — 매장기간판매수량 그룹과 별개, 65번째 칸)·
# BP(판매비중, 68번째 칸)에 해당. 구 BK~CM(69~97, 매장별 기간판매수량/금액/비교/소진예상기간
# 반복블록)은 이번 기준자료에서 전부 숨김 해제(가시화)됨 — 그대로 반영.
# (260826 SET가격 9컬럼: 구 29 이상 숨김열 +9 밀림 · 260826-2 물량등급: 구 41 이상 +1 추가 밀림 —
#  논리적으로 같은 컬럼들이 그대로 숨겨진다. 신규 SET 가격 9컬럼(29~37)·물량등급(41)은 숨김 대상 아님.)
_INV_HIDE_COLS = {6, 7, 10, 14, 16, 21, 38, 39, 40, 52, 61, 63, 64, 70, 71, 72, 75, 78}


def _inv_grade_one(s14, sd, X, Y):
    """사이즈 등급(AC) 판정 — size-grade-classifier 로직 (품절/품절근처/A~F/C-1~3)."""
    total = sum(s14.get(i, 0) for i in range(1, 15))
    if total == 0:
        return "품절"
    core = [i for i in sd["core"] if s14.get(i, 0) >= X]
    small = [i for i in sd["small"] if s14.get(i, 0) >= X]
    big = [i for i in sd["big"] if s14.get(i, 0) >= X]
    ok = [i for i in sd["all"] if s14.get(i, 0) >= X]
    if not ok:
        return "품절근처"
    nc = len(core)
    if sd.get("shoe_rule"):
        if nc >= 3:
            return "A"
        if nc == 2:
            return "B"
    else:
        if nc >= 3:
            return "A"
        if nc == 2:
            return "A" if len(ok) > nc else "B"
    if nc == 1:
        if len(ok) > 1:
            return "C-1"
        return "C-2" if s14.get(core[0], 0) >= Y else "C-3"
    if small and big:
        return "D"
    if big:
        return "E"
    if small:
        return "F"
    return "판정불가"


# 260816 신규(중태님 지시, 표 첨부) — 단품 사이즈 컨디션(AC)·SET 사이즈 컨디션(AE) 출력 문구를
# "핵심 개수/조건"이 그대로 드러나도록 라벨에 괄호 설명을 덧붙이는 방향으로 개편. 판정 로직
# (_inv_grade_one/_inv_set_grade의 내부 조건문)은 전혀 손대지 않고, 최종 반환값만 이 표로
# 치환한다 — 로직과 표시 문구를 분리해 두면 나중에 표시 문구만 또 바뀌어도 판정 로직은 안전.
#   · 구 C-2 → 신 "C-1(핵심1개만)", 구 C-3 → 신 "C-2(핵심1개만)" — 번호가 한 칸씩 당겨진다(중태님
#     제공 표 그대로). 구 C-1("핵심1개+@")과 신 "C-1(핵심1개만)"이 둘 다 "C-1"로 시작하지만 괄호
#     설명으로 구분되므로 혼동 없음 — 표에 그렇게 명시돼 있어 그대로 반영.
#   · 품절근처 → "전사이즈 {X}개이하" — AD(SET 가능여부)의 항목12 "상하모두 {X}개 이하"와 같은
#     맥락(이번 실행에 쓰인 OK 문턱값 X를 그대로 대입, 하드코딩 아님). 품절은 문구 변경 없음(그대로).
_INV_GRADE_LABEL = {
    "A": "A(핵심2개이상)", "B": "B(핵심2개만)",
    "C-1": "C-1(핵심1개+@)", "C-2": "C-1(핵심1개만)", "C-3": "C-2(핵심1개만)",
    "D": "D(빅&스몰)", "E": "E(빅만)", "F": "F(스몰만)",
    "품절": "품절",
}


def _inv_grade_display(code, X):
    """_inv_grade_one()의 원시 등급코드를 260816 개편 표시 문구로 치환. 표에 없는 코드
    (해당없음·FREE SIZE·판정불가 등)는 원문 그대로 통과."""
    if code == "품절근처":
        return f"전사이즈 {X}개이하"
    return _INV_GRADE_LABEL.get(code, code)


def _inv_set_grade(mq, Y, top_stock=None, core=None, small=None, big=None):
    """SET 등급(AE) 판정 — 매칭된 세트 사이즈(mq={상의사이즈: 세트가능수량}) 기준.

    260806: 기존 A를 A-1 / A-2로 분리 (스킬 최신본 반영).
      A-1 = 핵심 전부 성립 & (핵심이 2개뿐인 체계는 다른 사이즈 1개↑도 성립) & 핵심의 '상의 재고'가 전부 Y 이상
      A-2 = 위와 같으나 핵심 상의 재고가 하나라도 Y 미만
    판정 기준이 '세트 가능 수량'이 아니라 '상의 재고'이므로 top_stock({상의사이즈: 재고})을 받는다.
    top_stock이 없으면(구 호출부) 안전하게 A-2로 떨어뜨린다.

    260807 개정: 핵심 개수는 체계마다 다르다(A16=2개, A09=3개, 260807 개정으로 M·L→M·L·XL 확장).
    단품 등급 판정(_inv_grade_one)과 동일하게 핵심 성립 개수(nc) 기준으로 일반화한다.
      · nc ≥ 3 → 핵심만으로 이미 최상위 등급(다른 사이즈 무관), 상의 재고 Y 체크로 A-1/A-2
      · nc == 2 → 다른 사이즈 1개↑ 성립해야 A-1/A-2, 없으면 B
                  (핵심이 원래 2개뿐인 A16 체계에선 이 분기가 곧 '핵심 전부 성립' 분기라 기존 동작과 동일)
      · nc == 1 → 다른 사이즈 있으면 C-1 / 없으면 세트가능수량 Y 기준 C-2·C-3
      · nc == 0 → 스몰&빅 D / 빅만 E / 스몰만 F
    """
    if not mq:
        return "해당없음"
    CORE = core if core is not None else _INV_SET_CORE
    SMALL = small if small is not None else _INV_SET_SMALL
    BIG = big if big is not None else _INV_SET_BIG
    core = [s for s in mq if s in CORE]
    others = [s for s in mq if s not in CORE]
    sm = [s for s in mq if s in SMALL]
    bg = [s for s in mq if s in BIG]
    nc = len(core)
    ts = top_stock or {}
    if nc >= 3:
        return "A-1" if all(ts.get(s, 0) >= Y for s in core) else "A-2"
    if nc == 2:
        if not others:
            return "B"
        return "A-1" if all(ts.get(s, 0) >= Y for s in core) else "A-2"
    if nc == 1:
        if others:
            return "C-1"
        return "C-2" if mq[core[0]] >= Y else "C-3"
    if sm and bg:
        return "D"
    if bg:
        return "E"
    if sm:
        return "F"
    return "해당없음"


# 260816 신규(중태님 지시) — SET 사이즈 컨디션(AE)도 단품(AC)과 동일한 표시 문구 개편.
# _inv_set_grade()의 내부 판정 로직은 무수정, 반환값만 이 표로 치환한다.
#   · 260818 정정(중태님): 260816에 A-1/A-2를 **둘 다 "A-1(핵심2개이상)"으로 통합** 표시하게
#     했던 건 실수였음 → **다시 A-1 / A-2로 구분해서 출력**한다. 즉 A-2는 "A-2(핵심2개이상)".
#     (판정 로직은 260816에도 손댄 적이 없어 이번에도 무수정 — 표시 문구만 되돌림.
#      A-1/A-2를 가르는 기준은 "핵심 2개 사이즈의 **상의 실재고**가 둘 다 Y 이상이면 A-1,
#      하나라도 Y 미만이면 A-2" — 세트 가능 수량이 아니라 상의 재고 기준. 260806 5차 참고.)
#   · 구 C-2 → 신 "C-1(핵심1개만)", 구 C-3 → 신 "C-2(핵심1개만)" — 단품(AC)과 동일한 번호 이동.
#   · 해당없음(세트 불가·미매칭)은 표에 없는 값이라 원문 그대로 통과.
_INV_SET_GRADE_LABEL = {
    "A-1": "A-1(핵심2개이상)", "A-2": "A-2(핵심2개이상)",   # 260818: A-2 다시 분리 표기
    "B": "B(핵심2개만)",
    "C-1": "C-1(핵심1개+@)", "C-2": "C-1(핵심1개만)", "C-3": "C-2(핵심1개만)",
    "D": "D(빅&스몰)", "E": "E(빅만)", "F": "F(스몰만)",
}


def _inv_set_grade_display(code):
    """_inv_set_grade()의 원시 등급코드를 260816 개편 표시 문구로 치환. 표에 없는 값
    (해당없음 등)은 원문 그대로 통과."""
    return _INV_SET_GRADE_LABEL.get(code, code)
# ▲▲▲ size-grade-classifier 스킬 이식 블록 끝 (SYNC BLOCK) ▲▲▲
# ══════════════════════════════════════════════════════════════════════════


def _inv_num(v):
    if v is None:
        return None
    try:
        return float(str(v).replace(",", "").strip())
    except Exception:
        return None


def _inv_cutoff(p):
    return "A" if p <= 20 else "B" if p <= 50 else "C" if p <= 80 else "D"


def _inv_price_sim(mol, gijun, chojo=None):
    """260811(가격시뮬) 몰가격·기준판매가로 신규 가격 5컬럼 계산 — (네이버)상시가·(쿠폰진행)상시가·
    (쿠폰진행)행사가·(쿠폰X/무배)상시가·(쿠폰X/무배)행사가 순.

    mol=몰가격(숫자 또는 None), gijun=기준판매가(숫자 또는 None), chojo=최초가(숫자 또는 None,
    260814 추가) — 전부 _inv_num() 등으로 이미 숫자/None 변환된 값이어야 한다.
    - 몰가격이 없으면(공란) 5개 전부 None.
    - 계산식: 상시가(네이버)=몰가격×1.05 / 상시가(쿠폰진행)=몰가격×1.15 / 행사가(쿠폰진행)=몰가격×1.1 /
      상시가(쿠폰X/무배)=몰가격<30,000이면 (몰가격×1.05)+3,000 아니면 몰가격×1.05 /
      행사가(쿠폰X/무배)=몰가격<30,000이면 몰가격+3,000 아니면 몰가격.
    - 5개 전부 끝 3자리 기준 3단계 스냅(260813 변경): 000→그대로 · 001~500→500 · 501~999→900.
    - 캡핑(중간 단계 없이 1회 비교):
        · 기준판매가가 있으면 → 계산값이 기준판매가를 넘을 때 그 계산 하나만 기준판매가로 대체.
        · 기준판매가가 공란이면(260814 추가) → 최초가와 비교해서, 계산값이 최초가를 넘으면 그 계산
          하나만 최초가로 대체. 최초가는 상품 택에 붙는 가격표 가격이라, 어떤 계산이든 이 가격보다
          비싸게 나가면 안 되기 때문 — 기준판매가가 없을 때의 대체 안전장치.
        · 기준판매가·최초가 둘 다 없으면 캡핑 없이 계산값 그대로.
    (2026-08-11 팀장 지시, 2026-08-14 최초가 대체 캡핑 추가 · 채팅으로 실데이터 예시 전수 검증 후 확정)
    """
    if mol is None:
        return [None] * 5

    def _r100(v):
        # 260813(팀장 지시): 100원 단위 반올림 대신, 계산값의 끝 3자리를 보고 3단계로 스냅한다.
        #   · 끝 3자리 000 → 그대로
        #   · 끝 3자리 001~500 → 500으로 맞춤
        #   · 끝 3자리 501~999 → 900으로 맞춤
        # (예: 34,545→34,900 / 37,835→37,900 / 36,190→36,500)
        v = int(round(v))
        base = (v // 1000) * 1000
        rem = v - base
        if rem == 0:
            return v
        elif rem <= 500:
            return base + 500
        else:
            return base + 900

    def _cap(v):
        v = _r100(v)
        if gijun is not None:
            g = _r100(gijun)
            if v > g:
                v = g
        elif chojo is not None:
            # 260814: 기준판매가가 공란일 때만 최초가를 대체 캡핑 기준으로 쓴다.
            c = _r100(chojo)
            if v > c:
                v = c
        return v

    c1 = _cap(mol * 1.05)
    c2 = _cap(mol * 1.15)
    c3 = _cap(mol * 1.1)
    if mol < 30000:
        c4 = _cap(mol * 1.05 + 3000)
        c5 = _cap(mol + 3000)
    else:
        c4 = _cap(mol * 1.05)
        c5 = _cap(mol)
    return [c1, c2, c3, c4, c5]


def read_size_master_file(uploaded_file):
    """사이즈 마스터 엑셀 → DF[품번, 사이즈코드]. 원본 파이프라인과 동일하게
    첫 시트 2행부터, C열(품번)·D열(사이즈코드)을 읽고 품번 중복은 첫 값 유지."""
    import openpyxl
    wb = openpyxl.load_workbook(uploaded_file, read_only=True)
    seen, rows = set(), []
    for r in wb.active.iter_rows(min_row=2, values_only=True):
        if len(r) > 3 and r[2] is not None and str(r[2]).strip():
            pn = str(r[2]).strip()
            if pn in seen:
                continue
            seen.add(pn)
            rows.append((pn, str(r[3]).strip()))
    wb.close()
    m = pd.DataFrame(rows, columns=["품번", "사이즈코드"])
    if m.empty:
        raise ValueError("사이즈 마스터에서 품번을 읽지 못했어요 — C열(품번)·D열(사이즈코드) 구조를 확인하세요.")
    return m


def replace_size_master(m):
    eng = get_engine()
    with eng.begin() as conn:
        m.astype(str).to_sql(SIZE_MASTER_TABLE, conn, if_exists="replace", index=False)
    return len(m)


@st.cache_data(ttl=21600)
def load_size_master():
    """DB의 사이즈 마스터를 dict{품번: 사이즈코드}로 반환. 없으면 빈 dict."""
    eng = get_engine()
    try:
        with eng.connect() as conn:
            exists = conn.exec_driver_sql(
                "SELECT 1 FROM information_schema.tables WHERE table_name=%s"
                if eng.dialect.name == "postgresql" else
                "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
                (SIZE_MASTER_TABLE,)).fetchone()
            if not exists:
                return {}
            m = pd.read_sql(f'SELECT * FROM "{SIZE_MASTER_TABLE}"', conn)
    except Exception:
        return {}
    if m.empty or "품번" not in m.columns or "사이즈코드" not in m.columns:
        return {}
    return dict(zip(m["품번"].astype(str).str.strip(), m["사이즈코드"].astype(str).str.strip()))


def size_master_row_count():
    try:
        with get_engine().connect() as conn:
            return conn.exec_driver_sql(f'SELECT COUNT(*) FROM "{SIZE_MASTER_TABLE}"').scalar()
    except Exception:
        return 0


# ── 아이템 마스터 (아이템코드 → 아이템명·대카테고리·중카테고리·소카테고리·상하의구분) ──
# "AI 마스터파일 우리회사 품번 코드 체계" 워크북의 '아이템코드와 카테고리 구분' 시트를 그대로 업로드.
# 재고모니터링 중카테고리(_inv_cat_lookup)와 판매분석 아이템그룹(get_itemgroup_map)이 모두 여기서 나온다.
def read_item_master_file(uploaded_file):
    """아이템 마스터 엑셀에서 '아이템코드/아이템명/대카테고리/중카테고리/소카테고리/상하의 구분' 헤더가 있는
    시트를 자동으로 찾아 DF로 변환. (시트 순서·이름이 바뀌어도 헤더로 탐색하므로 안전)"""
    import openpyxl
    uploaded_file.seek(0)
    wb = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
    target_ws = None
    for ws in wb.worksheets:
        first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if first and len(first) >= 4 and str(first[0]).strip() == "아이템코드" \
                and str(first[1]).strip() == "아이템명":
            target_ws = ws
            break
    if target_ws is None:
        wb.close()
        raise ValueError("'아이템코드·아이템명·대카테고리·중카테고리·소카테고리' 헤더가 있는 시트를 "
                         "찾지 못했어요 — 시트 첫 행이 이 헤더로 시작해야 해요.")
    rows = []
    for r in target_ws.iter_rows(min_row=2, values_only=True):
        if not r or r[0] in (None, ""):
            continue
        code = str(r[0]).strip().upper()
        rows.append({
            "item_code": code,
            "item_name": str(r[1]).strip() if len(r) > 1 and r[1] is not None else "",
            "cat_large": str(r[2]).strip() if len(r) > 2 and r[2] is not None else "",
            "cat_mid": str(r[3]).strip() if len(r) > 3 and r[3] is not None else "",
            "cat_small": str(r[4]).strip() if len(r) > 4 and r[4] is not None else "",
            "top_bottom": str(r[5]).strip() if len(r) > 5 and r[5] is not None else "",
        })
    wb.close()
    if not rows:
        raise ValueError("아이템 마스터 시트에서 데이터 행을 찾지 못했어요.")
    return pd.DataFrame(rows).drop_duplicates(subset=["item_code"], keep="last")


def replace_item_master(m):
    eng = get_engine()
    with eng.begin() as conn:
        m.astype(str).to_sql(ITEM_MASTER_TABLE, conn, if_exists="replace", index=False)
    return len(m)


@st.cache_data(ttl=21600)
def load_item_master():
    """DB의 아이템 마스터를 dict{아이템코드: {name,large,mid,small,topbottom}}로 반환. 없으면 빈 dict."""
    eng = get_engine()
    try:
        with eng.connect() as conn:
            exists = conn.exec_driver_sql(
                "SELECT 1 FROM information_schema.tables WHERE table_name=%s"
                if eng.dialect.name == "postgresql" else
                "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
                (ITEM_MASTER_TABLE,)).fetchone()
            if not exists:
                return {}
            m = pd.read_sql(f'SELECT * FROM "{ITEM_MASTER_TABLE}"', conn)
    except Exception:
        return {}
    if m.empty or "item_code" not in m.columns:
        return {}
    out = {}
    for _, row in m.iterrows():
        out[str(row["item_code"]).strip().upper()] = {
            "name": row.get("item_name", "") or "",
            "large": row.get("cat_large", "") or "",
            "mid": row.get("cat_mid", "") or "",
            "small": row.get("cat_small", "") or "",
            "topbottom": row.get("top_bottom", "") or "",
        }
    return out


def item_master_row_count():
    try:
        with get_engine().connect() as conn:
            return conn.exec_driver_sql(f'SELECT COUNT(*) FROM "{ITEM_MASTER_TABLE}"').scalar()
    except Exception:
        return 0


def _inv_cat_lookup(item_code):
    """재고모니터링 중카테고리 조회 — 아이템 마스터(item_master) 우선, 없으면 구 하드코딩 폴백."""
    m = load_item_master()
    rec = m.get(item_code)
    if rec and rec["mid"]:
        return rec["mid"]
    return _INV_CAT_FALLBACK.get(item_code)


def _inv_cat_small_lookup(item_code):
    """재고모니터링 소카테고리 조회 — 아이템 마스터 우선, 마스터에 없는 코드는 소카테고리가
    따로 없으므로 중카테고리 폴백값을 그대로 대신 사용(최선 근사치, 화면에도 그렇게 표시됨)."""
    m = load_item_master()
    rec = m.get(item_code)
    if rec and rec["small"]:
        return rec["small"]
    return _INV_CAT_FALLBACK.get(item_code)


# 260803: 재고모니터링 C열·등급 모집단의 "카테고리 기준" 축 — 중카테고리(기본)/소카테고리/아이템코드 중 선택.
INV_CAT_LEVELS = ["중카테고리", "소카테고리", "아이템코드"]


def _inv_cat_level_value(item_code, cat_level):
    """cat_level 선택에 따라 C열에 표기·등급 모집단에 쓰일 값을 반환."""
    if cat_level == "소카테고리":
        return _inv_cat_small_lookup(item_code)
    if cat_level == "아이템코드":
        return item_code
    return _inv_cat_lookup(item_code)  # 기본: 중카테고리


# 판매분석 아이템그룹: 아이템 마스터 중카테고리를 기본으로 쓰되, 팀 요청으로 ACC에서 분리해온
# 신발·넥타이·벨트·양말은 마스터의 중카테고리 값과 무관하게 이 라벨을 그대로 덮어쓴다.
_ITEMGROUP_OVERRIDE_SPLIT = {"FW": "신발", "NT": "넥타이", "BE": "벨트", "SC": "양말"}


@st.cache_data(ttl=21600)
def get_itemgroup_map():
    """판매분석 아이템그룹 맵 — 아이템 마스터(item_master) 기준 + 팀 커스텀 분리 + 구 하드코딩 폴백."""
    m = load_item_master()
    out = dict(_ITEMGROUP_MAP_FALLBACK)  # 마스터에 없는 코드는 이 값 그대로 유지
    for code, rec in m.items():
        if code in _ITEMGROUP_OVERRIDE_SPLIT:
            out[code] = _ITEMGROUP_OVERRIDE_SPLIT[code]
        elif rec["mid"]:
            out[code] = rec["mid"]
    return out


# 복종별 판매비중 분석(2026-08-08) 전용 소카테고리 맵 — get_itemgroup_map()(중카테고리)과 짝을 이룬다.
# 아이템 마스터의 small 필드 우선, 없으면 중카테고리 값을 그대로 대신 사용(재고모니터링의
# _inv_cat_small_lookup과 동일한 폴백 원칙 — 소카테고리가 따로 없는 코드는 최선 근사치로 보여줌).
@st.cache_data(ttl=21600)
def get_itemgroup_map_small():
    m = load_item_master()
    out = dict(get_itemgroup_map())  # 폴백: 마스터에 없거나 small이 비어있으면 중카테고리 값 사용
    for code, rec in m.items():
        if code in _ITEMGROUP_OVERRIDE_SPLIT:
            out[code] = _ITEMGROUP_OVERRIDE_SPLIT[code]
        elif rec["small"]:
            out[code] = rec["small"]
        elif rec["mid"]:
            out[code] = rec["mid"]
    return out


# 복종별 판매비중 분석 카테고리 기준 — 중카테고리(기본)/소카테고리 중 선택.
CATMIX_CAT_LEVELS = ["중카테고리", "소카테고리"]


def _inv_peek_seasons(raw_file):
    """업로드된 로우데이터에서 실제로 쓰인 시즌 코드(컬럼17·raw[16])를 미리 훑어 목록으로 반환.

    Z/A/B/C/D(공통·봄·여름·가을·겨울) 알려진 순서를 먼저, 그 외 값은 뒤에 알파벳순.
    파일 포인터를 되돌리기 위해 호출 측에서 반드시 raw_file.seek(0)을 다시 해줘야 한다.
    """
    import openpyxl
    try:
        raw_file.seek(0)
    except Exception:
        pass
    wb = openpyxl.load_workbook(raw_file, read_only=True)
    ws = wb.worksheets[0]
    codes = []
    seen = set()
    for r in ws.iter_rows(min_row=3, values_only=True):
        if not r or r[0] in (None, ""):
            continue
        if len(r) <= 16 or r[16] is None:
            continue
        v = str(r[16]).strip()
        if v and v not in seen:
            seen.add(v); codes.append(v)
    wb.close()
    order = {"Z": 0, "A": 1, "B": 2, "C": 3, "D": 4, "E": 5}
    return sorted(codes, key=lambda c: (order.get(c, 99), c))


def _inv_peek_years(raw_file):
    """업로드된 로우데이터에서 실제로 쓰인 년도 코드(컬럼16·raw[15])를 미리 훑어 목록으로 반환.

    숫자로 해석되는 값은 오름차순, 그 외는 뒤에 알파벳순.
    파일 포인터를 되돌리기 위해 호출 측에서 반드시 raw_file.seek(0)을 다시 해줘야 한다.
    """
    import openpyxl
    try:
        raw_file.seek(0)
    except Exception:
        pass
    wb = openpyxl.load_workbook(raw_file, read_only=True)
    ws = wb.worksheets[0]
    codes = []
    seen = set()
    for r in ws.iter_rows(min_row=3, values_only=True):
        if not r or r[0] in (None, ""):
            continue
        if len(r) <= 15 or r[15] is None:
            continue
        v = str(r[15]).strip()
        if v and v not in seen:
            seen.add(v); codes.append(v)
    wb.close()

    def _key(c):
        try:
            return (0, float(c))
        except ValueError:
            return (1, c)
    return sorted(codes, key=_key)


def process_inventory(raw_file, master, template_path, X, Y, period, workdate,
                      season_group_map=None, year_group_map=None, cat_level="중카테고리",
                      vol_a=500, vol_b=300, vol_c=100):
    """재고 모니터링 로우데이터(94열, '사이즈구분' 컬럼 포함) → 123열 v3.5 가공 엑셀
    (process_260731 main() 1:1 이식 + 260811 사이즈코드 판정 소스 전환 + 가격 시뮬레이션 5컬럼 +
    260826 SET 가격 9컬럼 + 260826-2 물량등급 +
    기준판매가 비교컬럼 반영).

    raw_file=업로드 파일 객체, master=dict{품번:사이즈코드}(260811부터 판정에는 미사용, 참고 보고 전용),
    template_path=v3.3 서식 템플릿 경로.
    season_group_map={시즌코드: 비교대상군 라벨}, year_group_map={년도코드: 비교대상군 라벨}이면
    AA·AB 등급의 모집단(카테고리×년도×시즌)에서 시즌·년도 축을 각각 그 라벨 기준으로 묶어서 계산한다
    (예: season_group_map={"Z":"Z+A+C+D","A":"Z+A+C+D","C":"Z+A+C+D","D":"Z+A+C+D","B":"B"}
    → Z·A·C·D는 하나의 비교 대상군, B는 별도). None/미지정 코드는 원래 코드를 그대로 자기 자신의
    비교 대상군으로 사용(기존 동작과 동일 — 시즌·년도마다 따로 등급 매김).
    cat_level(INV_CAT_LEVELS 중 하나, 기본 "중카테고리")은 모집단의 카테고리 축과 C열 표기 기준을
    "중카테고리"/"소카테고리"/"아이템코드" 중 어느 세분화 레벨로 쓸지 정한다. C열 헤더 텍스트도
    이 선택에 맞춰 동적으로 바뀐다.
    vol_a/vol_b/vol_c(260826-2 물량등급): 온라인창고 재고 기준 A/B/C 등급 문턱(장, 이상 기준 —
    vol_a > vol_b > vol_c 이어야 함). D = vol_c 미만.
    반환: (엑셀 bytes, 리포트 dict). 규칙 위반·형식 오류는 ValueError로 중단.
    """
    import re
    from copy import copy
    from collections import defaultdict, Counter
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    fill_yellow = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
    fill_green = PatternFill(start_color="FF92D050", end_color="FF92D050", fill_type="solid")
    # 260815(헤더 개편): 단품 사이즈 컨디션·SET 가능여부·SET 사이즈 컨디션 3컬럼 = 연분홍, 수정일 = 회색.
    fill_pink = PatternFill(start_color="FFF2DCDB", end_color="FFF2DCDB", fill_type="solid")
    fill_gray = PatternFill(start_color="FFE9E9E9", end_color="FFE9E9E9", fill_type="solid")

    # 260826-2(물량등급): 기준 검증 — UI에서도 막지만 함수 단독 호출 대비 이중 방어.
    if not (vol_a > vol_b > vol_c > 0):
        raise ValueError(f"물량등급 기준이 A > B > C 순이 아니에요 (A={vol_a}, B={vol_b}, C={vol_c}) — "
                         f"숫자를 확인해 주세요.")

    try:
        raw_file.seek(0)
    except Exception:
        pass
    wb = openpyxl.load_workbook(raw_file, read_only=True)
    ws = wb.worksheets[0]
    _tail_col_ignored = ws.max_column == INV_RAW_COLS + 1
    if _tail_col_ignored:
        # 260811 이후 회사 ERP 추출 파일 일부에서 95번째(맨 끝) 열이 통째로 빈 채 딸려오는 경우가
        # 확인됨(엑셀 병합헤더 잔재로 추정 — row1 그룹헤더 '계'가 실제 재고계 칸보다 1칸 더 넓게
        # 병합돼 있고, row2 컬럼명·데이터는 전부 공란). 그 칸에 실제 값이 있으면(향후 로우데이터
        # 스펙이 진짜로 바뀐 걸 수 있으니) 중단하고 확인을 요청, 전부 공란이면 무시하고 94열로 처리
        # (아래 로직은 어차피 인덱스 0~93만 사용하므로 95번째 열은 있어도 그냥 안 쓰인다).
        _tail_vals = [r[INV_RAW_COLS] for r in ws.iter_rows(min_row=3, values_only=True)
                     if r and r[0] not in (None, "")]
        if any(v not in (None, "") for v in _tail_vals):
            raise ValueError(f"로우데이터 열 수 {ws.max_column} ≠ {INV_RAW_COLS} — 마지막(95번째) 열에 "
                             f"값이 들어있어요. 빈 꼬리 컬럼이면 자동으로 무시하는데, 이번엔 값이 있어서 "
                             f"로우데이터 스펙이 바뀐 건 아닌지 확인이 필요해요(담당팀에 문의해주세요).")
    elif ws.max_column != INV_RAW_COLS:
        raise ValueError(f"로우데이터 열 수 {ws.max_column} ≠ {INV_RAW_COLS} — 파일을 확인하세요 "
                         f"(93열이면 '사이즈구분' 컬럼이 없는 구형식이에요 · 48열이면 매출 파일이에요).")
    rows = [r for r in ws.iter_rows(min_row=3, values_only=True)]
    skipped = sum(1 for r in rows if r[0] in (None, ""))
    rows = [r for r in rows if r[0] not in (None, "")]
    wb.close()
    if not rows:
        raise ValueError("로우데이터에서 상품 행을 찾지 못했어요 (3행부터 읽어요).")

    recs = []
    for r in rows:
        pn = str(r[0]).strip(); item = str(r[14]).strip()
        season_raw = str(r[16]).strip()
        year_raw = str(r[15]).strip()
        # 260811: 사이즈코드는 사이즈 마스터 조회 대신 로우데이터의 '사이즈구분' 컬럼(raw 79번째 열,
        # 0-index 78)에서 직접 읽는다. 공란이면 마스터로 되돌아가 보완하지 않고 바로 '해당없음' 처리
        # (scode=None). 사이즈 마스터(master)는 더 이상 판정에 쓰지 않고, 새 컬럼값과의 불일치를
        # 참고 보고하는 용도로만 scode_master에 남겨둔다.
        scode_cell = r[78]
        scode = str(scode_cell).strip() if scode_cell not in (None, "") else None
        rec = dict(raw=r, pn=pn, item=item, cat=_inv_cat_lookup(item),
                   cat_level_val=_inv_cat_level_value(item, cat_level),  # C열·모집단에 실제로 쓰이는 값
                   year=year_raw, season=season_raw,
                   season_grp=(season_group_map or {}).get(season_raw, season_raw),  # 비교 대상군(묶음) 라벨
                   year_grp=(year_group_map or {}).get(year_raw, year_raw),          # 비교 대상군(묶음) 라벨
                   # 260811: 오프라인 판정 소스를 '이관구분'(raw22)에서 '수정일'(raw21)로 변경
                   # — 실제 오프라인/온라인/부분이관 값이 담긴 컬럼은 '수정일'이라는 이름으로 내려온다.
                   off=str(r[21]).strip() == "오프라인",
                   stock=_inv_num(r[39]) or 0, sales=_inv_num(r[45]) or 0, depl=_inv_num(r[47]),
                   size14={i + 1: int(_inv_num(r[79 + i]) or 0) for i in range(14)},
                   scode=scode, scode_master=master.get(pn))
        recs.append(rec)

    # 260811: 사이즈구분에 정식 6종(A16/A17/A09/A05/A06/A18) 외의 값이 있으면 중단하지 않고 경고만
    # 남긴다 — 오탈자이거나, 아직 등급 로직에 반영되지 않은 신규 사이즈체계일 수 있음. 해당 건은
    # _INV_SYSTEMS에 없으므로 AC 판정에서 자동으로 '해당없음' 처리된다.
    unk_size_codes = sorted({r["scode"] for r in recs
                             if r["scode"] and r["scode"] not in _INV_KNOWN_SIZE_CODES})
    # 260811: 사이즈구분 공란(미매칭) 건수 + 참고용 사이즈 마스터 대조(마스터가 있을 때만 의미 있음).
    scode_blank = sum(1 for r in recs if r["scode"] is None)
    scode_blank_had_master = sum(1 for r in recs if r["scode"] is None and r.get("scode_master"))
    scode_mismatch = [(r["pn"], r["scode"], r["scode_master"]) for r in recs
                      if master and r["scode"] and r.get("scode_master")
                      and r["scode"] != r["scode_master"]]

    _item_master = load_item_master()
    fallback_items = sorted({r["item"] for r in recs if r["item"] not in _item_master})

    # 260806: 중카테고리 매핑 없는 아이템 코드를 만나도 '중단'하지 않는다 (팀 요청).
    #   · 해당 상품 행은 결과 파일에 그대로 남긴다 — 재고가 눈앞에서 사라지면 총재고가 원본과
    #     안 맞고 "이 상품 왜 없지?" 사고가 나므로.
    #   · 다만 C열과 AA/AB 등급 모집단에서는 뺀다 — 카테고리를 모르는 상품을 남의 모집단에
    #     끼워 넣으면 그 모집단 전체의 랭킹이 오염되기 때문.
    #   · AA·AB·AF는 '미분류'로 표기. AC(사이즈 등급)·SET 판정은 사이즈 마스터 기준이라
    #     아이템 코드 매핑과 무관 → 그대로 정상 판정된다.
    #   · 아이템 마스터에 코드를 추가하고 다시 돌리면 자동으로 정상 등급을 받는다.
    unk_item = sorted({r["item"] for r in recs if r["cat"] is None})
    unk_rows = 0
    for rec in recs:
        rec["unmapped"] = rec["cat"] is None
        if rec["unmapped"]:
            unk_rows += 1
            if rec["cat_level_val"] in (None, ""):
                rec["cat_level_val"] = _INV_UNMAPPED

    # 260826-3(소카테고리 반영 확인, 중태님 문의): 선택한 카테고리 기준이 모집단에 실제로 몇 종의
    # 값으로 반영됐는지 + 소카테고리 선택 시 아이템 마스터에 소카테고리가 비어 있어 중카테고리로
    # 폴백된 아이템코드 목록을 리포트로 노출한다 — "소카테고리를 골라도 결과가 그대로"가
    # (a) 마스터의 소카테고리 공란/중카테고리와 동일값 때문인지 (b) 진짜 버그인지 화면에서 바로
    # 판별할 수 있게. (등급 모집단 키는 위 rec["cat_level_val"] — 선택 기준이 그대로 쓰인다.)
    cat_values = sorted({str(r["cat_level_val"]) for r in recs if not r["unmapped"]})
    cat_small_fallback = sorted({r["item"] for r in recs if not r["unmapped"]
                                 and not (_item_master.get(r["item"]) or {}).get("small")}) \
        if cat_level == "소카테고리" else []

    # K/L/M (세트 키)
    for rec in recs:
        pn, item = rec["pn"], rec["item"]
        rec["K"] = pn + ","
        rec["L"] = (pn[0] + ("SJ" if item in ("SJ", "SL") else "EJ") + pn[3:7] + "SET") \
            if item in ("SJ", "SL", "EJ", "EP") else ""
        rec["M"] = rec["L"] + "," if rec["L"] else ""

    # AA (기간판매 랭킹 등급) — 모집단 = 카테고리(cat_level 선택) × 년도 비교대상군(묶음) × 시즌 비교대상군(묶음)
    groups = defaultdict(list)
    for r in recs:
        if not r["unmapped"] and not r["off"] and r["stock"] >= 20 and r["sales"] > 0:
            groups[(r["cat_level_val"], r["year_grp"], r["season_grp"])].append(r)
    for g in groups.values():
        gs = sorted(g, key=lambda r: -r["sales"]); n = len(gs)
        for rank, r in enumerate(gs, 1):
            r["AA"] = _inv_cutoff(rank / n * 100)
    for r in recs:
        if r["off"]:
            r["AA"] = "오프라인"
        elif r["stock"] < 20:
            r["AA"] = "재고20미만"
        elif r["sales"] <= 0:
            r["AA"] = "E"
    for r in recs:                       # 260806: 미분류가 최우선 — 모집단에 없으므로 AA 키 자체가 없다
        if r["unmapped"]:
            r["AA"] = _INV_UNMAPPED

    # AB (소진 속도 등급) — 모집단 = AA와 동일 기준(카테고리(cat_level 선택) × 년도 비교대상군 × 시즌 비교대상군)
    groupsb = defaultdict(list)
    for r in recs:
        if not r["unmapped"] and not r["off"] and r["sales"] > 0:
            groupsb[(r["cat_level_val"], r["year_grp"], r["season_grp"])].append(r)
    INF = float("inf")
    for g in groupsb.values():
        gs = sorted(g, key=lambda r: r["depl"] if r["depl"] is not None else INF); n = len(gs)
        for rank, r in enumerate(gs, 1):
            r["AB"] = _inv_cutoff(rank / n * 100)
    for r in recs:
        if r["off"]:
            r["AB"] = "오프라인"
        elif r["sales"] <= 0:
            r["AB"] = "E"
    for r in recs:                       # 260806: AA와 동일 — 미분류 우선
        if r["unmapped"]:
            r["AB"] = _INV_UNMAPPED

    # AF 매트릭스 (AI제안방향)
    for r in recs:
        if r["unmapped"]:                # 260806: 카테고리를 모르면 제안 방향도 낼 수 없다
            r["AF"] = _INV_UNMAPPED; continue
        if r["off"]:
            r["AF"] = "오프라인"; continue
        if r["stock"] < 20:
            r["AF"] = "재고20미만"; continue
        aa, ab, s = r["AA"], r["AB"], r["stock"]
        if aa in "AB" and ab == "A":
            r["AF"] = "가격인상" if s < 200 else "가격유지"
        elif aa in "AB" and ab in "BC":
            r["AF"] = "가격유지"
        elif aa == "C" and ab in "AB":
            r["AF"] = "가격유지"
        elif aa in "CD" and ab in "CD":
            r["AF"] = "가격인하"
        elif aa in "AB" and ab == "D":
            r["AF"] = "자사타임특가"
        elif aa == "D" and ab in "AB":
            r["AF"] = "자사타임특가"
        elif aa == "E" and ab == "E":
            r["AF"] = "진열/가격확인"
        else:
            r["AF"] = "검증필요"
    af_bad = sum(1 for r in recs if r["AF"] == "검증필요")

    # AC 사이즈 등급
    for r in recs:
        c = r["scode"]
        if c is None or c == "A18":
            r["AC"] = "해당없음"
        elif c == "A06":
            # 260814: A06(FREE, 가방·모자·벨트 등 1사이즈 상품)은 문자 등급 대신 총재고 합계만 보고
            # X 기준 이상이면 판정 — 표기를 "OK"에서 "FREE SIZE"로 변경(팀장 지시, 상품 특성이 더
            # 잘 드러나도록). 미달 시 "품절근처"는 기존과 동일.
            r["AC"] = ("FREE SIZE" if sum(r["size14"].values()) >= X
                       else _inv_grade_display("품절근처", X))
        elif c in _INV_SYSTEMS:
            r["AC"] = _inv_grade_display(_inv_grade_one(r["size14"], _INV_SYSTEMS[c], X, Y), X)
        else:
            r["AC"] = "해당없음"

    # 260826-2 물량등급 — 온라인창고 재고(rec["stock"], 재고20미만 판정과 동일 소스) 숫자만으로
    # A/B/C/D 4등급. 모집단·카테고리와 무관한 단순 문턱 분류라 오프라인·미분류 행도 똑같이 매긴다.
    # 표기는 "물량A"~"물량D" — 옆의 기간판매·소진예상 등급(맨 A~E)과 필터·눈으로 바로 구분되게
    # (중태님 지시: A/B/C/D가 물량 등급임을 알 수 있게 표기).
    for r in recs:
        s = r["stock"]
        r["VOL"] = "물량" + ("A" if s >= vol_a else "B" if s >= vol_b else "C" if s >= vol_c else "D")

    # AD/AE SET 판정
    # 260814(중태님 지시): 기존 "해당없음" 하나였던 표기를 2가지로 세분화.
    #   · "단품아이템" — SET품번(L) 자체가 없는 상품. rec["L"]은 아이템코드가 SJ·SL·EJ·EP일 때만
    #     붙으므로(위 K/L/M 계산부 참고), 그 외 코드(ACC·KT·TS·JP·JA·CT·DJ·NT·SH·PA 등)는 애초에
    #     세트 판매 후보가 아니라 항상 이 값.
    #   · "SET구성실패" — SET품번(L)은 있는(=진짜 SJ/EJ/SL/EP 세트 상품인) 그룹인데 (a) 상/하 한쪽
    #     데이터가 이번 로우데이터에 없거나, (b) 사이즈코드 조합이 아직 미지원(예: A09×A17)이거나,
    #     (c) 상/하 데이터가 다 있고 지원 조합인데도 사이즈가 하나도 안 맞아떨어져 실제 매칭에 실패한
    #     경우. 실데이터 검증(1,410행)에서 이 케이스가 12건 확인됨 — 전부 진짜 SJ/EJ/SL/EP 세트 상품인데
    #     판정만 실패한 것이라 "단품아이템"과 구분해야 담당자가 "원래 단품"과 "세트인데 매칭 실패"를
    #     혼동하지 않는다.
    for r in recs:
        r["AD"] = "단품아이템" if not r["L"] else "SET구성실패"
        r["AE"] = "단품아이템" if not r["L"] else "SET구성실패"
    bysets = defaultdict(list)
    for r in recs:
        if r["L"]:
            bysets[r["L"]].append(r)
    pairs = nopair = 0

    # 260826(SET가격 9컬럼, 중태님 지시): 한 행의 단품가격 9칸(최초가·현판가·몰가격·기준판매가·
    # 가격시뮬5)을 계산해 리스트로 돌려주는 헬퍼 — 출력 루프의 20~28번째 칸과 정확히 같은 값.
    def _inv_row_price9(r_):
        raw_ = r_["raw"]
        _m = _inv_num(raw_[INV_RAW_MOLGA_COL])
        _g = _inv_num(raw_[INV_RAW_GIJUN_COL])
        _c = _inv_num(raw_[INV_RAW_CHOJOGA_COL])
        _h = _inv_num(raw_[INV_RAW_HYUNPAN_COL])
        return [_c, _h, _m, _g] + _inv_price_sim(_m, _g, _c)

    for g in bysets.values():
        # 260806: 상/하 판별을 사이즈코드(A16/A17)가 아니라 아이템 코드로 한다 — A09↔A09 세트업을
        #         잡으려면 필수. 지원 조합은 A16↔A17(숫자 정사이즈) · A09↔A09(문자) · A09↔A17(신규,
        #         260815 — 상의 문자/하의 숫자 혼합) 셋.
        #         그 외(A06 FREE · 단독행)는 이 그룹에 진입한
        #         시점에 이미 L(SET품번)이 있는 것이 확정이므로 '단품아이템'이 아니라 'SET구성실패'로
        #         남는다(260814, 위 기본값 설정부 참고). A18↔A18(아동)은 아래에서 별도로 '아동복'으로
        #         표기한다(260815).
        tops = [r for r in g if _inv_set_side(r) == "top"]
        bots = [r for r in g if _inv_set_side(r) == "bot"]
        if not tops or not bots:
            nopair += 1; continue
        ti, bi = tops[0], bots[0]
        # 260826(SET가격 9컬럼): 상/하 데이터가 둘 다 있는 그룹은 상의+하의 단품가격 9칸을 각각
        # 합산한 "SET 가격"을 그룹의 모든 행에 심는다(수기 샘플의 =T상의행+T하의행 수식과 동일).
        # 가격은 사이즈와 무관하므로 사이즈 매칭 성패·미지원 조합(아래 continue들)보다 먼저 처리 —
        # SET구성실패·아동복 행에도 SET 가격은 채워진다. 짝은 품번 끝 3자리(라인·패턴·색상)가 같은
        # 반대편 행을 우선 매칭하고, 없으면 반대편 첫 행으로 폴백(기존 ti/bi 관례와 동일). 한쪽
        # 값이 공란이면 엑셀 수식과 같게 0으로 보고 더하되, 양쪽 다 공란인 칸은 공란으로 둔다.
        for r in g:
            _pool = bots if _inv_set_side(r) == "top" else tops
            _mate = next((p for p in _pool if p["pn"][7:] == r["pn"][7:]), _pool[0])
            _a9, _b9 = _inv_row_price9(r), _inv_row_price9(_mate)
            r["SETP"] = [None if (a is None and b is None) else (a or 0) + (b or 0)
                         for a, b in zip(_a9, _b9)]
        sys_key = (ti["scode"], bi["scode"])
        # 260815 신규(중태님 지시): A18↔A18(아동)은 다른 미지원 조합과 묶어 "SET구성실패"로 뭉뚱그리지
        # 않고, 아동복이라는 실제 성격을 그대로 드러내는 "아동복"으로 별도 표기한다. SET품번은 있으나
        # (=SJ/EJ/SL/EP 아이템코드) 아동 사이즈체계(A18)라 애초에 등급 판정 대상이 아니라는 뜻.
        if sys_key == ("A18", "A18"):
            for r in g:
                r["AD"] = r["AE"] = "아동복"
            nopair += 1; continue
        if sys_key not in _INV_SET_SYS:
            nopair += 1; continue
        TMAP, BMAP, MTBL, S_CORE, S_SMALL, S_BIG = _INV_SET_SYS[sys_key]
        T_IDX = {v: k for k, v in TMAP.items()}
        B_IDX = {v: k for k, v in BMAP.items()}
        pairs += 1
        top_ok = [TMAP[k] for k in TMAP if ti["size14"][k] >= X]
        bot_ok = [BMAP[k] for k in BMAP if bi["size14"][k] >= X]
        matched = {}; used = set()
        excess_bot = set(); excess_top = set()
        for ts in top_ok:
            cand = [bs for bs in MTBL.get(ts, []) if bs in bot_ok]
            if cand:
                _top_stock = ti["size14"][T_IDX[ts]]
                matched[ts] = min(_top_stock, max(bi["size14"][B_IDX[b]] for b in cand))
                used.update(cand)
                # 260815 규칙(3): 상의 재고가 매칭되는 모든 후보 팬츠 재고 합계의 1.3배 이상이면
                # 상의가 과다재고 → 상의단품 신호로 추가.
                # 260816 3차 개정(중태님 지시): 이 합산 분모는 "세트 성립 판정용" cand(X문턱 게이트,
                # 위 라인)가 아니라 별도로 계산 — 매칭표(MTBL) 후보 중 이 사이즈 체계에 실재하는
                # 사이즈(B_IDX에 있음)이면서 재고 3장(_INV_SET_EXCESS_DEMAND_MIN) 이상이면 X문턱
                # 미만이라도 전부 포함한다. 예: 재킷100의 매칭표 후보는 78·80·82·84인데, 80은
                # 애초에 이 사이즈체계에 없는 사이즈(B_IDX에 없음)라 제외, 78(9장)·84(7장)는 X문턱
                # (10장) 미만이라도 3장은 넘으므로 포함 — "세트 성립"엔 못 쓰지만 과다재고 여부를
                # 가릴 땐 무시하기엔 너무 큰 재고라 분모에 반영.
                _cand_all = [bs for bs in MTBL.get(ts, []) if bs in B_IDX
                             and bi["size14"][B_IDX[bs]] >= _INV_SET_EXCESS_DEMAND_MIN]
                _cand_sum = sum(bi["size14"][B_IDX[b]] for b in _cand_all)
                if _top_stock >= _INV_SET_EXCESS_RATIO * _cand_sum:
                    excess_top.add(ts)
        # 260816 개정(중태님 지시) — 규칙(2) 합산기준: 팬츠 한 사이즈가 여러 상의 사이즈의 매칭
        # 후보로 동시에 걸릴 수 있어서(예: 팬츠 78 ↔ 재킷 95·100 둘 다), 종전처럼 그중 하나의
        # 상의 재고와만 비교하면 다른 상의의 수요를 무시한 오판이 나온다(하의78=93장이 재킷100=25장
        # 하나 기준으론 과다지만, 재킷95=120장 수요까지 합치면 전혀 과다가 아닌 사례로 실측 확인됨).
        # → 비교 분모를 "이 팬츠를 후보로 삼는 모든 매칭 상의 사이즈 재고의 합"으로 바꾸고,
        # 배율도 1.5배 → 1.3배로 낮춘다(합산기준은 분모가 커져 1.5배를 그대로 쓰면 과소검출됨 —
        # 1/1.2/1.3/1.5 실측 시뮬레이션 결과에 근거).
        # 260816 3차 개정: 규칙(3)과 동일한 이유로, 이 분모도 X문턱 게이트된 후보만이 아니라
        # 매칭표(MTBL) 전체 상의 사이즈(TMAP.values()) 중 이 팬츠를 후보로 삼는 것들을 재고 3장
        # 이상이면 X문턱 무관하게 전부 포함해서 계산한다.
        for b in used:
            _demand_sum = sum(ti["size14"][T_IDX[ts]] for ts in TMAP.values()
                               if b in MTBL.get(ts, []) and ti["size14"][T_IDX[ts]] >= _INV_SET_EXCESS_DEMAND_MIN)
            if bi["size14"][B_IDX[b]] >= _INV_SET_EXCESS_RATIO_BOT * _demand_sum:
                excess_bot.add(b)
        # 260815: 구 1.3배 그룹비율 게이트 폐지. tl/bl에 규칙(1)의 미매칭 잔여와 규칙(2)/(3)의
        # 사이즈별 과다재고 신호(excess_bot/excess_top)를 합쳐 최종 잔여로 삼는다.
        top_nomatch = set(top_ok) - set(matched)   # 규칙(1) — 애초에 매칭 후보 자체가 없는 진짜 미매칭
        bot_nomatch = set(bot_ok) - used            # 〃
        tl = top_nomatch | excess_top
        bl = bot_nomatch | excess_bot
        # 260815 추가(중태님 지시, 수정2): 이 행의 사이즈정보 14칸 중 "남는 상의/남는 하의"에
        # 해당하는 실제 사이즈 칸을 결과물에서 색으로 표시하기 위해, idx 위치(1~14)를
        # tops/bots 각 행에 심어둔다. 출력 시점(아래 for i, rec in enumerate(recs) 루프)에서
        # INV_SIZECODE_COL(사이즈구분) 바로 다음 14칸 중 이 idx에 해당하는 칸만 채색 처리.
        # 260816 3차 개정(중태님 지시): 지금까지는 "남는 사이즈"를 원인 구분 없이 전부 노란색
        # 하나로 칠했는데, 실제로는 원인이 둘로 나뉜다 — (a) 규칙(1): 애초에 매칭되는 반대쪽
        # 사이즈가 하나도 없어서 남는 경우(top_nomatch/bot_nomatch), (b) 규칙(2)/(3): 매칭은 됐지만
        # 1.3배 과다재고 규칙 때문에 추가로 남는 경우(excess_bot/excess_top). 이 둘은 서로 배타적
        # (excess_*는 matched에 들어간 사이즈에만 붙고, *_nomatch는 애초에 matched에 없는 사이즈라
        # 절대 안 겹침) — (a)는 초록색, (b)는 기존대로 노란색으로 나눠서, "진짜 매칭 상대가 없는
        # 사이즈"와 "재고 비율 규칙 때문에 단품 전환된 사이즈"를 결과물에서 한눈에 구분할 수 있게 한다.
        top_yellow_idx = {T_IDX[s] for s in excess_top}
        bot_yellow_idx = {B_IDX[s] for s in excess_bot}
        top_green_idx = {T_IDX[s] for s in top_nomatch}
        bot_green_idx = {B_IDX[s] for s in bot_nomatch}
        for r in tops:
            r["_yellow_idx"] = top_yellow_idx
            r["_green_idx"] = top_green_idx
        for r in bots:
            r["_yellow_idx"] = bot_yellow_idx
            r["_green_idx"] = bot_green_idx
        if matched:
            # 260807: 표기 문구 축약 (판정 로직은 그대로) — 세트만→SET만 / 세트&상하단품→SET+상하 /
            #         세트&상의단품→SET+상 / 세트&하의단품→SET+하
            # 260815 추가(중태님 지시, 수정1): "SET+상"/"SET+하"가 뜨는 이유가 실무에서 2가지로
            # 섞여 혼동됨 — (a) 애초에 매칭 후보 자체가 없어 남는 사이즈(규칙1)와 (b) 매칭은 됐지만
            # 과다재고라 추가로 남는 사이즈(규칙2·3)가 같은 라벨로 나갔음. 그 사이드(상/하)의
            # 잔여가 "전부" 배율 규칙 때문(=규칙1 잔여가 0개)일 때만 "(N배)" 태그를 붙여 구분한다.
            # 진짜 미매칭 사이즈가 하나라도 섞여 있으면(원인이 순수 배율 규칙이 아니면) 태그를 붙이지 않는다.
            # 260816 2차 개정: 규칙(2)·규칙(3) 둘 다 합산기준·1.3배로 통일됐으므로 태그도 다시
            # "(1.3배)"로 공용 표기(직전엔 규칙(3)이 1.5였던 짧은 기간 동안만 "(1.5배)"를 따로 썼음 —
            # 상수 _INV_SET_EXCESS_RATIO/_INV_SET_EXCESS_RATIO_BOT가 갈라지면 언제든 다시 나눌 수
            # 있도록 태그 문자열은 상수값을 그대로 참조).
            top_all_matched = not (set(top_ok) - set(matched))   # 상의 OK 사이즈가 전부 매칭됐는가
            bot_all_used = not (set(bot_ok) - used)               # 하의 OK 사이즈가 전부 매칭에 쓰였는가
            top_tag = f"({_INV_SET_EXCESS_RATIO}배)" if tl and top_all_matched else ""
            bot_tag = f"({_INV_SET_EXCESS_RATIO_BOT}배)" if bl and bot_all_used else ""
            stt = "SET만" if not tl and not bl else \
                (f"SET+상{top_tag}하{bot_tag}" if tl and bl
                 else (f"SET+상{top_tag}" if tl else f"SET+하{bot_tag}"))
            sg = _inv_set_grade_display(_inv_set_grade(matched, Y,
                                {s: ti["size14"][T_IDX[s]] for s in matched},
                                S_CORE, S_SMALL, S_BIG))
        else:
            # 260815(중태님 지시): SET 가능여부(AD)에서 "품절근처"라는 출력 조건을 완전히 제거.
            # 상·하 양쪽 다 OK 사이즈가 하나도 없는 경우(구 "품절근처") 대신, 실제 조건을 그대로
            # 드러내는 문구 "상하모두 {X}개 이하"로 표기(중태님 지시 — "사실 그대로 표현"). X는
            # 이번 실행에 쓰인 OK 문턱값 그대로 대입(시즌마다 값이 바뀌므로 하드코딩하지 않음).
            # 나머지 3라벨(상·하단품/상의단품/하의단품)은 260807 축약 표기 그대로 유지.
            # 260814: 이 분기는 상/하 데이터가 다 있고 지원 사이즈조합인데도(=진짜 SJ/EJ/SL/EP 세트
            # 후보) 사이즈가 하나도 안 맞아 세트 자체가 안 만들어진 경우라 SET등급은 "단품아이템"이
            # 아니라 "SET구성실패"(세트를 시도했으나 실패)로 표기한다.
            if tl and bl:
                stt = "상·하단품"
            elif tl:
                stt = "상의단품"
            elif bl:
                stt = "하의단품"
            else:
                stt = f"상하모두 {X}개 이하"
            sg = "SET구성실패"
        for r in g:
            r["AD"] = stt; r["AE"] = sg

    # ── 출력: 템플릿(v3 계열, 최종 122열로 자동 보정) 복제 후 데이터 교체 (상단 메타 6행 + 데이터 9행~) ──
    if not os.path.exists(template_path):
        raise ValueError("서식 템플릿(inventory_template.xlsx)이 저장소에 없어요 — "
                         "최신 v3.3 결과물을 inventory_template.xlsx로 GitHub에 올려주세요.")
    twb = openpyxl.load_workbook(template_path)
    tws = twb.active

    def _upgrade_insert(col, amount, style_src_col):
        """템플릿에 col 위치부터 amount개 컬럼을 삽입하고, 병합범위를 이동 재병합한 뒤
        style_src_col의 서식(폰트·테두리·정렬·표시형식·채우기·열너비)을 새 컬럼들에 복제한다.
        openpyxl의 insert_cols는 병합 셀 범위를 자동으로 밀어주지 않으므로(문서상 명시된 한계),
        삽입 지점 이상에 걸린 병합을 먼저 해제한 뒤 삽입하고, 같은 범위를 +amount칸 이동해 재병합한다.
        """
        affected = [m for m in list(tws.merged_cells.ranges) if m.max_col >= col]
        saved = [(m.min_row, m.max_row, m.min_col, m.max_col) for m in affected]
        for m in affected:
            tws.unmerge_cells(str(m))
        tws.insert_cols(col, amount=amount)
        for r1, r2, c1, c2 in saved:
            nc1 = c1 + amount if c1 >= col else c1
            nc2 = c2 + amount if c2 >= col else c2
            tws.merge_cells(start_row=r1, start_column=nc1, end_row=r2, end_column=nc2)
        for row in range(1, tws.max_row + 1):
            src = tws.cell(row, style_src_col)
            for k in range(amount):
                dst = tws.cell(row, col + k)
                dst.font = copy(src.font); dst.border = copy(src.border)
                dst.alignment = copy(src.alignment); dst.number_format = src.number_format
                dst.fill = copy(src.fill)
        src_w = tws.column_dimensions[get_column_letter(style_src_col)].width
        if src_w:
            for k in range(amount):
                tws.column_dimensions[get_column_letter(col + k)].width = src_w

    if tws.max_column == 106:
        # 260811: 구 106열 템플릿(사이즈구분 컬럼 도입 전) 자동 보정 — 신규 컬럼 1개 삽입(구 92번째,
        # 오른쪽 이웃이던 구 사이즈95 칸 서식을 복제). 106→107이 된 뒤 아래 107 분기로 이어서 처리된다.
        _upgrade_insert(92, 1, 93)
    if tws.max_column == 107:
        # 260811(가격시뮬 2차): 몰가격(22) 바로 뒤에 기준판매가 복제 컬럼 1개 먼저 삽입 — 107→108.
        # 서식은 몰가격 칸(왼쪽 이웃) 것을 그대로 복제. 108→113은 아래 108 분기로 이어서 처리된다.
        _upgrade_insert(INV_GIJUN_COPY_COL, 1, INV_GIJUN_COPY_COL - 1)
    if tws.max_column == 108:
        # 260811(가격시뮬 1차): 기준판매가 복제 컬럼(23) 바로 뒤에 신규 가격 5컬럼 삽입 — 108→113.
        _upgrade_insert(INV_PRICE_SIM_COL, INV_PRICE_SIM_N, INV_PRICE_SIM_COL - 1)
    if tws.max_column == 112:
        # 260811(가격시뮬 2차): 이미 가격5컬럼(구 23~27)까지만 반영된 112열 템플릿(직전 배포본)이면,
        # 몰가격 바로 뒤에 기준판매가 복제 컬럼 1개만 추가로 삽입 — 기존 가격5컬럼은 24~28로 밀림.
        _upgrade_insert(INV_GIJUN_COPY_COL, 1, INV_GIJUN_COPY_COL - 1)
    if tws.max_column == 113:
        # 260826(SET가격): 단품가격 블록 끝(28) 바로 뒤에 SET 가격 9컬럼 삽입 — 113→122.
        # 서식은 단품 최초가 칸(20번째)을 복제(헤더 회색·데이터 무채색·#,##0) — 초록(31,32)·
        # 노랑(33~37) 강제 지정은 아래 _INV_GREEN_COLS/_INV_YELLOW_COLS 루프가 처리한다.
        _upgrade_insert(INV_SET_PRICE_COL, INV_SET_PRICE_N, 20)
    if tws.max_column == 122:
        # 260826-2(물량등급): 최초출고일(40) 바로 뒤에 물량등급 1컬럼 삽입 — 122→123.
        # 서식은 삽입 후 오른쪽 이웃이 되는 기간판매수량분석 칸(구 41)을 복제(노란색·General).
        _upgrade_insert(INV_VOL_GRADE_COL, 1, INV_VOL_GRADE_COL + 1)
    if tws.max_column != INV_TOTAL_COLS:
        raise ValueError(f"템플릿 열 수 {tws.max_column} ≠ {INV_TOTAL_COLS}(또는 구버전 106·107·108·112·113·122) — "
                         f"v3 계열 결과물을 템플릿으로 지정하세요.")
    names_row = next((r for r in range(1, 12) if tws.cell(r, 10).value == "품번"), None)
    if names_row is None:
        raise ValueError("템플릿에서 컬럼명 행(품번)을 찾지 못했어요.")
    NAMES_R = 8; GROUP_R = 7; DATA_R = 9
    shift = NAMES_R - names_row
    if shift > 0:  # 구 레이아웃(4행 헤더) → 신 레이아웃 변환
        gr = names_row - 1
        merges = [(m.min_col, m.max_col) for m in list(tws.merged_cells.ranges) if m.min_row == gr]
        for m in list(tws.merged_cells.ranges):
            tws.unmerge_cells(str(m))
        tws.insert_rows(1, shift)
        for c1, c2 in merges:
            tws.merge_cells(start_row=GROUP_R, start_column=c1, end_row=GROUP_R, end_column=c2)
    # 시즌·년도 비교 대상군(묶음) 요약 — 등장 순서 기준으로 중복 없이 나열 (예: "Z+A+C+D · B")
    def _grp_summary(field, grp_field):
        _code_to_grp = {}
        for rec in recs:
            _code_to_grp.setdefault(rec[field], rec[grp_field])
        _seen_lbl, _grp_desc = set(), []
        for _code in _code_to_grp:
            _lbl = _code_to_grp[_code]
            if _lbl not in _seen_lbl:
                _seen_lbl.add(_lbl); _grp_desc.append(_lbl)
        return (" · ".join(_grp_desc) if _grp_desc else "–"), _code_to_grp

    season_group_summary, _season_code_to_grp = _grp_summary("season", "season_grp")
    year_group_summary, _year_code_to_grp = _grp_summary("year", "year_grp")

    # 상단 메타 6행 기입 (5행 + 시즌·년도 비교대상군 요약 1행)
    meta = [
        f"기간판매 기준: {period}",
        f"변수 X= {X}장 (사이즈 OK 기준)",
        f"변수 Y= {Y}장 (동일등급내에서 추가로 등급 나눌때 기준 수량)",
        f"작업일: {workdate}",
        f"가공기준: 260731 확정판 + 260811 사이즈구분 로우컬럼 직접반영 + 가격 시뮬레이션 5컬럼 "
        f"+ 260826 SET 가격 9컬럼 + 물량등급(A≥{vol_a}/B≥{vol_b}/C≥{vol_c}/D≤{vol_c - 1}장 · 온라인창고 재고) "
        f"(5등급 A~E · 재고20미만 · 오프라인 제외 · 모집단 {cat_level}×년도×시즌)",
        f"비교 대상군(등급 모집단 묶음) — 시즌: {season_group_summary} / 년도: {year_group_summary}",
    ]
    mf = Font(name="맑은 고딕", size=11, bold=True)
    for i, t in enumerate(meta, 1):
        cell = tws.cell(i, 1); cell.value = t; cell.font = mf
    for r in range(1, 7):
        if r in tws.row_dimensions:
            del tws.row_dimensions[r]
    tws.row_dimensions[GROUP_R].height = 24
    tws.row_dimensions[NAMES_R].height = 52.2
    # 260802 확정: 헤더 전체 채우기 강제 지정 (템플릿 상태와 무관하게 항상 적용)
    for c in _INV_YELLOW_COLS:
        tws.cell(NAMES_R, c).fill = fill_yellow
    for c in _INV_PINK_COLS:                                    # 260815(헤더 개편)
        tws.cell(NAMES_R, c).fill = fill_pink
    for c in _INV_GRAY_COLS:                                    # 260815(헤더 개편)
        tws.cell(NAMES_R, c).fill = fill_gray
    for c in _INV_GREEN_COLS:
        tws.cell(NAMES_R, c).fill = fill_green
    # 260803: C열 헤더는 cat_level 선택(중카테고리/소카테고리/아이템코드)에 맞춰 텍스트도 같이 바뀐다.
    tws.cell(NAMES_R, 3).value = cat_level
    # 260811: 신규 '사이즈구분' 컬럼 헤더명 기입 (그룹행은 비워둠 — 사이즈정보 그룹과는 별개의
    # 단일 컬럼, J·Z·AJ열과 같은 성격). 로우파일 값을 그대로 통과시키는 열이라 노란색 대상 아님.
    tws.cell(NAMES_R, INV_SIZECODE_COL).value = "사이즈구분"
    # 260811(가격시뮬): 신규 가격 5컬럼 헤더명 기입 (그룹행은 비워둠 — 사이즈구분과 동일 방식).
    for _k, _h in enumerate(INV_PRICE_SIM_HEADERS):
        tws.cell(NAMES_R, INV_PRICE_SIM_COL + _k).value = _h
    # 260826(SET가격): SET 가격 9컬럼 헤더명 기입 (그룹행 "SET 가격" 병합은 아래 그룹헤더 재구성부에서).
    for _k, _h in enumerate(INV_SET_PRICE_HEADERS):
        tws.cell(NAMES_R, INV_SET_PRICE_COL + _k).value = _h
    # 260811 추가 개정(2): 몰가격 바로 옆에 '기준판매가' 복제 컬럼 헤더명 기입 (그룹행은 비워둠 —
    # 사이즈구분과 동일 방식). 원본 기준판매가 컬럼(뒤쪽 패스스루 구간)은 헤더/값 모두 그대로 유지된다.
    tws.cell(NAMES_R, INV_GIJUN_COPY_COL).value = "기준판매가"

    # 260811 추가 개정(3): '수정일'을 로우데이터 순서대로 '이관구분' 바로 왼쪽(41번째 칸)으로 옮기며
    # AA~변경후할인율 블록이 한 칸씩 앞당겨졌다 — 템플릿에 박혀있던 32~42번째 칸 헤더명(NAMES_R)을
    # 새 순서에 맞게 코드로 다시 써준다(안 그러면 헤더 텍스트와 실제 값이 어긋난다).
    # 260815(헤더 개편, 중태님 확정): 컬럼명 3개 변경(사이즈 등급→단품 사이즈 컨디션 / SET 상태 구분→
    # SET 가능여부 / SET 등급→SET 사이즈 컨디션) + AI제안방향을 사이즈/세트 3컬럼 앞(34번째)으로 이동.
    # 260826-2(물량등급): 블록 맨 앞에 신규 "물량등급" 헤더 추가(41번째부터 기입).
    _INV_COL32_42_HEADERS = ["물량등급",
                             "기간판매수량분석", "소진예상기간분석", "AI제안방향",
                             "단품\n사이즈 컨디션", "SET\n가능여부", "SET\n사이즈 컨디션",
                             "휴먼의사결정", "변동가격", "변경후할인율",
                             "수정일", "이관구분"]
    # 260826(SET가격): 이 블록의 시작 위치가 32 → 41로 +9 밀림.
    for _k, _h in enumerate(_INV_COL32_42_HEADERS):
        tws.cell(NAMES_R, INV_VOL_GRADE_COL + _k).value = _h

    # 260815(헤더 개편): 위 재배치에 맞춰 GROUP_R(7행) 상위 그룹 병합 범위도 조정.
    #   · '기본사항' 그룹 14~31 (변경 없음)
    #   · 구 '분석·의사결정' 단일 그룹(32~40)을 3개 그룹으로 분할:
    #     - 32~34 "판매진도에 따른 가격 변화 결정" (노랑)
    #     - 35~37 "사이즈가 정상 인지? 세트가 되는지?" (연분홍)
    #     - 38~40 "사람이 최종 의사 결정" (노랑)
    #   · 41번째 칸(수정일)은 사이즈구분·가격시뮬 컬럼과 동일하게 그룹헤더 없는 단일 컬럼으로 둔다.
    def _inv_unmerge_group(old_min, old_max):
        """GROUP_R 행에서 (old_min~old_max) 병합 범위를 찾아 해제하고 그 라벨을 지운다.
        해제 후에는 old_min~old_max 구간의 모든 칸이 다시 독립 셀로 돌아와 재병합 가능해진다."""
        for m in list(tws.merged_cells.ranges):
            if m.min_row <= GROUP_R <= m.max_row and m.min_col == old_min and m.max_col == old_max:
                tws.unmerge_cells(start_row=m.min_row, start_column=m.min_col,
                                   end_row=m.max_row, end_column=m.max_col)
                tws.cell(GROUP_R, old_min).value = None
                break

    def _inv_set_group(new_min, new_max, label):
        """이미 독립 셀 상태인 new_min~new_max 구간에 새 그룹 라벨을 쓰고 병합한다."""
        tws.cell(GROUP_R, new_min).value = label
        if new_max > new_min:
            tws.merge_cells(start_row=GROUP_R, start_column=new_min, end_row=GROUP_R, end_column=new_max)

    # 260826(SET가격)·260826-2(물량등급): 그룹헤더 재구성 — 어떤 세대 템플릿이 들어와도(업그레이드
    # 직후의 광폭 병합, 구 레이아웃의 14~32/33~41 병합, 신 123열 결과물의 완성 병합 모두)
    # 14~51 구간의 GROUP_R 병합을 전부 해제·초기화한 뒤 새 레이아웃으로 다시 병합한다:
    #   14~19 기본사항 · 20~28 단품가격(검정 바탕) · 29~37 SET 가격(진파랑 바탕) ·
    #   38~40 그룹없음(할인율·최초입고일·최초출고일) · 41~44 판매진도(물량등급 포함, 260826-2로
    #   3→4칸 확장) / 45~47 / 48~50 (기존 3그룹) · 51 수정일(그룹없음)
    for m in list(tws.merged_cells.ranges):
        if m.min_row <= GROUP_R <= m.max_row and m.max_col >= 14 and m.min_col <= 51:
            tws.unmerge_cells(str(m))
    for _gc in range(14, 52):
        tws.cell(GROUP_R, _gc).value = None
    _inv_set_group(14, 19, "기본사항")
    _inv_set_group(20, 28, "단품가격")
    _inv_set_group(29, 37, "SET 가격")
    _inv_set_group(41, 44, "판매진도에 따른 가격 변화 결정")
    _inv_set_group(45, 47, "사이즈가 정상 인지? 세트가 되는지?")
    _inv_set_group(48, 50, "사람이 최종 의사 결정")
    for _gc in (14, 20, 29, 41, 45, 48):
        tws.cell(GROUP_R, _gc).alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    # 260803 확정: 위 초록(GREEN) 컬럼들의 상위 그룹 헤더(GROUP_R, 병합 셀)도 같은 초록으로.
    # 병합 셀은 좌상단 셀에만 실제로 서식이 저장되므로, 열이 속한 병합범위의 좌상단 열을 찾아 그 칸에만 칠한다
    # (병합이 없으면 그 열 자체가 좌상단이므로 그대로 칠해짐).
    def _group_header_fill_col(col):
        for m in tws.merged_cells.ranges:
            if m.min_row <= GROUP_R <= m.max_row and m.min_col <= col <= m.max_col:
                return m.min_col
        return col
    for c in _INV_GREEN_COLS:
        tws.cell(GROUP_R, _group_header_fill_col(c)).fill = fill_green
    # 260815(헤더 개편): 새로 분할된 3그룹의 그룹헤더도 각각 노랑·연분홍·노랑으로 강제 지정
    # (템플릿 상속에 의존하지 않고 GREEN과 동일하게 항상 명시적으로 칠한다).
    tws.cell(GROUP_R, 41).fill = fill_yellow
    tws.cell(GROUP_R, 45).fill = fill_pink
    tws.cell(GROUP_R, 48).fill = fill_yellow
    # 260826(SET가격): 새 그룹헤더 강제 지정 — 기본사항은 초록(구 레이아웃에서 초록이던 톤 유지),
    # 단품가격은 검정 바탕·흰 글씨, SET 가격은 진파랑 바탕·흰 글씨(중태님 수기 샘플 파일과 동일).
    tws.cell(GROUP_R, 14).fill = fill_green
    _fill_black = PatternFill(start_color="FF000000", end_color="FF000000", fill_type="solid")
    _fill_navy = PatternFill(start_color="FF2F5597", end_color="FF2F5597", fill_type="solid")
    _font_white = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFFFF")
    tws.cell(GROUP_R, 20).fill = _fill_black; tws.cell(GROUP_R, 20).font = _font_white
    tws.cell(GROUP_R, 29).fill = _fill_navy; tws.cell(GROUP_R, 29).font = _font_white
    dstyle = {c: (copy(tws.cell(DATA_R, c).font), copy(tws.cell(DATA_R, c).border),
                  copy(tws.cell(DATA_R, c).alignment), tws.cell(DATA_R, c).number_format,
                  copy(tws.cell(DATA_R, c).fill)) for c in range(1, INV_TOTAL_COLS + 1)}
    if tws.max_row >= DATA_R:
        tws.delete_rows(DATA_R, tws.max_row - DATA_R + 1)

    # 260811(가격시뮬 + 기준판매가 복제): 구 range(19,24)·range(37,...) → 가격5컬럼(24~28) +
    # 기준판매가 복제(23) 포함해 +6 밀림. 260826(SET가격): SET 가격 9컬럼(29~37)도 숫자 —
    # 가격 구간이 19~37로 늘고 할인율(38)까지 숫자. 260826-2(물량등급): 물량등급(41)은 문자 —
    # 패스스루 숫자 구간이 +1 더 밀려 53부터.
    NUM_COLS = set(range(19, 39)) | set(range(53, INV_TOTAL_COLS + 1))
    # '변동가격' 칸(49번째, = AW열) 참조. 260815 헤더 개편은 순서 재배치뿐이라 휴먼의사결정·변동가격·
    # 변경후할인율 상대 위치는 불변 — 계산식 대신 고정값으로 명시한다. (260826 SET가격 +9 ·
    # 260826-2 물량등급 +1로 39 → 49)
    _INV_AH_COL_LETTER = get_column_letter(49)

    def to_num(v):
        if v is None or v == "":
            return None
        s = str(v).strip().replace(",", "")
        if re.fullmatch(r"-?\d+", s):
            return int(s)
        if re.fullmatch(r"-?\d*\.\d+", s):
            return float(s)
        return v

    for i, rec in enumerate(recs):
        rr = i + DATA_R; raw = rec["raw"]; vals = {}
        vals[1], vals[2], vals[3] = raw[13], raw[14], rec["cat_level_val"]
        for j, rc in enumerate([15, 16, 17, 18, 19, 20]):
            vals[4 + j] = raw[rc]
        vals[10], vals[11] = rec["pn"], rec["K"]
        vals[12] = rec["L"] or None; vals[13] = rec["M"] or None
        for j in range(9):
            vals[14 + j] = raw[1 + j]                          # 14~22 = raw1~9 (색상..몰가격)
        # 260811 추가 개정(2): 몰가격(22) 바로 뒤에 '기준판매가' 복제 컬럼(23) 1칸 삽입 — 원본
        # 기준판매가 컬럼은 뒤쪽 패스스루 구간에 그대로 남아 있고, 이건 순수 참고용 복제일 뿐이다.
        _mol = _inv_num(raw[INV_RAW_MOLGA_COL])
        _gijun = _inv_num(raw[INV_RAW_GIJUN_COL])
        _chojo = _inv_num(raw[INV_RAW_CHOJOGA_COL])   # 260814: 기준판매가 공란 시 대체 캡핑 기준
        vals[INV_GIJUN_COPY_COL] = raw[INV_RAW_GIJUN_COL]
        # 260811(가격시뮬): 그 다음에 신규 가격 5컬럼(24~28) 삽입 — 이후 컬럼은 전부 +6(가격시뮬5 + 기준판매가복제1).
        for _k, _v in enumerate(_inv_price_sim(_mol, _gijun, _chojo)):
            vals[INV_PRICE_SIM_COL + _k] = _v
        # 260826(SET가격): SET 가격 9컬럼(29~37) — 위 SET 판정 루프에서 상+하 합산해 심어둔 값.
        # 단품아이템·짝 없는 세트상품 행은 rec에 SETP가 없어 9칸 전부 공란.
        for _k, _v in enumerate(rec.get("SETP") or [None] * INV_SET_PRICE_N):
            vals[INV_SET_PRICE_COL + _k] = _v
        vals[38], vals[39], vals[40] = raw[10], raw[11], raw[12]   # 할인율·최초입고일·최초출고일(구29~31 → +9)
        # 260811: '수정일'을 이 자리(구26)에서 빼서 로우데이터 원래 순서대로 '이관구분' 바로 왼쪽
        # (41번째 칸)으로 옮긴다 — 아래 AA~변경후할인율 블록이 그만큼 한 칸씩 앞으로 당겨진다.
        # 260815(헤더 개편): AI제안방향(AF)을 34번째로 앞당기고, 단품 사이즈 컨디션·SET 가능여부·
        # SET 사이즈 컨디션(AC·AD·AE)을 35~37로 밀어 하나로 묶는다.
        vals[INV_VOL_GRADE_COL] = rec["VOL"]                        # 260826-2: 물량등급(41)
        vals[42], vals[43], vals[44] = rec["AA"], rec["AB"], rec["AF"]
        vals[45], vals[46], vals[47] = rec["AC"], rec["AD"], rec["AE"]
        vals[48] = vals[49] = None
        vals[50] = f"={_INV_AH_COL_LETTER}{rr}/T{rr}"                # 변경후할인율 = 변동가격(AW)÷최초가(T)
        vals[51] = raw[21]                                          # 수정일(이관구분 바로 왼쪽)
        vals[52] = raw[22]                                          # 이관구분
        # 260811: 패스스루 구간이 raw24~93(70열) → raw24~94(71열)로 1열 확장. 신규 '사이즈구분'이
        # raw79 자리에 자연스럽게 끼어 있어 이 구간 안에서 함께 넘어간다(출력 108번째 칸에 그대로 안착).
        # 260826(SET가격) +9 · 260826-2(물량등급) +1: 시작 위치 43 → 53.
        for j in range(71):
            vals[53 + j] = raw[23 + j]
        for c in range(1, INV_TOTAL_COLS + 1):
            v = vals.get(c)
            if v == "":
                v = None
            if c in NUM_COLS:
                v = to_num(v)
            nc = tws.cell(rr, c, v)
            f, b, al, nf, fl = dstyle[c]
            nc.font = f; nc.border = b; nc.alignment = al; nc.number_format = nf
            if c in _INV_YELLOW_COLS:
                nc.fill = fill_yellow
            elif c in _INV_GREEN_COLS:
                nc.fill = fill_green
            elif c in _INV_PINK_COLS:                               # 260815(헤더 개편)
                nc.fill = fill_pink
            elif c in _INV_GRAY_COLS:                                # 260815(헤더 개편)
                nc.fill = fill_gray
            elif (INV_SIZECODE_COL < c <= INV_SIZECODE_COL + 14
                  and (c - INV_SIZECODE_COL) in rec.get("_yellow_idx", ())):
                # 260815 추가(수정2): 사이즈정보 14칸 중 "남는 상의/남는 하의"에 해당하는 사이즈 칸 —
                # 260816 3차 개정으로 이제 이 노란색은 "매칭은 됐지만 1.3배 과다재고 규칙 때문에
                # 남은" 경우(규칙2·3)만 의미한다. 애초에 매칭 후보가 없어 남은 경우는 아래 초록색 분기.
                nc.fill = fill_yellow
            elif (INV_SIZECODE_COL < c <= INV_SIZECODE_COL + 14
                  and (c - INV_SIZECODE_COL) in rec.get("_green_idx", ())):
                # 260816 신설(중태님 지시, 수정3): 애초에 매칭되는 반대쪽 사이즈가 하나도 없어서
                # 단품 판매가 추천된 사이즈(규칙1) — 재고비율 규칙(노란색)과 시각적으로 구분.
                nc.fill = fill_green
            else:
                nc.fill = fl
        tws.row_dimensions[rr].height = 20.25

    tws.auto_filter.ref = f"A{NAMES_R}:{get_column_letter(INV_TOTAL_COLS)}{NAMES_R + len(recs)}"
    tws.freeze_panes = f"K{DATA_R}"
    # 260802 확정: BK~CM 항상 숨김. 260815(헤더 개편) 보강: 그 외 열은 템플릿에 사람이 실수로
    # 숨겨둔 상태가 남아있어도 매번 명시적으로 숨김 해제한다(안 그러면 템플릿의 숨김 상태가 결과물에
    # 그대로 전염된다).
    for c in range(1, INV_TOTAL_COLS + 1):
        tws.column_dimensions[get_column_letter(c)].hidden = (c in _INV_HIDE_COLS)
    buf = io.BytesIO()
    twb.save(buf)

    def dist(k):
        return dict(Counter(r[k] for r in recs))
    unmatched = [r["pn"] for r in recs if r["scode"] is None]
    report = {
        "rows": len(recs), "skipped": skipped, "X": X, "Y": Y,
        "AA": dist("AA"), "AB": dist("AB"), "AF": dist("AF"), "AC": dist("AC"),
        "VOL": dist("VOL"), "vol_thr": (vol_a, vol_b, vol_c),      # 260826-2 물량등급 분포·기준
        "cat_values_n": len(cat_values),                            # 260826-3 모집단 카테고리 값 종수
        "cat_small_fallback": cat_small_fallback,                   # 260826-3 소카테고리 공란 폴백 코드

        "SET": dict(Counter(r["AD"] for r in recs if r["L"])),
        "pairs": pairs, "nopair": nopair, "af_bad": af_bad,
        "unmatched": unmatched,
        "small_groups": {" × ".join(map(str, k)): len(v) for k, v in groups.items() if len(v) <= 4},
        "season_group_summary": season_group_summary,
        "season_group_map": dict(_season_code_to_grp),
        "year_group_summary": year_group_summary,
        "year_group_map": dict(_year_code_to_grp),
        "fallback_items": fallback_items,
        # 260806: 매핑 없는 아이템 코드 — 중단 대신 리포트로 알린다(마스터 보강 신호)
        "unmapped_items": unk_item,
        "unmapped_rows": unk_rows,
        "unmapped_pns": sorted({r["pn"] for r in recs if r["unmapped"]}),
        "cat_level": cat_level,
        # 260811: 사이즈코드 소스가 '사이즈구분' 로우컬럼으로 바뀐 데 따른 참고 리포트.
        "unknown_size_codes": unk_size_codes,
        "scode_blank": scode_blank,
        "scode_blank_had_master": scode_blank_had_master,
        "scode_mismatch": scode_mismatch,
        "has_size_master": bool(master),
        # 260811: 로우데이터 맨 끝에 빈 95번째 열이 딸려와서 94열로 취급하고 무시했는지 여부.
        "tail_col_ignored": _tail_col_ignored,
    }
    return buf.getvalue(), report


def _inv_group_ui(codes, key_prefix):
    """codes(문자열 리스트) 각각에 그룹 셀렉트박스를 한 줄로 렌더링하고, 선택 결과로부터
    {코드: 라벨(같은 그룹끼리 "+"로 합친 문자열)} 맵과, 2개 이상 묶인 그룹 목록(list of code-list)을 반환한다.
    기본 선택값은 코드마다 자기 자신의 순번 그룹 — 즉 아무것도 안 바꾸면 전부 별도(묶지 않음)가 된다.
    """
    from collections import defaultdict as _defaultdict
    n = len(codes)
    opts = [f"그룹{i + 1}" for i in range(n)]
    cols = st.columns(n)
    assign = {}
    for i, code in enumerate(codes):
        assign[code] = cols[i].selectbox(code, opts, index=i, key=f"{key_prefix}_{code}")
    by_val = _defaultdict(list)
    for code, val in assign.items():
        by_val[val].append(code)
    group_map = {}
    for val, codes_list in by_val.items():
        codes_sorted = [c for c in codes if c in codes_list]
        lbl = "+".join(codes_sorted)
        for c in codes_sorted:
            group_map[c] = lbl
    merged = [g for g in by_val.values() if len(g) > 1]
    return group_map, merged


def render_inventory():
    """🏷️ 재고 가공 메뉴 — 로우데이터 업로드 → 가공 → v3.5 엑셀 다운로드 (전 팀원 사용 가능)."""
    # 260826-2(중태님 지시): 제목 옆에 붙던 "1차 (260731 확정 기준 · …)" 버전 꼬리표는 더 이상
    # 노출하지 않는다 — 상세 기준은 아래 캡션에만 남긴다.
    st.subheader("🏷️ 쇼핑몰 재고 가공")
    st.caption("재고모니터링 로우데이터(94열, '사이즈구분' 컬럼 포함)를 올리면 AA·AB 5등급, AF(AI제안방향), "
               "단품 사이즈 컨디션(AC), SET 가능여부·SET 사이즈 컨디션(AD·AE), 기준판매가 비교컬럼, 가격 시뮬레이션 5컬럼, "
               "SET 가격 9컬럼, 물량등급(온라인창고 재고 기준 물량A~물량D — 기준 숫자는 아래에서 직접 입력)을 부여한 "
               "123열 v3.5 엑셀을 만들어 드려요. "
               "재고 데이터는 DB에 저장하지 않아요(가공 → 다운로드만). "
               "260811부터 사이즈코드는 로우데이터의 '사이즈구분' 컬럼값을 그대로 사용해요(마스터 조회 안 함). "
               "몰가격 바로 뒤에 기준판매가를 그대로 복제한 컬럼이 초록색으로 1개 추가되고(원본 기준판매가 컬럼은 "
               "뒤쪽 그대로 유지, 비교하기 편하도록 옆에 나란히 표시), 그 뒤로 (네이버)상시가·(쿠폰진행)상시가·(쿠폰진행)행사가·"
               "(쿠폰X/무배)상시가·(쿠폰X/무배)행사가 5컬럼이 노란색으로 추가돼요"
               "(끝 3자리 기준 000→그대로·001~500→500·501~999→900 스냅, 기준판매가 넘으면 자동 캡핑 "
               "— 기준판매가가 공란이면 대신 최초가를 넘지 않도록 캡핑). "
               "단품가격 블록 바로 뒤 'SET 가격' 9컬럼(진파랑 그룹헤더)에는 SET품번으로 짝지어진 상의+하의의 "
               "최초가·현판가·몰가격·기준판매가·가격시뮬 5컬럼을 각각 합산한 세트 가격이 짝 양쪽 행에 "
               "동일하게 들어가요(단품아이템·짝이 없는 세트상품은 공란).")

    master = load_size_master()
    n_master = len(master)
    n_item_master = item_master_row_count()
    tpl_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), INV_TEMPLATE_FILE)
    c_info1, c_info2, c_info3 = st.columns(3)
    c_info1.caption(f"📏 사이즈 마스터(참고용, 판정엔 미사용): **{n_master:,}개** 품번"
                    if n_master else "📏 사이즈 마스터(참고용): 미등록 — 없어도 판정엔 지장 없어요")
    c_info2.caption("🧾 서식 템플릿: " + ("**동봉됨** (inventory_template.xlsx)" if os.path.exists(tpl_path)
                                          else "⚠️ **없음** — inventory_template.xlsx를 GitHub에 올려야 해요"))
    c_info3.caption(f"🗂️ 아이템 마스터: **{n_item_master:,}개** 코드"
                    + ("" if n_item_master else " — 미등록 시 구 기준(니트류·티셔츠류 분리)으로 자동 대체"))
    st.caption("ℹ️ **260811부터 사이즈 마스터는 판정에 쓰지 않아요.** 단품 사이즈 컨디션(AC)·SET 가능여부·SET 사이즈 컨디션(AD·AE)은 "
               "로우데이터에 함께 오는 '사이즈구분' 컬럼값을 그대로 사용해요. 사이즈 마스터를 등록해두면 "
               "새 컬럼값과 다른 건수만 참고로 알려드려요(결과에는 영향 없음).")

    # ── 가공 옵션 (X·Y는 변수 — 시즌 시점 따라 조정, 하드코딩 금지 원칙) ──
    o1, o2, o3, o4 = st.columns(4)
    X = o1.number_input("변수 X — 사이즈 OK 기준(장)", min_value=1, max_value=999, value=10, key="inv_x")
    # 260811 재확인: size-grade-classifier 스킬(classify.py/set_classify.py) 기본값이 Y=30이라
    # 스킬과 맞춤(과거 이 화면의 기본값 20은 기준 문서와 어긋나 있었음).
    Y = o2.number_input("변수 Y — 동일등급 내 세분 기준(장)", min_value=1, max_value=999, value=30, key="inv_y")
    # 260811: 같은 날 여러 번 다운로드해도 파일이 안 겹치도록 기본값에 시간(시:분)까지 포함.
    # 사용자가 직접 지우고 다시 입력할 수도 있는 텍스트칸이라 강제는 아님(수정하면 엑셀 안
    # "작업일" 표기·파일명 둘 다 그 값을 그대로 따라감).
    # 260811(2차 수정): datetime.now()는 배포 서버(UTC로 도는 경우)의 시각을 그대로 쓰기 때문에
    # 실제 한국 시간보다 9시간 늦게 표시되는 버그가 있었음 → now_kst()로 교체.
    workdate = o3.text_input("작업일", value=now_kst().strftime("%y.%m.%d %H:%M"), key="inv_workdate")
    period = o4.text_input("기간판매 조회 기준", placeholder="예: 26.07.20~26.07.31", key="inv_period")

    # ── 260826-2 물량등급 기준 (중태님 지시) — 온라인창고 재고 숫자로 A/B/C/D 구분.
    #    A/B/C 문턱(이상 기준)만 직접 입력하고, D는 C 기준이 정해지면 "C-1장 이하"로 자동 표기.
    v1, v2, v3, v4 = st.columns(4)
    vol_a = v1.number_input("물량등급 A — (장) 이상", min_value=1, max_value=999999, value=500,
                            key="inv_vol_a")
    vol_b = v2.number_input("물량등급 B — (장) 이상", min_value=1, max_value=999999, value=300,
                            key="inv_vol_b")
    vol_c = v3.number_input("물량등급 C — (장) 이상", min_value=1, max_value=999999, value=100,
                            key="inv_vol_c")
    # D 칸은 입력이 아니라 자동 표기 전용 — key를 주지 않아 C 값이 바뀌면 즉시 따라 바뀐다.
    v4.text_input("물량등급 D — 자동 지정", value=f"{int(vol_c) - 1}장 이하", disabled=True)
    _vol_ok = int(vol_a) > int(vol_b) > int(vol_c)
    if _vol_ok:
        st.caption(f"📦 물량등급 기준(온라인창고 재고): **A** {int(vol_a):,}장 이상 · "
                   f"**B** {int(vol_b):,}장 이상 · **C** {int(vol_c):,}장 이상 · "
                   f"**D** {int(vol_c) - 1:,}장 이하(자동) — 결과물엔 물량A~물량D로 표기돼요.")
    else:
        st.error("물량등급 기준은 A > B > C 순으로 커야 해요 — 숫자를 확인해 주세요.")

    # ── 카테고리 기준 설정 — AA·AB 등급 모집단·C열 표기에 쓸 카테고리 세분화 레벨을 고른다.
    #    기본=중카테고리(기존과 동일). C열 헤더 텍스트도 이 선택을 그대로 따라간다.
    st.markdown("##### 📁 카테고리 기준 설정 (기간판매등급/소진예상등급 · C열 표기)")
    st.caption("등급 모집단(중카테고리×년도×시즌)의 '카테고리' 축을 어느 세분화 레벨로 쓸지 선택해요. "
               "결과 엑셀 C열의 헤더·값도 이 선택을 그대로 따라가요. 기본값은 중카테고리(기존과 동일)예요.")
    cat_level = st.selectbox("카테고리 기준", INV_CAT_LEVELS, index=0, key="inv_catlevel")

    up = st.file_uploader("재고모니터링 로우데이터 업로드 (94열 엑셀 1개 · '사이즈구분' 컬럼 포함)",
                          type=["xlsx"], accept_multiple_files=False, key="inv_up")

    # ── 시즌·년도 비교 대상군(묶음) 설정 — AA·AB 등급 모집단(카테고리×년도×시즌)의 '시즌'·'년도' 축을
    #    이번 실행에서 어떻게 묶을지 정한다. 기본=코드마다 따로(원래 260731 확정판과 동일).
    #    업로드된 파일에서 실제로 쓰인 코드를 미리 훑어 그 코드로만 선택지를 만든다.
    season_group_map = None
    year_group_map = None
    if up is not None:
        try:
            seasons_found = _inv_peek_seasons(up)
        except Exception:
            seasons_found = []
        finally:
            try:
                up.seek(0)
            except Exception:
                pass
        try:
            years_found = _inv_peek_years(up)
        except Exception:
            years_found = []
        finally:
            try:
                up.seek(0)
            except Exception:
                pass

        if seasons_found:
            st.markdown("##### 🧩 시즌 비교 대상군 설정 (기간판매등급/소진예상등급)")
            st.caption("등급은 **중카테고리 × 년도 × 시즌**이 모두 같은 상품끼리만 비교해요(동기 그룹). "
                       "아래에서 이번 가공에 한해 시즌을 묶을 수 있어요 — **같은 그룹을 고른 시즌끼리 하나의 "
                       "비교 대상군**이 돼요. 기본값은 시즌마다 따로(묶지 않음)예요.")
            season_group_map, sz_merged = _inv_group_ui(seasons_found, "inv_szgrp")
            if sz_merged:
                merged_txt = " · ".join("+".join(c for c in seasons_found if c in g) for g in sz_merged)
                st.caption(f"✅ 지금 설정: **{merged_txt}** 묶음 적용 (나머지는 시즌별 개별 유지)")
            else:
                st.caption("현재 설정: 묶음 없음 — 시즌 각각 별도 비교 대상군(기존과 동일).")

        if years_found:
            st.markdown("##### 📅 년도 비교 대상군 설정 (기간판매등급/소진예상등급)")
            st.caption("등급은 **중카테고리 × 년도 × 시즌**이 모두 같은 상품끼리만 비교해요(동기 그룹). "
                       "아래에서 이번 가공에 한해 년도를 묶을 수 있어요 — **같은 그룹을 고른 년도끼리 하나의 "
                       "비교 대상군**이 돼요. 기본값은 년도마다 따로(묶지 않음)예요.")
            year_group_map, yr_merged = _inv_group_ui(years_found, "inv_yrgrp")
            if yr_merged:
                merged_txt_y = " · ".join("+".join(c for c in years_found if c in g) for g in yr_merged)
                st.caption(f"✅ 지금 설정: **{merged_txt_y}** 묶음 적용 (나머지는 년도별 개별 유지)")
            else:
                st.caption("현재 설정: 묶음 없음 — 년도 각각 별도 비교 대상군(기존과 동일).")

    if up is not None:
        if st.button("⚙️ 1차 가공 실행", type="primary", use_container_width=True, key="inv_run"):
            if not period.strip():
                st.error("'기간판매 조회 기준'을 입력해 주세요 (결과물 상단 메타에 들어가요).")
            elif not _vol_ok:
                st.error("물량등급 기준(A > B > C)을 먼저 맞춰 주세요.")
            else:
                try:
                    with st.spinner("가공 중… (등급 판정 → 세트 매칭 → 서식 적용)"):
                        xls, rep = process_inventory(up, master, tpl_path, int(X), int(Y),
                                                     period.strip(), workdate.strip(),
                                                     season_group_map=season_group_map,
                                                     year_group_map=year_group_map,
                                                     cat_level=cat_level,
                                                     vol_a=int(vol_a), vol_b=int(vol_b),
                                                     vol_c=int(vol_c))
                    st.session_state["inv_result"] = {
                        "bytes": xls, "report": rep,
                        "fname": f"재고가공_{_safe_name(workdate.strip() or 'result')}.xlsx"}
                except ValueError as ex:
                    st.session_state.pop("inv_result", None)
                    st.error(f"[중단] {ex}")
                except Exception as ex:
                    st.session_state.pop("inv_result", None)
                    st.error(f"가공 오류: {ex}")

    res = st.session_state.get("inv_result")
    if res:
        rep = res["report"]
        st.success(f"가공 완료 ✅ 상품 {rep['rows']:,}행 (합계행 {rep['skipped']}개 제외) · "
                   f"X={rep['X']} · Y={rep['Y']} · 세트그룹 {rep['pairs']}쌍 (짝없음 {rep['nopair']})")
        st.download_button("⬇ 가공 결과 엑셀 다운로드", res["bytes"], file_name=res["fname"],
                           mime=XLSX_MIME, type="primary", use_container_width=True, key="inv_dl")
        if rep.get("tail_col_ignored"):
            st.caption("ℹ️ 로우데이터 맨 끝에 빈 95번째 열이 딸려있어서 자동으로 무시하고 94열로 처리했어요"
                      "(엑셀 병합헤더 잔재로 보여요 — 값이 있었다면 에러로 중단됐을 거예요).")
        if rep.get("VOL"):
            _va, _vb, _vc = rep.get("vol_thr", ("?", "?", "?"))
            _vtxt = " · ".join(f"{k} {rep['VOL'].get(k, 0):,}행"
                               for k in ("물량A", "물량B", "물량C", "물량D"))
            st.caption(f"📦 물량등급 분포 — {_vtxt} (기준 A≥{_va}장 / B≥{_vb}장 / C≥{_vc}장 · 온라인창고 재고)")
        if rep.get("season_group_summary") or rep.get("year_group_summary") or rep.get("cat_level"):
            _catn = f" (모집단 카테고리 값 {rep['cat_values_n']}종)" if rep.get("cat_values_n") else ""
            st.caption(f"🧩 적용된 비교 대상군 — 카테고리 기준: **{rep.get('cat_level', '중카테고리')}**{_catn} · "
                       f"시즌: **{rep.get('season_group_summary', '–')}** · "
                       f"년도: **{rep.get('year_group_summary', '–')}**")
        # 260826-3(중태님 문의): 소카테고리를 골랐는데 마스터에 소카테고리가 비어 있는 코드는
        # 중카테고리 값으로 폴백돼 결과가 중카테고리 기준과 같아 보일 수 있다 — 그 사실을 명시.
        if rep.get("cat_small_fallback"):
            _fb = rep["cat_small_fallback"]
            st.warning(f"⚠️ 소카테고리 기준을 선택했지만, 아이템 마스터에 소카테고리가 비어 있어 "
                       f"**중카테고리 값으로 대체된 아이템코드 {len(_fb)}종**: {', '.join(_fb)} — "
                       f"이 코드들은 소카테고리를 골라도 중카테고리 기준과 같은 모집단으로 계산돼요. "
                       f"아이템 마스터의 '소카테고리' 칸을 채워 다시 업로드하면 반영됩니다.")
        if rep["af_bad"]:
            st.warning(f"⚠️ AF 매트릭스 미커버 {rep['af_bad']}건 — '검증필요'로 표시했어요. 규칙 점검이 필요해요.")
        if rep["unmatched"]:
            st.warning(f"⚠️ '사이즈구분' 공란(미매칭) {len(rep['unmatched'])}건 — AC는 '해당없음', AD·AE는 "
                       f"SET품번 유무에 따라 '단품아이템'/'SET구성실패'로 처리. "
                       f"예: {', '.join(rep['unmatched'][:10])}"
                       + (" …" if len(rep["unmatched"]) > 10 else ""))
            if rep.get("has_size_master") and rep.get("scode_blank_had_master"):
                st.caption(f"↳ 이 중 과거 사이즈 마스터엔 값이 있었던 건 **{rep['scode_blank_had_master']}건** "
                           f"— 로우데이터 생성 단계에서 '사이즈구분'을 못 채운 건일 수 있어요(신규 상품이라 "
                           f"마스터에도 없는 것과는 성격이 달라요).")
        if rep.get("unknown_size_codes"):
            st.warning(f"⚠️ '사이즈구분'에 정식 6종(A16/A17/A09/A05/A06/A18) 외의 값 "
                       f"{len(rep['unknown_size_codes'])}종 발견: **{', '.join(rep['unknown_size_codes'])}** "
                       f"— 오탈자이거나 신규 사이즈체계일 수 있어요. 해당 건은 AC '해당없음', AD·AE는 "
                       f"SET품번 유무에 따라 '단품아이템'/'SET구성실패'로 처리했어요.")
        if rep.get("has_size_master") and rep.get("scode_mismatch"):
            mm = rep["scode_mismatch"]
            st.info(f"ℹ️ '사이즈구분' 값이 등록된 사이즈 마스터와 다른 건 {len(mm)}건 — "
                    f"결과에는 항상 '사이즈구분' 값을 그대로 반영했어요(참고용). "
                    f"예: {', '.join(f'{pn}({new}≠{old})' for pn, new, old in mm[:5])}"
                    + (" …" if len(mm) > 5 else ""))
        if rep.get("unmapped_items"):
            st.warning(
                f"⚠️ 중카테고리 매핑 없는 아이템 코드 {len(rep['unmapped_items'])}종 "
                f"(**{', '.join(rep['unmapped_items'])}**) · 상품 {rep['unmapped_rows']:,}행 — "
                f"가공은 정상 완료했고, 이 상품들만 C열·AA·AB·AF를 '{_INV_UNMAPPED}'로 표기했어요. "
                f"단품 사이즈 컨디션(AC)·SET 판정은 로우데이터 '사이즈구분' 컬럼 기준이라 아이템 마스터와 무관하게 정상입니다. "
                f"아이템 마스터에 코드를 추가하고 다시 돌리면 정상 등급을 받아요.")
            with st.expander(f"🔎 '{_INV_UNMAPPED}' 처리된 품번 {len(rep.get('unmapped_pns', []))}건 보기"):
                st.dataframe(pd.DataFrame({"품번": rep.get("unmapped_pns", [])}),
                             use_container_width=True, height=240)
        if rep.get("fallback_items"):
            st.info(f"ℹ️ 아이템 마스터에 없어서 구 기준(폴백)으로 중카테고리를 처리한 아이템 코드: "
                   f"{', '.join(rep['fallback_items'])} — 실사용 코드라면 아이템 마스터에 추가해주세요.")
        with st.expander("📊 판정 분포 리포트 (AA · AB · AF · AC · SET)"):
            r1, r2 = st.columns(2)
            r1.markdown("**AA (기간판매 랭킹)**")
            r1.dataframe(pd.Series(rep["AA"], name="건수").rename_axis("등급"), use_container_width=True)
            r1.markdown("**AB (소진 속도)**")
            r1.dataframe(pd.Series(rep["AB"], name="건수").rename_axis("등급"), use_container_width=True)
            r1.markdown("**AF (AI제안방향)**")
            r1.dataframe(pd.Series(rep["AF"], name="건수").rename_axis("제안"), use_container_width=True)
            r2.markdown("**AC (단품 사이즈 컨디션)**")
            r2.dataframe(pd.Series(rep["AC"], name="건수").rename_axis("등급"), use_container_width=True)
            r2.markdown("**SET 가능여부 (세트키 보유 상품)**")
            if rep["SET"]:
                r2.dataframe(pd.Series(rep["SET"], name="건수").rename_axis("상태"), use_container_width=True)
            else:
                r2.caption("세트키(수트·셋업) 상품이 없어요.")
            if rep["small_groups"]:
                st.caption("AA 소형 모집단(≤4개): " +
                           " · ".join(f"{k} {v}개" for k, v in rep["small_groups"].items()))
    st.caption("※ 판정 규칙(260731 확정판 + 260811 사이즈구분 로우컬럼 반영): AA/AB 모집단=카테고리(선택한 "
               "기준)×년도×시즌(카테고리는 아이템 마스터 기준·소형그룹 예외 없음) · 선판정: 오프라인→'오프라인', "
               "온라인창고<20→'재고20미만'(AB 미적용) · AF 재고 분기 200 · AC/SET은 로우데이터 '사이즈구분' "
               "컬럼(A16 상의/A17 하의/A09 M-L-X/A05 신발/A06 FREE/A18 아동) 기준, 공란은 '해당없음' 처리.")


# ==============================================================================
# 추세분석  ─ 메뉴1: 주별 시즌상품 판매 변화 (2026-08-03 신설 · 중태님 지시)
# ==============================================================================
# "언제 붙기 시작하고, 언제 피크를 찍고, 언제 꺾이는가"를 주 단위로 보는 시즌 타이밍 분석 화면.
#  · 기준축(계열) 4종: 중카테고리 · 소카테고리 · 아이템코드 · 시즌(Z/A/B/C/D — 품번 5번째 자리)
#  · 지표 토글 4종: 실판매금액(백만) · 판매수량 · 판가율(%) · 주간 비중(%)
#  · 조회기간 최대 1년(53주) · 전년 동기간 점선 오버레이 On/Off(52주=364일 시프트 정렬)
#  · 공통 필터: 브랜드 · 시즌 · 중카테고리 · 소카테고리 (빈칸=전체)
#  · 시점 자동판정: 본격 상승 시점 / 피크 시점 / 피크아웃 시점 (피크 대비 기준선 % 조절)
#  · 룰11: 주별 데이터 표 + 시점 요약표 각각 ⬇엑셀 기본 제공
TREND_AXES = ["중카테고리", "소카테고리", "아이템코드", "시즌 (Z/A/B/C/D)"]
TREND_METRICS = ["실판매금액(백만)", "판매수량", "판가율(%)", "주간 비중(%)"]
TREND_MAX_WEEKS = 53                 # 조회 가능 최대 기간 = 1년(53주)
TREND_MAX_DAYS = TREND_MAX_WEEKS * 7
TREND_PREV_SHIFT_DAYS = 364          # 전년 동기간 정렬: 52주 시프트(요일 정렬 유지)
TREND_DEFAULT_TOPN = 7               # 처음 화면에 자동 선택되는 계열 수

# 차트 글자 크기 프리셋 (2026-08-03 중태님 요청: 가로축 날짜·범례가 너무 작음 → 기본 '크게').
#  x=가로축 날짜 · y=세로축 눈금 · lg=상단 범례 · ax=축 제목 · pk=피크 라벨 · h=차트 높이(px)
TREND_FONT_PRESETS = {
    "보통":      {"x": 11, "y": 12, "lg": 12, "ax": 13, "pk": 11, "h": 470},
    "크게":      {"x": 20, "y": 16, "lg": 17, "ax": 16, "pk": 14, "h": 620},
    "아주 크게": {"x": 25, "y": 20, "lg": 21, "ax": 19, "pk": 16, "h": 700},
}
TREND_FONT_DEFAULT = "크게"


# 추세 다듬기(스무딩) — 2026-08-04 중태님 요청.
#  ※ 왜 '7일 이동평균'이 아니라 '주 단위'인가:
#     이 화면은 이미 월~일 7일치를 한 점으로 합친 '주별 집계'다. 일별 7일 이동평균을 주 간격으로
#     찍으면 (주별 합계 ÷ 7)과 수학적으로 같은 모양이 나와서 변동성이 전혀 줄지 않는다.
#     주 사이의 들쭉날쭉함을 줄이려면 '여러 주'를 묶어 평균/중앙값을 내야 한다.
#  ※ center=True(중앙 정렬): 뒤로 미는 방식(trailing)을 쓰면 피크 시점이 뒤로 밀려버린다.
#     우리는 이미 지난 데이터를 보는 사후 분석이므로 앞뒤를 같이 평균 내 피크 위치를 보존한다.
#  ※ 중앙값(median)은 하루짜리 대형 주문·이벤트로 한 주만 튄 경우를 사실상 무시해 준다.
TREND_SMOOTH = {
    "없음 (원본 주별)": None,
    "3주 이동평균 (권장)": ("mean", 3),
    "4주 이동평균": ("mean", 4),
    "5주 이동평균 (큰 흐름만)": ("mean", 5),
    "3주 중앙값 (단발 폭증에 강함)": ("median", 3),
}
TREND_SMOOTH_DEFAULT = "3주 이동평균 (권장)"


def _trend_smooth(M, mode):
    """주별 매트릭스에 이동평균/이동중앙값을 적용(중앙 정렬 · 양끝은 있는 만큼만 사용)."""
    if not mode or M is None or getattr(M, "empty", True):
        return M
    kind, win = mode
    r = M.rolling(window=win, center=True, min_periods=1)
    return r.median() if kind == "median" else r.mean()


def _trend_smooth_tag(key):
    """제목·표에 붙일 짧은 꼬리표 (스무딩 없음이면 빈 문자열)."""
    return "" if TREND_SMOOTH.get(key) is None else f" · {key.split(' (')[0]}"


def _trend_tick_step(n_weeks, x_font, plot_px=1150):
    """가로축 라벨을 글자 크기에 맞춰 몇 주 간격으로 찍을지 계산(겹침 방지).

    -60° 회전 기준 라벨 하나가 가로로 차지하는 폭 ≈ 글자크기 × 1.45.
    글자를 키우면 53주를 전부 찍을 수 없으므로 2주·3주 간격으로 자동으로 벌린다.
    """
    if n_weeks < 2:
        return 1
    spacing = plot_px / (n_weeks - 1)
    need = x_font * 1.45
    return max(1, int(math.ceil(need / spacing)))

# 계열 색상 팔레트 — 올해 실선과 전년 점선이 같은 색을 쓰도록 고정 배정
TREND_PALETTE = ["#E8743B", "#7B52AB", "#EFB700", "#2F8FD6", "#4C9A52",
                 "#1F3864", "#C62828", "#009B8E", "#B5651D", "#8E44AD",
                 "#2E7D32", "#5D6D7E"]

# 시즌 축 라벨 — 품번 5번째 자리 코드 + 시즌명 (중태님 지시: Z/A/B/C/D로 보이게)
_TREND_SEASON_LABEL = {"공통": "Z (공통)", "봄": "A (봄)", "여름": "B (여름)",
                       "가을": "C (가을)", "겨울": "D (겨울)", "RUNNING": "E (RUNNING)"}
_TREND_SEASON_ORDER = ["Z (공통)", "A (봄)", "B (여름)", "C (가을)", "D (겨울)", "E (RUNNING)"]


@st.cache_data(ttl=21600)
def _trend_cat_maps():
    """아이템코드 → (중카테고리, 소카테고리, 아이템명) 맵.

    진짜 기준은 DB 아이템 마스터(item_master). 마스터에 없는 코드만 구 하드코딩(_INV_CAT_FALLBACK,
    ITEM_MAP)으로 폴백한다 — 재고모니터링·판매분석과 완전히 같은 단일 기준을 쓴다.
    """
    m = load_item_master()
    mid, small, name = {}, {}, {}
    for code, rec in m.items():
        if rec.get("mid"):
            mid[code] = rec["mid"]
        if rec.get("small"):
            small[code] = rec["small"]
        if rec.get("name"):
            name[code] = rec["name"]
    for code, cat in _INV_CAT_FALLBACK.items():      # 마스터 미등록 코드 폴백
        mid.setdefault(code, cat)
        small.setdefault(code, cat)
    for code, tup in ITEM_MAP.items():
        name.setdefault(code, tup[0])
    return mid, small, name


def _trend_prep(frame):
    """조회 대상 프레임에 추세분석용 파생 컬럼(_중카·_소카·_아이템축·_시즌축)을 붙여 반환.

    날짜로 이미 좁힌 뒤에 호출해서 메모리 사용을 줄인다(전체 50만 행에 문자열 컬럼을 붙이지 않음).
    """
    mid_map, small_map, name_map = _trend_cat_maps()
    out = frame.copy()
    if "아이템" in out.columns:
        ic = out["아이템"].astype(str).str.strip().str.upper()
    elif "품번" in out.columns:
        ic = out["품번"].astype(str).str.strip().str.upper().str[1:3]
    else:
        ic = pd.Series("", index=out.index, dtype="object")
    out["_중카"] = ic.map(mid_map).fillna("기타")
    out["_소카"] = ic.map(small_map).fillna("기타")
    out["_아이템축"] = ic + " (" + ic.map(name_map).fillna("미등록") + ")"
    if "시즌명" in out.columns:
        out["_시즌축"] = out["시즌명"].astype(str).map(_TREND_SEASON_LABEL).fillna("기타")
    else:
        out["_시즌축"] = "기타"
    return out


def _trend_axis_col(axis):
    """기준축 선택값 → 실제 사용할 컬럼명."""
    return {"중카테고리": "_중카", "소카테고리": "_소카",
            "아이템코드": "_아이템축", "시즌 (Z/A/B/C/D)": "_시즌축"}[axis]


def _trend_long(frame, axis_col):
    """주(월요일 시작) × 계열 단위 집계 long DataFrame [week, label, rev, orig, qty]."""
    cols = ["week", "label", "rev", "orig", "qty"]
    if frame is None or frame.empty:
        return pd.DataFrame(columns=cols)
    dt = frame["_판매일"]
    week = (dt - pd.to_timedelta(dt.dt.weekday, unit="D")).dt.normalize()
    g = pd.DataFrame({
        "week": week.values,
        "label": frame[axis_col].astype(str).values,
        "rev": pd.to_numeric(frame["_매출액"], errors="coerce").fillna(0.0).values,
        "orig": pd.to_numeric(frame["_최초가매출"], errors="coerce").fillna(0.0).values,
        "qty": pd.to_numeric(frame["_수량"], errors="coerce").fillna(0.0).values,
    })
    return g.groupby(["week", "label"], as_index=False)[["rev", "orig", "qty"]].sum()


def _trend_matrix(long_df, metric):
    """long 집계 → 행=주, 열=계열 매트릭스(선택 지표 기준)."""
    if long_df is None or long_df.empty:
        return pd.DataFrame()

    def piv(v):
        return long_df.pivot_table(index="week", columns="label", values=v, aggfunc="sum")

    if metric == "판가율(%)":
        M = (piv("rev") / piv("orig").replace(0, np.nan)) * 100.0   # 판매 없는 주는 공백(선 끊김)
    elif metric == "판매수량":
        M = piv("qty").fillna(0.0)
    elif metric == "주간 비중(%)":
        r = piv("rev").fillna(0.0)
        M = r.div(r.sum(axis=1).replace(0, np.nan), axis=0) * 100.0
    else:                                                            # 실판매금액(백만)
        M = piv("rev").fillna(0.0) / 1e6
    return M.sort_index()


def _trend_timing(weeks, vals, ratio):
    """한 계열의 시즌 타이밍 자동 판정 — 본격 상승 / 피크 / 피크아웃.

    · 피크 = 기간 내 최대값 주.
    · 본격 상승 시점 = 피크 이전 최저점 이후, 처음으로 '피크 × ratio' 선을 넘어선 주.
    · 피크아웃 시점 = 피크 이후, 처음으로 '피크 × ratio' 선 아래로 내려온 주.
    ratio(기준선)를 올리면 더 늦게 붙고 더 빨리 꺾인 것으로 잡힌다(민감도 조절).
    값이 전부 0/결측이면 None.
    """
    v = np.asarray(vals, dtype="float64")
    if v.size == 0 or not np.isfinite(v).any():
        return None
    vmax = np.nanmax(v)
    if not np.isfinite(vmax) or vmax <= 0:
        return None
    pi = int(np.nanargmax(v))
    thr = vmax * ratio
    pre = v[:pi + 1].copy()
    pre[~np.isfinite(pre)] = np.inf
    ti = int(np.argmin(pre)) if pre.size else 0
    rise = None
    for i in range(ti, pi + 1):
        if np.isfinite(v[i]) and v[i] >= thr:
            rise = i
            break
    out = None
    for i in range(pi + 1, v.size):
        if np.isfinite(v[i]) and v[i] < thr:
            out = i
            break
    return {"peak_i": pi, "peak_v": float(vmax), "rise_i": rise, "out_i": out,
            "rise_w": (weeks[rise] if rise is not None else None),
            "peak_w": weeks[pi],
            "out_w": (weeks[out] if out is not None else None)}


def _trend_wk(ts):
    """주 시작일(월요일)을 표 라벨용 문자열로 — 룰2에 맞춰 연도 2자리."""
    return pd.Timestamp(ts).strftime("%y-%m-%d")


def _trend_metric_fmt(metric, v):
    if v is None or (isinstance(v, float) and not np.isfinite(v)):
        return "–"
    if metric == "실판매금액(백만)":
        return f"{v:,.1f}"
    if metric == "판매수량":
        return f"{v:,.0f}"
    return f"{v:.1f}%"


def render_trend_weekly(df):
    """추세분석 · 메뉴1 — 주별 시즌상품 판매 변화."""
    if df is None or df.empty or "_판매일" not in df.columns or df["_판매일"].notna().sum() == 0:
        st.info("데이터를 먼저 적재하세요.")
        return
    d = df[df["_판매일"].notna()]
    dmin, dmax = d["_판매일"].min().date(), d["_판매일"].max().date()

    st.caption("주(월요일 시작) 단위로 계열별 판매 흐름을 그려서 **언제 붙기 시작하고, 언제 피크를 찍고, "
               "언제 꺾이는지**를 잡아내는 화면이에요. 조회 기간은 최대 1년(53주)까지 선택할 수 있어요.")

    # ── 조건 폼 (2026-08-06): 폼 안 위젯은 바꿔도 계산 안 함, 🔍 조회 때 1번만 계산 ──
    with st.form("tr_form"):
        # ── 기준축 · 지표 · 기간 ─────────────────────────────────────
        st.markdown("##### ① 무엇을 볼지 정하기")
        c1, c2, c3 = st.columns([1.15, 1.0, 1.15])
        axis = c1.selectbox("📊 기준 — 그래프의 **선을 무엇으로 나눌지** (여기를 바꿔야 선이 바뀝니다)",
                            TREND_AXES, index=0, key="tr_axis",
                            help="아이템그룹(중카테고리)별로 선을 보고 싶으면 여기를 '중카테고리'로 바꾸세요. "
                                 "아래쪽 '중카테고리'는 조회 대상을 걸러내는 필터라서 선을 나누지 않아요.")
        metric = c2.selectbox("지표 (세로축)", TREND_METRICS, index=0, key="tr_metric")
        smooth_key = c3.selectbox("〰️ 추세 다듬기 (스무딩)", list(TREND_SMOOTH.keys()),
                                  index=list(TREND_SMOOTH).index(TREND_SMOOTH_DEFAULT), key="tr_smooth",
                                  help="한 주만 튀는 프로모션·단발 대형주문 때문에 선이 들쭉날쭉할 때 씁니다. "
                                       "앞뒤 주를 함께 평균 내는 '중앙 정렬'이라 피크 시점이 뒤로 밀리지 않아요. "
                                       "한 주짜리 폭증이 심하면 '3주 중앙값'이 가장 강하게 눌러줍니다. "
                                       "정확한 원본 수치는 아래 '주별 데이터' 표에 그대로 남아 있어요.")
        smooth_mode = TREND_SMOOTH[smooth_key]
        smooth_tag = _trend_smooth_tag(smooth_key)

        default_start = max(pd.to_datetime(dmax) - pd.Timedelta(days=TREND_MAX_DAYS - 7), pd.to_datetime(dmin)).date()
        rng = st.date_input(f"조회기간 (최대 1년 · 기본 = 최근 {TREND_MAX_WEEKS}주)",
                            value=(default_start, dmax), min_value=dmin, max_value=dmax, key="tr_rng")

        o0, o1, o2, o3, o4 = st.columns([1.05, 0.95, 0.95, 1.5, 0.95])
        show_total = o0.checkbox("전체 실판가 배경선", value=True, key="tr_total",
                                 help="**회사 전체** 실판매금액을 굵은 회색 반투명 선으로 뒤에 깔아줘요. "
                                      "필터·기준축을 어떻게 바꿔도 이 선은 그대로라서, 특정 아이템·시즌이 "
                                      "전체 흐름과 얼마나 다르게 움직이는지 바로 비교할 수 있어요. "
                                      "숫자 크기가 달라 다른 선이 눌리지 않도록 별도 보조축을 씁니다.")
        show_prev = o1.checkbox("전년 동기간 점선 비교", value=False, key="tr_prev",
                                help="같은 색 점선으로 전년 같은 주(52주 전)를 겹쳐 그려요 — "
                                     "'작년 이맘때보다 빠른가/늦은가'를 볼 수 있어요.")
        show_peak = o2.checkbox("피크 시점 자동 표시", value=True, key="tr_peak",
                                help="계열마다 최고점 주에 마커와 '○○ 피크' 라벨을 찍어줘요.")
        thr_pct = o3.slider("시점 판정 기준선 (피크 대비 %)", min_value=30, max_value=80, value=50, step=5,
                            key="tr_thr",
                            help="이 선을 처음 넘은 주 = 본격 상승 시점, 피크 뒤 처음 내려온 주 = 피크아웃 시점.")
        font_key = o4.selectbox("글자 크기", list(TREND_FONT_PRESETS.keys()),
                                index=list(TREND_FONT_PRESETS).index(TREND_FONT_DEFAULT), key="tr_font",
                                help="회의 때 화면에 띄우면 '아주 크게'가 잘 보여요. 글자를 키우면 "
                                     "가로축 날짜가 겹치지 않게 라벨 간격(1주→2주)이 자동으로 벌어져요.")
        FS = TREND_FONT_PRESETS[font_key]

        # ── 🌡️ 서울 기온 겹쳐보기 (2026-08-03) ─────────────────────
        #    "아침 최저기온이 20도 아래로 떨어진 주에 니트가 붙는다" 같은 임계 온도를 눈으로도, 표로도 확인.
        n_wx = weather_row_count()
        show_wx, wx_lines, wx_key = False, [], "최저기온"
        if n_wx:
            wd0, wd1 = weather_span()
            g1, g2, g3 = st.columns([1, 1.5, 1.2])
            show_wx = g1.checkbox("🌡️ 서울 기온 겹쳐보기", value=False, key="tr_wx",
                                  help=f"기상청 ASOS 일자료 {wd0}~{wd1} 적재됨. 주 평균으로 보조축(오른쪽)에 겹쳐 그려요.")
            wx_lines = g2.multiselect("겹쳐 그릴 기온", WEATHER_KINDS, default=["최저기온", "최고기온"],
                                      key="tr_wxlines", placeholder="선택")
            wx_key = g3.selectbox("임계 기온 기준", WEATHER_KINDS, index=1, key="tr_wxkey",
                                  help="시점 요약표의 '그때 기온' 컬럼에 쓸 기준이에요. "
                                       "가을 시즌 진입은 보통 아침 최저기온이 가장 잘 맞아요.")
        else:
            st.caption("🌡️ 서울 기온을 겹쳐 보려면 관리자가 사이드바 **기온 데이터 업로드**에서 "
                       "기상자료개방포털 ASOS 일자료를 올리거나, 기상청 API로 자동 수집하면 돼요.")

        # ── 데이터 준비: 필요한 날짜 구간만 잘라서 파생 컬럼 부착 (메모리 절약) ──
        #    폼 안에서는 ② 필터의 선택지(브랜드·시즌·카테고리 목록)를 만들기 위해서만 자른다.
        #    기간이 아직 미완성(시작일만 선택 등)이면 기본 기간으로 대체해 선택지만 구성하고,
        #    실제 기간 검증·오류 안내는 폼 밖(조회 버튼 이후)에서 한다.
        _rng_ok = (isinstance(rng, (list, tuple)) and len(rng) == 2
                   and pd.to_datetime(rng[1]) >= pd.to_datetime(rng[0])
                   and (pd.to_datetime(rng[1]) - pd.to_datetime(rng[0])).days <= TREND_MAX_DAYS)
        if _rng_ok:
            s, e = pd.to_datetime(rng[0]), pd.to_datetime(rng[1])
        else:
            s, e = pd.to_datetime(default_start), pd.to_datetime(dmax)
        s_load = (s - pd.Timedelta(days=TREND_PREV_SHIFT_DAYS)) if show_prev else s
        base = d[(d["_판매일"] >= s_load) & (d["_판매일"] <= e)]
        if not base.empty:
            base = _trend_prep(base)

        # ── 공통 필터 (빈칸=전체) — 브랜드 · 시즌 · 중카테고리 · 소카테고리 ──
        #    ⚠️ 여기는 '조회 대상을 걸러내는' 필터일 뿐, 그래프의 선을 나누지 않는다(선을 나누는 건 위 ① 기준).
        #    실제로 중카테고리 필터를 걸어놓고 "왜 아이템그룹별로 안 보이지?" 하는 혼동이 있었어서(260803),
        #    라벨을 ②로 번호 붙이고, 아래에 '기준 바꾸기' 원클릭 버튼 안내를 띄운다.
        st.markdown("##### ② 조회 대상 좁히기 (필터 · 빈칸=전체)"
                    "<span style='color:#888;font-weight:400;font-size:0.78rem;'> — 여기는 데이터를 "
                    "걸러내기만 해요. 선을 나누는 건 위 ①의 **기준**입니다.</span>", unsafe_allow_html=True)
        f1, f2, f3, f4 = st.columns(4)
        brands = sorted(base["브랜드명"].dropna().astype(str).unique()) if "브랜드명" in base.columns else []
        seasons = ([x for x in _TREND_SEASON_ORDER if x in set(base["_시즌축"])] +
                   sorted(x for x in set(base["_시즌축"]) if x not in _TREND_SEASON_ORDER)) \
            if "_시즌축" in base.columns else []
        mids = sorted(set(base["_중카"])) if "_중카" in base.columns else []
        smalls = sorted(set(base["_소카"])) if "_소카" in base.columns else []
        selb = f1.multiselect("브랜드", brands, default=[], placeholder="전체", key="tr_fb")
        sels = f2.multiselect("시즌", seasons, default=[], placeholder="전체", key="tr_fs")
        selm = f3.multiselect("중카테고리", mids, default=[], placeholder="전체", key="tr_fm")
        selsm = f4.multiselect("소카테고리", smalls, default=[], placeholder="전체", key="tr_fsm")
        run = st.form_submit_button("🔍 조회", type="primary")

    if _need_search("tr_go", run):
        st.caption(f"🔍 조회하면 **{axis}별 {metric} 추세 그래프**(주 단위)와 그 아래 "
                   "**시점요약표·주별데이터표**가 나와요. 위 ①에서 기준·지표·스무딩을 먼저 정해 보세요.")
        return
    if not (isinstance(rng, (list, tuple)) and len(rng) == 2):
        st.info("기간(시작~끝)을 선택한 뒤 🔍 조회를 눌러 주세요.")
        return
    _s0, _e0 = pd.to_datetime(rng[0]), pd.to_datetime(rng[1])
    if _e0 < _s0:
        st.error("종료일이 시작일보다 앞서요. 기간을 다시 선택해 주세요.")
        return
    if (_e0 - _s0).days > TREND_MAX_DAYS:
        st.error(f"조회 기간은 최대 1년({TREND_MAX_WEEKS}주)까지예요. "
                 f"현재 선택 {(_e0 - _s0).days + 1}일 — 기간을 줄여 주세요.")
        return
    if base.empty:
        st.info("선택한 기간에 매출 데이터가 없어요.")
        return
    if selb and "브랜드명" in base.columns:
        base = base[base["브랜드명"].astype(str).isin(selb)]
    if sels:
        base = base[base["_시즌축"].isin(sels)]
    if selm:
        base = base[base["_중카"].isin(selm)]
    if selsm:
        base = base[base["_소카"].isin(selsm)]

    # 💡 필터를 건 축과 그래프 기준이 다르면(예: 중카테고리 필터 + 시즌 기준) 원클릭 전환 버튼 안내.
    #    필터를 건 축 자체(= 그 항목들끼리 비교)와 한 단계 더 잘게 쪼갠 축(= 그 안을 들여다보기) 둘 다 제안.
    _filtered_dims = [nm for nm, sel in (("중카테고리", selm), ("소카테고리", selsm),
                                         ("시즌 (Z/A/B/C/D)", sels)) if sel]
    _sugg = []
    if selm:
        _sugg += ["중카테고리", "아이템코드"]      # 카테고리끼리 비교 / 그 안의 아이템별로 쪼개기
    if selsm:
        _sugg += ["소카테고리", "아이템코드"]
    if sels:
        _sugg += ["시즌 (Z/A/B/C/D)"]
    _swap = [x for x in dict.fromkeys(_sugg) if x != axis][:2]
    if _filtered_dims and _swap:
        _hc = st.columns([3.0] + [1.5] * len(_swap))
        _hc[0].caption(f"💡 **{' · '.join(_filtered_dims)}** 로 걸러내는 중인데, 그래프는 "
                       f"**{axis}** 기준으로 선을 나누고 있어요. 아래 버튼으로 기준을 바로 바꿀 수 있어요.")
        for _i, _nm in enumerate(_swap):
            if _hc[_i + 1].button(f"→ {_nm}로 선 나누기", key=f"tr_swap_{_i}", use_container_width=True):
                st.session_state["tr_axis"] = _nm
                st.rerun()

    if base.empty:
        st.info("필터 조건에 맞는 데이터가 없어요. 조건을 넓혀 보세요.")
        return

    axis_col = _trend_axis_col(axis)
    cur = base[(base["_판매일"] >= s) & (base["_판매일"] <= e)]
    if cur.empty:
        st.info("선택한 기간·조건에 매출 데이터가 없어요.")
        return

    M = _trend_matrix(_trend_long(cur, axis_col), metric)
    if M.empty:
        st.info("집계 결과가 비어 있어요.")
        return

    # ── 계열 선택 (기본 = 기간 매출 상위 N개) ────────────────────────
    rev_rank = (cur.groupby(cur[axis_col].astype(str), observed=True)["_매출액"].sum()
                .sort_values(ascending=False))
    if axis_col == "_시즌축":                                    # 시즌은 Z→A→B→C→D 고정 순서로
        ordered = [x for x in _TREND_SEASON_ORDER if x in M.columns] + \
                  [x for x in rev_rank.index if x in M.columns and x not in _TREND_SEASON_ORDER]
    else:
        ordered = [x for x in rev_rank.index if x in M.columns]
    ordered += [c for c in M.columns if c not in ordered]
    default_sel = ordered if axis_col == "_시즌축" else ordered[:TREND_DEFAULT_TOPN]
    picked = st.multiselect(f"③ 표시할 {axis} (기본 = 매출 상위 {TREND_DEFAULT_TOPN}개 · 비우면 기본값)",
                            ordered, default=default_sel, key=f"tr_pick_{axis_col}",
                            placeholder=f"비워두면 기본값(매출 상위 {TREND_DEFAULT_TOPN}개)으로 표시돼요")
    if not picked:
        picked = default_sel
    picked = [c for c in ordered if c in picked]
    # Mraw = 원본 주별 수치(아래 '주별 데이터' 표·엑셀은 항상 이 값 = 사실 데이터)
    # M    = 스무딩 적용본(그래프와 시점 판정용 = 추세 해석)
    Mraw = M[picked]
    M = _trend_smooth(Mraw, smooth_mode)

    # 가로축은 '주 시작일(월요일)' 실제 날짜를 그대로 씀 — MM-DD 문자열을 쓰면 1년(53주)을 꽉 채웠을 때
    # 첫 주와 마지막 주의 MM-DD가 겹쳐 두 점이 한 칸으로 합쳐지는 사고가 나기 때문(날짜축이면 안전).
    weeks = list(M.index)
    color_of = {name: TREND_PALETTE[i % len(TREND_PALETTE)] for i, name in enumerate(picked)}

    # ── 전년 동기간 (52주=364일 시프트로 주 라벨 정렬) ──────────────
    Mp = pd.DataFrame()
    if show_prev:
        sp, ep = s - pd.Timedelta(days=TREND_PREV_SHIFT_DAYS), e - pd.Timedelta(days=TREND_PREV_SHIFT_DAYS)
        prev = base[(base["_판매일"] >= sp) & (base["_판매일"] <= ep)]
        Mp = _trend_matrix(_trend_long(prev, axis_col), metric)
        if not Mp.empty:
            Mp.index = Mp.index + pd.Timedelta(days=TREND_PREV_SHIFT_DAYS)
            Mp = Mp.reindex(index=M.index, columns=picked)

    # ── 차트 ─────────────────────────────────────────────────────────
    ttl = f"주별 {axis} 판매 변화 — {metric}{smooth_tag}"
    st.markdown(f"**{ttl}**  <span style='color:#888;font-size:0.8rem;'>"
                f"({s.date()} → {e.date()} · {len(weeks)}주 · 주 시작일=월요일)</span>"
                + (_NOTE_FLOAT if metric == "실판매금액(백만)" else ""), unsafe_allow_html=True)
    fig = go.Figure()

    # ── ① 전체 실판가 배경선 (2026-08-03 중태님 지시 · 260803 수정) ──
    #    "전체 매출은 이렇게 움직이는데, 이 아이템/시즌은 어떻게 움직이나"를 한 화면에서 보기 위한 기준선.
    #    ⚠️ 중요: 이 선은 **어떤 조회 조건에도 영향받지 않는 '회사 전체' 매출**이다.
    #       기준축·필터(브랜드/시즌/중·소카테고리)·계열 선택을 아무리 바꿔도 항상 같은 선이 유지돼야
    #       "신발만 봤을 때 전체 흐름과 어떻게 다른가"를 판단할 수 있기 때문 → 필터 적용 전 원본(d)에서 집계.
    #    전체 합계는 개별 계열보다 자릿수가 커서 같은 축에 두면 나머지 선이 바닥에 눌린다 → 별도 보조축(y3).
    #    맨 먼저 그려야 다른 선들 뒤(배경)에 깔린다. 지표 선택과 무관하게 항상 '실판매금액(백만)' 기준.
    tot_ser = pd.Series(dtype="float64")
    tot_company = 0.0
    if show_total:
        all_cur = d[(d["_판매일"] >= s) & (d["_판매일"] <= e)]      # ← 필터 적용 전 = 회사 전체
        tot_company = float(pd.to_numeric(all_cur["_매출액"], errors="coerce").fillna(0.0).sum())
        _dtc = all_cur["_판매일"]
        _wkc = (_dtc - pd.to_timedelta(_dtc.dt.weekday, unit="D")).dt.normalize()
        tot_ser = (pd.Series(pd.to_numeric(all_cur["_매출액"], errors="coerce").fillna(0.0).values)
                   .groupby(_wkc.values).sum().reindex(M.index).fillna(0.0) / 1e6)
        tot_raw_sum = float(tot_ser.sum()) * 1e6      # 스무딩 전 실제 합계(요약표 금액용)
        tot_ser = _trend_smooth(tot_ser.to_frame("v"), smooth_mode)["v"]   # 계열과 같은 스무딩 적용
        fig.add_scatter(x=weeks, y=[float(v) for v in tot_ser], name="전체 실판가 (회사 전체)",
                        mode="lines", yaxis="y3",
                        line=dict(color="rgba(126,131,138,0.32)", width=8, shape="spline"),
                        hovertemplate="전체 실판가(회사 전체) %{y:,.1f}<extra></extra>")

    if show_prev and not Mp.empty:
        for name in picked:
            if name not in Mp.columns:
                continue
            fig.add_scatter(x=weeks, y=[None if pd.isna(v) else float(v) for v in Mp[name]],
                            name=f"{name} (전년)", mode="lines", legendgroup=name,
                            line=dict(color=color_of[name], width=1.6, dash="dot"),
                            opacity=0.55, hovertemplate=f"{name}(전년) " + "%{y:,.1f}<extra></extra>")
    timing = {}
    for name in picked:
        ys = [None if pd.isna(v) else float(v) for v in M[name]]
        fig.add_scatter(x=weeks, y=ys, name=name, mode="lines", legendgroup=name,
                        line=dict(color=color_of[name], width=2.6),
                        hovertemplate=f"{name} " + "%{y:,.1f}<extra></extra>")
        t = _trend_timing(weeks, M[name].to_numpy(dtype="float64"), thr_pct / 100.0)
        timing[name] = t
        if t and show_peak:
            fig.add_scatter(x=[weeks[t["peak_i"]]], y=[t["peak_v"]], mode="markers",
                            marker=dict(color=color_of[name], size=10, symbol="diamond",
                                        line=dict(color="#fff", width=1.4)),
                            showlegend=False, legendgroup=name, hoverinfo="skip")
            fig.add_annotation(x=weeks[t["peak_i"]], y=t["peak_v"], text=f"{name} 피크",
                               showarrow=True, arrowhead=2, arrowsize=1, arrowwidth=1.6,
                               arrowcolor=color_of[name], ax=0, ay=-30,
                               bgcolor=color_of[name], bordercolor=color_of[name],
                               font=dict(color="#ffffff", size=FS["pk"]))
    # ── 기온 보조축(오른쪽) 겹쳐 그리기 ──
    wxw = weather_weekly(weeks) if (show_wx and wx_lines) else pd.DataFrame()
    if not wxw.empty:
        for kind in wx_lines:
            if kind not in wxw.columns:
                continue
            fig.add_scatter(x=weeks, y=[None if pd.isna(v) else float(v) for v in wxw[kind]],
                            name=f"🌡 {kind}", mode="lines", yaxis="y2",
                            line=dict(color=_WEATHER_LINE.get(kind, "#888"), width=1.8, dash="longdash"),
                            opacity=0.75, hovertemplate=f"{kind} " + "%{y:.1f}℃<extra></extra>")
    ytitle = {"실판매금액(백만)": "백만원", "판매수량": "수량(점)",
              "판가율(%)": "판가율(%)", "주간 비중(%)": "비중(%)"}[metric]
    step_w = _trend_tick_step(len(weeks), FS["x"])
    fig.update_layout(height=FS["h"], margin=dict(t=34, b=10, l=0, r=10),
                      legend=dict(orientation="h", y=1.12, font=dict(size=FS["lg"])),
                      yaxis_title=ytitle, hovermode="x unified",
                      hoverlabel=dict(font_size=max(13, FS["y"])))
    # 오른쪽 보조축이 2개(기온 + 전체 실판가)면 서로 겹치지 않게 그래프 폭을 살짝 줄이고 축을 나란히 배치
    has_tot = show_total and not tot_ser.empty
    n_right = int(not wxw.empty) + int(has_tot)
    if n_right > 1:
        fig.update_layout(xaxis=dict(domain=[0.0, 0.93]))
    if not wxw.empty:
        fig.update_layout(yaxis2=dict(title="기온(℃)", overlaying="y", side="right",
                                      showgrid=False, zeroline=True, zerolinecolor="#ddd",
                                      tickfont=dict(size=FS["y"]), title_font=dict(size=FS["ax"])))
    if has_tot:
        y3 = dict(title="회사 전체 실판가(백만)", overlaying="y", side="right", showgrid=False,
                  rangemode="tozero", tickfont=dict(size=FS["y"], color="#7E838A"),
                  title_font=dict(size=FS["ax"], color="#7E838A"))
        if n_right > 1:                      # 기온 축 바깥쪽(맨 오른쪽)에 한 칸 띄워 배치
            y3.update(anchor="free", position=1.0)
        fig.update_layout(yaxis3=y3)
    fig.update_xaxes(type="date", tickformat="%m-%d", dtick=step_w * 7 * 24 * 3600 * 1000,
                     tickangle=-60, tickfont=dict(size=FS["x"]), automargin=True)
    fig.update_yaxes(tickfont=dict(size=FS["y"]), title_font=dict(size=FS["ax"]), automargin=True)
    st.plotly_chart(fig, use_container_width=True)
    st.caption("※ 가로축 = 주 시작일(월요일) MM-DD"
               + (f" · 글자 크기 '{font_key}'라 라벨은 **{step_w}주 간격**으로 표시(점·선은 매주 그대로)"
                  if step_w > 1 else "")
               + ". 마우스를 올리면 그 주의 모든 계열 값이 한 번에 떠요. "
               "범례를 클릭하면 해당 계열만 켜고 끌 수 있어요."
               + ("  전년 점선은 52주 전 같은 주와 맞춰 그린 값이에요." if show_prev else "")
               + ("  〰️ **" + smooth_key.split(" (")[0] + "** 적용 — 앞뒤 주를 함께 평균 낸 "
                  "중앙 정렬 방식이라 피크 시점은 밀리지 않아요. 정확한 원본 수치는 아래 '주별 데이터' 표에 있어요."
                  if smooth_mode else "")
               + ("  **굵은 회색 배경선 = 회사 전체 실판가**(오른쪽 회색 눈금 · 백만원). "
                  "**기준축·필터·계열 선택을 바꿔도 이 선은 항상 회사 전체로 고정**돼요 — "
                  "그래야 '신발만 봤을 때 전체 흐름과 어떻게 다른가'를 비교할 수 있으니까요. "
                  "개별 계열과 자릿수가 달라 별도 축을 쓰니 **선의 높낮이가 아니라 오르내리는 모양**을 비교해 주세요."
                  if has_tot else ""))

    # ── 시점 요약표 (본격 상승 · 피크 · 피크아웃) ────────────────────
    tot_all = float(cur["_매출액"].sum())
    trows, tidx = [], []
    gseries = M[picked].sum(axis=1) if metric in ("실판매금액(백만)", "판매수량") else None
    wxt = wxw if (show_wx and not wxw.empty and wx_key in getattr(wxw, "columns", [])) else None
    if has_tot:
        # 전체 실판가 행은 '회사 전체'라 조회 조건과 모집단이 달라 비중 칸은 '–'로 둔다(tot_all=0 → '–').
        # 값은 지표 선택과 무관하게 항상 백만원 기준(첫 행 노란 강조).
        tt = _trend_timing(weeks, tot_ser.to_numpy(dtype="float64"), thr_pct / 100.0)
        tidx.append("■ 전체 실판가 (회사 전체)")
        trows.append(_trend_timing_row(tt, "실판매금액(백만)", 0.0, rev=tot_company,
                                       wxw=wxt, wx_key=wx_key))
    if gseries is not None:
        gt = _trend_timing(weeks, gseries.to_numpy(dtype="float64"), thr_pct / 100.0)
        tidx.append("G.TOTAL (선택 계열 합)")
        trows.append(_trend_timing_row(gt, metric, tot_all, agg=True, wxw=wxt, wx_key=wx_key))
    for name in picked:
        tidx.append(name)
        trows.append(_trend_timing_row(timing.get(name), metric, tot_all,
                                       rev=float(rev_rank.get(name, 0.0)),
                                       wxw=wxt, wx_key=wx_key))
    T = pd.DataFrame(trows, index=tidx)
    T.index.name = axis

    t1, t2 = st.columns([5, 1])
    t1.markdown("##### 🕒 계열별 시즌 타이밍 요약"
                + (f" <span style='color:#888;font-weight:400;font-size:0.78rem;'>"
                   f"({smooth_key.split(' (')[0]} 기준)</span>" if smooth_mode else "")
                + _NOTE_FLOAT, unsafe_allow_html=True)
    t2.download_button("⬇ 엑셀", table_excel_bytes(T, "시점요약"),
                       file_name=f"추세분석_시점요약_{_safe_name(axis)}_{e.date()}.xlsx",
                       mime=XLSX_MIME, key="tr_dl_timing", use_container_width=True)
    render_styled_table(T.style.set_properties(**{"text-align": "right"}))
    _tail = ""
    if has_tot:
        _share = (tot_all / tot_company * 100) if tot_company else None
        _tail = ("  **■ 전체 실판가 행은 필터와 무관한 '회사 전체' 기준**이라 모집단이 달라 비중 칸은 '–'예요"
                 + (f" — 지금 조회 조건은 회사 전체 매출의 **{_share:.1f}%**({_mm(tot_all):,.1f} / "
                    f"{_mm(tot_company):,.1f} 백만)입니다." if _share is not None else "."))
    st.caption(f"※ 기준선 = 피크 대비 **{thr_pct}%**. **본격 상승 시점** = 피크 직전 최저점 이후 이 선을 "
               "처음 넘어선 주 · **피크아웃 시점** = 피크 뒤 이 선 아래로 처음 내려온 주(기간 안에서 아직 "
               "안 꺾였으면 '–'). 조회 기간이 짧으면 실제 피크가 기간 밖일 수 있으니 1년으로 넓혀서도 확인해 주세요."
               + _tail)

    # ── 🌡️ 임계 온도 자동 코멘트 (보고서에 그대로 붙일 수 있는 문장) ──
    if wxt is not None:
        cmts = []
        for name in picked:
            t = timing.get(name)
            if not t:
                continue
            tr_ = _wx_at(wxt, wx_key, t["rise_i"])
            to_ = _wx_at(wxt, wx_key, t["out_i"])
            if tr_ is None and to_ is None:
                continue
            seg = f"<b>{name}</b> — "
            if tr_ is not None:
                seg += (f"주간 {wx_key} <b>{tr_:.1f}℃</b> 구간({_trend_wk(t['rise_w'])} 주)에서 붙기 시작")
            if to_ is not None:
                seg += (" · " if tr_ is not None else "") + \
                       f"{to_:.1f}℃ 구간({_trend_wk(t['out_w'])} 주)에서 꺾임"
            cmts.append(seg)
        if cmts:
            st.markdown("##### 🌡️ 임계 온도 자동 코멘트")
            body = "<br>".join(f"{i + 1}. {t}" for i, t in enumerate(cmts))
            st.markdown("<div style='background:#f7f9fc;border:1px solid #dde6f0;border-radius:8px;"
                        f"padding:12px 16px;font-size:0.88rem;line-height:1.9;'>{body}</div>",
                        unsafe_allow_html=True)
            st.caption(f"※ 해당 주의 일자료 평균 {wx_key} 기준(서울 108 지점). 매년 같은 온도대에서 "
                       "반복되는지 확인하면, 다음 시즌 상품 투입·프로모션 시점을 온도로 잡을 수 있어요.")

    # ── 주별 데이터 표 + 엑셀 (룰11) ─────────────────────────────────
    disp = pd.DataFrame({name: [_trend_metric_fmt(metric, v) for v in Mraw[name]] for name in picked},
                        index=[_trend_wk(w) for w in weeks])
    if metric in ("실판매금액(백만)", "판매수량"):
        disp.insert(0, "합계", [_trend_metric_fmt(metric, v) for v in Mraw[picked].sum(axis=1)])
        head = {"합계": _trend_metric_fmt(metric, float(Mraw[picked].sum(axis=1).sum()))}
        head.update({name: _trend_metric_fmt(metric, float(Mraw[name].sum())) for name in picked})
    else:
        head = {name: _trend_metric_fmt(metric, float(Mraw[name].mean(skipna=True))) for name in picked}
    lbl = "기간 합계" if metric in ("실판매금액(백만)", "판매수량") else "기간 평균"
    disp = pd.concat([pd.DataFrame([head], index=[lbl]), disp])
    disp.index.name = "주 시작일"

    w1, w2 = st.columns([5, 1])
    w1.markdown(f"##### 📋 주별 데이터 ({metric})"
                + (" <span style='color:#888;font-weight:400;font-size:0.78rem;'>"
                   "— 스무딩 적용 안 된 <b>원본 수치</b></span>" if smooth_mode else "")
                + (_NOTE_FLOAT if metric == "실판매금액(백만)" else ""), unsafe_allow_html=True)
    w2.download_button("⬇ 엑셀", table_excel_bytes(disp, "주별데이터"),
                       file_name=f"추세분석_주별_{_safe_name(axis)}_{e.date()}.xlsx",
                       mime=XLSX_MIME, key="tr_dl_weekly", use_container_width=True)
    render_styled_table(disp.style.set_properties(**{"text-align": "right"}))
    st.caption(f"※ 첫 행 = {lbl}(노란 강조). 주 시작일은 월요일 기준이라 마지막 주는 조회 종료일까지만 "
               "집계된 '미완성 주'일 수 있어요 — 끝부분이 갑자기 낮으면 이것 때문일 수 있으니 참고하세요."
               + ("  이 표와 엑셀은 **스무딩을 적용하지 않은 원본 수치**예요(그래프·시점 판정에만 적용)."
                  if smooth_mode else ""))


def _trend_timing_row(t, metric, tot_all=0.0, rev=None, agg=False, wxw=None, wx_key="최저기온"):
    """시점 요약표의 한 행 만들기 — 본격 상승 / 피크 / 피크아웃 + 소요 주수 + 기간 매출·비중.

    wxw(주간 기온표)를 넘기면 각 시점 주의 기온(wx_key 기준) 컬럼 3개가 함께 붙는다
    — "니트류는 주간 최저기온 12℃ 구간에서 붙기 시작" 같은 임계 온도를 표에서 바로 읽기 위한 것.
    """
    if t is None:
        row = {"본격 상승 시점": "–", "피크 시점": "–", "피크값": "–", "피크아웃 시점": "–",
               "상승→피크(주)": "–", "피크→피크아웃(주)": "–"}
        if wxw is not None:
            row.update({f"상승 시점 {wx_key}": "–", f"피크 시점 {wx_key}": "–",
                        f"피크아웃 {wx_key}": "–"})
    else:
        row = {
            "본격 상승 시점": _trend_wk(t["rise_w"]) if t["rise_w"] is not None else "–",
            "피크 시점": _trend_wk(t["peak_w"]),
            "피크값": _trend_metric_fmt(metric, t["peak_v"]),
            "피크아웃 시점": _trend_wk(t["out_w"]) if t["out_w"] is not None else "–",
            "상승→피크(주)": (f"{t['peak_i'] - t['rise_i']}주" if t["rise_i"] is not None else "–"),
            "피크→피크아웃(주)": (f"{t['out_i'] - t['peak_i']}주" if t["out_i"] is not None else "–"),
        }
        if wxw is not None:
            row[f"상승 시점 {wx_key}"] = _wx_fmt(_wx_at(wxw, wx_key, t["rise_i"]))
            row[f"피크 시점 {wx_key}"] = _wx_fmt(_wx_at(wxw, wx_key, t["peak_i"]))
            row[f"피크아웃 {wx_key}"] = _wx_fmt(_wx_at(wxw, wx_key, t["out_i"]))
    if agg:
        row["기간 실판매(백만)"] = "–"
        row["기간 비중"] = "100.0%"
    else:
        row["기간 실판매(백만)"] = f"{(rev or 0.0) / 1e6:,.1f}"
        row["기간 비중"] = f"{(rev or 0.0) / tot_all * 100:.1f}%" if tot_all else "–"
    return row


TREND_MENUS = ["📈 주별 시즌상품 판매 변화"]


def render_trend(df):
    """📉 추세분석 탭 — 메뉴 선택 후 해당 화면 렌더(메뉴는 앞으로 계속 추가 예정)."""
    st.subheader("📉 추세분석")
    menu = st.radio("메뉴", TREND_MENUS, horizontal=True,
                    label_visibility="collapsed", key="trend_menu")
    st.divider()
    if menu == TREND_MENUS[0]:
        render_trend_weekly(df)


# ==============================================================================
# 서울 기온 데이터  ─ 추세분석 기온 겹쳐보기용 (2026-08-03 신설 · 중태님 지시)
# ==============================================================================
# "아침 최저기온이 20도 아래로 떨어진 주에 니트가 붙기 시작한다" 같은 임계 온도를 잡아내기 위해
# 기상청 종관기상관측(ASOS) 일자료를 DB에 쌓아두고, 추세분석 차트에 보조축으로 겹쳐 그린다.
#  · 적재 방법 2가지 (둘 다 관리자 전용 · 사이드바)
#    ① 파일 업로드 — 기상자료개방포털(data.kma.go.kr) → 종관기상관측(ASOS) → 일자료 CSV/엑셀 그대로
#    ② API 자동 수집 — 공공데이터포털 '기상청_지상(종관, ASOS) 일자료 조회서비스' 인증키가 있을 때
#  · 저장 단위는 '일자료'. 화면에서 쓸 때 주(월요일 시작) 평균으로 자동 집계한다.
#  · 개인정보와 무관한 공공데이터라 PII 방어벽 대상이 아니다.
WEATHER_TABLE = "weather_daily"
WEATHER_STN_DEFAULT = "108"          # 108 = 서울
WEATHER_KINDS = ["평균기온", "최저기온", "최고기온"]
_WEATHER_COL = {"평균기온": "avg_ta", "최저기온": "min_ta", "최고기온": "max_ta"}
_WEATHER_LINE = {"평균기온": "#4CAF50", "최저기온": "#2F6BB0", "최고기온": "#C62828"}
KMA_ASOS_URL = "https://apis.data.go.kr/1360000/AsosDalyInfoService/getWthrDataList"
# 공공데이터포털 인증 관련 오류코드 → 팀원이 바로 조치할 수 있는 한글 안내로 변환
_KMA_ERR = {
    "01": "제공기관 서비스 오류예요. 잠시 뒤 다시 시도해 주세요.",
    "03": "해당 기간·지점에 자료가 없어요. 기간이나 지점번호를 확인해 주세요.",
    "04": "요청값(HTTP) 오류예요.",
    "12": "폐기된 서비스예요. 포털에서 서비스 상태를 확인해 주세요.",
    "20": "접근이 거부됐어요 — 포털에서 이 API를 '활용신청' 했는지 확인해 주세요.",
    "22": "오늘 요청 한도를 초과했어요(개발계정 1일 10,000건). 내일 다시 시도해 주세요.",
    "30": "등록되지 않은 인증키예요 — ① '일반 인증키(Decoding)'를 복사했는지 "
          "② 활용신청 직후라면 1시간 정도 뒤에 다시 시도해 주세요.",
    "31": "기한이 만료된 인증키예요. 포털에서 연장 신청해 주세요.",
    "32": "등록되지 않은 도메인/IP예요.",
}


def _wx_num(v):
    try:
        f = float(str(v).strip().replace(",", ""))
        return None if pd.isna(f) else f
    except Exception:
        return None


def read_weather_file(uploaded_file):
    """기상자료개방포털 ASOS 일자료 파일(CSV/엑셀) → DF[date, stn, avg_ta, min_ta, max_ta].

    포털에서 받은 CSV는 인코딩이 CP949인 경우가 많고 위쪽에 주석/안내 줄이 붙기도 해서,
    '일시(또는 날짜/tm)' 헤더가 있는 줄을 찾아 그 줄부터 읽는다. 컬럼은 이름 키워드로 잡되
    '최저기온시각'처럼 시각(hhmi) 컬럼은 제외한다.
    """
    import csv as _csv
    name = str(uploaded_file.name).lower()
    if name.endswith((".xlsx", ".xls")):
        raw = pd.read_excel(uploaded_file, header=None, dtype=str)
    else:
        # 포털 CSV는 위쪽에 '# 종관기상관측…' 같은 안내 줄이 붙어 줄마다 칸 수가 달라서
        # pandas 기본 파서가 실패한다 → 직접 읽어 최대 칸 수에 맞춰 채운 뒤 표로 만든다.
        try:
            uploaded_file.seek(0)
        except Exception:
            pass
        data = uploaded_file.read()
        text = None
        if isinstance(data, str):
            text = data
        else:
            for enc in ("utf-8-sig", "cp949", "euc-kr", "utf-8", "latin-1"):
                try:
                    text = data.decode(enc)
                    break
                except Exception:
                    text = None
        if not text:
            raise ValueError("CSV를 읽지 못했어요 — 파일 인코딩(UTF-8/CP949)을 확인해 주세요.")
        head = text[:4000]
        delim = "\t" if head.count("\t") > head.count(",") else ","
        rows = [r for r in _csv.reader(io.StringIO(text), delimiter=delim) if any(str(x).strip() for x in r)]
        if not rows:
            raise ValueError("CSV 내용이 비어 있어요.")
        ncol = max(len(r) for r in rows)
        raw = pd.DataFrame([r + [""] * (ncol - len(r)) for r in rows], dtype=str)

    hrow = None
    for i in range(min(30, len(raw))):
        vals = [str(v).strip() for v in raw.iloc[i].tolist()]
        if any(v in ("일시", "날짜", "tm", "TM") for v in vals) and any("기온" in v or "Ta" in v for v in vals):
            hrow = i
            break
    if hrow is None:
        raise ValueError("'일시'와 '기온' 컬럼이 있는 헤더 줄을 찾지 못했어요 — "
                         "기상자료개방포털 ASOS 일자료 원본 파일인지 확인해 주세요.")
    header = [str(v).strip() for v in raw.iloc[hrow].tolist()]
    body = raw.iloc[hrow + 1:].copy()
    body.columns = header
    body = body.dropna(how="all")

    def pick(*keys, exclude=("시각", "hhmi", "HHMI")):
        for c in header:
            if any(k in c for k in keys) and not any(x in c for x in exclude):
                return c
        return None

    c_date = pick("일시", "날짜", "tm", "TM", exclude=())
    c_avg = pick("평균기온", "avgTa")
    c_min = pick("최저기온", "minTa")
    c_max = pick("최고기온", "maxTa")
    c_stn = pick("지점", "stnId", exclude=("명",))
    if c_date is None or not any([c_avg, c_min, c_max]):
        raise ValueError("일시 또는 기온(평균/최저/최고) 컬럼을 찾지 못했어요.")

    out = pd.DataFrame({
        "date": pd.to_datetime(body[c_date], errors="coerce"),
        "stn": (body[c_stn].astype(str).str.strip() if c_stn else WEATHER_STN_DEFAULT),
        "avg_ta": body[c_avg].map(_wx_num) if c_avg else None,
        "min_ta": body[c_min].map(_wx_num) if c_min else None,
        "max_ta": body[c_max].map(_wx_num) if c_max else None,
    })
    out = out[out["date"].notna()]
    if out.empty:
        raise ValueError("날짜를 인식한 행이 없어요 — 파일 내용을 확인해 주세요.")
    out["date"] = out["date"].dt.strftime("%Y-%m-%d")
    return out.drop_duplicates(subset=["date"], keep="last").reset_index(drop=True)


def fetch_weather_kma(start_date, end_date, service_key, stn=WEATHER_STN_DEFAULT):
    """공공데이터포털 ASOS 일자료 API로 기간 기온을 받아 DF[date, stn, avg_ta, min_ta, max_ta] 반환.

    엔드포인트: /1360000/AsosDalyInfoService/getWthrDataList (dataCd=ASOS · dateCd=DAY).
    한 번에 최대 999행씩 페이징하며, 인증키 오류·응답 형식 오류는 ValueError로 알려준다.
    """
    import re
    import requests
    rows, page, PER = [], 1, 999
    while page <= 30:
        params = {"serviceKey": service_key, "numOfRows": PER, "pageNo": page,
                  "dataType": "JSON", "dataCd": "ASOS", "dateCd": "DAY",
                  "startDt": pd.to_datetime(start_date).strftime("%Y%m%d"),
                  "endDt": pd.to_datetime(end_date).strftime("%Y%m%d"), "stnIds": str(stn)}
        r = requests.get(KMA_ASOS_URL, params=params, timeout=25)
        r.raise_for_status()
        try:
            js = r.json()
        except Exception:
            # 인증키가 잘못되면 dataType=JSON이어도 XML 에러문서가 돌아온다 → 코드를 뽑아 한글 안내로 변환
            txt = r.text or ""
            m = re.search(r"<returnReasonCode>(\d+)</returnReasonCode>", txt) or \
                re.search(r"<errMsg>(.*?)</errMsg>", txt)
            hit = m.group(1) if m else ""
            if hit in _KMA_ERR:
                raise ValueError(f"[{hit}] {_KMA_ERR[hit]}")
            if "SERVICE_KEY_IS_NOT_REGISTERED" in txt or "SERVICE KEY IS NOT REGISTERED" in txt:
                raise ValueError(f"[30] {_KMA_ERR['30']}")
            raise ValueError("응답이 JSON이 아니에요(인증키 오류일 가능성이 커요): "
                             + txt[:200].replace("\n", " "))
        body = (js.get("response", {}) or {}).get("body", {}) or {}
        head = (js.get("response", {}) or {}).get("header", {}) or {}
        code = str(head.get("resultCode", "")).zfill(2) if str(head.get("resultCode", "")) else ""
        if code not in ("", "00"):
            raise ValueError(f"[{code}] " + _KMA_ERR.get(code, head.get("resultMsg", "알 수 없는 오류")))
        items = ((body.get("items") or {}).get("item")) or []
        if isinstance(items, dict):
            items = [items]
        rows.extend(items)
        if len(items) < PER:
            break
        page += 1
    if not rows:
        raise ValueError("받아온 데이터가 없어요 — 기간·지점번호를 확인해 주세요.")
    out = pd.DataFrame({
        "date": pd.to_datetime([r.get("tm") for r in rows], errors="coerce"),
        "stn": [str(r.get("stnId", stn)) for r in rows],
        "avg_ta": [_wx_num(r.get("avgTa")) for r in rows],
        "min_ta": [_wx_num(r.get("minTa")) for r in rows],
        "max_ta": [_wx_num(r.get("maxTa")) for r in rows],
    })
    out = out[out["date"].notna()]
    out["date"] = out["date"].dt.strftime("%Y-%m-%d")
    return out.drop_duplicates(subset=["date"], keep="last").reset_index(drop=True)


def merge_weather(new_df):
    """기온 일자료를 기존 데이터와 합쳐 저장(같은 날짜는 새 값으로 교체). 저장 후 총 행수 반환.

    전체 교체가 아니라 '병합'이라, 매년 새 기간만 올려도 과거 기온이 지워지지 않는다.
    (일자료라 10년치를 다 담아도 4천 행 수준이라 통째 재기록이 더 안전하고 빠르다)
    """
    old = load_weather(raw=True)
    both = new_df if (old is None or old.empty) else pd.concat([old, new_df], ignore_index=True)
    both = both.dropna(subset=["date"]).drop_duplicates(subset=["date"], keep="last")
    both = both.sort_values("date").reset_index(drop=True)
    eng = get_engine()
    with eng.begin() as conn:
        both.astype(object).where(both.notna(), None).to_sql(
            WEATHER_TABLE, conn, if_exists="replace", index=False)
    return len(both)


@st.cache_data(ttl=21600)
def load_weather(raw=False):
    """DB의 기온 일자료 DF[date, stn, avg_ta, min_ta, max_ta]. 없으면 빈 DF."""
    eng = get_engine()
    try:
        with eng.connect() as conn:
            exists = conn.exec_driver_sql(
                "SELECT 1 FROM information_schema.tables WHERE table_name=%s"
                if eng.dialect.name == "postgresql" else
                "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
                (WEATHER_TABLE,)).fetchone()
            if not exists:
                return pd.DataFrame()
            w = pd.read_sql(f'SELECT * FROM "{WEATHER_TABLE}"', conn)
    except Exception:
        return pd.DataFrame()
    if w.empty:
        return w
    for c in ("avg_ta", "min_ta", "max_ta"):
        if c in w.columns:
            w[c] = pd.to_numeric(w[c], errors="coerce")
    w["date"] = w["date"].astype(str).str.slice(0, 10)
    return w if raw else w.sort_values("date").reset_index(drop=True)


def weather_row_count():
    try:
        with get_engine().connect() as conn:
            return conn.exec_driver_sql(f'SELECT COUNT(*) FROM "{WEATHER_TABLE}"').scalar()
    except Exception:
        return 0


def weather_span():
    """적재된 기온 데이터의 (시작일, 종료일) 문자열. 없으면 (None, None)."""
    w = load_weather()
    if w is None or w.empty:
        return None, None
    return str(w["date"].min()), str(w["date"].max())


def weather_weekly(week_index):
    """일자료 → 주(월요일 시작) 평균 기온. week_index(주 시작일 목록)에 맞춰 정렬해 반환.

    반환: DataFrame(index=week_index, columns=[평균기온, 최저기온, 최고기온]). 데이터 없으면 빈 DF.
    """
    w = load_weather()
    if w is None or w.empty or len(week_index) == 0:
        return pd.DataFrame()
    dt = pd.to_datetime(w["date"], errors="coerce")
    ok = dt.notna()
    if not ok.any():
        return pd.DataFrame()
    wk = (dt[ok] - pd.to_timedelta(dt[ok].dt.weekday, unit="D")).dt.normalize()
    g = pd.DataFrame({"week": wk.values,
                      "평균기온": w.loc[ok, "avg_ta"].values if "avg_ta" in w.columns else np.nan,
                      "최저기온": w.loc[ok, "min_ta"].values if "min_ta" in w.columns else np.nan,
                      "최고기온": w.loc[ok, "max_ta"].values if "max_ta" in w.columns else np.nan})
    g = g.groupby("week")[WEATHER_KINDS].mean()
    return g.reindex(pd.DatetimeIndex(week_index))


def _wx_at(wxw, kind, i):
    """주간 기온표에서 i번째 주의 kind 기온 — 값이 없으면 None."""
    if wxw is None or wxw.empty or kind not in wxw.columns or i is None:
        return None
    try:
        v = float(wxw[kind].iloc[i])
    except Exception:
        return None
    return v if np.isfinite(v) else None


def _wx_fmt(v):
    return "–" if v is None else f"{v:.1f}℃"


def render_weather_admin():
    """사이드바(관리자) — 서울 기온 일자료 적재 UI. 업로드 / API 자동 수집 2가지."""
    n = weather_row_count()
    d0, d1 = weather_span()
    st.caption(f"🌡️ 서울 기온 일자료: 현재 **{n:,}일**"
               + (f" ({d0} ~ {d1})" if d0 else " — 추세분석 '기온 겹쳐보기'에 쓰여요"))
    wup = st.file_uploader("기온 데이터 업로드 (기상자료개방포털 ASOS 일자료 CSV/엑셀)",
                           type=["csv", "xlsx", "xls"], accept_multiple_files=False, key="wx_up")
    if wup is not None:
        if st.button("🌡️ 기온 데이터 적재(같은 날짜는 교체)", use_container_width=True, key="wx_apply"):
            try:
                total = merge_weather(read_weather_file(wup))
                load_weather.clear()
                st.success(f"기온 데이터 갱신 완료 ✅ 누적 {total:,}일")
            except Exception as ex:
                st.error(f"기온 데이터 오류: {ex}")
    with st.expander("🔗 기상청 API로 자동 수집 (인증키 필요)"):
        st.caption("공공데이터포털 → '기상청_지상(종관, ASOS) 일자료 조회서비스' 활용신청 후 "
                   "받은 **일반 인증키(Decoding)** 를 넣으면 기간을 지정해 바로 받아와요. "
                   "Streamlit Secrets에 `[kma] service_key = \"...\"` 로 넣어두면 매번 안 넣어도 돼요.")
        try:
            _sk_default = str(st.secrets.get("kma", {}).get("service_key", "") or "")
        except Exception:
            _sk_default = ""
        key = st.text_input("인증키", value=_sk_default, type="password", key="wx_key")
        cA, cB = st.columns(2)
        # 260811: 날짜 기본값도 서버 시간대(UTC) 영향을 받지 않도록 now_kst() 사용.
        d_from = cA.date_input("시작일", value=(now_kst() - pd.Timedelta(days=760)).date(), key="wx_from")
        d_to = cB.date_input("종료일", value=now_kst().date(), key="wx_to")
        stn = st.text_input("지점번호 (108=서울)", value=WEATHER_STN_DEFAULT, key="wx_stn")
        if st.button("🔗 기상청에서 받아오기", use_container_width=True, key="wx_fetch"):
            if not key.strip():
                st.error("인증키를 입력해 주세요.")
            elif d_to < d_from:
                st.error("종료일이 시작일보다 앞서요.")
            else:
                try:
                    with st.spinner("기상청에서 받아오는 중…"):
                        got = fetch_weather_kma(d_from, d_to, key.strip(), stn.strip() or WEATHER_STN_DEFAULT)
                        total = merge_weather(got)
                    load_weather.clear()
                    st.success(f"수집 완료 ✅ {len(got):,}일 수신 · 누적 {total:,}일")
                except Exception as ex:
                    st.error(f"수집 실패: {ex}")


# ==============================================================================
# 반품률 분석  ─ 품번별 판매수량 대비 반품수량 이상치 탐지 (2026-08-04 신설 · 중태님 지시)
# ==============================================================================
# 매출 로우데이터에는 반품이 별도 컬럼이 아니라 '판매수량이 음수인 행'으로 섞여 있다(ERP 원본 방식).
# 품번 단위로 판매(양수)·반품(음수)을 나눠 반품률(반품수량÷판매수량)을 구하고, 비교 기준(전체/중카테고리/
# 소카테고리/아이템코드) 평균 대비 이상치 판정 방식(배수/표준편차)을 화면에서 그때그때 바꿔가며 볼 수 있게 한다.
# 목적: 반품률이 유난히 높은 상품을 찾아 품질 문제·상품정보(사이즈/색상/사진/설명) 오류 여부를 점검하는 것.
RR_BASIS_OPTIONS = ["전체 상품 평균", "중카테고리(아이템그룹) 평균", "소카테고리 평균", "아이템코드 평균"]
RR_METHOD_OPTIONS = ["평균 대비 배수", "평균 + 표준편차"]
# 260804(중태님 지시): 매장코드 SD185(SD 쿠팡(그로스))는 유통채널 특이성 때문에 기계적인 판매·반품이
# 반복 발생해서 반품률 통계를 왜곡시킴 — 반품률 분석 전체(판매수량·반품수량·소계·이상치 판정 전부)에서
# 이 매장의 매출은 통째로 제외한다. 다른 화면(채널랭킹 등)의 CH_MERGE(쿠팡토탈=SD185+SD184 합산)와는
# 목적이 달라서 별도 상수로 관리 — CH_MERGE를 건드리면 안 됨.
RR_EXCLUDE_STORES = {"SD185": "SD 쿠팡(그로스)"}


def render_return_rate(df):
    """🔄 반품률 분석 — 기간 내 품번별 판매수량 대비 반품수량이 평균보다 훨씬 많은 상품(품질/상품정보
    오류 의심 후보)을 찾아내는 화면.

    비교 기준(전체 / 중카테고리 / 소카테고리 / 아이템코드)을 화면에서 바꿀 수 있게 한 이유: 아이템 종류마다
    원래 반품률 수준 자체가 다르다(예: 신발·정장은 사이즈 이슈로 반품률이 구조적으로 높을 수 있음) —
    "평균보다 훨씬 높다"의 '평균'을 같은 종류 상품끼리로 좁혀야 공정한 비교가 된다.
    """
    st.subheader("🔄 반품률 분석")
    if df is None or df.empty or "_판매일" not in df.columns or df["_판매일"].notna().sum() == 0:
        st.info("데이터를 먼저 적재하세요.")
        return
    d = df[df["_판매일"].notna()].copy()
    if "품번" not in d.columns or "판매수량" not in d.columns:
        st.info("품번·판매수량 컬럼이 없어 반품률을 계산할 수 없어요.")
        return

    # 260804(중태님 지시): SD185(SD 쿠팡(그로스)) 매장은 유통채널 특이성으로 기계적인 판매·반품이
    # 반복 발생해 반품률 통계를 왜곡시키므로, 반품률 분석에서는 이 매장의 매출을 통째로 제외한다.
    # 화면에서도 바로 보이게 눈에 띄는 안내를 띄운다(중태님 지시: "이 내용을 화면에도 표기해줘").
    rr_excl_n = 0
    if "매장코드" in d.columns:
        _store_code = d["매장코드"].astype(str).str.strip().str.upper()
        _excl_mask = _store_code.isin(RR_EXCLUDE_STORES.keys())
        rr_excl_n = int(_excl_mask.sum())
        d = d[~_excl_mask]

    st.caption("매출 로우데이터에서 **반품은 판매수량이 음수인 행**으로 잡혀요(ERP 원본 방식) — 이 화면은 "
               "품번별로 판매수량 대비 반품수량 비율(반품률)을 구해서, 비교 기준 평균보다 유난히 반품률이 "
               "높은 상품을 걸러내요. 품질 문제·상품정보(사이즈·색상·사진·설명) 오류를 의심해볼 후보를 "
               "찾는 용도예요.")
    _excl_names = ", ".join(f"{v}({k})" for k, v in RR_EXCLUDE_STORES.items())
    st.info(f"🚫 **{_excl_names}** 매장 매출은 이 화면(반품률 분석) 전체에서 제외하고 계산해요 — 유통채널 "
            f"특이성으로 기계적인 판매·반품이 반복 발생해 반품률 통계를 왜곡시키기 때문이에요"
            + (f" (적재된 데이터 전체 기준 제외 {rr_excl_n:,}행 — 아래 조회기간과 무관하게 미리 제외됨)."
               if rr_excl_n else "."))

    dmin, dmax = d["_판매일"].min().date(), d["_판매일"].max().date()
    default_start = max(pd.to_datetime(dmax) - pd.Timedelta(days=89), pd.to_datetime(dmin)).date()
    # ── 조건 폼 (2026-08-06): 조건 변경 중엔 계산 안 함, 🔍 조회 때 1번만 ──
    with st.form("rr_form"):
        rng = st.date_input("조회기간 (기본: 최근 90일)", value=(default_start, dmax),
                            min_value=dmin, max_value=dmax, key="rr_rng")
        # 공통 필터 (빈칸=전체) — 브랜드 · 시즌 · 중카테고리(아이템그룹)
        f1, f2, f3 = st.columns(3)
        brands = sorted(d["브랜드명"].dropna().astype(str).unique()) if "브랜드명" in d.columns else []
        seasons = sorted(d["시즌명"].dropna().astype(str).unique()) if "시즌명" in d.columns else []
        groups = [g for g in ITEMGROUP_ORDER if g in set(d.get("아이템그룹", pd.Series(dtype=str)).astype(str))]
        selb = f1.multiselect("브랜드", brands, default=[], placeholder="전체", key="rr_fb")
        sels = f2.multiselect("시즌", seasons, default=[], placeholder="전체", key="rr_fs")
        selg = f3.multiselect("중카테고리(아이템그룹)", groups, default=[], placeholder="전체", key="rr_fg")
        run = st.form_submit_button("🔍 조회", type="primary")
    if _need_search("rr_go", run):
        st.caption("🔍 조회하면 **품번별 반품률 상세표**가 나와요 — 컬럼: 품번·아이템명·품명·"
                   "아이템그룹·소카테고리·시즌명·판매수량·반품수량·**반품률**·비교기준평균·"
                   "그룹내상품수·이상치기준값·반품금액·이상치 여부.")
        return
    if not (isinstance(rng, (list, tuple)) and len(rng) == 2):
        st.info("기간(시작~끝)을 선택한 뒤 🔍 조회를 눌러 주세요.")
        return
    s, e = pd.to_datetime(rng[0]), pd.to_datetime(rng[1])
    if e < s:
        st.error("종료일이 시작일보다 앞서요. 기간을 다시 선택해 주세요.")
        return

    # 카테고리 파생(중카테고리는 load_db가 이미 만들어 둔 '아이템그룹' 그대로 사용, 소카테고리는
    # 추세분석과 동일한 아이템 마스터 기준 맵(_trend_cat_maps)을 그대로 재사용 — 단일 소스 유지)
    # ※ 2026-08-06: 조회 게이트 뒤로 이동 — 조회 전 화면 진입만으로는 이 파생 계산이 돌지 않게.
    mid_map, small_map, name_map = _trend_cat_maps()
    if "아이템" in d.columns:
        _ic = d["아이템"].astype(str).str.strip().str.upper()
    elif "품번" in d.columns:
        _ic = d["품번"].astype(str).str.strip().str.upper().str[1:3]
    else:
        _ic = pd.Series("", index=d.index, dtype="object")
    d["_소카"] = _ic.map(small_map).fillna(d.get("아이템그룹", "기타"))
    d["_아이템코드"] = _ic

    base = d[(d["_판매일"] >= s) & (d["_판매일"] <= e)]
    if selb and "브랜드명" in base.columns:
        base = base[base["브랜드명"].astype(str).isin(selb)]
    if sels and "시즌명" in base.columns:
        base = base[base["시즌명"].astype(str).isin(sels)]
    if selg and "아이템그룹" in base.columns:
        base = base[base["아이템그룹"].astype(str).isin(selg)]
    if base.empty:
        st.info("선택한 기간·조건에 매출 데이터가 없어요.")
        return

    qty = pd.to_numeric(base["판매수량"], errors="coerce").fillna(0.0)
    rev = pd.to_numeric(base["_매출액"], errors="coerce").fillna(0.0)
    tmp = pd.DataFrame({
        "품번": base["품번"].astype(str).str.strip(),
        "아이템명": base["아이템명"].astype(str) if "아이템명" in base.columns else "",
        # 260804(중태님 지시): 표에서 브랜드명 대신 품명(매출 로우데이터의 품번 옆 실제 상품명)을
        # 보여준다. 브랜드 자체는 위쪽 "브랜드" 필터로 이미 좁힐 수 있어 표 안에서는 굳이 다시 안
        # 보여줘도 되고, 어떤 상품인지 바로 알아볼 수 있는 품명이 더 유용하다는 판단.
        "품명": base["품명"].astype(str) if "품명" in base.columns else "",
        "아이템그룹": base["아이템그룹"].astype(str) if "아이템그룹" in base.columns else "기타",
        "소카테고리": base["_소카"].astype(str),
        "아이템코드": base["_아이템코드"].astype(str),
        "시즌명": base["시즌명"].astype(str) if "시즌명" in base.columns else "",
        "판매수량_gross": np.where(qty > 0, qty, 0.0),
        "반품수량": np.where(qty < 0, -qty, 0.0),
        "반품금액": np.where(rev < 0, -rev, 0.0),
    })
    agg = tmp.groupby("품번").agg(
        아이템명=("아이템명", "first"), 품명=("품명", "first"),
        아이템그룹=("아이템그룹", "first"), 소카테고리=("소카테고리", "first"),
        아이템코드=("아이템코드", "first"), 시즌명=("시즌명", "first"),
        판매수량=("판매수량_gross", "sum"), 반품수량=("반품수량", "sum"),
        반품금액=("반품금액", "sum")).reset_index()

    # 이번 기간 판매는 0인데 반품만 잡힌 품번(전기 판매분 반품 가능성) — 반품률 정의 불가라 별도 분리
    zero_sale = agg[(agg["판매수량"] <= 0) & (agg["반품수량"] > 0)]
    agg = agg[agg["판매수량"] > 0].copy()
    agg["반품률"] = agg["반품수량"] / agg["판매수량"]
    if agg.empty:
        st.info("선택한 조건에서 판매수량이 있는 품번이 없어요.")
        return

    # 노이즈 필터 — 최소 판매수량 (이 수량 미만은 반품률이 소수 케이스로 왜곡되기 쉬움)
    # 260804: 슬라이더 최대값을 데이터 95분위로 자동 제한하던 방식은, 상품별 판매수량이
    # 원래 10~18개 정도로 작은 경우(예: 11개 팔고 9개 반품=81.8%처럼 소수 케이스에 비율이
    # 확 튀는 구조) 정작 "20개, 50개 이상 팔린 상품만 놓고 봐도 여전히 심각한지" 확인하려는
    # 값을 슬라이더로 못 올리는 문제가 있었음 → 직접 숫자를 입력하는 방식으로 변경(중태님 요청).
    max_qty_data = int(agg["판매수량"].max()) if len(agg) else 10
    min_qty = st.number_input(
        "최소 판매수량 (이 수량 미만 상품은 분석에서 제외 — 20, 50처럼 원하는 값을 직접 입력해서 "
        "\"표본이 작아서 반품률이 튄 것\"인지 \"진짜 심각한 문제\"인지 확인해보세요)",
        min_value=0, max_value=max(max_qty_data, 1000), value=10, step=5, key="rr_minqty")
    min_qty = int(min_qty)
    excluded_n = int((agg["판매수량"] < min_qty).sum())
    pool = agg[agg["판매수량"] >= min_qty].copy()
    if pool.empty:
        st.info("최소 판매수량 조건을 만족하는 품번이 없어요. 기준을 낮춰 보세요.")
        return

    # 비교 기준 · 이상치 판정 방식 — 아이템 종류마다 반품률 수준 자체가 다를 수 있어 비교 기준을
    # 전체/중카테고리/소카테고리/아이템코드 중 골라서 "같은 종류 상품끼리" 평균을 낼 수 있게 한다.
    o1, o2, o3 = st.columns([1.3, 1.2, 1.6])
    basis = o1.selectbox("비교 기준 (반품률 '평균'을 어느 범위에서 낼지)", RR_BASIS_OPTIONS,
                         index=2, key="rr_basis",
                         help="아이템 종류마다 원래 반품률 수준이 달라요(예: 신발·정장은 사이즈 이슈로 "
                              "구조적으로 높을 수 있음) — 범위를 좁힐수록 '같은 종류 상품끼리' 공정하게 "
                              "비교하지만, 그룹 표본이 작아지면 평균이 흔들릴 수 있어요(표의 '그룹내 상품수' 참고).")
    method = o2.radio("이상치 판정 방식", RR_METHOD_OPTIONS, key="rr_method")
    if method == "평균 대비 배수":
        mult = o3.slider("평균의 몇 배 이상을 이상치로 볼지", min_value=1.2, max_value=5.0, value=2.0, step=0.1,
                         key="rr_mult")
    else:
        kstd = o3.slider("평균 + 표준편차 × 배수", min_value=0.5, max_value=4.0, value=2.0, step=0.5,
                         key="rr_kstd")

    _basis_col = {"전체 상품 평균": None, "중카테고리(아이템그룹) 평균": "아이템그룹",
                  "소카테고리 평균": "소카테고리", "아이템코드 평균": "아이템코드"}[basis]
    if _basis_col is None:
        avg_all = float(pool["반품률"].mean(skipna=True))
        std_all = float(pool["반품률"].std(skipna=True) or 0.0)
        pool["_avg"] = avg_all
        pool["_std"] = std_all
        pool["_그룹내상품수"] = len(pool)
    else:
        gstat = pool.groupby(_basis_col)["반품률"].agg(["mean", "std", "count"]).rename(
            columns={"mean": "_avg", "std": "_std", "count": "_그룹내상품수"})
        pool = pool.merge(gstat, left_on=_basis_col, right_index=True, how="left")
    pool["_std"] = pool["_std"].fillna(0.0)

    if method == "평균 대비 배수":
        pool["_기준값"] = pool["_avg"] * mult
    else:
        pool["_기준값"] = pool["_avg"] + pool["_std"] * kstd
    pool["이상치"] = pool["반품률"] >= pool["_기준값"]
    pool = pool.sort_values(["이상치", "반품률"], ascending=[False, False])

    n_out = int(pool["이상치"].sum())
    k1, k2, k3, k4 = st.columns(4)
    k1.metric("분석 대상 품번", f"{len(pool):,}개")
    k2.metric("전체 평균 반품률", f"{pool['반품률'].mean(skipna=True)*100:.1f}%")
    k3.metric("이상치 후보", f"{n_out:,}개")
    k4.metric("최소판매수량 미달 제외", f"{excluded_n:,}개")
    if len(zero_sale):
        st.caption(f"ℹ️ 이번 기간 판매는 0인데 반품만 발생한 품번 {len(zero_sale):,}개는 반품률 정의가 "
                   "안 돼(전기 판매분 반품 가능성) 위 분석에서 제외했어요.")

    disp = pd.DataFrame({
        "품번": pool["품번"], "아이템명": pool["아이템명"], "품명": pool["품명"],
        "아이템그룹": pool["아이템그룹"], "소카테고리": pool["소카테고리"], "시즌명": pool["시즌명"],
        "판매수량": pool["판매수량"].map(lambda v: f"{v:,.0f}"),
        "반품수량": pool["반품수량"].map(lambda v: f"{v:,.0f}"),
        "반품률": pool["반품률"].map(lambda v: f"{v*100:.1f}%" if pd.notnull(v) else "–"),
        "비교기준평균": pool["_avg"].map(lambda v: f"{v*100:.1f}%" if pd.notnull(v) else "–"),
        "그룹내상품수": pool["_그룹내상품수"].map(lambda v: f"{int(v):,}" if pd.notnull(v) else "–"),
        "이상치기준값": pool["_기준값"].map(lambda v: f"{v*100:.1f}%" if pd.notnull(v) else "–"),
        "반품금액(백만)": pool["반품금액"].map(lambda v: f"{v/1e6:,.2f}"),
        "이상치": pool["이상치"].map(lambda v: "⚠️ 의심" if v else ""),
    })
    disp = disp.set_index("품번")

    # ── 표 맨 위 소계 (2026-08-04 저녁 6차 수정 · 중태님 확인)
    # 처음 버전은 비교 기준이 "전체 상품 평균"이 아닐 때 그룹별 소계만 보여줬는데, 그러면 진짜
    # "전체 합계"(전 카테고리 총계) 숫자가 아예 안 보이는 문제가 있었고("전체 평균은 아예 숫자가
    # 안 들어가 있는데"), 게다가 이 앱 다른 표들의 공통 규칙(룰6: 표의 첫 행=G.TOTAL/합계는 항상
    # 노란 강조, _TBL_CSS의 `tbody tr:first-child` 규칙)과 충돌해서 그룹 소계 중 우연히 반품률이
    # 가장 높아 맨 위로 온 행(예: 수트류)만 노란색으로 보이고 나머지는 파란색으로 보이는 것처럼
    # 되어 있었음("슈트류만 컬러가 다르게 표현되고 있고") — 사실은 색이 "달라진" 게 아니라, 매 순간
    # 맨 위(1등)에 오는 그룹이 이 공통 규칙 때문에 강제로 노란색이 되는 것이었음.
    # → 지금은 비교 기준과 무관하게 "■ 전체 합계"(전체 필터 결과 총계) 행을 항상 맨 위 1줄로 고정
    #   해서 이 앱의 공통 규칙(첫 행=노란 G.TOTAL)과 그대로 맞춘다. 비교 기준이 중카테고리/소카테고리/
    #   아이템코드처럼 "그룹별로 나눠 보기"인 경우에는, 그 아래에 그룹별 소계(파란 배경, 반품률 높은
    #   순 정렬)를 이어서 보여준다.
    tot_sale = float(pool["판매수량"].sum())
    tot_ret = float(pool["반품수량"].sum())
    tot_amt = float(pool["반품금액"].sum())
    tot_rate = (tot_ret / tot_sale) if tot_sale else np.nan
    if _basis_col is None:
        tot_avg = float(pool["_avg"].iloc[0]) if len(pool) else np.nan
        tot_thr = float(pool["_기준값"].iloc[0]) if len(pool) else np.nan
    else:
        # 그룹별로 평균·기준값이 제각각이라 "전체 합계" 행 하나로는 단일 기준값을 못 정함 —
        # 비교기준평균은 전체 품번 반품률의 단순평균(참고용)으로 보여주고, 이상치기준값은 그룹마다
        # 달라서 "–"(해당 없음)로 비워 혼동을 방지한다.
        tot_avg = float(pool["반품률"].mean(skipna=True))
        tot_thr = np.nan
    tot_out = int(pool["이상치"].sum())
    sub_rows = [{
        "품번": "■ 전체 합계", "아이템명": f"전체 ({len(pool):,}개 품번)",
        "품명": "", "아이템그룹": "", "소카테고리": "", "시즌명": "",
        "판매수량": f"{tot_sale:,.0f}", "반품수량": f"{tot_ret:,.0f}",
        "반품률": f"{tot_rate*100:.1f}%" if pd.notnull(tot_rate) else "–",
        "비교기준평균": f"{tot_avg*100:.1f}%" if pd.notnull(tot_avg) else "–",
        "그룹내상품수": f"{len(pool):,}",
        "이상치기준값": f"{tot_thr*100:.1f}%" if pd.notnull(tot_thr) else "–",
        "반품금액(백만)": f"{tot_amt/1e6:,.2f}",
        "이상치": f"{tot_out:,}건 의심" if tot_out else "0건",
    }]
    if _basis_col is not None:
        grp_rows = []
        for name, gdf in pool.groupby(_basis_col):
            g_sale = float(gdf["판매수량"].sum())
            g_ret = float(gdf["반품수량"].sum())
            g_amt = float(gdf["반품금액"].sum())
            g_rate = (g_ret / g_sale) if g_sale else np.nan
            avg_v = float(gdf["_avg"].iloc[0]) if len(gdf) else np.nan
            thr_v = float(gdf["_기준값"].iloc[0]) if len(gdf) else np.nan
            n_out_g = int(gdf["이상치"].sum())
            row = {
                "품번": f"■ 소계 · {name}", "아이템명": f"{name} ({len(gdf):,}개 품번)",
                "품명": "", "아이템그룹": "", "소카테고리": "", "시즌명": "",
                "판매수량": f"{g_sale:,.0f}", "반품수량": f"{g_ret:,.0f}",
                "반품률": f"{g_rate*100:.1f}%" if pd.notnull(g_rate) else "–",
                "비교기준평균": f"{avg_v*100:.1f}%" if pd.notnull(avg_v) else "–",
                "그룹내상품수": f"{len(gdf):,}",
                "이상치기준값": f"{thr_v*100:.1f}%" if pd.notnull(thr_v) else "–",
                "반품금액(백만)": f"{g_amt/1e6:,.2f}",
                "이상치": f"{n_out_g:,}건 의심" if n_out_g else "0건",
                "_sort": g_rate if pd.notnull(g_rate) else -1.0,
            }
            if _basis_col == "아이템그룹":
                row["아이템그룹"] = str(name)
            elif _basis_col == "소카테고리":
                row["소카테고리"] = str(name)
            grp_rows.append(row)
        grp_rows.sort(key=lambda r: r["_sort"], reverse=True)
        for r in grp_rows:
            del r["_sort"]
        sub_rows.extend(grp_rows)
    sub_df = pd.DataFrame(sub_rows).set_index("품번")
    disp_full = pd.concat([sub_df, disp])

    h1, h2 = st.columns([5, 1])
    h1.markdown(f"**품번별 반품률 상세**  <span style='color:#888;font-size:0.8rem;'>"
                f"({s.date()} → {e.date()} · 비교기준={basis} · 판정방식={method})</span>"
                "<span style='float:right;color:#888;font-weight:400;font-size:0.78rem;'>"
                "[반품금액: 백만원 · 반품률=반품수량÷판매수량]</span>", unsafe_allow_html=True)
    h2.download_button("⬇ 엑셀", table_excel_bytes(disp_full, "반품률분석"),
                       file_name=f"반품률분석_{e.date()}.xlsx", mime=XLSX_MIME,
                       key="rr_dl", use_container_width=True)

    _out_flag = pool.set_index("품번")["이상치"]

    def _hl(row):
        name = str(row.name)
        if name == "■ 전체 합계":
            # 260804 저녁 6차: 맨 첫 행 = 전체 합계는 이 앱 공통 규칙(룰6: 표의 첫 행=G.TOTAL은
            # 항상 노란 강조, _TBL_CSS의 tr:first-child 규칙)에 맞춰 노란 배경으로 통일한다.
            # 아래 인라인 스타일은 참고용이고, 실제 화면 색은 _TBL_CSS의 !important 규칙이 최종
            # 적용됨(다른 표들과 똑같은 "첫 행=노란 총계" 느낌을 그대로 가져오기 위함).
            return ["background-color:#fff2b8;font-weight:700;border-bottom:2px solid #29508c" for _ in row]
        if name.startswith("■"):
            # 그룹별 소계(예: ■ 소계 · 수트류) = 파란 배경 + 아래쪽 굵은 경계선. 바로 밑에 개별
            # 품번(빨간 이상치 행 포함)이 이어지다 보니 "소계가 여러 줄인가?"로 오해될 수 있어
            # (중태님 확인), 소계 블록과 상세 목록을 시각적으로 분리하는 경계선을 추가.
            return ["background-color:#e3ecf7;font-weight:700;border-bottom:2px solid #29508c" for _ in row]
        flag = bool(_out_flag.get(row.name, False))
        return ["background-color:#ffe3e3;font-weight:600" if flag else "" for _ in row]

    sty = disp_full.style.apply(_hl, axis=1).set_properties(**{"text-align": "right"})
    render_styled_table(sty)
    n_grp = len(sub_df) - 1
    st.caption(f"※ 맨 위 **노란 행(■ 전체 합계)**은 지금 필터 조건 전체를 합산한 총계예요(이 앱 다른 표들과 "
               "동일하게 첫 행=노란 총계). "
               + (f"그 아래 **파란 행({n_grp}줄)**은 지금 선택한 비교 기준(**{basis}**)별 소계이고, "
                  "그룹 반품률이 높은 순으로 정렬돼요. " if n_grp else
                  "지금은 비교 기준이 '전체 상품 평균'이라 그룹별 소계 없이 전체 총계 1줄만 나와요. ")
               + "판매수량·반품수량은 합산 기준 반품률(볼륨가중)이에요. 노란·파란 선 아래부터는 "
               "**개별 품번** 목록이고, 그중 빨간 배경 행은 총계·소계와 별개로 그 품번 하나하나가 "
               "이상치로 잡혔다는 표시예요(총계·소계 줄이 아니라 낱개 상품 경고입니다). 반품률 = 해당 "
               "기간 반품수량 ÷ 판매수량(양수). ⚠️ 의심 = 선택한 비교 기준·판정 방식에서 이상치로 잡힌 "
               "품번 — 상품 품질, 사이즈/색상 등 상품정보 표기 오류, 사진·설명 오인 소지 등을 우선적으로 "
               "점검해 볼 후보예요. 그룹내상품수가 적으면(예: 5개 미만) 평균·표준편차가 표본 부족으로 "
               "흔들릴 수 있으니 참고만 하세요. 최소 판매수량 미만 품번은 위 입력란에 20, 50처럼 직접 값을 "
               "넣어 제외 기준을 조정해서 노이즈를 줄일 수 있어요.")

    # 산점도: 판매수량 vs 반품률 (버블 크기=반품수량)
    st.markdown("##### 📍 판매수량 대비 반품률 분포")
    sc = pool.copy()
    sc["_size"] = sc["반품수량"].clip(lower=1)
    normal = sc[~sc["이상치"]]
    outl = sc[sc["이상치"]]
    fig = go.Figure()
    fig.add_scatter(x=normal["판매수량"], y=normal["반품률"] * 100, mode="markers",
                    name="정상 범위", text=normal["품번"] + " · " + normal["아이템명"],
                    marker=dict(size=(normal["_size"] ** 0.5) * 3 + 4, color="#8fb3d9", opacity=0.6),
                    hovertemplate="%{text}<br>판매수량 %{x:,.0f} · 반품률 %{y:.1f}%<extra></extra>")
    fig.add_scatter(x=outl["판매수량"], y=outl["반품률"] * 100, mode="markers",
                    name="⚠️ 이상치 후보", text=outl["품번"] + " · " + outl["아이템명"],
                    marker=dict(size=(outl["_size"] ** 0.5) * 3 + 4, color="#c62828", opacity=0.85,
                                line=dict(color="#7a0000", width=1)),
                    hovertemplate="%{text}<br>판매수량 %{x:,.0f} · 반품률 %{y:.1f}%<extra></extra>")
    fig.update_layout(height=420, margin=dict(t=10, b=0, l=0, r=0),
                      xaxis_title="판매수량", yaxis_title="반품률(%)",
                      legend=dict(orientation="h", y=1.1))
    st.plotly_chart(fig, use_container_width=True)
    st.caption("※ 점 크기 = 반품수량. 오른쪽 위(판매도 많은데 반품률도 높음)일수록 실제 영향(반품 처리비용·"
               "재고 손실)이 크니 우선순위로 확인하세요.")


# ==============================================================================
# 🧩 SET/단품 판매 분석 (수트·셋업) — 260804 신규
# 설계 문서: claude/수트_세트판매가능성_사이즈매칭_설계검토.md (부속 작업, v3 + 3-1절 + 6절)
# ==============================================================================
_SUITSET_LOOKBACK_DAYS = 60   # 3-1절: 그룹 정체성 판정용 룩업 윈도우(조회 시작일 이전). 성능 부담 크면 30으로.


def _suitset_classify(j, p):
    """세트그룹/단품그룹/제외 판정 — 순자켓·순하의 둘 다 0이 아니면 세트(부호 무관, 260804 실데이터
    검증 중 확장: 반품만 있어 순수량이 음수인 그룹도 정확히 분류하려면 '>0'이 아니라 '!=0' 기준이어야 함)."""
    if j != 0 and p != 0:
        return "세트그룹"
    elif j != 0 or p != 0:
        return "단품그룹"
    return "제외"


def render_suitset(df):
    """🧩 SET/단품 판매 분석 — 수트(SJ/SL)·셋업(EJ/EP)이 매장별로 몇 피스가 '세트로' 팔렸고 몇 피스가
    '단품으로' 팔렸는지 집계. 판정 단위 = 주문번호+SET품번(품번 10자리 중 아이템 자리를 SJ/EJ 대표코드로,
    라인·패턴·색상 3자리를 'SET'으로 치환한 공식 조인키 — set-product-code-generator 스킬/STO
    sto_config.json 규칙) 그룹, 판매·반품은 그룹 순수량(net)으로 합쳐서 계산(반품이 원 주문의 세트/단품
    소속을 그대로 물려받음). SD163(SD댄블)은 세트로 사도 라인마다 다른(인접) 주문번호가 찍히는 채널
    특성이 있어 '인접 주문번호+SET품번'으로 별도 그룹핑. 그룹 성격(세트/단품) 판정은 조회 시작일 이전
    최대 60일까지 넓혀서 봐서, 반품의 원 판매가 화면 기간 밖에 있어도 놓치지 않게 하되, 화면에 실제
    더해지는 실적은 조회기간 '안'에서 발생한 판매·반품 델타만 반영(이중집계 방지, 설계문서 3-1절).
    """
    st.subheader("🧩 SET/단품 판매 분석 (수트·셋업)")
    if df is None or df.empty or "_판매일" not in df.columns or df["_판매일"].notna().sum() == 0:
        st.info("데이터를 먼저 적재하세요.")
        return
    need_cols = {"아이템", "매장코드", "매장명", "판매수량", "_매출액", "_판매일", "품번"}
    if not need_cols.issubset(df.columns):
        st.info("이 리포트에 필요한 컬럼(아이템·매장코드·품번·판매수량 등)이 없어요.")
        return
    if not {"순번", "주문번호"}.issubset(df.columns):
        st.warning("이 화면은 순번·주문번호 컬럼이 있어야 정확히 계산돼요. 과거에 적재한 원본 파일에 "
                   "이 컬럼이 없으면(구버전 업로드분) 최신 원본 파일로 다시 적재해 주세요.")
        return

    st.caption(
        "슈트(SJ/SL)·셋업(EJ/EP)이 **같은 주문번호 + SET품번**(품번 10자리 중 아이템 자리를 SJ/EJ "
        "대표코드로, 라인·패턴·색상 3자리를 'SET'으로 치환한 공식 조인키 — 예: SSJVB01CSK·SSLVB01CSK "
        "→ 둘 다 SSJVB01SET)**으로 묶여 팔렸으면 "
        "'세트로판매', 한쪽만 팔렸으면 '단품판매'로 집계해요. 판매·반품은 그룹 단위 순수량(net)으로 "
        "합쳐서 계산하고(반품은 원래 세트/단품 소속을 그대로 물려받음), 전량 반품된 그룹은 세트·단품 "
        "어디에도 안 잡히고 조용히 빠져요. SD163(SD댄블)은 세트로 사도 라인마다 다른(인접) 주문번호가 "
        "찍히는 채널 특성이 있어 '인접 주문번호+SET품번'으로 따로 묶어요. 그룹이 세트인지 단품인지 "
        f"판정할 땐 조회 시작일 이전 최대 {_SUITSET_LOOKBACK_DAYS}일까지 넓혀서 봐서(성능상 제한 — 실무상 "
        "반품은 구매 후 1개월을 잘 안 넘긴다는 전제), 반품의 원 판매가 화면 기간 밖에 있어도 정확히 "
        "판정해요. 다만 화면에 더해지는 실적 자체는 **조회기간 안에서 실제로 발생한 판매·반품만** — "
        "그래서 '이번 기간엔 반품만 있었다' 같은 경우 세트로판매·단품판매가 마이너스로 보일 수 있는데, "
        "이건 오류가 아니라 그 기간의 순감소를 있는 그대로 보여주는 정상적인 결과예요."
    )

    d = df[df["아이템"].astype(str).str.strip().str.upper().isin(["SJ", "SL", "EJ", "EP"])].copy()
    if d.empty:
        st.info("슈트/셋업(SJ·SL·EJ·EP) 판매 데이터가 없어요.")
        return

    dmin, dmax = d["_판매일"].min().date(), d["_판매일"].max().date()
    default_start = max(pd.to_datetime(dmax) - pd.Timedelta(days=6), pd.to_datetime(dmin)).date()
    # ── 조건 폼 (2026-08-06): 조건 변경 중엔 계산 안 함, 🔍 조회 때 1번만 ──
    with st.form("ss_form"):
        rng = st.date_input("조회기간 (기본: 최근 7일)", value=(default_start, dmax),
                            min_value=dmin, max_value=dmax, key="ss_rng")
        # 공통룰10: 브랜드·시즌 필터(빈칸=전체). 디자인키에 브랜드·시즌이 이미 포함되므로 매칭 조건과
        # 충돌 없이 미리 좁혀도 안전 — 세트를 이루는 자켓·하의는 항상 브랜드·시즌이 같기 때문.
        f1, f2 = st.columns(2)
        brands = sorted(d["브랜드명"].dropna().astype(str).unique()) if "브랜드명" in d.columns else []
        seasons = sorted(d["시즌명"].dropna().astype(str).unique()) if "시즌명" in d.columns else []
        selb = f1.multiselect("브랜드", brands, default=[], placeholder="전체", key="ss_fb")
        sels = f2.multiselect("시즌", seasons, default=[], placeholder="전체", key="ss_fs")
        run = st.form_submit_button("🔍 조회", type="primary")
    if _need_search("ss_go", run):
        st.caption("🔍 조회하면 **매장별 SET 판매 분석표**(자켓·팬츠 판매량, 세트로판매/단품판매, "
                   "세트매출/단품매출, 세트비중 등)와 그 아래 **특정매장 SET 품번별 분석**이 나와요.")
        return
    if not (isinstance(rng, (list, tuple)) and len(rng) == 2):
        st.info("기간(시작~끝)을 선택한 뒤 🔍 조회를 눌러 주세요.")
        return
    s, e = pd.to_datetime(rng[0]), pd.to_datetime(rng[1])
    if e < s:
        st.error("종료일이 시작일보다 앞서요. 기간을 다시 선택해 주세요.")
        return

    lb_start = s - pd.Timedelta(days=_SUITSET_LOOKBACK_DAYS)
    uni = d[(d["_판매일"] >= lb_start) & (d["_판매일"] <= e)].copy()
    if selb and "브랜드명" in uni.columns:
        uni = uni[uni["브랜드명"].astype(str).isin(selb)]
    if sels and "시즌명" in uni.columns:
        uni = uni[uni["시즌명"].astype(str).isin(sels)]
    if uni.empty:
        st.info("선택한 조건에 슈트/셋업 데이터가 없어요.")
        return

    # ── SET품번(공식 조인키, set-product-code-generator 스킬/STO sto_config.json 규칙) & 상하의 구분
    # 260804 심야: 처음엔 브랜드+연도+시즌+순번을 이어붙인 임시 "_디자인키"를 썼으나, 재고 마스터의
    # 진짜 SET품번 생성 규칙(품번 10자리 중 아이템 자리를 그룹 대표코드로, 라인·패턴·색상 3자리를
    # 'SET' 리터럴로 치환 — 예: SSJVB01CSK/SSLVB01CSK → 둘 다 SSJVB01SET)을 확인해서 그대로 적용.
    # 회사 process_260731.py K/L/M 로직과 동일 결과(검증 완료). 이게 이제 진짜 "SET 품번"이라 표2에서
    # 별도 대표품번 추정 없이 이 값 자체를 그대로 라벨로 쓸 수 있다.
    item_u = uni["아이템"].astype(str).str.strip().str.upper()
    uni["_상하의"] = item_u.map({"SJ": "상의", "EJ": "상의", "SL": "하의", "EP": "하의"})
    _pn = uni["품번"].astype(str).str.strip().str.upper()
    _item_repr = _pn.str.slice(1, 3).map({"SJ": "SJ", "SL": "SJ", "EJ": "EJ", "EP": "EJ"})
    _pn_ok = (_pn.str.len() == 10) & _item_repr.notna()
    _set_official = _pn.str.slice(0, 1) + _item_repr.fillna("") + _pn.str.slice(3, 7) + "SET"
    # 폴백(품번이 10자리가 아닌 이례적 데이터일 때만): 브랜드명+연도+시즌명+순번 기반 대체 키.
    # 그룹핑은 그대로 되지만 'SET품번'다운 표기는 아니므로 화면엔 "FB-" 접두어를 붙여 구분한다.
    _brand = uni["브랜드명"].astype(str) if "브랜드명" in uni.columns else ""
    _year = uni["연도"].astype(str) if "연도" in uni.columns else ""
    _season = uni["시즌명"].astype(str) if "시즌명" in uni.columns else ""
    _seq = pd.to_numeric(uni["순번"], errors="coerce").fillna(-1).astype(int).astype(str)
    _fallback_key = "FB-" + _brand + "|" + _year + "|" + _season + "|" + _seq
    uni["_디자인키"] = np.where(_pn_ok, _set_official, _fallback_key)
    uni["_수량"] = pd.to_numeric(uni["판매수량"], errors="coerce").fillna(0.0)
    uni["_매출"] = pd.to_numeric(uni["_매출액"], errors="coerce").fillna(0.0)
    uni["_주문번호_num"] = pd.to_numeric(uni["주문번호"], errors="coerce")

    store_code = uni["매장코드"].astype(str).str.strip().str.upper()
    is_sd163 = store_code == "SD163"

    # 일반 매장: 매장코드+주문번호+디자인키 그대로 그룹키.
    normal_key = store_code + "‖" + uni["주문번호"].astype(str) + "‖" + uni["_디자인키"]

    # SD163(설계문서 6절): 정렬 후 벡터화 diff로 "인접 주문번호 + 디자인키 일치" 클러스터링.
    grp_id = pd.Series(index=uni.index, dtype="object")
    grp_id.loc[~is_sd163] = normal_key.loc[~is_sd163]
    sd_idx = uni.index[is_sd163]
    if len(sd_idx):
        sub = uni.loc[sd_idx].sort_values(["매장코드", "_주문번호_num"])
        prev_num = sub["_주문번호_num"].shift()
        prev_key = sub["_디자인키"].shift()
        same_as_prev = ((sub["_주문번호_num"] - prev_num == 1) & (sub["_디자인키"] == prev_key)
                        & sub["_주문번호_num"].notna() & prev_num.notna())
        cluster_id = (~same_as_prev).cumsum()
        grp_id.loc[sub.index] = ("SD163‖G" + cluster_id.astype(str)).values
    uni["_그룹ID"] = grp_id

    # ── ① 그룹 정체성(넓은 룩업 윈도우 전체) — 세트그룹/단품그룹 판정 ──────────
    ident = uni.groupby("_그룹ID").apply(lambda g: pd.Series({
        "_순상의": g.loc[g["_상하의"] == "상의", "_수량"].sum(),
        "_순하의": g.loc[g["_상하의"] == "하의", "_수량"].sum(),
    }))
    ident["_분류"] = [ _suitset_classify(j, p) for j, p in zip(ident["_순상의"], ident["_순하의"]) ]
    uni = uni.merge(ident[["_분류"]], left_on="_그룹ID", right_index=True, how="left")

    # ── ② 조회기간 "안"에서 발생한 델타만 화면 실적에 반영 (3-1절, 이중집계 방지) ──
    per = uni[(uni["_판매일"] >= s) & (uni["_판매일"] <= e)].copy()
    if per.empty:
        st.info("선택한 조회기간 안에 슈트/셋업 판매·반품이 없어요.")
        return

    g = per.groupby(["매장코드", "매장명", "_그룹ID", "_분류"]).apply(lambda x: pd.Series({
        "_상의델타": x.loc[x["_상하의"] == "상의", "_수량"].sum(),
        "_하의델타": x.loc[x["_상하의"] == "하의", "_수량"].sum(),
        "_매출델타": x["_매출"].sum(),
    })).reset_index()

    def _rowmetrics(r):
        j, p, cls = r["_상의델타"], r["_하의델타"], r["_분류"]
        out = {"세트수": 0.0, "팬츠추가": 0.0, "자켓추가": 0.0, "세트로판매": 0.0, "단품판매": 0.0,
               "세트매출": 0.0, "단품매출": 0.0, "_세트상의": 0.0, "_세트하의": 0.0}
        if cls == "세트그룹":
            out["세트수"] = (np.sign(j) * min(abs(j), abs(p))) if (j != 0 and p != 0) else 0.0
            out["팬츠추가"] = max(0.0, p - j) if abs(p) > abs(j) else 0.0
            out["자켓추가"] = max(0.0, j - p) if abs(j) > abs(p) else 0.0
            out["세트로판매"] = j + p
            out["세트매출"] = r["_매출델타"]
            out["_세트상의"], out["_세트하의"] = j, p
        elif cls == "단품그룹":
            # 260804: 그룹 성격(세트/단품) 판정은 넓은 룩업 윈도우 기준이라, 조회기간(period)만 보면
            # '단품그룹'인데도 그 기간 안에 상의·하의가 둘 다 델타를 갖는 경우가 생길 수 있음
            # (예: 자켓은 룩업기간 내 다른 시점에 팔려 순상의=0이라 단품 판정, 팬츠는 기간 중 반품되어
            # 하의델타가 발생한 케이스). 이때 "j if j!=0 else p"로 하나만 집계하면 팬츠판매량 등
            # 기간총판매엔 반영되는데 단품판매엔 안 잡혀 세트비중/단품비중 합이 100%를 벗어나는
            # 버그가 있었음 — j+p로 둘 다 반영해서 기간총판매(세트로판매+단품판매)와 항상 일치시킴.
            out["단품판매"] = j + p
            out["단품매출"] = r["_매출델타"]
        return pd.Series(out)

    metrics = g.apply(_rowmetrics, axis=1)
    g = pd.concat([g, metrics], axis=1)
    g["자켓판매량"] = g["_상의델타"]
    g["팬츠판매량"] = g["_하의델타"]

    store = g.groupby(["매장코드", "매장명"], observed=True).agg(
        자켓판매량=("자켓판매량", "sum"), 팬츠판매량=("팬츠판매량", "sum"),
        세트수=("세트수", "sum"), 팬츠추가=("팬츠추가", "sum"), 자켓추가=("자켓추가", "sum"),
        세트로판매=("세트로판매", "sum"), 단품판매=("단품판매", "sum"),
        세트매출=("세트매출", "sum"), 단품매출=("단품매출", "sum"),
        세트그룹상의=("_세트상의", "sum"), 세트그룹하의=("_세트하의", "sum"),
    ).reset_index()
    store["매장코드"] = store["매장코드"].astype(str)
    store["매장명"] = store["매장명"].astype(str)
    store = store[(store["자켓판매량"] != 0) | (store["팬츠판매량"] != 0)]  # 조회기간에 실적 없는 매장은 표에서 생략
    if store.empty:
        st.info("선택한 조회기간·조건에 슈트/셋업 실적이 있는 매장이 없어요.")
        return
    store["기간총판매"] = store["자켓판매량"] + store["팬츠판매량"]
    store["기간매출"] = store["세트매출"] + store["단품매출"]
    _tot = store["기간총판매"].replace(0, np.nan)
    store["세트비중"] = store["세트로판매"] / _tot * 100
    store["단품비중"] = store["단품판매"] / _tot * 100
    _denom = store["세트그룹상의"].replace(0, np.nan)
    store["자켓팬츠비율"] = store["세트그룹하의"] / _denom
    store = store.sort_values("매장코드").reset_index(drop=True)

    # ── 화면/엑셀 표시용 DataFrame — 룰1(백만원)·룰6(첫행 총계 노랑)·룰8/룰11/룰13 ─────
    def _int(v):
        return "–" if pd.isna(v) else f"{v:,.0f}"

    def _money_mm(v):
        return "–" if pd.isna(v) else f"{v/1e6:,.2f}"

    def _pct(v):
        return "–" if pd.isna(v) else f"{v:,.1f}%"

    def _ratio(v):
        return "–" if pd.isna(v) else f"{v:,.2f}"

    cols = [("", "매장명"),
            ("기간판매량", "자켓판매량"), ("기간판매량", "팬츠판매량"),
            ("단품 기준 분석", "기간총판매"), ("단품 기준 분석", "기간매출"),
            ("단품 기준 분석", "세트로판매"), ("단품 기준 분석", "단품판매"),
            ("단품 기준 분석", "세트매출"), ("단품 기준 분석", "단품매출"),
            ("단품 기준 분석", "세트비중"), ("단품 기준 분석", "단품비중"),
            ("SET 분석", "세트수"), ("SET 분석", "팬츠추가"), ("SET 분석", "자켓추가"),
            ("SET 분석", "자켓:팬츠비율")]

    def _fmt_row(r):
        return [r["매장명"], _int(r["자켓판매량"]), _int(r["팬츠판매량"]),
                _int(r["기간총판매"]), _money_mm(r["기간매출"]),
                _int(r["세트로판매"]), _int(r["단품판매"]),
                _money_mm(r["세트매출"]), _money_mm(r["단품매출"]),
                _pct(r["세트비중"]), _pct(r["단품비중"]),
                _int(r["세트수"]), _int(r["팬츠추가"]), _int(r["자켓추가"]),
                _ratio(r["자켓팬츠비율"])]

    tot_row = {
        "매장명": f"전체 ({len(store):,}개 매장)",
        "자켓판매량": store["자켓판매량"].sum(), "팬츠판매량": store["팬츠판매량"].sum(),
        "세트수": store["세트수"].sum(), "팬츠추가": store["팬츠추가"].sum(),
        "자켓추가": store["자켓추가"].sum(),
        "세트로판매": store["세트로판매"].sum(), "단품판매": store["단품판매"].sum(),
        "세트매출": store["세트매출"].sum(), "단품매출": store["단품매출"].sum(),
        "세트그룹상의": store["세트그룹상의"].sum(), "세트그룹하의": store["세트그룹하의"].sum(),
    }
    tot_row["기간총판매"] = tot_row["자켓판매량"] + tot_row["팬츠판매량"]
    tot_row["기간매출"] = tot_row["세트매출"] + tot_row["단품매출"]
    _tt = tot_row["기간총판매"] if tot_row["기간총판매"] else np.nan
    tot_row["세트비중"] = tot_row["세트로판매"] / _tt * 100 if _tt else np.nan
    tot_row["단품비중"] = tot_row["단품판매"] / _tt * 100 if _tt else np.nan
    _td = tot_row["세트그룹상의"] if tot_row["세트그룹상의"] else np.nan
    tot_row["자켓팬츠비율"] = tot_row["세트그룹하의"] / _td if _td else np.nan

    rows = [_fmt_row(pd.Series(tot_row))] + [_fmt_row(r) for _, r in store.iterrows()]
    idx = ["■ 전체 합계"] + list(store["매장코드"])
    disp = pd.DataFrame(rows, index=idx, columns=pd.MultiIndex.from_tuples(cols))

    h1, h2 = st.columns([5, 1])
    h1.markdown(f"**매장별 SET 판매 분석**{_NOTE_FLOAT}", unsafe_allow_html=True)
    h2.download_button("⬇ 엑셀", table_excel_bytes(disp, "SET판매분석"),
                       file_name=f"SET판매분석_{e.date()}.xlsx", mime=XLSX_MIME,
                       key="ss_dl", use_container_width=True)

    def _hl(row):
        if str(row.name) == "■ 전체 합계":
            return ["background-color:#fff2b8;font-weight:700" for _ in row]
        return ["" for _ in row]

    sty = disp.style.apply(_hl, axis=1).set_properties(**{"text-align": "right"})
    sty = block_border(block_border(block_border(sty, 1), 3), 11)   # 룰12 활용: 매장명|기간판매량|단품기준분석|SET분석 블록 경계선
    render_styled_table(sty)
    st.caption("※ 자켓추가 컬럼(항상 0에 가까움)은 세트그룹인데 팬츠보다 자켓이 더 남는 이례 케이스 "
               "발견용으로 항상 표시해요. 자켓:팬츠비율 = 세트그룹 순하의수량 ÷ 세트그룹 순자켓수량 "
               "(정상적인 상의1:하의2 구성이면 2.00 근처).")

    # ══════════════════════════════════════════════════════════════════════
    # 특정매장 SET 품번별 분석 — 260804 심야 추가
    # ══════════════════════════════════════════════════════════════════════
    st.divider()
    st.markdown(f"**특정매장 SET 품번별 분석**{_NOTE_FLOAT}", unsafe_allow_html=True)
    st.caption(
        "SET 품번 = 품번 10자리 중 아이템 자리를 세트 대표코드(SJ 또는 EJ)로, 라인·패턴·색상 3자리를 "
        "'SET'으로 치환한 값이에요(예: SSJVB01CSK·SSLVB01CSK → 둘 다 SSJVB01SET) — 재고 마스터에서 "
        "쓰는 것과 같은 공식 SET품번 생성 규칙이라, 자켓·팬츠가 자동으로 같은 코드로 묶여요."
    )
    store_opts = sorted(uni["매장코드"].astype(str).unique())
    if not store_opts:
        st.info("선택 가능한 매장이 없어요.")
        return
    _name_lookup = dict(zip(uni["매장코드"].astype(str), uni["매장명"].astype(str)))
    sel_store = st.selectbox(
        "매장선택", store_opts,
        format_func=lambda c: f"{c} · {_name_lookup.get(c, '')}", key="ss_store_pick")

    g2 = g[g["매장코드"].astype(str) == sel_store].copy()
    id2design = uni.drop_duplicates("_그룹ID").set_index("_그룹ID")["_디자인키"]
    g2["_디자인키"] = g2["_그룹ID"].map(id2design)
    g2 = g2[g2["_디자인키"].notna()]
    if g2.empty:
        st.info(f"{sel_store} 매장은 선택한 조회기간에 슈트/셋업 실적이 없어요.")
        return

    style = g2.groupby("_디자인키", observed=True).agg(
        자켓판매량=("자켓판매량", "sum"), 팬츠판매량=("팬츠판매량", "sum"),
        세트수=("세트수", "sum"), 팬츠추가=("팬츠추가", "sum"), 자켓추가=("자켓추가", "sum"),
        세트로판매=("세트로판매", "sum"), 단품판매=("단품판매", "sum"),
        세트매출=("세트매출", "sum"), 단품매출=("단품매출", "sum"),
        세트그룹상의=("_세트상의", "sum"), 세트그룹하의=("_세트하의", "sum"),
    ).reset_index()
    style = style[(style["자켓판매량"] != 0) | (style["팬츠판매량"] != 0)]
    if style.empty:
        st.info(f"{sel_store} 매장은 선택한 조회기간에 슈트/셋업 실적이 없어요.")
        return
    style["SET 품번"] = style["_디자인키"]
    style["기간총판매"] = style["자켓판매량"] + style["팬츠판매량"]
    style["기간매출"] = style["세트매출"] + style["단품매출"]
    _tot2 = style["기간총판매"].replace(0, np.nan)
    style["세트비중"] = style["세트로판매"] / _tot2 * 100
    style["단품비중"] = style["단품판매"] / _tot2 * 100
    _den2 = style["세트그룹상의"].replace(0, np.nan)
    style["자켓팬츠비율"] = style["세트그룹하의"] / _den2
    style = style.sort_values("SET 품번").reset_index(drop=True)

    def _fmt_row2(r):
        return [sel_store, r["SET 품번"], _int(r["자켓판매량"]), _int(r["팬츠판매량"]),
                _int(r["기간총판매"]), _money_mm(r["기간매출"]),
                _int(r["세트로판매"]), _int(r["단품판매"]),
                _money_mm(r["세트매출"]), _money_mm(r["단품매출"]),
                _pct(r["세트비중"]), _pct(r["단품비중"]),
                _int(r["세트수"]), _int(r["팬츠추가"]), _int(r["자켓추가"]),
                _ratio(r["자켓팬츠비율"])]

    tot2 = {
        "자켓판매량": style["자켓판매량"].sum(), "팬츠판매량": style["팬츠판매량"].sum(),
        "세트수": style["세트수"].sum(), "팬츠추가": style["팬츠추가"].sum(),
        "자켓추가": style["자켓추가"].sum(),
        "세트로판매": style["세트로판매"].sum(), "단품판매": style["단품판매"].sum(),
        "세트매출": style["세트매출"].sum(), "단품매출": style["단품매출"].sum(),
        "세트그룹상의": style["세트그룹상의"].sum(), "세트그룹하의": style["세트그룹하의"].sum(),
    }
    tot2["기간총판매"] = tot2["자켓판매량"] + tot2["팬츠판매량"]
    tot2["기간매출"] = tot2["세트매출"] + tot2["단품매출"]
    _tt2 = tot2["기간총판매"] if tot2["기간총판매"] else np.nan
    tot2["세트비중"] = tot2["세트로판매"] / _tt2 * 100 if _tt2 else np.nan
    tot2["단품비중"] = tot2["단품판매"] / _tt2 * 100 if _tt2 else np.nan
    _td2 = tot2["세트그룹상의"] if tot2["세트그룹상의"] else np.nan
    tot2["자켓팬츠비율"] = tot2["세트그룹하의"] / _td2 if _td2 else np.nan
    tot2["SET 품번"] = f"■ {sel_store} 합계 ({len(style):,}개 SET품번)"

    cols2 = [("", "매장코드"), ("", "SET 품번"),
             ("기간판매량", "자켓판매량"), ("기간판매량", "팬츠판매량"),
             ("단품 기준 분석", "기간총판매"), ("단품 기준 분석", "기간매출"),
             ("단품 기준 분석", "세트로판매"), ("단품 기준 분석", "단품판매"),
             ("단품 기준 분석", "세트매출"), ("단품 기준 분석", "단품매출"),
             ("단품 기준 분석", "세트비중"), ("단품 기준 분석", "단품비중"),
             ("SET 분석", "세트수"), ("SET 분석", "팬츠추가"), ("SET 분석", "자켓추가"),
             ("SET 분석", "자켓:팬츠비율")]
    rows2 = [_fmt_row2(pd.Series(tot2))] + [_fmt_row2(r) for _, r in style.iterrows()]
    idx2 = ["■ 합계"] + list(style["SET 품번"])
    disp2 = pd.DataFrame(rows2, index=idx2, columns=pd.MultiIndex.from_tuples(cols2))

    h3, h4 = st.columns([5, 1])
    h3.markdown(f"**{sel_store} · SET 품번별 분석**", unsafe_allow_html=True)
    h4.download_button("⬇ 엑셀", table_excel_bytes(disp2, "SET품번분석"),
                       file_name=f"SET품번분석_{sel_store}_{e.date()}.xlsx", mime=XLSX_MIME,
                       key="ss_dl2", use_container_width=True)

    def _hl2(row):
        if str(row.name).startswith("■"):
            return ["background-color:#fff2b8;font-weight:700" for _ in row]
        return ["" for _ in row]

    sty2 = disp2.style.apply(_hl2, axis=1).set_properties(**{"text-align": "right"})
    sty2 = block_border(block_border(block_border(sty2, 2), 4), 12)
    render_styled_table(sty2)


# ==============================================================================
# 로그인 / 사용자 관리  ─ 개인 계정 + 역할(admin=관리자 / viewer=뷰어)
# ==============================================================================
USERS_TABLE = "app_users"


def _make_hash(password):
    """PBKDF2-SHA256 해시 문자열 'pbkdf2$iter$salt$hash' 생성. 평문 비번은 저장하지 않는다."""
    salt = os.urandom(16).hex()
    it = 200_000
    h = hashlib.pbkdf2_hmac("sha256", str(password).encode("utf-8"), bytes.fromhex(salt), it).hex()
    return f"pbkdf2${it}${salt}${h}"


def _verify_pw(password, stored):
    try:
        _algo, it, salt, h = str(stored).split("$")
        calc = hashlib.pbkdf2_hmac("sha256", str(password).encode("utf-8"), bytes.fromhex(salt), int(it)).hex()
        return hmac.compare_digest(calc, h)
    except Exception:
        return False


def ensure_users_table():
    eng = get_engine()
    with eng.begin() as conn:
        conn.exec_driver_sql(
            f'CREATE TABLE IF NOT EXISTS "{USERS_TABLE}" ('
            '"username" TEXT PRIMARY KEY, "display_name" TEXT, "pw" TEXT, '
            '"role" TEXT DEFAULT \'viewer\', "active" TEXT DEFAULT \'Y\', "created_at" TEXT)')


def get_user(username):
    eng = get_engine()
    with eng.connect() as conn:
        r = conn.execute(text(
            f'SELECT "username","display_name","pw","role","active" FROM "{USERS_TABLE}" WHERE "username"=:u'),
            {"u": username}).fetchone()
    if not r:
        return None
    return {"username": r[0], "display_name": r[1], "pw": r[2], "role": r[3], "active": r[4]}


def list_users():
    eng = get_engine()
    try:
        with eng.connect() as conn:
            rows = conn.execute(text(
                f'SELECT "username","display_name","role","active" FROM "{USERS_TABLE}" ORDER BY "role","username"')).fetchall()
    except Exception:
        return []
    return [{"username": r[0], "display_name": r[1], "role": r[2], "active": r[3]} for r in rows]


def user_count():
    ensure_users_table()
    with get_engine().connect() as conn:
        return conn.exec_driver_sql(f'SELECT COUNT(*) FROM "{USERS_TABLE}"').scalar()


def upsert_user(username, display_name, password, role, active="Y"):
    """계정 생성/수정. password가 비어 있으면(None) 비번은 그대로 두고 나머지만 갱신."""
    ensure_users_table()
    eng = get_engine()
    exists = get_user(username) is not None
    with eng.begin() as conn:
        if exists:
            if password:
                conn.execute(text(f'UPDATE "{USERS_TABLE}" SET "display_name"=:d,"pw"=:p,"role"=:r,"active"=:a WHERE "username"=:u'),
                             {"d": display_name, "p": _make_hash(password), "r": role, "a": active, "u": username})
            else:
                conn.execute(text(f'UPDATE "{USERS_TABLE}" SET "display_name"=:d,"role"=:r,"active"=:a WHERE "username"=:u'),
                             {"d": display_name, "r": role, "a": active, "u": username})
        else:
            conn.execute(text(f'INSERT INTO "{USERS_TABLE}" ("username","display_name","pw","role","active","created_at") '
                              'VALUES (:u,:d,:p,:r,:a,:c)'),
                         {"u": username, "d": display_name, "p": _make_hash(password or os.urandom(8).hex()),
                          "r": role, "a": active, "c": datetime.now().strftime("%Y-%m-%d")})


def delete_user(username):
    with get_engine().begin() as conn:
        conn.execute(text(f'DELETE FROM "{USERS_TABLE}" WHERE "username"=:u'), {"u": username})


# ── 로그인 유지(쿠키 세션) ────────────────────────────────────────────
# 로그인 상태를 브라우저 메모리(st.session_state)에만 두면 새로고침·잠깐 방치로
# 바로 재로그인이 필요했음 → 로그인 시 토큰을 발급해 브라우저 쿠키 + DB(app_sessions,
# 해시만 저장)에 두고, '마지막 사용 후 6시간'까지는 자동으로 로그인을 이어준다.
# (2026-08-07: 2시간→6시간 상향 + 쿠키 기록 방식 보강 — "몇 분만 방치해도 재로그인" 개선)
SESSIONS_TABLE = "app_sessions"
AUTH_COOKIE = "erp_auth"
IDLE_LIMIT_HOURS = 6          # 이 시간 동안 사용이 없으면 자동 만료 → 재로그인
_TS_FMT = "%Y-%m-%d %H:%M:%S"


def _hash_token(token):
    return hashlib.sha256(str(token).encode("utf-8")).hexdigest()


def ensure_sessions_table():
    with get_engine().begin() as conn:
        conn.exec_driver_sql(
            f'CREATE TABLE IF NOT EXISTS "{SESSIONS_TABLE}" ('
            '"token_hash" TEXT PRIMARY KEY, "username" TEXT, "last_seen" TEXT)')


def create_session(username):
    """새 로그인 토큰 발급. DB에는 해시만 저장(유출 대비). 오래된 세션은 청소."""
    ensure_sessions_table()
    token = os.urandom(32).hex()
    now = datetime.now()
    with get_engine().begin() as conn:
        cutoff = (now - pd.Timedelta(days=7)).strftime(_TS_FMT)
        conn.execute(text(f'DELETE FROM "{SESSIONS_TABLE}" WHERE "last_seen" < :c'), {"c": cutoff})
        conn.execute(text(f'INSERT INTO "{SESSIONS_TABLE}" ("token_hash","username","last_seen") '
                          'VALUES (:t,:u,:s)'),
                     {"t": _hash_token(token), "u": username, "s": now.strftime(_TS_FMT)})
    return token


def _session_user(token):
    """쿠키 토큰 검증: 마지막 사용이 IDLE_LIMIT_HOURS 이내면 사용자명 반환(+시간 갱신), 아니면 None."""
    if not token:
        return None
    th = _hash_token(token)
    now = datetime.now()
    try:
        ensure_sessions_table()
        with get_engine().begin() as conn:
            r = conn.execute(text(
                f'SELECT "username","last_seen" FROM "{SESSIONS_TABLE}" WHERE "token_hash"=:t'),
                {"t": th}).fetchone()
            if not r:
                return None
            try:
                last = datetime.strptime(str(r[1]), _TS_FMT)
            except Exception:
                last = None
            if last is None or (now - last).total_seconds() > IDLE_LIMIT_HOURS * 3600:
                conn.execute(text(f'DELETE FROM "{SESSIONS_TABLE}" WHERE "token_hash"=:t'), {"t": th})
                return None
            conn.execute(text(f'UPDATE "{SESSIONS_TABLE}" SET "last_seen"=:s WHERE "token_hash"=:t'),
                         {"s": now.strftime(_TS_FMT), "t": th})
        return r[0]
    except Exception:
        return None


def touch_session():
    """화면을 쓸 때마다 마지막 사용시간 갱신 → IDLE_LIMIT_HOURS 카운트가 계속 리셋됨."""
    tok = st.session_state.get("auth_token")
    if not tok:
        return
    try:
        with get_engine().begin() as conn:
            conn.execute(text(f'UPDATE "{SESSIONS_TABLE}" SET "last_seen"=:s WHERE "token_hash"=:t'),
                         {"s": datetime.now().strftime(_TS_FMT), "t": _hash_token(tok)})
    except Exception:
        pass


def drop_session():
    """로그아웃: DB 세션 즉시 삭제(쿠키가 남아 있어도 더 이상 못 씀)."""
    tok = st.session_state.get("auth_token")
    if not tok:
        return
    try:
        with get_engine().begin() as conn:
            conn.execute(text(f'DELETE FROM "{SESSIONS_TABLE}" WHERE "token_hash"=:t'), {"t": _hash_token(tok)})
    except Exception:
        pass


def _write_cookie(value, max_age):
    """브라우저 쿠키 기록/삭제(로그인 유지용). 0높이 컴포넌트로 JS만 실행 — 화면·로그인 흐름에 영향 없음.

    ※ 쿠키 기록이 실패해도 로그인은 st.session_state로 이미 성공한 상태라 사용엔 지장 없음
      (그 경우 하드새로고침 때만 재로그인 — 기존과 동일).

    2026-08-07: 컴포넌트 iframe 안에서 document.cookie만 쓰면 브라우저·상황에 따라
    실제 저장이 누락되는 경우가 있어(=잦은 재로그인 원인 추정), 부모 문서(window.parent)에도
    동일하게 한 번 더 기록해 저장 성공률을 높임. 각각 독립적으로 try/catch 처리해 하나가
    막혀도 다른 하나·기존 로그인 흐름에는 영향이 없도록 함.
    """
    cookie_str = f"{AUTH_COOKIE}={value}; path=/; max-age={int(max_age)}; SameSite=Lax"
    components.html(
        "<script>"
        f'try {{ document.cookie = "{cookie_str}"; }} catch(e) {{}}'
        f'try {{ window.parent.document.cookie = "{cookie_str}"; }} catch(e) {{}}'
        "</script>", height=0)


def _cookie_token():
    """브라우저가 보낸 로그인 유지 쿠키 읽기(구버전 Streamlit이면 None)."""
    try:
        return st.context.cookies.get(AUTH_COOKIE)
    except Exception:
        return None


def _render_login():
    st.title("🔐 온라인팀 미니 ERP")
    if user_count() == 0:
        st.info("최초 관리자 계정을 만들어 주세요. (계정이 하나도 없을 때만 나오는 화면이에요)")
        with st.form("bootstrap_admin"):
            u = st.text_input("관리자 ID")
            dn = st.text_input("이름 (표시용)")
            p1 = st.text_input("비밀번호", type="password")
            p2 = st.text_input("비밀번호 확인", type="password")
            ok = st.form_submit_button("관리자 계정 만들기", type="primary")
        if ok:
            if not u.strip() or not p1:
                st.error("ID와 비밀번호를 입력하세요.")
            elif p1 != p2:
                st.error("비밀번호가 일치하지 않아요.")
            else:
                upsert_user(u.strip(), dn.strip() or u.strip(), p1, "admin")
                st.success("관리자 계정 생성 완료! 이제 아래에서 로그인하세요.")
                st.rerun()
        return
    with st.form("login"):
        u = st.text_input("ID")
        p = st.text_input("비밀번호", type="password")
        ok = st.form_submit_button("로그인", type="primary")
    if ok:
        rec = get_user(u.strip())
        if rec and str(rec["active"]).upper() != "N" and _verify_pw(p, rec["pw"]):
            # 로그인은 세션에 즉시 반영(쿠키 성공 여부와 무관하게 항상 성공).
            # 로그인 유지 쿠키는 로그인 후 화면(사이드바)에서 매 실행 기록/갱신.
            token = create_session(rec["username"])
            st.session_state["auth_user"] = rec["username"]
            st.session_state["auth_name"] = rec["display_name"] or rec["username"]
            st.session_state["auth_role"] = rec["role"]
            st.session_state["auth_token"] = token
            st.rerun()
        else:
            st.error("ID·비밀번호가 올바르지 않거나 비활성화된 계정이에요.")
    st.caption("계정이 필요하면 관리자(팀장)에게 요청하세요.")


# ==============================================================================
# SECTION P. 💰 최저가 관리 — 외부몰 최저가 행사 원장 + 캘린더 + 최저가 체크 + 네이버 체크
#   (2026-08-20 신설 — "가격관리 메뉴 개발 정의.docx" + 중태님 4메뉴 구조 확정)
#   좌측 메뉴 "💰 최저가 관리" 안에 4개 메뉴(탭):
#   · 1️⃣ 외부몰 행사 확정: MD가 확정 행사 폼 업로드 → promo_events 테이블(팀 공동 원장)에 누적
#   · 2️⃣ 행사 진행 캘린더: 원장을 간트차트로 조회 — 기간 설정 + [브랜드/년도/시즌/아이템] 필터
#       또는 품번 직접 입력(1개 상품) 조회. 오늘 기준선 표시.
#   · 3️⃣ 외부몰 행사 최저가 체크: 기획 폼 업로드 → 원장과 [품번 동일 × 기간 겹침] 비교
#       → 최저가 여부 OK/NO + 겹치는 경쟁 행사 블록(오른쪽 반복)을 채운 엑셀 다운로드
#   · 4️⃣ 네이버 최저가 체크: 네이버 폼 업로드 → "오늘 진행 중" 외부몰 행사와 비교
#       → 위반(인상필요) 품번만 + 목표가(외부몰최저가+100원)를 채운 엑셀 다운로드
#   확정 룰 (중태님, 2026-08-20):
#   · 실질판매가 = 쿠폰적용가(숫자이고 0 초과)가 있으면 쿠폰적용가, 없으면 행사가
#   · 최저가 체크: 겹치는 기간에 타 채널 실질판매가 ≤ 내 기획 실질판매가 → NO (동일가도 NO)
#   · 네이버: 업로드 당일 진행 중인 타 채널 행사만 비교. 네이버가격 ≤ 외부몰최저가 → "인상필요"
#     (낮거나 같으면 위반 — 동일가 포함). 목표가 = 외부몰최저가 + 100원. 문제 없는 품번은 결과 미표시.
#   · 겹치는 행사가 여러 개면 결과 엑셀 오른쪽으로 9컬럼 블록을 실질판매가 낮은 순으로 반복
#   · 테이블은 첫 사용 시 자동 생성(CREATE TABLE IF NOT EXISTS) — Supabase 별도 작업 불필요
# ==============================================================================
PROMO_TABLE = "promo_events"
PROMO_COLS = ["_pkey", "행사시작", "행사종료", "매장코드", "매장명", "품번",
              "최초가", "행사가", "쿠폰율", "쿠폰적용가", "최종할인율",
              "실질판매가", "행사명", "등록자", "등록자명", "등록시각"]
# 260821: "행사명" 컬럼 추가 — 확정 폼 업로드 시 "행사명은 무엇입니까?" 1회 입력받아
# 그 업로드의 전 행에 저장(확정 행사 스케쥴·조회 표기용). 기존 운영 중인 테이블에는
# _pm_ensure_table()이 자동으로 컬럼을 추가하므로 별도 DB 작업 불필요(기존 행은 공란).
_PM_FORM_HEADERS = ["행사시작", "행사종료", "매장코드", "매장명", "품번",
                    "최초가", "행사가", "쿠폰율", "쿠폰적용가", "최종할인율"]
_PM_NAVER_HEADERS = ["매장코드", "매장명", "품번", "최초가", "네이버가격", "할인율"]
# 260820 수정2: 결과 블록의 매장명 옆에 담당자 컬럼 추가(매장 마스터 자동 매핑) → 블록 10컬럼
_PM_BLOCK_CHECK = ["매장코드", "최저가 매장", "담당자", "행사시작", "행사종료",
                   "최초가", "행사가", "쿠폰율", "쿠폰적용가", "최종할인율"]
_PM_BLOCK_NAVER = ["매장코드", "동시점 행사 매장", "담당자", "행사시작", "행사종료",
                   "최초가", "행사가", "쿠폰율", "쿠폰적용가", "최종할인율"]


def _pm_num(v):
    """가격류 숫자 파싱 — 실패·공란은 None."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        if isinstance(v, float) and math.isnan(v):
            return None
        return float(v)
    s = str(v).strip().replace(",", "").replace("원", "")
    if not s or s.lower() in ("nan", "none", "-"):
        return None
    try:
        return float(s)
    except Exception:
        return None


def _pm_rate(v):
    """쿠폰율·할인율 파싱 — '20%'→0.2, 20→0.2, 0.2→0.2. 실패·공란은 None."""
    if v is None:
        return None
    s = str(v).strip()
    pct = s.endswith("%")
    n = _pm_num(s[:-1] if pct else s)
    if n is None:
        return None
    if pct:
        return n / 100.0
    return n / 100.0 if n > 1 else n


def _pm_date(v):
    """행사시작/종료 파싱 → datetime.date. 260801(YYMMDD)·'26.08.01'·'2026-08-01'·엑셀 날짜셀 지원."""
    import datetime as _d
    if v is None:
        return None
    if isinstance(v, _d.datetime):
        return v.date()
    if isinstance(v, _d.date):
        return v
    s = str(v).strip()
    if not s or s.lower() in ("nan", "none"):
        return None
    if s.endswith(".0"):
        s = s[:-2]
    digits = "".join(ch for ch in s if ch.isdigit())
    try:
        if len(digits) == 6:      # YYMMDD (예: 260801)
            return _d.date(2000 + int(digits[:2]), int(digits[2:4]), int(digits[4:6]))
        if len(digits) == 8:      # YYYYMMDD
            return _d.date(int(digits[:4]), int(digits[4:6]), int(digits[6:8]))
    except Exception:
        return None
    return None


def _pm_yymmdd(iso):
    """ISO 'YYYY-MM-DD' → 260801(int) — 폼과 동일한 표기로 엑셀에 기입."""
    try:
        return int(str(iso)[2:4] + str(iso)[5:7] + str(iso)[8:10])
    except Exception:
        return None


def _pm_eff_price(hangsa, coupon_applied):
    """실질판매가 — 쿠폰적용가(0 초과)가 있으면 그 값, 없으면 행사가 (중태님 확정 룰)."""
    if coupon_applied is not None and coupon_applied > 0:
        return coupon_applied
    return hangsa


def _pm_cell(v):
    """엑셀 셀 기입용 정리 — NaN→None, 딱 떨어지는 float→int."""
    if v is None:
        return None
    if isinstance(v, float):
        if math.isnan(v):
            return None
        if v.is_integer():
            return int(v)
    return v


def _pm_norm(s):
    return str(s).strip().replace(" ", "") if s is not None else ""


def _pm_find_sheet(wb, headers, prefer_kw=None):
    """헤더(첫 N컬럼)가 일치하는 시트·헤더행을 찾는다. prefer_kw가 시트명에 있으면 우선.
    (마스터 폼 파일 하나에 3개 양식 시트가 함께 들어 있어도 올바른 시트를 집도록.)"""
    want = [_pm_norm(h) for h in headers]
    hits = []
    for ws in wb.worksheets:
        for hr in range(1, 6):
            if ws.max_row < hr:
                break
            row = [_pm_norm(c.value) for c in ws[hr][:len(want)]]
            if row == want:
                hits.append((ws, hr))
                break
    if not hits:
        return None, None
    if prefer_kw:
        for ws, hr in hits:
            if prefer_kw in _pm_norm(ws.title):
                return ws, hr
    return hits[0]


def _pm_read_form(uploaded, kind):
    """업로드 엑셀에서 폼 데이터 파싱. kind: 'confirm'(행사 확정) | 'check'(최저가 체크, 앞 10컬럼만)
    | 'naver'(6컬럼). return (rows, errors)."""
    import openpyxl
    wb = openpyxl.load_workbook(uploaded, data_only=True)
    if kind == "naver":
        ws, hr = _pm_find_sheet(wb, _PM_NAVER_HEADERS, prefer_kw="네이버")
        ncol, want = len(_PM_NAVER_HEADERS), _PM_NAVER_HEADERS
    else:
        prefer = "최저가체크" if kind == "check" else "행사확정"
        ws, hr = _pm_find_sheet(wb, _PM_FORM_HEADERS, prefer_kw=prefer)
        ncol, want = len(_PM_FORM_HEADERS), _PM_FORM_HEADERS
    if ws is None:
        return [], [f"필요한 헤더({' · '.join(want)})로 시작하는 시트를 찾지 못했어요. 폼 양식 그대로 올려주세요."]
    rows, errors = [], []
    for i, vals in enumerate(ws.iter_rows(min_row=hr + 1, values_only=True), start=hr + 1):
        vals = list(vals[:ncol]) + [None] * max(0, ncol - len(vals))
        if all(v is None or str(v).strip() == "" for v in vals):
            continue
        if kind == "naver":
            code, name, pn, chojo, nprice, drate = vals
            pn = str(pn).strip() if pn is not None else ""
            nprice_n = _pm_num(nprice)
            if not pn or nprice_n is None:
                errors.append(f"{i}행: 품번 또는 네이버가격이 비어 있어 건너뛰었어요.")
                continue
            rows.append({"매장코드": str(code).strip() if code is not None else "",
                         "매장명": str(name).strip() if name is not None else "",
                         "품번": pn, "최초가": _pm_num(chojo),
                         "네이버가격": nprice_n, "할인율": _pm_rate(drate)})
        else:
            s, e, code, name, pn, chojo, hangsa, crate, cprice, frate = vals
            sd, ed = _pm_date(s), _pm_date(e)
            pn = str(pn).strip() if pn is not None else ""
            hangsa_n = _pm_num(hangsa)
            if not pn or sd is None or ed is None or hangsa_n is None:
                errors.append(f"{i}행: 품번·행사시작·행사종료·행사가 중 비었거나 형식을 읽지 못해 건너뛰었어요.")
                continue
            if ed < sd:
                errors.append(f"{i}행: 행사종료({ed})가 행사시작({sd})보다 빨라 건너뛰었어요.")
                continue
            cp = _pm_num(cprice)
            rows.append({"행사시작": sd.isoformat(), "행사종료": ed.isoformat(),
                         "매장코드": str(code).strip() if code is not None else "",
                         "매장명": str(name).strip() if name is not None else "",
                         "품번": pn, "최초가": _pm_num(chojo), "행사가": hangsa_n,
                         "쿠폰율": _pm_rate(crate), "쿠폰적용가": cp,
                         "최종할인율": _pm_rate(frate),
                         "실질판매가": _pm_eff_price(hangsa_n, cp)})
    return rows, errors


def _pm_ensure_table(conn):
    q = '"'
    defs = ", ".join(f'{q}{c}{q} TEXT' for c in PROMO_COLS if c != "_pkey")
    conn.exec_driver_sql(
        f'CREATE TABLE IF NOT EXISTS {q}{PROMO_TABLE}{q} ({q}_pkey{q} TEXT PRIMARY KEY, {defs})')
    # 260821: 이미 운영 중인 테이블에 새 컬럼(행사명 등)이 없으면 자동 추가 — 기존 원장 데이터 보존
    try:
        if conn.engine.dialect.name == "postgresql":
            have = {r[0] for r in conn.exec_driver_sql(
                "SELECT column_name FROM information_schema.columns WHERE table_name=%s",
                (PROMO_TABLE,)).fetchall()}
        else:
            have = {r[1] for r in conn.exec_driver_sql(
                f'PRAGMA table_info("{PROMO_TABLE}")').fetchall()}
        for c in PROMO_COLS:
            if c not in have:
                conn.exec_driver_sql(f'ALTER TABLE {q}{PROMO_TABLE}{q} ADD COLUMN {q}{c}{q} TEXT')
    except Exception:
        pass


def _pm_key(r):
    base = "|".join(str(r.get(k, "")) for k in ("행사시작", "행사종료", "매장코드", "품번", "행사가", "쿠폰적용가"))
    return hashlib.md5(base.encode("utf-8")).hexdigest()


def promo_insert(rows, event_name=""):
    """행사 확정 원장 적재 — 동일 행사(기간+매장+품번+가격이 같은 행)는 건너뜀(재업로드 안전).
    event_name(행사명, 260821): 업로드 시 1회 입력받아 이 업로드의 전 행에 저장.
    (중복 판정 키에는 행사명이 안 들어가므로, 같은 행사를 다른 이름으로 재업로드해도 원본이 유지됨.)"""
    if not rows:
        return {"inserted": 0, "skipped": 0}
    user = st.session_state.get("auth_user", "") or ""
    name = st.session_state.get("auth_name", "") or ""
    ts = now_kst().strftime("%Y-%m-%d %H:%M")
    eng = get_engine()
    ins = skip = 0
    with eng.begin() as conn:
        _pm_ensure_table(conn)
        existing = set(r[0] for r in conn.exec_driver_sql(
            f'SELECT "_pkey" FROM "{PROMO_TABLE}"').fetchall())
        ph = "%s" if eng.dialect.name == "postgresql" else "?"
        cols_sql = ", ".join('"' + c + '"' for c in PROMO_COLS)
        sql = f'INSERT INTO "{PROMO_TABLE}" ({cols_sql}) VALUES ({", ".join([ph] * len(PROMO_COLS))})'
        seen = set()
        for r in rows:
            k = _pm_key(r)
            if k in existing or k in seen:
                skip += 1
                continue
            seen.add(k)
            r = dict(r)
            r["행사명"] = str(r.get("행사명") or event_name or "").strip()
            vals = [k] + [("" if r.get(c) is None else str(r.get(c))) for c in PROMO_COLS[1:-3]] \
                 + [user, name, ts]
            conn.exec_driver_sql(sql, tuple(vals))
            ins += 1
    return {"inserted": ins, "skipped": skip}


def promo_load():
    """원장 전체 로드 → DataFrame(파싱 숫자컬럼 _최초가.._실질판매가 포함). 테이블 없으면 빈 DF.
    (행사 원장은 수백 행 규모라 캐시 없이 매번 읽어 항상 최신을 보여준다 — 팀 공동 입력 특성.)"""
    eng = get_engine()
    try:
        with eng.begin() as conn:
            _pm_ensure_table(conn)
            df = pd.read_sql(f'SELECT * FROM "{PROMO_TABLE}"', conn)
    except Exception:
        return pd.DataFrame(columns=PROMO_COLS)
    if df.empty:
        return df
    if "행사명" not in df.columns:          # 컬럼 신설(260821) 이전에 읽힌 캐시/구 스키마 방어
        df["행사명"] = ""
    df["행사명"] = df["행사명"].fillna("").astype(str).replace("None", "")
    for c in ("최초가", "행사가", "쿠폰율", "쿠폰적용가", "최종할인율"):
        df["_" + c] = pd.to_numeric(df[c].map(_pm_num), errors="coerce")
    # 실질판매가는 항상 재계산(룰이 한 곳에만 살도록): 쿠폰적용가>0 → 쿠폰적용가, 아니면 행사가
    df["_실질판매가"] = np.where(df["_쿠폰적용가"].fillna(0) > 0, df["_쿠폰적용가"], df["_행사가"])
    return df


def promo_delete(pkeys):
    if not pkeys:
        return 0
    eng = get_engine()
    n = 0
    with eng.begin() as conn:
        _pm_ensure_table(conn)
        ph = "%s" if eng.dialect.name == "postgresql" else "?"
        for k in pkeys:
            res = conn.exec_driver_sql(f'DELETE FROM "{PROMO_TABLE}" WHERE "_pkey" = {ph}', (k,))
            n += res.rowcount or 0
    return n


def _pm_overlaps(ledger, pn, sd_iso, ed_iso, exclude_code=None):
    """원장에서 [품번 동일 × 기간 겹침] 행사만 추출 (exclude_code 매장 = 자기 채널은 제외),
    실질판매가 낮은 순 정렬. ISO 날짜 문자열은 사전순 비교가 곧 날짜 비교라 안전."""
    if ledger.empty:
        return ledger
    m = (ledger["품번"].astype(str).str.strip() == str(pn).strip()) \
        & (ledger["행사시작"] <= ed_iso) & (ledger["행사종료"] >= sd_iso)
    if exclude_code:
        m &= ledger["매장코드"].astype(str).str.strip() != str(exclude_code).strip()
    return ledger[m].sort_values("_실질판매가", na_position="last")


def _pm_check_rows(plans, ledger):
    """최저가 체크 판정 — NO 조건: 겹치는 타 채널 실질판매가 ≤ 내 기획가 (동일가 NO, 중태님 확정)."""
    out = []
    for p in plans:
        comp = _pm_overlaps(ledger, p["품번"], p["행사시작"], p["행사종료"],
                            exclude_code=p["매장코드"] or None)
        lows = comp["_실질판매가"].dropna()
        lowest = float(lows.min()) if len(lows) else None
        ok = (lowest is None) or (lowest > p["실질판매가"])
        out.append({"plan": p, "ok": ok, "lowest": lowest, "comp": comp})
    return out


def _pm_naver_rows(rows, ledger, today_iso):
    """네이버 금일최저가 판정 — 오늘 진행 중인 타 채널 행사만 비교(중태님 확정).
    위반 조건: 네이버가격 ≤ 외부몰 최저 실질판매가 (동일가 포함). 목표가 = 외부몰최저가 + 100원."""
    out = []
    for r in rows:
        comp = _pm_overlaps(ledger, r["품번"], today_iso, today_iso,
                            exclude_code=r["매장코드"] or None)
        lows = comp["_실질판매가"].dropna()
        lowest = float(lows.min()) if len(lows) else None
        if lowest is None:          # 오늘 진행 중인 외부몰 행사 없음 → 문제 없음(결과 미표시)
            continue
        if r["네이버가격"] <= lowest:
            out.append({"row": r, "lowest": lowest, "target": lowest + 100, "comp": comp})
    return out


def _pm_mgr_map():
    """매장코드 → 담당자 매핑 (매장 마스터 · 유통별 세부 분석과 동일 소스). 마스터 없으면 빈 dict."""
    m = load_master()
    if not m.empty and "담당자" in m.columns and "매장코드" in m.columns:
        return dict(zip(m["매장코드"].astype(str).str.strip(),
                        m["담당자"].astype(str).str.strip()))
    return {}


def _pm_block_vals(e, mgr):
    """원장 행(Series) → 결과 블록 10칸 값(폼 표기: 날짜 YYMMDD, 담당자 = 매장 마스터 매핑)."""
    return [e["매장코드"], e["매장명"], mgr.get(str(e["매장코드"]).strip(), ""),
            _pm_yymmdd(e["행사시작"]), _pm_yymmdd(e["행사종료"]),
            _pm_cell(e["_최초가"]), _pm_cell(e["_행사가"]), _pm_cell(e["_쿠폰율"]),
            _pm_cell(e["_쿠폰적용가"]), _pm_cell(e["_최종할인율"])]


def _pm_style_header(ws, heads, n_gray, red_idx, block_from):
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    fill_gray = PatternFill("solid", fgColor="D9D9D9")
    fill_red = PatternFill("solid", fgColor="C00000")
    fill_pink = PatternFill("solid", fgColor="F2DCDB")
    center = Alignment(horizontal="center")
    for j, h in enumerate(heads, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.alignment = center
        if j in red_idx:
            c.fill, c.font = fill_red, Font(bold=True, color="FFFFFF")
        elif j <= n_gray:
            c.fill, c.font = fill_gray, Font(bold=True)
        elif j >= block_from:
            c.fill, c.font = fill_pink, Font(bold=True)
    ws.freeze_panes = "A2"
    for col in range(1, len(heads) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 11.5


def _pm_check_excel(results):
    """최저가 체크 결과 엑셀 — 폼 10컬럼 + 최저가 여부 + 경쟁 행사 10컬럼 블록(실질판매가 낮은 순 반복)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = "외부몰 행사 최저가 체크"
    mgr = _pm_mgr_map()
    max_blocks = max([len(r["comp"]) for r in results] + [1])
    heads = list(_PM_FORM_HEADERS) + ["최저가 여부"] + _PM_BLOCK_CHECK * max_blocks
    _pm_style_header(ws, heads, n_gray=10, red_idx={11}, block_from=12)
    center = Alignment(horizontal="center")
    rr = 2
    for res in results:
        p = res["plan"]
        vals = [_pm_yymmdd(p["행사시작"]), _pm_yymmdd(p["행사종료"]), p["매장코드"], p["매장명"],
                p["품번"], _pm_cell(p["최초가"]), _pm_cell(p["행사가"]), _pm_cell(p["쿠폰율"]),
                _pm_cell(p["쿠폰적용가"]), _pm_cell(p["최종할인율"]),
                "OK" if res["ok"] else "NO"]
        for _, e in res["comp"].iterrows():
            vals += _pm_block_vals(e, mgr)
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=rr, column=j, value=v)
            if j in (6, 7, 9):
                c.number_format = "#,##0"
            elif j in (8, 10):
                c.number_format = "0%"
            elif j == 11:
                c.font = Font(bold=True, color=("C00000" if v == "NO" else "1F7A33"))
                c.alignment = center
            elif j >= 12:
                k = (j - 12) % 10
                if k in (5, 6, 8):
                    c.number_format = "#,##0"
                elif k in (7, 9):
                    c.number_format = "0%"
        rr += 1
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _pm_naver_excel(violations, today):
    """네이버 최저가 체크 결과 엑셀 — 위반(인상필요) 품번만. 폼 6컬럼 + 최저가 위반 여부('인상필요')
    + 가격인하 필요(=목표가, 외부몰최저가+100원) + 동시점 행사 9컬럼 블록(낮은 순 반복)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = "네이버 최저가 체크"
    mgr = _pm_mgr_map()
    max_blocks = max([len(v["comp"]) for v in violations] + [1])
    heads = list(_PM_NAVER_HEADERS) + ["최저가 위반 여부", "가격인하 필요"] + _PM_BLOCK_NAVER * max_blocks
    _pm_style_header(ws, heads, n_gray=6, red_idx={7, 8}, block_from=9)
    center = Alignment(horizontal="center")
    red_bold = Font(bold=True, color="C00000")
    rr = 2
    for vv in violations:
        r = vv["row"]
        vals = [r["매장코드"], r["매장명"], r["품번"], _pm_cell(r["최초가"]),
                _pm_cell(r["네이버가격"]), _pm_cell(r["할인율"]),
                "인상필요", _pm_cell(float(vv["target"]))]
        for _, e in vv["comp"].iterrows():
            vals += _pm_block_vals(e, mgr)
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=rr, column=j, value=v)
            if j in (4, 5, 8):
                c.number_format = "#,##0"
            elif j == 6:
                c.number_format = "0%"
            elif j >= 9:
                k = (j - 9) % 10
                if k in (5, 6, 8):
                    c.number_format = "#,##0"
                elif k in (7, 9):
                    c.number_format = "0%"
            if j in (7, 8):
                c.font = red_bold
                if j == 7:
                    c.alignment = center
        rr += 1
    ws.cell(row=1, column=len(heads) + 2,
            value=f"기준일: {today} (업로드 당일 진행 중 행사와 비교)")
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _pm_fmt_won(v):
    return "" if v is None or (isinstance(v, float) and math.isnan(v)) else f"{int(round(v)):,}"


def _pm_fmt_rate(v):
    return "" if v is None or (isinstance(v, float) and math.isnan(v)) else f"{v:.0%}"


def _pm_status(sd_iso, ed_iso, today_iso):
    if ed_iso < today_iso:
        return "종료"
    if sd_iso > today_iso:
        return "예정"
    return "진행중"


def _pm_decorate(ledger):
    """원장에 품번 해독 컬럼(_브랜드/_연도/_시즌/_아이템)을 붙인다 — 캘린더 필터용.
    해독은 기존 decode_stco(STCO 10자리 품번 규칙) 재사용, 해독 불가 값은 '기타'."""
    if ledger.empty:
        return ledger
    out = ledger.copy()
    info = {}
    for pn in out["품번"].astype(str).str.strip().unique():
        try:
            info[pn] = decode_stco(pn)
        except Exception:
            info[pn] = {}
    def _g(pn, key):
        v = info.get(str(pn).strip(), {}).get(key)
        return str(v) if v not in (None, "", "None") else "기타"
    out["_브랜드"] = out["품번"].map(lambda p: _g(p, "브랜드명"))
    out["_연도"] = out["품번"].map(lambda p: _g(p, "연도"))
    out["_시즌"] = out["품번"].map(lambda p: _g(p, "시즌명"))
    out["_아이템"] = out["품번"].map(lambda p: _g(p, "중카테고리"))
    return out


def _pm_gantt(view, d_from, d_to, today_iso):
    """행사 진행 캘린더(간트차트) — 중태님 예시 엑셀 이미지 기준:
    행 = [아이템코드 · 품번] (같은 품번이 겹치는 기간에 복수 채널이면 품번이 반복되며 아래로 나열),
    막대 = 행사 기간(채널별 색), 막대 안 글자 = '채널명 실질판매가원', 노란 세로선 = 오늘."""
    g = view.copy()
    g["_s"] = pd.to_datetime(g["행사시작"])
    g["_e"] = pd.to_datetime(g["행사종료"]) + pd.Timedelta(days=1)   # 종료일 '포함'으로 보이게
    g = g.sort_values(["품번", "_s", "매장명"]).reset_index(drop=True)
    _item = g["품번"].astype(str).str.strip().str.upper().str[1:3]   # 품번 2~3번째 자리 = 아이템코드
    g["행라벨"] = _item + " · " + g["품번"].astype(str)
    # 행 1개 = 행사 1건 — 같은 품번·같은 기간이라도 채널마다 별도 행(아래로 나열)이 되도록 고유 키 부여
    g["_rowid"] = g["행라벨"] + "|" + g.index.astype(str)
    g["가격"] = g["_실질판매가"].map(lambda v: "" if pd.isna(v) else f"{int(v):,}원")
    g["막대표기"] = g["매장명"].astype(str) + " " + g["가격"]
    g["기간"] = g["행사시작"] + " ~ " + g["행사종료"]
    fig = px.timeline(
        g, x_start="_s", x_end="_e", y="_rowid", color="매장명", text="막대표기",
        hover_data={"품번": True, "매장명": True, "기간": True, "가격": True,
                    "_s": False, "_e": False, "_rowid": False})
    # 260821 수정: 막대 안 글씨는 항상 동일 크기(13px) — 막대가 짧아 안 들어가면 밖으로 표기
    fig.update_traces(textposition="auto", insidetextanchor="middle",
                      textfont=dict(size=13),
                      marker_line_color="rgba(0,0,0,0.25)", marker_line_width=0.5)
    fig.update_layout(uniformtext=dict(minsize=13, mode="show"))
    fig.update_yaxes(autorange="reversed", title=None, tickmode="array",
                     tickvals=g["_rowid"].tolist(), ticktext=g["행라벨"].tolist(),
                     categoryorder="array", categoryarray=g["_rowid"].tolist())
    fig.update_xaxes(title=None, side="top",
                     range=[pd.Timestamp(d_from), pd.Timestamp(d_to) + pd.Timedelta(days=1)],
                     dtick="D1", tickformat="%m/%d", tickfont=dict(size=10), showgrid=True,
                     gridcolor="rgba(0,0,0,0.06)")
    # 오늘 기준선 (레퍼런스 이미지의 노란 세로선)
    t0 = pd.Timestamp(today_iso) + pd.Timedelta(hours=12)
    fig.add_shape(type="line", x0=t0, x1=t0, y0=0, y1=1, yref="paper",
                  line=dict(color="#f4d03f", width=3))
    fig.add_annotation(x=t0, y=1.02, yref="paper", text="오늘", showarrow=False,
                       font=dict(size=11, color="#b7950b"))
    fig.update_layout(
        height=max(340, 34 * len(g) + 150),
        margin=dict(l=10, r=10, t=60, b=10),
        legend=dict(orientation="h", yanchor="bottom", y=1.05, title=None),
        plot_bgcolor="white")
    return fig


def _pm_sched_gantt(view, d_from, d_to, today_iso, mgr):
    """📆 확정 행사 스케쥴 간트(메뉴1, 260821 신설 — 중태님 예시 엑셀 이미지 기준):
    행 = [매장명 | 담당자 | 행사명], 막대 1개 = 행사(같은 매장·행사명·기간 묶음),
    막대 안 = '최대할인율 XX%, {품번} {가격}원' — 그 행사에서 할인율이 가장 큰 품번과
    그 품번의 실질판매가(중태님 확정: 최대할인율 품번 기준). 노란 세로선 = 오늘."""
    g = view.copy()
    # 할인율 = 최종할인율(있으면), 없으면 1 - 실질판매가/최초가 로 계산
    alt = 1 - g["_실질판매가"] / g["_최초가"]
    g["_할인율계"] = g["_최종할인율"].fillna(alt)
    rows = []
    for (code, name, ev, s, e), grp in g.groupby(
            ["매장코드", "매장명", "행사명", "행사시작", "행사종료"], dropna=False):
        if grp["_할인율계"].notna().any():
            top = grp.loc[grp["_할인율계"].idxmax()]
        elif grp["_실질판매가"].notna().any():
            top = grp.loc[grp["_실질판매가"].idxmin()]
        else:
            top = grp.iloc[0]
        dtxt = f"{top['_할인율계']:.0%}" if pd.notna(top["_할인율계"]) else "-"
        ptxt = f" {int(top['_실질판매가']):,}원" if pd.notna(top["_실질판매가"]) else ""
        rows.append({
            "매장명": str(name), "행사명": (str(ev).strip() or "(행사명 미입력)"),
            "_s": pd.Timestamp(s), "_e": pd.Timestamp(e) + pd.Timedelta(days=1),
            "행라벨": f"{name} | {mgr.get(str(code).strip(), '')} | {str(ev).strip() or '(행사명 미입력)'}",
            "막대표기": f"최대할인율 {dtxt}, {top['품번']}{ptxt}",
            "기간": f"{s} ~ {e}", "품번수": len(grp)})
    sched = pd.DataFrame(rows).sort_values(["_s", "매장명", "행사명"]).reset_index(drop=True)
    sched["_rowid"] = sched["행라벨"] + "|" + sched.index.astype(str)
    fig = px.timeline(
        sched, x_start="_s", x_end="_e", y="_rowid", color="매장명", text="막대표기",
        hover_data={"행사명": True, "기간": True, "품번수": True,
                    "_s": False, "_e": False, "_rowid": False})
    # 260821 수정: 막대 안 글씨는 항상 동일 크기(13px) — 막대가 짧아 안 들어가면 밖으로 표기
    fig.update_traces(textposition="auto", insidetextanchor="middle",
                      textfont=dict(size=13),
                      marker_line_color="rgba(0,0,0,0.25)", marker_line_width=0.5)
    fig.update_layout(uniformtext=dict(minsize=13, mode="show"))
    fig.update_yaxes(autorange="reversed", title=None, tickmode="array",
                     tickvals=sched["_rowid"].tolist(), ticktext=sched["행라벨"].tolist(),
                     categoryorder="array", categoryarray=sched["_rowid"].tolist())
    fig.update_xaxes(title=None, side="top",
                     range=[pd.Timestamp(d_from), pd.Timestamp(d_to) + pd.Timedelta(days=1)],
                     dtick="D1", tickformat="%m/%d", tickfont=dict(size=10), showgrid=True,
                     gridcolor="rgba(0,0,0,0.06)")
    t0 = pd.Timestamp(today_iso) + pd.Timedelta(hours=12)
    fig.add_shape(type="line", x0=t0, x1=t0, y0=0, y1=1, yref="paper",
                  line=dict(color="#f4d03f", width=3))
    fig.add_annotation(x=t0, y=1.02, yref="paper", text="오늘", showarrow=False,
                       font=dict(size=11, color="#b7950b"))
    fig.update_layout(
        height=max(320, 38 * len(sched) + 150),
        margin=dict(l=10, r=10, t=60, b=10),
        showlegend=False,   # 260821 수정: 범례 삭제 — 행 라벨에 매장명이 이미 있어 불필요
        plot_bgcolor="white")
    return fig


def render_price_mgmt():
    """💰 최저가 관리 메뉴 — 전 팀원 입력 가능(공동 원장), 삭제는 본인 등록분 또는 관리자만."""
    st.subheader("💰 최저가 관리 — 외부몰 최저가 행사")
    st.caption("외부 유통 채널의 최저가 보장 행사 예약을 팀 공동 원장에 쌓고, 캘린더로 조회하고, "
               "새 행사 기획과 네이버 브랜드 스토어 가격이 기존 최저가 약속과 충돌하는지 자동으로 "
               "체크해요. 모든 비교는 **실질판매가**(쿠폰적용가가 있으면 쿠폰적용가, 없거나 0이면 "
               "행사가) 기준이에요.")
    is_admin = st.session_state.get("auth_role") == "admin"
    me = st.session_state.get("auth_user", "") or ""
    today = now_kst().date()
    today_iso = today.isoformat()
    if st.session_state.pop("pm_flash", None):
        st.success(st.session_state.pop("pm_flash_msg", "완료됐어요 ✅"))
    ledger = promo_load()
    # 담당자 매핑 (매장 마스터 · 매장코드 기준) — 3·4번 메뉴 결과 표시용 (260820 수정2:
    # 1번 메뉴는 본인이 본인 행사를 올리는 것이라 담당자 표기 불필요, 3·4번 결과에만 표기)
    _mgr_d = _pm_mgr_map()
    def _mgr(code):
        return _mgr_d.get(str(code).strip(), "")
    n_all = len(ledger)
    n_act = 0 if ledger.empty else int(((ledger["행사시작"] <= today_iso)
                                        & (ledger["행사종료"] >= today_iso)).sum())
    # 260820: 4개 탭을 '폴더 탭' 스타일로 — 글씨는 아래 h5(#####) 수준으로 키우고, 탭마다
    #   색을 달리해 선택 탭은 올라와 보이고 나머지는 흐리게. 이 페이지에만 탭이 있으므로
    #   stTabs 전역 셀렉터를 써도 다른 메뉴에 영향 없음.
    st.markdown("""
<style>
/* Streamlit 버전별 마크업 차이(button / div[role=tab], baseweb / react-aria) 모두 커버 */
div[data-testid="stTabs"] [role="tablist"]{
  gap:6px; border-bottom:3px solid #f2b705; padding:0 4px; align-items:flex-end;
  overflow-x:auto; overflow-y:visible; margin-top:6px;}
div[data-testid="stTabs"] [role="tab"]{
  position:relative; padding:10px 20px 9px 18px; height:auto; border:none; cursor:pointer;
  border-radius:14px 14px 0 0; clip-path:polygon(0 0, calc(100% - 14px) 0, 100% 100%, 0 100%);
  color:#fff !important; opacity:.7; filter:saturate(.85); margin-bottom:0;
  transition:opacity .15s, padding .15s;}
div[data-testid="stTabs"] [role="tab"] p{
  font-size:1.1rem !important; font-weight:700 !important; letter-spacing:-0.01em;
  white-space:nowrap; color:#fff !important;}
div[data-testid="stTabs"] [role="tab"]:nth-child(1){background:#f2b705;}
div[data-testid="stTabs"] [role="tab"]:nth-child(2){background:#f36f21;}
div[data-testid="stTabs"] [role="tab"]:nth-child(3){background:#e8427c;}
div[data-testid="stTabs"] [role="tab"]:nth-child(4){background:#7b4bd6;}
div[data-testid="stTabs"] [role="tab"]:hover{opacity:.85;}
div[data-testid="stTabs"] [role="tab"][aria-selected="true"]{
  opacity:1; filter:none; padding-top:15px; padding-bottom:12px;}
div[data-testid="stTabs"] [data-baseweb="tab-highlight"],
div[data-testid="stTabs"] [data-baseweb="tab-border"],
div[data-testid="stTabs"] .react-aria-SelectionIndicator{display:none !important;}
div[data-testid="stTabs"] [role="tabpanel"]{padding-top:18px;}
/* 탭 안 소제목(#####)과 바로 아래 설명글 사이 여백 확보 (260820 요청) */
div[data-testid="stTabs"] [role="tabpanel"] h5{
  margin-bottom:14px !important; padding-bottom:0 !important; font-size:1.25rem !important;}
div[data-testid="stTabs"] [role="tabpanel"] [data-testid="stCaptionContainer"]{line-height:1.6;}
</style>""", unsafe_allow_html=True)
    tab1, tab2, tab3, tab4 = st.tabs(
        ["1️⃣ 외부몰 행사 확정", "2️⃣ 행사 진행 캘린더", "3️⃣ 외부몰 행사 최저가 체크", "4️⃣ 네이버 최저가 체크"])

    # ── 메뉴1. 외부몰 행사 확정 — 폼 업로드 → 원장 적재 + 원장 조회·삭제 ─────────────
    with tab1:
        st.markdown("##### 1️⃣ 외부몰 행사 확정 등록")
        st.caption("유통 채널에서 최저가 행사가 확정되면 **'외부몰 행사 확정' 폼**(행사시작·행사종료·"
                   "매장코드·매장명·품번·최초가·행사가·쿠폰율·쿠폰적용가·최종할인율 10컬럼)을 올리고 "
                   "등록 버튼을 눌러 주세요. 등록된 행사는 로우데이터(원장)로 쌓여서 '행사 진행 캘린더'와 "
                   "최저가 체크의 비교 기준이 돼요. 같은 행사(기간+매장+품번+가격 동일)는 다시 올려도 "
                   "중복으로 쌓이지 않아요.")
        c_m1, c_m2 = st.columns(2)
        c_m1.metric("원장 등록 행사", f"{n_all:,} 건")
        c_m2.metric("오늘 진행 중", f"{n_act:,} 건")
        up1 = st.file_uploader("① '외부몰 행사 확정' 폼 업로드", type=["xlsx"],
                               accept_multiple_files=False, key="pm_confirm_up")
        if up1 is not None:
            rows1, errs1 = _pm_read_form(up1, "confirm")
            for msg in errs1:
                st.warning("⚠️ " + msg)
            if rows1:
                prev = pd.DataFrame([{
                    "행사시작": r["행사시작"], "행사종료": r["행사종료"],
                    "매장코드": r["매장코드"], "매장명": r["매장명"], "품번": r["품번"],
                    "최초가": _pm_fmt_won(r["최초가"]), "행사가": _pm_fmt_won(r["행사가"]),
                    "쿠폰율": _pm_fmt_rate(r["쿠폰율"]), "쿠폰적용가": _pm_fmt_won(r["쿠폰적용가"]),
                    "실질판매가": _pm_fmt_won(r["실질판매가"]),
                } for r in rows1])
                st.caption(f"📄 읽은 행사: **{len(rows1)}건** — 아래 내용 확인 후 등록해 주세요. "
                           "(실질판매가 = 쿠폰적용가 있으면 쿠폰적용가, 없으면 행사가)")
                st.dataframe(prev, use_container_width=True, hide_index=True)
                # 260821: 행사명 1회 입력 → 이 업로드의 전 행에 저장(확정 행사 스케쥴·조회 표기용)
                ev_name = st.text_input("행사명은 무엇입니까?", key="pm_event_name",
                                        placeholder="예: 무신사 8월 최저가 위크")
                if not ev_name.strip():
                    st.info("👆 행사명을 입력하면 등록 버튼이 활성화돼요 — 아래 '확정 행사 스케쥴' "
                            "간트차트에 이 이름으로 표기돼요.")
                if st.button("② 원장에 행사 등록", type="primary", use_container_width=True,
                             key="pm_confirm_btn", disabled=not ev_name.strip()):
                    res = promo_insert(rows1, ev_name.strip())
                    st.session_state["pm_flash"] = True
                    st.session_state["pm_flash_msg"] = (
                        f"행사 등록 완료 ✅ 신규 {res['inserted']:,}건 · "
                        f"중복 건너뜀 {res['skipped']:,}건")
                    st.rerun()
            elif not errs1:
                st.info("폼에서 읽을 데이터 행이 없어요 — 회색 영역을 채워서 올려주세요.")

        st.divider()
        # ── 📆 확정 행사 스케쥴 (260821 신설 — 금일 최저가 현황 위) ─────────────────
        st.markdown("##### 📆 확정 행사 스케쥴")
        st.caption("등록 완료된 행사 계획을 간트차트로 보여줘요 — 행 = **매장명 | 담당자 | 행사명**, "
                   "막대 안 = 그 행사의 **최대할인율**과 최대할인 품번·실질판매가, 노란 세로선 = 오늘. "
                   "같은 매장·행사명·기간으로 등록된 품번들은 막대 1개로 묶여요(품번 수는 마우스 오버로 확인).")
        if ledger.empty:
            st.info("아직 등록된 행사가 없어요 — 위에서 '외부몰 행사 확정' 폼을 올려 시작해 주세요.")
        else:
            sc1, sc2 = st.columns(2)
            # 260821 수정: 기본 표시 기간 = 조회일부터 3주 (길게 잡으면 날짜 칸이 좁아져 글씨가 작아짐)
            sch_from = sc1.date_input("스케쥴 시작일", value=today, key="pm_sch_from")
            sch_to = sc2.date_input("스케쥴 종료일", value=today + timedelta(days=21), key="pm_sch_to")
            if sch_to < sch_from:
                st.warning("스케쥴 종료일이 시작일보다 빨라요 — 기간을 다시 선택해 주세요.")
            else:
                sub = ledger[(ledger["행사시작"] <= sch_to.isoformat())
                             & (ledger["행사종료"] >= sch_from.isoformat())]
                if sub.empty:
                    st.info("선택한 기간에 걸치는 행사가 없어요.")
                else:
                    st.plotly_chart(_pm_sched_gantt(sub, sch_from, sch_to, today_iso, _mgr_d),
                                    use_container_width=True)

        st.divider()
        st.markdown("##### 📍 금일 품번별 최저가 현황")
        st.caption(f"기준일 **{today_iso}** — 오늘 진행 중인 행사만 모아 품번별 최저 실질판매가와 "
                   "그 채널을 보여줘요. 시점별 최저가를 바로 파악하는 용도예요.")
        if ledger.empty:
            st.info("아직 등록된 행사가 없어요 — 위에서 '외부몰 행사 확정' 폼을 올려 시작해 주세요.")
        else:
            act = ledger[(ledger["행사시작"] <= today_iso) & (ledger["행사종료"] >= today_iso)]
            if act.empty:
                st.info("오늘 진행 중인 행사가 없어요.")
            else:
                best = act.loc[act.groupby("품번")["_실질판매가"].idxmin()].sort_values("품번")
                cnt = act.groupby("품번").size()
                st.dataframe(pd.DataFrame([{
                    "품번": b["품번"], "최저가 채널": b["매장명"],
                    "실질판매가": _pm_fmt_won(b["_실질판매가"]),
                    "행사가": _pm_fmt_won(b["_행사가"]),
                    "쿠폰적용가": _pm_fmt_won(b["_쿠폰적용가"]),
                    "행사기간": f'{b["행사시작"]} ~ {b["행사종료"]}',
                    "진행 행사 수": int(cnt.get(b["품번"], 1)),
                } for _, b in best.iterrows()]), use_container_width=True, hide_index=True)

        st.divider()
        st.markdown("##### 🗂️ 행사 원장 조회")
        if ledger.empty:
            st.caption("등록된 행사가 없어요.")
        else:
            f1, f2, f3 = st.columns([1.2, 1.4, 1])
            q_pn = f1.text_input("품번 검색(부분 일치)", key="pm_q_pn")
            stores = sorted(ledger["매장명"].astype(str).str.strip().unique().tolist())
            q_st = f2.multiselect("매장(채널)", stores, key="pm_q_store")
            q_stat = f3.selectbox("상태", ["전체", "진행중", "예정", "종료"], key="pm_q_status")
            view = ledger.copy()
            view["상태"] = [_pm_status(s, e, today_iso)
                          for s, e in zip(view["행사시작"], view["행사종료"])]
            if q_pn.strip():
                view = view[view["품번"].astype(str).str.contains(q_pn.strip(), case=False, na=False)]
            if q_st:
                view = view[view["매장명"].astype(str).str.strip().isin(q_st)]
            if q_stat != "전체":
                view = view[view["상태"] == q_stat]
            view = view.sort_values(["행사시작", "품번"], ascending=[False, True])
            st.caption(f"조회 결과: **{len(view):,}건**")
            st.dataframe(pd.DataFrame([{
                "상태": v["상태"], "행사시작": v["행사시작"], "행사종료": v["행사종료"],
                "매장코드": v["매장코드"], "매장명": v["매장명"],
                "행사명": v["행사명"], "품번": v["품번"],
                "최초가": _pm_fmt_won(v["_최초가"]), "행사가": _pm_fmt_won(v["_행사가"]),
                "쿠폰율": _pm_fmt_rate(v["_쿠폰율"]), "쿠폰적용가": _pm_fmt_won(v["_쿠폰적용가"]),
                "실질판매가": _pm_fmt_won(v["_실질판매가"]),
                "등록자": v["등록자명"] or v["등록자"], "등록시각": v["등록시각"],
            } for _, v in view.iterrows()]), use_container_width=True, hide_index=True)

            # 삭제 — 본인 등록분만(관리자는 전체). 잘못 등록/취소된 행사 정리용.
            dele = ledger if is_admin else ledger[ledger["등록자"].astype(str) == me]
            with st.expander(f"🗑️ 행사 삭제 ({'관리자 — 전체' if is_admin else '내가 등록한 행사만'} "
                             f"{len(dele):,}건)"):
                if dele.empty:
                    st.caption("삭제할 수 있는 행사가 없어요.")
                else:
                    opts = {f'{(v["행사명"] or "행사명없음")} · {v["품번"]} · {v["매장명"]} · {v["행사시작"]}~{v["행사종료"]} · '
                            f'{_pm_fmt_won(v["_실질판매가"])}원 · {v["등록자명"] or v["등록자"]}': v["_pkey"]
                            for _, v in dele.sort_values("행사시작", ascending=False).iterrows()}
                    sel = st.multiselect("삭제할 행사 선택", list(opts.keys()), key="pm_del_sel")
                    if sel and st.button(f"선택한 {len(sel)}건 삭제", type="secondary", key="pm_del_btn"):
                        n = promo_delete([opts[s] for s in sel])
                        st.session_state["pm_flash"] = True
                        st.session_state["pm_flash_msg"] = f"행사 삭제 완료 ✅ {n:,}건"
                        st.rerun()

    # ── 메뉴2. 행사 진행 캘린더 (간트차트) — 기간 설정 + 필터/품번 직접 조회 ─────────
    with tab2:
        st.markdown("##### 2️⃣ 행사 진행 캘린더 (간트차트)")
        st.caption("원장에 등록된 행사를 기간 막대(간트차트)로 보여줘요 — 행 = 품번 · 채널, "
                   "막대 색 = 채널, 막대 안 숫자 = 실질판매가, 노란 세로선 = 오늘. "
                   "브랜드/년도/시즌/아이템 필터로 조회하거나, 품번 하나를 직접 입력해 조회할 수 있어요.")
        if ledger.empty:
            st.info("아직 등록된 행사가 없어요 — '1️⃣ 외부몰 행사 확정'에서 행사를 먼저 등록해 주세요.")
        else:
            deco = _pm_decorate(ledger)
            c_d1, c_d2 = st.columns(2)
            d_from = c_d1.date_input("조회 시작일", value=today - timedelta(days=7), key="pm_cal_from")
            d_to = c_d2.date_input("조회 종료일", value=today + timedelta(days=45), key="pm_cal_to")
            if d_to < d_from:
                st.warning("조회 종료일이 시작일보다 빨라요 — 기간을 다시 선택해 주세요.")
            else:
                mode = st.radio("조회 방식", ["필터로 조회 (브랜드/년도/시즌/아이템)", "품번 직접 입력 (1개 상품)"],
                                horizontal=True, key="pm_cal_mode")
                view = deco[(deco["행사시작"] <= d_to.isoformat())
                            & (deco["행사종료"] >= d_from.isoformat())].copy()
                if mode.startswith("품번"):
                    pn_in = st.text_input("품번 입력 (정확히 일치, 대소문자 무관)", key="pm_cal_pn",
                                          placeholder="예: SDSVC09STR")
                    if pn_in.strip():
                        view = view[view["품번"].astype(str).str.strip().str.upper()
                                    == pn_in.strip().upper()]
                        if view.empty:
                            st.info(f"조회 기간 내 '{pn_in.strip().upper()}' 품번의 행사가 없어요.")
                    else:
                        view = view.iloc[0:0]
                        st.caption("👆 품번을 입력하면 해당 상품의 행사만 캘린더로 보여드려요.")
                else:
                    ff1, ff2, ff3, ff4 = st.columns(4)
                    sel_br = ff1.multiselect("브랜드", sorted(view["_브랜드"].unique().tolist()),
                                             key="pm_cal_br")
                    sel_yr = ff2.multiselect("년도", sorted(view["_연도"].unique().tolist()),
                                             key="pm_cal_yr")
                    sel_ss = ff3.multiselect("시즌", sorted(view["_시즌"].unique().tolist()),
                                             key="pm_cal_ss")
                    sel_it = ff4.multiselect("아이템", sorted(view["_아이템"].unique().tolist()),
                                             key="pm_cal_it")
                    if sel_br:
                        view = view[view["_브랜드"].isin(sel_br)]
                    if sel_yr:
                        view = view[view["_연도"].isin(sel_yr)]
                    if sel_ss:
                        view = view[view["_시즌"].isin(sel_ss)]
                    if sel_it:
                        view = view[view["_아이템"].isin(sel_it)]
                    st.caption("필터를 비워두면 전체가 보여요 (여러 개 선택 = OR 조건).")
                if not view.empty:
                    st.caption(f"조회 결과: **{len(view):,}건** "
                               f"(품번 {view['품번'].nunique():,}개 · 채널 {view['매장명'].nunique():,}개) · "
                               f"기간 {d_from.isoformat()} ~ {d_to.isoformat()}")
                    st.plotly_chart(_pm_gantt(view, d_from, d_to, today_iso),
                                    use_container_width=True)
                elif not mode.startswith("품번"):
                    st.info("조회 조건에 맞는 행사가 없어요.")

    # ── 메뉴3. 외부몰 행사 최저가 체크 — 기획 폼 업로드 → OK/NO 판정 엑셀 ────────────
    with tab3:
        st.markdown("##### 3️⃣ 외부몰 행사 최저가 체크")
        st.caption("행사를 **기획하는 단계**에서 '외부몰 행사 최저가 체크' 폼의 회색 영역(앞 10컬럼)을 "
                   "채워 올리면, 원장에 등록된 [같은 품번 × 기간이 겹치는] 다른 채널 행사와 비교해서 "
                   "**최저가 여부(OK/NO)** 와 겹치는 행사 정보를 채운 엑셀을 만들어 드려요. "
                   "다른 채널 실질판매가가 내 기획가보다 낮거나 **같아도 NO**예요(단독 최저가 기준). "
                   "겹치는 행사가 여러 개면 실질판매가 낮은 순으로 오른쪽에 계속 붙어요 — "
                   "NO면 첫 블록이 나를 이긴 행사, OK면 참고용 차순위 행사예요.")
        if ledger.empty:
            st.warning("원장에 등록된 행사가 아직 없어요 — 비교 대상이 없어 모든 기획이 OK로 나와요.")
        up2 = st.file_uploader("'외부몰 행사 최저가 체크' 폼 업로드 (회색 영역 작성)", type=["xlsx"],
                               accept_multiple_files=False, key="pm_check_up")
        if up2 is not None:
            rows2, errs2 = _pm_read_form(up2, "check")
            for msg in errs2:
                st.warning("⚠️ " + msg)
            if rows2:
                results = _pm_check_rows(rows2, ledger)
                n_ok = sum(1 for r in results if r["ok"])
                n_no = len(results) - n_ok
                (st.success if n_no == 0 else st.error)(
                    f"판정 완료 — 기획 {len(results)}건 중 ✅ OK {n_ok}건 · ❌ NO {n_no}건"
                    + ("" if n_no == 0 else " (더 낮거나 같은 가격의 행사가 이미 잡혀 있어요)"))
                st.dataframe(pd.DataFrame([{
                    "최저가 여부": "OK" if r["ok"] else "NO",
                    "품번": r["plan"]["품번"],
                    "기획 채널": r["plan"]["매장명"],
                    "기획 기간": f'{r["plan"]["행사시작"]} ~ {r["plan"]["행사종료"]}',
                    "기획 실질판매가": _pm_fmt_won(r["plan"]["실질판매가"]),
                    "최저 경쟁 채널": ("" if r["comp"].empty else str(r["comp"].iloc[0]["매장명"])),
                    "담당자": ("" if r["comp"].empty else _mgr(r["comp"].iloc[0]["매장코드"])),
                    "경쟁 실질판매가": _pm_fmt_won(r["lowest"]),
                    "겹치는 행사 수": len(r["comp"]),
                } for r in results]), use_container_width=True, hide_index=True)
                st.download_button(
                    "⬇ 최저가 체크 결과 엑셀 다운로드",
                    _pm_check_excel(results),
                    file_name=f"외부몰행사_최저가체크_{now_kst().strftime('%y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True, key="pm_check_dl")
            elif not errs2:
                st.info("폼에서 읽을 데이터 행이 없어요 — 회색 영역을 채워서 올려주세요.")

    # ── 메뉴4. 네이버 최저가 체크 — 네이버 폼 업로드 → 인상필요 품번만 엑셀 ─────────────
    with tab4:
        st.markdown("##### 4️⃣ 네이버 최저가 체크 (금일최저가맞추기)")
        st.caption(f"네이버 브랜드 스토어 담당자가 매일 '네이버 최저가 체크' 폼의 회색 영역(매장코드·"
                   f"매장명·품번·최초가·네이버가격·할인율)을 채워 올리면, **오늘({today_iso}) 진행 중인** "
                   "외부몰 행사와 비교해서 네이버 가격이 외부몰 최저가보다 **낮거나 같은 품번만** 뽑아 "
                   "드려요. 외부몰이 최저가 행사를 진행하는 동안에는 네이버 가격이 더 높아야 하니까요. "
                   "'가격인하 필요' 칸에는 **외부몰최저가 + 100원**으로 맞춘 조정 목표가가 들어가요. "
                   "문제 없는 품번은 결과에 나오지 않아요.")
        if ledger.empty:
            st.warning("원장에 등록된 행사가 아직 없어요 — 비교 대상이 없어 위반이 나올 수 없어요.")
        up3 = st.file_uploader("'네이버 최저가 체크' 폼 업로드 (회색 영역 작성)", type=["xlsx"],
                               accept_multiple_files=False, key="pm_naver_up")
        if up3 is not None:
            rows3, errs3 = _pm_read_form(up3, "naver")
            for msg in errs3:
                st.warning("⚠️ " + msg)
            if rows3:
                viols = _pm_naver_rows(rows3, ledger, today_iso)
                if not viols:
                    st.success(f"👍 업로드한 {len(rows3)}개 품번 모두 오늘 기준 최저가 위반이 없어요.")
                else:
                    st.error(f"업로드 {len(rows3)}개 품번 중 ⚠️ **인상필요 {len(viols)}건** — "
                             "네이버 가격이 외부몰 행사 최저가보다 낮거나 같아요.")
                    st.dataframe(pd.DataFrame([{
                        "품번": v["row"]["품번"],
                        "네이버가격": _pm_fmt_won(v["row"]["네이버가격"]),
                        "외부몰 최저가": _pm_fmt_won(v["lowest"]),
                        "최저가 채널": str(v["comp"].iloc[0]["매장명"]),
                        "담당자": _mgr(v["comp"].iloc[0]["매장코드"]),
                        "가격인하 필요(목표가)": _pm_fmt_won(v["target"]),
                        "동시점 행사 수": len(v["comp"]),
                    } for v in viols]), use_container_width=True, hide_index=True)
                    st.download_button(
                        "⬇ 인상필요 품번 엑셀 다운로드",
                        _pm_naver_excel(viols, today_iso),
                        file_name=f"네이버_최저가체크_{now_kst().strftime('%y%m%d_%H%M')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True, key="pm_naver_dl")
            elif not errs3:
                st.info("폼에서 읽을 데이터 행이 없어요 — 회색 영역을 채워서 올려주세요.")


def render_user_admin():
    """관리자용 사용자 관리: 목록/활성토글/삭제 + 추가·비번 재설정."""
    me = st.session_state.get("auth_user")
    users = list_users()
    n_admin_active = sum(1 for u in users
                         if u["role"] == "admin" and str(u["active"]).upper() != "N")
    for us in users:
        act = str(us["active"]).upper() != "N"
        is_adm = us["role"] == "admin"
        c = st.columns([2.5, 1.2, 1.1, 0.9])
        c[0].caption(f"{'🟢' if act else '⚪'} **{us['display_name']}** ({us['username']}) · "
                     f"{'관리자' if is_adm else '뷰어'}")
        # 역할 전환 버튼 (2026-07-31 추가): 목록에서 바로 관리자↔뷰어 변경
        if c[1].button("뷰어로" if is_adm else "관리자로", key=f"role_{us['username']}", use_container_width=True):
            if is_adm and us["username"] == me:
                st.warning("본인 계정의 역할은 스스로 낮출 수 없어요. (관리자 잠금 방지)")
            elif is_adm and n_admin_active <= 1:
                st.warning("마지막 관리자는 뷰어로 바꿀 수 없어요.")
            else:
                upsert_user(us["username"], us["display_name"], None,
                            "viewer" if is_adm else "admin", us["active"])
                st.rerun()
        if c[2].button("비활성" if act else "활성화", key=f"tgl_{us['username']}", use_container_width=True):
            # 잠금 사고 방지: 본인 계정·마지막 관리자 계정은 비활성 불가 (2026-07-31 잠금 사고 재발 방지)
            if act and us["username"] == me:
                st.warning("본인 계정은 비활성화할 수 없어요. (잠금 사고 방지)")
            elif act and is_adm and n_admin_active <= 1:
                st.warning("마지막 관리자 계정은 비활성화할 수 없어요.")
            else:
                upsert_user(us["username"], us["display_name"], None, us["role"], "N" if act else "Y")
                st.rerun()
        if c[3].button("삭제", key=f"del_{us['username']}", use_container_width=True):
            if us["username"] == me:
                st.warning("본인 계정은 삭제할 수 없어요.")
            else:
                delete_user(us["username"])
                st.rerun()
    st.caption("— 계정 추가 / 비밀번호 재설정 (기존 ID면 갱신) —")
    with st.form("add_user", clear_on_submit=True):
        nu = st.text_input("ID")
        nn = st.text_input("이름")
        npw = st.text_input("비밀번호", type="password")
        nr = st.selectbox("역할", ["viewer", "admin"],
                          format_func=lambda x: "관리자" if x == "admin" else "뷰어")
        ok = st.form_submit_button("저장", use_container_width=True)
    if ok:
        if not nu.strip():
            st.error("ID를 입력하세요.")
        elif get_user(nu.strip()) is None and not npw:
            st.error("새 계정은 비밀번호가 필요해요.")
        elif nu.strip() == me and nr != "admin":
            st.error("본인 계정의 역할은 뷰어로 바꿀 수 없어요. (관리자 잠금 방지)")
        else:
            # 기존 계정 갱신 시 이름을 비워 두면 기존 표시 이름 유지 (ID로 덮어쓰지 않음)
            prev = get_user(nu.strip())
            keep_name = nn.strip() or (prev["display_name"] if prev else nu.strip())
            upsert_user(nu.strip(), keep_name, npw or None, nr)
            st.success(f"저장 완료 ✅ {nu.strip()}")
            st.rerun()


def main():
    st.set_page_config(page_title="온라인팀 미니 ERP", page_icon="📊", layout="wide")
    # 전역 여백 축소 + Apple 스타일 테마(2026-08-05) — 배경/폰트/버튼/사이드바 메뉴
    st.markdown(_APPLE_CSS, unsafe_allow_html=True)
    _copy_shortcut_guard()   # 260818: 표에서 Ctrl+C로 복사할 때 'Clear caches'가 뜨는 것 차단
    # ── 로그인 게이트 ──────────────────────────────────────────────
    ensure_users_table()
    if not st.session_state.get("auth_user"):
        # 1) 쿠키(로그인 유지 토큰)로 자동 로그인 — 마지막 사용 IDLE_LIMIT_HOURS 이내면 유지
        tok = _cookie_token()
        uname = _session_user(tok) if tok else None
        rec = get_user(uname) if uname else None
        if rec and str(rec["active"]).upper() != "N":
            st.session_state["auth_user"] = rec["username"]
            st.session_state["auth_name"] = rec["display_name"] or rec["username"]
            st.session_state["auth_role"] = rec["role"]
            st.session_state["auth_token"] = tok
        else:
            # 2) 유효한 토큰이 없으면 로그인 화면
            _render_login()
            return
    touch_session()   # 사용 중엔 매 동작마다 IDLE_LIMIT_HOURS 카운트 리셋
    is_admin = st.session_state.get("auth_role") == "admin"

    # 타이틀 (2026-08-05 확정): 애플 스타일 2톤 — 검정 볼드 + 회색 서브카피
    st.markdown(
        "<div style='margin:2px 0 6px;'>"
        "<div style='font-size:2.3rem;font-weight:700;letter-spacing:-0.035em;"
        "color:#1d1d1f;line-height:1.15;'>온라인팀 ERP</div>"
        "<div style='font-size:2.3rem;font-weight:600;letter-spacing:-0.03em;"
        "color:#6e6e73;line-height:1.3;margin-top:2px;'>"
        "Data to Insight, Insight to Action !</div></div>",
        unsafe_allow_html=True)
    fresh_slot = st.container()   # 타이틀 바로 아래: 매출 데이터 최종 업데이트 일자 표기 자리

    with st.sidebar:
        # 로그인 유지 쿠키를 매 실행 기록/갱신 (수명 30일 — 실제 만료는 서버가
        # '마지막 사용 IDLE_LIMIT_HOURS'로 판정하므로 쿠키 자체 수명은 넉넉히 둠)
        if st.session_state.get("auth_token"):
            _write_cookie(st.session_state["auth_token"], 30 * 24 * 3600)
        st.caption(f"👋 **{st.session_state.get('auth_name','')}**님 "
                   f"· {'관리자' if is_admin else '뷰어'}")
        if st.button("🚪 로그아웃", use_container_width=True):
            drop_session()   # DB 세션 삭제 — 쿠키가 브라우저에 남아 있어도 즉시 무효
            for _k in ("auth_user", "auth_name", "auth_role", "auth_token"):
                st.session_state.pop(_k, None)
            _write_cookie("", 0)   # 쿠키 삭제 시도
            st.rerun()
        st.metric("현재 DB 누적", f"{db_row_count():,} 건")
        if st.button("🔄 새로고침(캐시 비우기)", use_container_width=True):
            load_db.clear(); load_master.clear(); load_plan.clear(); load_priority.clear()
            load_size_master.clear(); load_item_master.clear(); load_weather.clear()
            get_itemgroup_map.clear(); get_itemgroup_map_small.clear(); _trend_cat_maps.clear()
            st.rerun()

        # ── 조회 메뉴 (탭 대체) ──────────────────────────────────────
        st.divider()
        st.caption("📂 **조회 메뉴**")
        menu = render_nav_menu()   # 260820: 3개 카테고리(Analysis/노가다 금지/궁금한 것) 그룹 radio

        if not is_admin:
            st.divider()
            st.caption("🔒 조회 전용(뷰어) 계정이에요. 데이터 업로드·적재·삭제는 관리자만 할 수 있어요.")

        if is_admin:
            st.divider()
            st.header("⚙️ 데이터 관리")
            st.caption(f"저장소: **{backend_name()}**")
            ups = st.file_uploader("① 로우데이터 업로드 (여러 개 한 번에 가능)",
                                   type=["xlsx", "xls", "csv"], accept_multiple_files=True)
            if ups:
                st.caption(f"{len(ups)}개 파일 선택됨")
                overwrite = st.checkbox(
                    "♻️ 덮어쓰기 모드 (파일에 있는 날짜는 먼저 삭제 후 적재)",
                    help="당일 매출처럼 ERP 값이 바뀌는 경우 켜세요. 업로드한 파일에 포함된 '날짜'의 "
                         "기존 데이터를 먼저 지우고 새로 넣습니다(수정·취소분까지 정확히 교체). "
                         "파일에 없는 다른 날짜는 그대로 둡니다. 끄면 기존 방식(중복 건너뛰고 추가만).")
                if st.button("② DB에 적재하기", type="primary", use_container_width=True):
                    tn = ts = dn = 0; last = db_row_count()
                    deleted_dates = set()
                    prog = st.progress(0.0)
                    status = st.empty()
                    for i, f in enumerate(ups):
                        try:
                            status.caption(f"⏳ ({i+1}/{len(ups)}) {f.name} 처리 중…")
                            clean = add_row_key(enrich(read_raw_file(f)))
                            if overwrite and "_판매일" in clean.columns:
                                fdates = sorted(clean["_판매일"].dropna().dt.strftime("%Y-%m-%d").unique())
                                todo = [d for d in fdates if d not in deleted_dates]
                                if todo:
                                    dn += delete_dates(todo)
                                    deleted_dates.update(todo)
                            res = append_to_db(clean)
                            tn += res["inserted"]; ts += res["skipped"]; last = res["total_after"]
                            del clean, res            # 파일별 메모리 즉시 해제 (OOM 방지)
                            gc.collect()
                        except Exception as ex:
                            st.error(f"{f.name} 오류: {ex}")
                            gc.collect()
                        prog.progress((i + 1) / len(ups))
                    status.empty()
                    load_db.clear()
                    if overwrite:
                        st.success(f"덮어쓰기 적재 완료 ✅ 삭제 {dn:,} · 신규 {tn:,} / 중복 {ts:,} · DB 총 {last:,}건")
                        if deleted_dates:
                            st.caption("교체된 날짜: " + ", ".join(sorted(deleted_dates)))
                    else:
                        st.success(f"적재 완료 ✅ 신규 {tn:,} / 중복 {ts:,} · DB 총 {last:,}건")

            st.divider()
            st.caption(f"🏬 매장 기준정보(태그): 현재 **{master_row_count():,}개** 매장")
            mup = st.file_uploader("매장 기준정보 업로드 (담당자·유통성격·채널소유·채널스토리)",
                                   type=["xlsx", "xls", "csv"], accept_multiple_files=False, key="master_up")
            if mup is not None:
                if st.button("🏬 매장 기준정보 적용(전체 교체)", use_container_width=True):
                    try:
                        n = replace_master(read_master_file(mup))
                        load_master.clear()
                        st.success(f"매장 기준정보 갱신 완료 ✅ {n}개 매장")
                    except Exception as ex:
                        st.error(f"매장 기준정보 오류: {ex}")

            st.divider()
            st.caption(f"🎯 사업계획(월별 목표): 현재 **{plan_row_count():,}행**")
            pup = st.file_uploader("사업계획 업로드 (매장별·브랜드별 월별 목표)",
                                   type=["xlsx", "xls"], accept_multiple_files=False, key="plan_up")
            if pup is not None:
                if st.button("🎯 사업계획 적용(전체 교체)", use_container_width=True):
                    try:
                        n = replace_plan(read_plan_file(pup))
                        load_plan.clear()
                        st.success(f"사업계획 갱신 완료 ✅ {n:,}행")
                    except Exception as ex:
                        st.error(f"사업계획 오류: {ex}")

            st.divider()
            st.caption(f"📌 온라인팀 우선순위(당월·금주): 현재 **{priority_row_count():,}건** "
                       "· 종합 대시보드 상단 표시 + 주간현황 분석 '⬇ 엑셀'의 우선순위 칸 자동 채움에 쓰여요")
            wup = st.file_uploader("온라인팀 우선순위 업로드 (주간 업무 보고 FORM · 당월·금주)",
                                   type=["xlsx"], accept_multiple_files=False, key="priority_up")
            if wup is not None:
                if st.button("📌 우선순위 적용(전체 교체)", use_container_width=True):
                    try:
                        _pdf = read_priority_file(wup)
                        n = replace_priority(_pdf)
                        load_priority.clear()
                        _cnt = _pdf.groupby("section")["content"].count().to_dict()
                        st.success(f"우선순위 갱신 완료 ✅ {n}건 "
                                   f"(당월 {_cnt.get('당월', 0)} · 금주 {_cnt.get('금주', 0)})")
                    except Exception as ex:
                        st.error(f"우선순위 오류: {ex}")

            st.divider()
            st.caption(f"📏 사이즈 마스터(품번→사이즈코드): 현재 **{size_master_row_count():,}개** 품번 "
                       "· 260811부터 재고 가공 판정에는 안 쓰고, 로우데이터 '사이즈구분' 컬럼값과의 "
                       "불일치를 참고 보고하는 용도로만 쓰여요(없어도 무방).")
            sup = st.file_uploader("사이즈 마스터 업로드 (참고용 · C열=품번, D열=사이즈코드)",
                                   type=["xlsx"], accept_multiple_files=False, key="sizemaster_up")
            if sup is not None:
                if st.button("📏 사이즈 마스터 적용(전체 교체)", use_container_width=True):
                    try:
                        n = replace_size_master(read_size_master_file(sup))
                        load_size_master.clear()
                        st.success(f"사이즈 마스터 갱신 완료 ✅ {n:,}개 품번")
                    except Exception as ex:
                        st.error(f"사이즈 마스터 오류: {ex}")

            st.divider()
            st.caption(f"🗂️ 아이템 마스터(아이템코드→대/중/소카테고리): 현재 **{item_master_row_count():,}개** 코드"
                       " · 재고 가공 중카테고리·판매분석 아이템그룹이 모두 여기서 나와요(단일 기준)")
            iup = st.file_uploader("아이템 마스터 업로드 ('아이템코드와 카테고리 구분' 시트 포함 워크북)",
                                   type=["xlsx"], accept_multiple_files=False, key="itemmaster_up")
            if iup is not None:
                if st.button("🗂️ 아이템 마스터 적용(전체 교체)", use_container_width=True):
                    try:
                        n = replace_item_master(read_item_master_file(iup))
                        load_item_master.clear()
                        get_itemgroup_map.clear()
                        st.success(f"아이템 마스터 갱신 완료 ✅ {n:,}개 코드 "
                                  "(재고 가공·판매분석에 즉시 반영돼요)")
                    except Exception as ex:
                        st.error(f"아이템 마스터 오류: {ex}")

            st.divider()
            render_weather_admin()   # 🌡️ 서울 기온 일자료 (추세분석 기온 겹쳐보기용)

            st.divider()
            with st.expander("👤 사용자 관리 (계정 추가·권한·비활성)"):
                render_user_admin()

    df = load_db()
    # 타이틀 아래 최종 업데이트 일자 (매출 로우데이터의 마지막 판매일자 = 데이터가 채워진 마지막 날)
    if not df.empty and "_판매일" in df.columns and df["_판매일"].notna().any():
        _last = df["_판매일"].max()
        fresh_slot.caption(
            f"🗓️ **매출 로우데이터 최종 업데이트 일자 : {_last.year}년 {_last.month:02d}월 {_last.day:02d}일**"
            "  (이 날짜까지의 매출이 입력되어 있어요)")
    if df.empty:
        st.info("👈 사이드바에서 매출 로우데이터를 업로드하고 [DB에 적재하기]를 눌러 시작하세요."
                "  (🏷️ 재고 가공은 매출 데이터 없이도 바로 쓸 수 있어요)")
        # 재고 가공·최저가 관리는 매출 DB와 무관하므로 매출 데이터가 없어도 사용 가능하게 유지
        if menu == MENU_PRICE:
            render_price_mgmt()
        else:
            render_inventory()
        return

    # 사이드바에서 고른 메뉴 '1개만' 실행 (탭 방식은 8개가 매번 전부 계산돼 느렸다)
    if menu == MENU_DASH:
        render_dashboard(df)
    elif menu == MENU_WEEK:
        render_weekly_report(df)
    elif menu == MENU_FLAG:
        render_flagship(df)
    elif menu == MENU_CHAN:
        render_channel_brand(df)
    elif menu == MENU_CATMIX:
        render_category_mix(df)
    elif menu == MENU_INV:
        render_inventory()
    elif menu == MENU_TRND:
        render_trend(df)
    elif menu == MENU_RTN:
        render_return_rate(df)
    elif menu == MENU_SET:
        render_suitset(df)
    elif menu == MENU_PRICE:
        render_price_mgmt()
    else:
        render_dashboard(df)


if __name__ == "__main__":
    main()
